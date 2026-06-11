"""
增值税申报表 API 路由
使用 FastAPI APIRouter，在 main.py 中 include_router 加载。
"""
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, Form, Request
from pydantic import BaseModel
from sqlalchemy.orm import Session
from sqlalchemy import func, or_, and_
from datetime import datetime
import calendar
import json
from typing import Optional

from database import (
    VATDeclaration, Company, SalesInvoice, PurchaseInvoice,
    InputVATDeduction, JournalEntry, SalaryRecord, get_db
)

# VAT计算使用的科目编码常量
VAT_ACCOUNT_CODES = {
    "revenue_main": "6001",       # 主营业务收入
    "revenue_other": "6051",      # 其他业务收入
    "cost_main": "6401",          # 主营业务成本
    "cost_other": "6402",         # 其他业务成本
    "tax_surcharge": "6403",      # 税金及附加
    "sales_expense": "6601",      # 销售费用
    "admin_expense": "6602",      # 管理费用
    "rd_expense": "6604",         # 研发费用
    "fin_expense": "6603",        # 财务费用
    "fin_interest": "660301",     # 财务费用-利息
    "invest_income": "6111",      # 投资收益
    "credit_loss": "6701",        # 信用减值损失
    "asset_loss": "6702",         # 资产减值损失
    "asset_disposal": "6712",     # 资产处置收益
    "non_op_income": "6301",      # 营业外收入
    "non_op_expense": "6711",     # 营业外支出
    "income_tax": "6801",         # 所得税费用
}

def _end_of_month(period: str) -> str:
    """返回期间的月末日期，如 '2025-02' -> '2025-02-28'"""
    y, m = period.split("-")
    last_day = calendar.monthrange(int(y), int(m))[1]
    return f"{y}-{m}-{last_day:02d}"

router = APIRouter(prefix="/api/vat", tags=["增值税申报"])



# ==================== 外部导入申报表 ====================

@router.post("/declarations/import")
async def import_vat_declaration(
    company_id: int = Form(...),
    file: UploadFile = File(...),
    declaration_id: Optional[str] = Form(None),
    db: Session = Depends(get_db)
):
    """
    外部导入增值税申报表（Excel文件）
    接收用户上传的已填好的申报表Excel文件，解析并保存到数据库。
    """
    import tempfile, os
    from openpyxl import load_workbook
    from datetime import datetime
    
    # 保存上传的文件到临时目录
    suffix = file.filename.rsplit('.', 1)[-1] if '.' in file.filename else 'xlsx'
    with tempfile.NamedTemporaryFile(suffix='.' + suffix, delete=False) as tmp:
        content = await file.read()
        tmp.write(content)
        tmp_path = tmp.name
    
    try:
        wb = load_workbook(tmp_path, data_only=True)
        
        # 获取或创建申报表记录
        decl = None
        if declaration_id:
            try:
                decl_id = int(declaration_id)
                decl = db.query(VATDeclaration).filter(
                    VATDeclaration.id == decl_id,
                    VATDeclaration.company_id == company_id
                ).first()
            except:
                pass
        
        if not decl:
            # 从Excel中读取期间信息（扫描主表前30行，找 "所属期" 相关字段）
            period = None
            import re
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(values_only=True))
                for i, row in enumerate(rows[:30]):
                    for j, cell in enumerate(row):
                        if cell and isinstance(cell, str) and ('所属期' in cell or '申报期' in cell):
                            # 先检查同行右侧的单元格
                            for k in range(j + 1, min(j + 4, len(row))):
                                val = row[k]
                                if val:
                                    val_str = str(val)
                                    m = re.search(r'(\d{4})[年\-](\d{1,2})', val_str)
                                    if m:
                                        period = f"{m.group(1)}-{m.group(2).zfill(2)}"
                                        break
                            if not period and i + 1 < len(rows):
                                # 检查下一行同列
                                next_row = rows[i + 1]
                                if j < len(next_row) and next_row[j]:
                                    val_str = str(next_row[j])
                                    m = re.search(r'(\d{4})[年\-](\d{1,2})', val_str)
                                    if m:
                                        period = f"{m.group(1)}-{m.group(2).zfill(2)}"
                        if period:
                            break
                    if period:
                        break
                if period:
                    break
            
            if not period:
                # 从文件名推断期间
                match = re.search(r'(\d{4})[年\-](\d{1,2})', file.filename)
                if match:
                    period = f"{match.group(1)}-{match.group(2).zfill(2)}"
                else:
                    from datetime import datetime
                    now = datetime.now()
                    period = f"{now.year}-{now.month:02d}"
            
            # 创建新申报表
            max_id = db.query(func.max(VATDeclaration.id)).filter(
                VATDeclaration.company_id == company_id
            ).scalar() or 0
            
            decl = VATDeclaration(
                id=max_id + 1,
                company_id=company_id,
                period=period,
                status="已导入",
                created_at=datetime.now(),
                updated_at=datetime.now()
            )
            db.add(decl)
            db.flush()
        
        # 解析Excel文件，提取申报表数据
        form_main = None
        form_sales = None
        form_input = None
        form_deduction = None
        form_credit = None
        form_surcharge = None
        form_reduction = None
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # 将工作表转换为2D数组
            data = []
            for row in ws.iter_rows(values_only=True):
                data.append(list(row))
            
            # 根据工作表名称判断类型（精确匹配优先，避免"附列资料一"也含"申报表"导致覆盖主表）
            sn = sheet_name
            if '附列资料（一）' in sn or '附列资料一' in sn or '销项' in sn:
                form_sales = data
            elif '附列资料（二）' in sn or '附列资料二' in sn or '进项' in sn:
                form_input = data
            elif '附列资料（三）' in sn or '附列资料三' in sn or '扣除' in sn or 'deduction' in sn.lower():
                form_deduction = data
            elif '附列资料（四）' in sn or '附列资料四' in sn or '减免' in sn or 'reduction' in sn.lower():
                form_reduction = data
            elif '附列资料（五）' in sn or '附列资料五' in sn or '附加税费申报表附列资料' in sn or ('附加' in sn and '申报表' not in sn) or 'surcharge' in sn.lower():
                form_surcharge = data
            elif '抵扣' in sn or 'credit' in sn.lower():
                form_credit = data
            elif '主表' in sn or '申报表' in sn:
                # 最后匹配主表，避免被附列资料名字覆盖
                form_main = data
        
        # 保存到数据库
        if form_main:
            decl.form_main = json.dumps(form_main, ensure_ascii=False, default=str)
        if form_sales:
            decl.form_sales = json.dumps(form_sales, ensure_ascii=False, default=str)
        if form_input:
            decl.form_input = json.dumps(form_input, ensure_ascii=False, default=str)
        if form_deduction:
            decl.form_deduction = json.dumps(form_deduction, ensure_ascii=False, default=str)
        if form_credit:
            decl.form_credit = json.dumps(form_credit, ensure_ascii=False, default=str)
        if form_surcharge:
            decl.form_surcharge = json.dumps(form_surcharge, ensure_ascii=False, default=str)
        if form_reduction:
            decl.form_reduction = json.dumps(form_reduction, ensure_ascii=False, default=str)
        
        decl.status = "已导入"
        decl.updated_at = datetime.now()
        
        db.commit()
        
        return {
            "success": True,
            "message": "申报表导入成功",
            "declaration_id": decl.id,
            "period": decl.period
        }
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=400, detail=f"文件解析失败: {str(e)}")
    finally:
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)


@router.get("/declarations")
def list_vat_declarations(company_id: int, period: str = Query(None), db: Session = Depends(get_db)):
    q = db.query(VATDeclaration).filter(VATDeclaration.company_id == company_id)
    if period:
        q = q.filter(VATDeclaration.period == period)
    items = q.order_by(VATDeclaration.period.desc()).all()
    return [{
        "id": v.id, "company_id": v.company_id, "period": v.period,
        "taxpayer_name": v.taxpayer_name, "status": v.status,
        "fill_date": str(v.fill_date) if v.fill_date else None,
        "submitted_at": v.submitted_at.isoformat() if v.submitted_at else None,
        "created_at": v.created_at.isoformat() if v.created_at else None,
        "micro_enterprise": v.micro_enterprise,
        "six_tax_reduction": v.six_tax_reduction,
    } for v in items]


# ===== 小型微利企业自动校验 =====
SMALL_MICRO_INCOME_LIMIT  = 3_000_000.00   # 年度应纳税所得额 ≤ 300万元
SMALL_MICRO_EMPLOYEE_LIMIT = 300            # 从业人数 ≤ 300人
SMALL_MICRO_ASSET_LIMIT    = 50_000_000.00  # 资产总额 ≤ 5000万元
# 法律依据：财政部 税务总局公告2023年第12号、公告2022年第10号
# 六税两费减征50%，政策有效期 2023-01-01 至 2027-12-31


def _validate_small_micro_enterprise(db: Session, company_id: int, period: str) -> dict:
    """
    自动校验小型微利企业三大指标，返回完整校验报告。
    应纳税所得额：利润表 1-12月 净利润
    从业人数：工资记录 1-12月 身份证去重
    资产总额：资产负债表 期末资产总计
    """
    year = int(period[:4])
    p_from = f"{year}-01"
    p_to   = f"{year}-12"

    # ---- 1. 汇算科目余额（整年 1-12月）----
    # 使用 SQL 聚合，避免全量加载到内存
    balances = {}
    from sqlalchemy import func
    for row in db.query(
        JournalEntry.account_code,
        func.sum(JournalEntry.debit_amount).label('total_debit'),
        func.sum(JournalEntry.credit_amount).label('total_credit')
    ).filter(
        JournalEntry.company_id == company_id,
        JournalEntry.period >= p_from,
        JournalEntry.period <= p_to,
    ).group_by(JournalEntry.account_code).all():
        balances[row.account_code] = {
            "debit": float(row.total_debit or 0),
            "credit": float(row.total_credit or 0)
        }

    def _pl(code_prefix, is_credit=True):
        dr = sum(v["debit"]  for k, v in balances.items() if k.startswith(code_prefix))
        cr = sum(v["credit"] for k, v in balances.items() if k.startswith(code_prefix))
        return round(cr - dr, 2) if is_credit else round(dr - cr, 2)

    def _bs(code_prefix, is_debit=True):
        dr = sum(v["debit"]  for k, v in balances.items() if k.startswith(code_prefix))
        cr = sum(v["credit"] for k, v in balances.items() if k.startswith(code_prefix))
        return round(dr - cr, 2) if is_debit else round(cr - dr, 2)

    # ---- 1. 利润总额 & 净利润（≈ 应纳税所得额基础）----
    AC = VAT_ACCOUNT_CODES
    revenue     = _pl(AC["revenue_main"]) + _pl(AC["revenue_other"])
    cost        = _pl(AC["cost_main"], False) + _pl(AC["cost_other"], False)
    tax_sur     = _pl(AC["tax_surcharge"], False)
    period_exp  = _pl(AC["sales_expense"], False) + _pl(AC["admin_expense"], False) + _pl(AC["rd_expense"], False) + _pl(AC["fin_expense"], False)
    invest      = _pl(AC["invest_income"])
    interest_in = _pl(AC["fin_interest"])
    credit_l    = _pl(AC["credit_loss"], False)
    asset_l     = _pl(AC["asset_loss"], False)
    disposal    = _pl(AC["asset_disposal"])
    non_op_in   = _pl(AC["non_op_income"])
    non_op_out  = _pl(AC["non_op_expense"], False)
    income_tax  = _pl(AC["income_tax"], False)

    gross_profit       = round(revenue - cost - tax_sur, 2)
    operating_profit   = round(gross_profit - period_exp + invest + interest_in - credit_l - asset_l + disposal, 2)
    profit_total       = round(operating_profit + non_op_in - non_op_out, 2)
    net_profit         = round(profit_total - income_tax, 2)

    # ---- 2. 从业人数（年度有工资记录的去重人数）----
    employee_count = db.query(func.count(func.distinct(SalaryRecord.id_number))).filter(
        SalaryRecord.company_id == company_id,
        SalaryRecord.period >= p_from,
        SalaryRecord.period <= p_to,
        SalaryRecord.id_number.isnot(None),
        SalaryRecord.id_number != "",
    ).scalar() or 0

    # ---- 3. 资产总额 = 流动资产合计 + 非流动资产合计 ----
    cur = (
        _bs("1001") + _bs("1002") + _bs("1003") +   # 货币资金
        _bs("1101") + _bs("1121") + _bs("1122") +   # 交易性金融资产 + 应收票据 + 应收账款
        _bs("1124") + _bs("1123") + _bs("1221") +   # 应收款项融资 + 预付款项 + 其他应收款
        _bs("1403") + _bs("1405") + _bs("1406") + _bs("1408") + _bs("1411") +  # 存货
        _bs("1401") + _bs("1501") + _bs("1502") + _bs("1503")  # 合同资产 + 持有待售 + 一年内到期 + 其他流动
    )
    nc = (
        _bs("1504") + _bs("1505") + _bs("1511") + _bs("1512") +
        _bs("1513") + _bs("1514") + _bs("1521") +
        _bs("1601") + _bs("1602", False) + _bs("1604") +
        _bs("1621") + _bs("1631") + _bs("1641") +
        _bs("1701") + _bs("1702") + _bs("1711") +
        _bs("1801") + _bs("1811") + _bs("1901")
    )
    total_assets = round(cur + nc, 2)

    # ---- 判定 ----
    income_ok   = net_profit <= SMALL_MICRO_INCOME_LIMIT if net_profit > 0 else True
    employee_ok = employee_count <= SMALL_MICRO_EMPLOYEE_LIMIT
    asset_ok    = total_assets <= SMALL_MICRO_ASSET_LIMIT
    all_ok      = income_ok and employee_ok and asset_ok

    warnings = []
    if not income_ok:
        warnings.append(f"年度净利润 ¥{net_profit:,.2f} 超过 ¥{SMALL_MICRO_INCOME_LIMIT/10000:.0f}万标准")
    if not employee_ok:
        warnings.append(f"从业人数 {employee_count} 人 超过 {SMALL_MICRO_EMPLOYEE_LIMIT} 人标准")
    if not asset_ok:
        warnings.append(f"资产总额 ¥{total_assets:,.2f} 超过 ¥{SMALL_MICRO_ASSET_LIMIT/10000:.0f}万标准")
    if not all_ok and warnings:
        warnings.append("⚠️ 不符合小型微利企业认定标准，如勾选六税两费减征可能产生税务风险")

    return {
        "profit_total":     profit_total,
        "net_profit":       net_profit,
        "employee_count":   employee_count,
        "total_assets":     total_assets,
        "income_ok":        income_ok,
        "employee_ok":      employee_ok,
        "asset_ok":         asset_ok,
        "all_ok":           all_ok,
        "warnings":         warnings,
        "standards": {
            "income":   SMALL_MICRO_INCOME_LIMIT,
            "employees": SMALL_MICRO_EMPLOYEE_LIMIT,
            "assets":   SMALL_MICRO_ASSET_LIMIT,
        },
    }


@router.get("/check-micro-enterprise")
def check_micro_enterprise(company_id: int = Query(), period: str = Query(), db: Session = Depends(get_db)):
    """预校验：创建申报表前先查看是否符合小型微利企业标准"""
    return _validate_small_micro_enterprise(db, company_id, period)


class VATDeclarationCreate(BaseModel):
    period: str = ""
    taxpayer_name: str = ""
    industry: str = ""
    register_type: str = ""
    bank_account: str = ""
    phone: str = ""
    micro_enterprise: bool = False
    six_tax_reduction: bool = False


@router.post("/declarations")
def create_vat_declaration(data: VATDeclarationCreate, company_id: int = Query(), db: Session = Depends(get_db)):
    period = data.period
    if not period:
        raise HTTPException(400, detail="税款所属期不能为空")
    company = db.query(Company).filter(Company.id == company_id).first()
    if not company:
        raise HTTPException(404, detail="公司不存在")

    micro = data.micro_enterprise
    six_tax = data.six_tax_reduction

    # ===== 自动校验小型微利企业三大指标 =====
    validation = _validate_small_micro_enterprise(db, company_id, period)

    legal_warnings = []
    if micro and six_tax:
        period_year = int(period[:4])
        if period_year > 2027:
            legal_warnings.append(f"⚠️ 六税两费减征政策有效期至2027-12-31，当前期间{period}已超出政策有效期，减征可能不适用")
        if not validation["all_ok"]:
            legal_warnings.extend(validation["warnings"])

    reduction_start = None
    reduction_end = None
    if micro and six_tax:
        reduction_start = period + "-01"
        reduction_end = _end_of_month(period)

    vd = VATDeclaration(
        company_id=company_id, period=period,
        taxpayer_name=data.taxpayer_name or company.name,
        taxpayer_id=company.uscc or "",
        industry=data.industry,
        register_type=data.register_type,
        legal_representative=company.legal_representative or "",
        address=company.address or "",
        bank_account=data.bank_account,
        phone=data.phone,
        micro_enterprise=micro,
        six_tax_reduction=six_tax,
        reduction_start=reduction_start,
        reduction_end=reduction_end,
    )
    db.add(vd)
    db.flush()
    _compute_vat_forms(db, vd)
    db.commit()
    db.refresh(vd)
    result = {
        "id": vd.id, "period": vd.period, "status": vd.status,
        "msg": "申报表已创建并计算完成",
        "validation": validation,
    }
    if legal_warnings:
        result["legal_warnings"] = legal_warnings
    return result


@router.get("/declarations/{declaration_id}")
def get_vat_declaration(declaration_id: int, company_id: int = Query(), db: Session = Depends(get_db)):
    vd = db.query(VATDeclaration).filter(VATDeclaration.id == declaration_id, VATDeclaration.company_id == company_id).first()
    if not vd:
        raise HTTPException(404, detail="申报表不存在")
    return {
        "id": vd.id, "company_id": vd.company_id, "period": vd.period,
        "taxpayer_name": vd.taxpayer_name, "taxpayer_id": vd.taxpayer_id,
        "industry": vd.industry, "register_type": vd.register_type,
        "legal_representative": vd.legal_representative, "address": vd.address,
        "bank_account": vd.bank_account, "phone": vd.phone,
        "fill_date": str(vd.fill_date) if vd.fill_date else None,
        "micro_enterprise": vd.micro_enterprise,
        "six_tax_reduction": vd.six_tax_reduction,
        "reduction_start": vd.reduction_start, "reduction_end": vd.reduction_end,
        "city_maintenance_tax": vd.city_maintenance_tax,
        "education_surcharge": vd.education_surcharge,
        "local_education_surcharge": vd.local_education_surcharge,
        "status": vd.status,
        "submitted_at": vd.submitted_at.isoformat() if vd.submitted_at else None,
        "form_main": json.loads(vd.form_main) if vd.form_main else {},
        "form_sales": json.loads(vd.form_sales) if vd.form_sales else {},
        "form_input": json.loads(vd.form_input) if vd.form_input else {},
        "form_deduction": json.loads(vd.form_deduction) if vd.form_deduction else {},
        "form_credit": json.loads(vd.form_credit) if vd.form_credit else {},
        "form_surcharge": json.loads(vd.form_surcharge) if vd.form_surcharge else {},
        "form_reduction": json.loads(vd.form_reduction) if vd.form_reduction else {},
        "created_at": vd.created_at.isoformat() if vd.created_at else None,
        "updated_at": vd.updated_at.isoformat() if vd.updated_at else None,
    }


@router.put("/declarations/{declaration_id}")
def update_vat_declaration(declaration_id: int, data: dict, company_id: int = Query(), db: Session = Depends(get_db)):
    vd = db.query(VATDeclaration).filter(VATDeclaration.id == declaration_id, VATDeclaration.company_id == company_id).first()
    if not vd:
        raise HTTPException(404, detail="申报表不存在")
    for key in ["form_main", "form_sales", "form_input", "form_deduction", "form_credit", "form_surcharge", "form_reduction"]:
        if key in data:
            setattr(vd, key, json.dumps(data[key], ensure_ascii=False))
    for field in ["taxpayer_name", "taxpayer_id", "industry", "register_type",
                   "legal_representative", "address", "bank_account", "phone",
                   "fill_date", "micro_enterprise", "six_tax_reduction",
                   "reduction_start", "reduction_end"]:
        if field in data:
            setattr(vd, field, data[field])
    if data.get("status") in ("草稿", "已申报", "已缴税"):
        vd.status = data["status"]
        if vd.status == "已申报" and not vd.submitted_at:
            vd.submitted_at = datetime.now()
    db.commit()
    return {"msg": "保存成功"}


@router.delete("/declarations/{declaration_id}")
def delete_vat_declaration(declaration_id: int, company_id: int = Query(), db: Session = Depends(get_db)):
    vd = db.query(VATDeclaration).filter(VATDeclaration.id == declaration_id, VATDeclaration.company_id == company_id).first()
    if not vd:
        raise HTTPException(404, detail="申报表不存在")
    db.delete(vd)
    db.commit()
    return {"msg": "删除成功"}


@router.post("/declarations/{declaration_id}/recompute")
def recompute_vat_declaration(declaration_id: int, company_id: int = Query(), db: Session = Depends(get_db)):
    vd = db.query(VATDeclaration).filter(VATDeclaration.id == declaration_id, VATDeclaration.company_id == company_id).first()
    if not vd:
        raise HTTPException(404, detail="申报表不存在")
    _compute_vat_forms(db, vd)
    db.commit()
    return {"msg": "重新计算完成"}

# ========== 计算逻辑 ==========

def _compute_vat_forms(db: Session, vd: VATDeclaration):
    """从序时账自动计算7张申报表的数据"""
    company_id = vd.company_id
    period = vd.period

    entries = db.query(JournalEntry).filter(
        JournalEntry.company_id == company_id,
        JournalEntry.period == period
    ).all()

    # ===== 销项税额 =====
    # 优先从序时账取数（贷方 2221 应交增值税-销项税额）
    output_tax = float(sum(
        e.credit_amount or 0
        for e in entries
        if e.account_code and "2221" in e.account_code and "销项" in (e.account_name or "")
    ))

    # 销项发票：按税收分类编码前两位自动分类，填入主表各栏
    # 税收分类编码规则：10=货物，11=矿产品，12=电力，20=加工修理修配劳务，
    # 30=服务，40=不动产，50=无形资产
    sales_invoices = db.query(SalesInvoice).filter(
        SalesInvoice.company_id == company_id,
        SalesInvoice.invoice_date >= period + "-01",
        SalesInvoice.invoice_date <= _end_of_month(period),
        SalesInvoice.status == "正常",
        SalesInvoice.is_positive == True,
    ).all()

    def _inv_category(inv):
        """根据税收分类编码前两位判断类别：goods/service/realestate/intangible"""
        code = (inv.tax_category_code or "").strip()
        prefix = code[:2]
        if prefix in ("10", "11", "12", "13", "14", "15"):
            return "goods"
        if prefix == "20":
            return "labor"
        if prefix in ("30",):
            return "service"
        if prefix in ("40",):
            return "realestate"
        if prefix in ("50",):
            return "intangible"
        # fallback: 根据goods_name关键词判断
        name = (inv.goods_name or "").lower()
        if any(k in name for k in ["货物", "产品", "材料", "设备"]):
            return "goods"
        if any(k in name for k in ["服务", "咨询", "软件", "技术", "租赁"]):
            return "service"
        return "goods"  # 默认货物

    def _is_exempt(inv):
        """判断是否免税：税率为0且有免税标志，或invoice_category含'免税'"""
        rt = inv.tax_rate
        return (rt is not None and rt == 0) or "免税" in (inv.invoice_category or "")

    def _is_simple(inv):
        """判断是否简易计税：征收率3%、5%等（税率存储为百分比，如3=3%）"""
        rt = inv.tax_rate
        return rt is not None and rt in (3, 5)

    # 按类别+税率分类汇总
    category_sales = {}   # (category, tax_rate) -> amount
    category_tax = {}     # (category, tax_rate) -> tax_amount
    exempt_sales = 0.0    # 免税销售额
    simple_sales = 0.0    # 简易计税销售额
    simple_tax = 0.0      # 简易计税税额
    goods_sales = 0.0      # 应税货物销售额（第2栏）
    labor_sales = 0.0      # 应税劳务销售额（第3栏）
    service_sales = 0.0    # 应税服务销售额
    goods_tax = 0.0
    labor_tax = 0.0
    service_tax = 0.0

    for inv in sales_invoices:
        if _is_exempt(inv):
            exempt_sales += float(inv.amount or 0)
            continue
        if _is_simple(inv):
            simple_sales += float(inv.amount or 0)
            simple_tax += float(inv.tax_amount or 0)
            continue
        cat = _inv_category(inv)
        rate = inv.tax_rate or 0
        key = (cat, rate)
        category_sales[key] = category_sales.get(key, 0) + float(inv.amount or 0)
        category_tax[key] = category_tax.get(key, 0) + float(inv.tax_amount or 0)
        if cat == "goods":
            goods_sales += float(inv.amount or 0)
            goods_tax += float(inv.tax_amount or 0)
        elif cat == "labor":
            labor_sales += float(inv.amount or 0)
            labor_tax += float(inv.tax_amount or 0)
        elif cat in ("service", "realestate", "intangible"):
            service_sales += float(inv.amount or 0)
            service_tax += float(inv.tax_amount or 0)

    # 销售额 = 不含税金额（amount），不是价税合计（total_amount）
    sales_total = sum(float(i.amount or 0) for i in sales_invoices)
    sales_total_inclusive = sum(float(i.total_amount or 0) for i in sales_invoices)
    # 销项税额：优先序时账，兜底发票
    if output_tax == 0:
        output_tax = sum(float(i.tax_amount or 0) for i in sales_invoices)

    # ===== 进项税额 =====
    # 优先从序时账取数（借方 2221 应交增值税-进项税额）
    input_tax_journal = float(sum(
        e.debit_amount or 0
        for e in entries
        if e.account_code and "2221" in e.account_code and "进项" in (e.account_name or "")
    ))
    # 进项抵扣表：按抵扣所属期取数，仅正常发票
    # 抵扣所属期为空时，fallback 到开票日期（invoice_date）的年月
    input_deductions = db.query(InputVATDeduction).filter(
        InputVATDeduction.company_id == company_id,
        or_(
            InputVATDeduction.deduction_period == period,
            and_(InputVATDeduction.deduction_period == None,
                 func.strftime('%Y-%m', InputVATDeduction.invoice_date) == period),
            and_(InputVATDeduction.deduction_period == "",
                 func.strftime('%Y-%m', InputVATDeduction.invoice_date) == period),
        ),
        InputVATDeduction.invoice_status == "正常",
    ).all()
    input_tax = sum(float(d.deductible_tax_amount or 0) for d in input_deductions)
    if input_tax_journal > 0:
        input_tax = input_tax_journal

    tax_payable = max(0, output_tax - input_tax)

    # 附加税费
    city_rate = 0.07
    edu_rate = 0.03
    local_edu_rate = 0.02
    city_tax = round(tax_payable * city_rate, 2)
    edu_surcharge = round(tax_payable * edu_rate, 2)
    local_edu = round(tax_payable * local_edu_rate, 2)

    if vd.micro_enterprise and vd.six_tax_reduction:
        city_tax = round(city_tax * 0.5, 2)
        edu_surcharge = round(edu_surcharge * 0.5, 2)
        local_edu = round(local_edu * 0.5, 2)

    vd.city_maintenance_tax = city_tax
    vd.education_surcharge = edu_surcharge
    vd.local_education_surcharge = local_edu

    # 解析当前期间（后续多处使用）
    period_date = datetime.strptime(period + "-01", "%Y-%m-%d")

    # ====== 主表 —— 本年累计(YTD)计算 ======
    # 取本年1月至当前月所有申报表，汇总各栏次数值
    year_str = period[:4]
    ytd_periods = [f"{year_str}-{m:02d}" for m in range(1, period_date.month + 1)]

    # WARNING: 查询本年所有已保存申报表（不含当前正在计算的本条，因为 form_main 尚未入库）
    ytd_declarations = db.query(VATDeclaration).filter(
        VATDeclaration.company_id == company_id,
        VATDeclaration.period.in_(ytd_periods),
        VATDeclaration.id != vd.id,
    ).all()

    def _parse_main(decl):
        """安全解析申报表 form_main JSON"""
        if not decl or not decl.form_main:
            return {}
        if isinstance(decl.form_main, str):
            return json.loads(decl.form_main)
        return decl.form_main

    # 需要累计的字段列表（主表所有数字栏次）
    ytd_fields = [
        "row1_sales", "row2_other_invoice", "row3_no_invoice", "row4_tax_check",
        "row5_simple_method", "row6_exempt_sales", "row7_export_exempt",
        "row8_tax_free", "row9_exempt_goods", "row10_exempt_service",
        "row11_output_tax", "row12_input_tax", "row13_prior_credit",
        "row14_input_transfer_out", "row15_exempt_refund", "row16_actual_deduct_by_item",
        "row17_total_deductible", "row18_actual_deduct", "row19_tax_payable",
        "row20_end_credit", "row21_simple_tax", "row22_simple_tax_reduction",
        "row23_reduction", "row24_tax_payable_total",
        "row25_prior_unpaid", "row26_real_paid_during", "row27_installment_prepaid",
        "row28_export_tax_refund", "row29_remote_prepaid", "row30_already_paid_total",
        "row31_should_pay_refund", "row32_check_tax_should", "row33_check_prepaid",
        "row34_should_check", "row36_prior_unpaid_check", "row37_check_paid",
        "row38_end_check",
        "row39_city_maintenance_tax", "row40_education_surcharge", "row41_local_education_surcharge",
    ]
    # 累加历史各月数据
    ytd_sums = {f: 0.0 for f in ytd_fields}
    for d in ytd_declarations:
        m = _parse_main(d)
        for f in ytd_fields:
            ytd_sums[f] += m.get(f, 0.0)

    # ====== 主表（会企02号）—— 41行完整字段 ======
    # 取上期留抵：从同公司上期申报表取期末留抵
    # 上期 = 当前期间往前1个月
    prev_year = period_date.year
    prev_month = period_date.month - 1
    if prev_month == 0:
        prev_year -= 1
        prev_month = 12
    prev_period = f"{prev_year}-{prev_month:02d}"
    prior_vd = db.query(VATDeclaration).filter(
        VATDeclaration.company_id == company_id,
        VATDeclaration.period == prev_period
    ).first()
    prior_credit = 0.0
    if prior_vd and prior_vd.form_main:
        prior_main = json.loads(prior_vd.form_main) if isinstance(prior_vd.form_main, str) else prior_vd.form_main
        prior_credit = prior_main.get("row20_end_credit", 0.0)

    # ===== 主表栏次赋值（按官方填写说明）=====
    # 第1栏 按适用税率计税销售额 = 一般计税所有货物+服务销售额
    row1_sales = round(sales_total - exempt_sales - simple_sales, 2) if sales_total > 0 else 0.0

    # 第2栏 应税货物销售额（税收编码前两位10/11/12/13/14/15）
    row2_goods = round(goods_sales, 2)
    # 第3栏 应税劳务销售额（税收编码前两位20）
    row3_labor = round(labor_sales, 2)
    # 第4栏 纳税检查调整的销售额（从序时账"纳税检查"关键词提取）
    row4_tax_check = round(float(
        sum(e.credit_amount or 0 for e in entries
             if e.account_code and "2221" in e.account_code
             and "销项" in (e.account_name or "")
             and "检查" in (e.summary or ""))), 2
    )
    # 第5栏 简易计税销售额
    row5_simple = round(simple_sales, 2)
    # 第6栏 免税销售额（其中：纳税检查调整，暂为0）
    row6_exempt_check = 0.0
    # 第7栏 免、抵、退销售额（从出口销售模块取数，暂为0）
    row7_export = 0.0
    # 第8栏 免税销售额
    row8_exempt = round(exempt_sales, 2)
    # 第9栏 免税货物销售额（从免税销售额中再分货物）
    row9_exempt_goods = round(
        sum(float(i.amount or 0) for i in sales_invoices
             if _is_exempt(i) and _inv_category(i) == "goods"), 2
    )
    # 第10栏 免税劳务销售额
    row10_exempt_service = round(row8_exempt - row9_exempt_goods, 2)

    # 第11栏 销项税额（已从前序计算）
    row11_output_tax = round(output_tax, 2)

    # 第12栏 进项税额（已从前序计算）
    row12_input_tax = round(input_tax, 2)

    # 第13栏 上期留抵税额
    row13_prior_credit = round(prior_credit, 2)

    # 第14栏 进项税额转出 = 附表二第13栏合计
    # 从进项发票中筛选进项税额转出的记录（红字或标注"转出"）
    # 第14栏 进项税额转出：从序时账取"进项税额转出"类分录
    # 会计科目2221下的贷方发生额，或摘要含"转出"、"进项转出"、"免税项目"、"集体福利"等
    input_transfer_out = round(float(
        sum(e.credit_amount or 0 for e in entries
             if e.account_code and "2221" in e.account_code
             and (e.credit_amount or 0) > 0
             and any(k in (e.summary or "") + " " + (e.account_name or "")
                     for k in ["转出", "进项转出", "免税项目", "集体福利", "个人消费", "非正常损失"])) +
        sum(d.deductible_tax_amount or 0 for d in input_deductions
             if "转出" in (d.remark or "") or (d.invoice_status or "") == "红字")),
        2
    )
    # 第15栏 免、抵、退应退税额（从出口退税模块取，无则为0）
    exempt_refund = 0.0
    # 第16栏 按适用税率计算的纳税检查应补缴税额（无税务检查则为0）
    row16_tax_check = 0.0

    # 第17栏 应抵扣税额合计 = 第12+第13-第14-第15+第16
    row17_total_deduct = round(row12_input_tax + row13_prior_credit - input_transfer_out - exempt_refund + row16_tax_check, 2)

    # 第18栏 实际抵扣税额 = min(第17栏, 第11栏)
    row18_actual_deduct = min(round(row17_total_deduct, 2), row11_output_tax)

    # 第19栏 应纳税额 = 第11栏 - 第18栏
    row19_tax_payable = max(0, round(row11_output_tax - row18_actual_deduct, 2))

    # 第20栏 期末留抵税额 = 第17栏 - 第18栏
    row20_end_credit = round(row17_total_deduct - row18_actual_deduct, 2)

    # 第21栏 简易计税办法计算的应纳税额
    row21_simple_tax = round(simple_tax, 2)

    # 第22栏 简易计税纳税检查应补缴税额（无税务检查则为0）
    row22_simple_check = 0.0

    # 第23栏 应纳税额减征额（从税收优惠政策，暂为0）
    row23_reduction = 0.0

    # 第24栏 应纳税额合计 = 第19+第21-第23
    row24_tax_total = round(row19_tax_payable + row21_simple_tax - row23_reduction, 2)

    # 第25栏 期初未缴税额 = 上期申报表第32栏
    row25_prior_unpaid = round(prior_main.get("row32_end_unpaid", 0.0) if prior_vd and prior_vd.form_main else 0.0, 2)

    # 第26-31栏 本期已缴税额（从银行流水取数）
    # 第26栏 本期已缴税额（本期缴纳前期应纳税额）
    row26_paid_prior = round(float(
        sum(e.debit_amount or 0 for e in entries
             if e.account_code and "2221" in e.account_code
             and "已缴" in (e.summary or "")
             and "前期" in (e.summary or ""))), 2
    )
    # 第27栏 本期已缴税额（本期预缴税额）
    row27_prepay = round(float(
        sum(e.debit_amount or 0 for e in entries
             if e.account_code and "2221" in e.account_code
             and "预缴" in (e.summary or ""))), 2
    )
    # 第28栏 本期已缴税额（本期缴纳查补税额）
    row28_paid_check = round(float(
        sum(e.debit_amount or 0 for e in entries
             if e.account_code and "2221" in e.account_code
             and "查补" in (e.summary or ""))), 2
    )
    # 第30栏 本期已缴税额合计 = 第26+第27+第28
    row30_paid_total = round(row26_paid_prior + row27_prepay + row28_paid_check, 2)

    # 第31栏 期末未缴税额 = 第25+第24-第30
    row31_end_unpaid = round(row25_prior_unpaid + row24_tax_total - row30_paid_total, 2)

    # 第32-38栏 查补税额（暂为0）
    row32_check_tax = 0.0
    row33_check_paid = 0.0
    row34_check_end = 0.0

    # 第39-41栏 附加税费（已计算）
    row39_city_tax = round(city_tax, 2)
    row40_edu_surcharge = round(edu_surcharge, 2)
    row41_local_edu = round(local_edu, 2)


    # ====== 附列资料（一）：本期销售情况明细 ======
    # 按官方填写说明重新设计：区分"货物及加工修理修配劳务"和"服务、不动产和无形资产"
    # 第1行：13%税率的货物及加工修理修配劳务
    # 第2行：13%税率的服务、不动产和无形资产
    # 第3行：9%税率的货物及加工修理修配劳务
    # 第4行：9%税率的服务、不动产和无形资产
    # 第5行：6%税率（服务、不动产和无形资产）
    # 第6行：即征即退货物及加工修理修配劳务
    # 第7行：即征即退服务、不动产和无形资产
    # 第8行：简易计税 6%征收率
    # 第9a行：简易计税 5%征收率（货物）
    # 第9b行：简易计税 5%征收率（服务）
    # 第10行：简易计税 4%征收率
    # 第11行：简易计税 3%征收率（货物）
    # 第12行：简易计税 3%征收率（服务）
    
    # 按税率分类汇总
    sales_by_rate = {}
    for inv in sales_invoices:
        rate = int(inv.tax_rate or 13)
        # 判断是货物劳务还是服务不动产
        # 根据发票内容判断（这里简化处理，实际应根据商品名称判断）
        is_service = "服务" in (inv.goods_name or "") or "不动产" in (inv.goods_name or "") or "无形资产" in (inv.goods_name or "")
        
        key = f"rate_{rate}_{'service' if is_service else 'goods'}"
        if key not in sales_by_rate:
            sales_by_rate[key] = {"amount": 0, "tax": 0, "inclusive": 0}
        sales_by_rate[key]["amount"] += float(inv.amount or 0)
        sales_by_rate[key]["tax"] += float(inv.tax_amount or 0)
        sales_by_rate[key]["inclusive"] += float(inv.total_amount or 0)
    
    def _sr(rate_key):
        """安全取税率汇总数据"""
        return sales_by_rate.get(rate_key, {"amount": 0, "tax": 0, "inclusive": 0})
    
    # 计算各行数据（第1-12列）
    # 第1列：开具增值税专用发票
    # 第2列：开具增值税专用发票销项税额
    # 第3列：开具其他发票销售额
    # 第4列：开具其他发票销项税额
    # 第5列：未开具发票销售额
    # 第6列：未开具发票销项税额
    # 第7列：纳税检查调整销售额
    # 第8列：纳税检查调整销项税额
    # 第9列：合计销售额
    # 第10列：合计销项税额
    # 第11列：合计价税合计额
    # 第12列：服务、不动产和无形资产扣除项目本期实际扣除金额
    # 第13列：扣除后含税(免税)销售额
    # 第14列：扣除后销项(应纳)税额
    
    # 简化处理：当前系统所有数据都填入"开具增值税专用发票"列
    # 后续可扩展：根据发票类型分别填入不同列
    
    form_sales = {
        "period": period,
        # 第1行：13%税率的货物及加工修理修配劳务
        "row1_13_goods_special_sales": round(_sr("rate_13_goods")["amount"], 2),
        "row1_13_goods_special_tax": round(_sr("rate_13_goods")["tax"], 2),
        "row1_13_goods_other_sales": 0.0, "row1_13_goods_other_tax": 0.0,
        "row1_13_goods_no_invoice_sales": 0.0, "row1_13_goods_no_invoice_tax": 0.0,
        "row1_13_goods_check_sales": 0.0, "row1_13_goods_check_tax": 0.0,
        "row1_13_goods_total_sales": round(_sr("rate_13_goods")["amount"], 2),
        "row1_13_goods_total_tax": round(_sr("rate_13_goods")["tax"], 2),
        
        # 第2行：13%税率的服务、不动产和无形资产
        "row2_13_service_special_sales": round(_sr("rate_13_service")["amount"], 2),
        "row2_13_service_special_tax": round(_sr("rate_13_service")["tax"], 2),
        "row2_13_service_other_sales": 0.0, "row2_13_service_other_tax": 0.0,
        "row2_13_service_no_invoice_sales": 0.0, "row2_13_service_no_invoice_tax": 0.0,
        "row2_13_service_check_sales": 0.0, "row2_13_service_check_tax": 0.0,
        "row2_13_service_total_sales": round(_sr("rate_13_service")["amount"], 2),
        "row2_13_service_total_tax": round(_sr("rate_13_service")["tax"], 2),
        
        # 第3行：9%税率的货物及加工修理修配劳务
        "row3_9_goods_special_sales": round(_sr("rate_9_goods")["amount"], 2),
        "row3_9_goods_special_tax": round(_sr("rate_9_goods")["tax"], 2),
        "row3_9_goods_other_sales": 0.0, "row3_9_goods_other_tax": 0.0,
        "row3_9_goods_no_invoice_sales": 0.0, "row3_9_goods_no_invoice_tax": 0.0,
        "row3_9_goods_check_sales": 0.0, "row3_9_goods_check_tax": 0.0,
        "row3_9_goods_total_sales": round(_sr("rate_9_goods")["amount"], 2),
        "row3_9_goods_total_tax": round(_sr("rate_9_goods")["tax"], 2),
        
        # 第4行：9%税率的服务、不动产和无形资产
        "row4_9_service_special_sales": round(_sr("rate_9_service")["amount"], 2),
        "row4_9_service_special_tax": round(_sr("rate_9_service")["tax"], 2),
        "row4_9_service_other_sales": 0.0, "row4_9_service_other_tax": 0.0,
        "row4_9_service_no_invoice_sales": 0.0, "row4_9_service_no_invoice_tax": 0.0,
        "row4_9_service_check_sales": 0.0, "row4_9_service_check_tax": 0.0,
        "row4_9_service_total_sales": round(_sr("rate_9_service")["amount"], 2),
        "row4_9_service_total_tax": round(_sr("rate_9_service")["tax"], 2),
        
        # 第5行：6%税率（服务、不动产和无形资产）
        "row5_6_service_special_sales": round(_sr("rate_6_service")["amount"], 2),
        "row5_6_service_special_tax": round(_sr("rate_6_service")["tax"], 2),
        "row5_6_service_other_sales": 0.0, "row5_6_service_other_tax": 0.0,
        "row5_6_service_no_invoice_sales": 0.0, "row5_6_service_no_invoice_tax": 0.0,
        "row5_6_service_check_sales": 0.0, "row5_6_service_check_tax": 0.0,
        "row5_6_service_total_sales": round(_sr("rate_6_service")["amount"], 2),
        "row5_6_service_total_tax": round(_sr("rate_6_service")["tax"], 2),
        
        # 第6行：即征即退货物及加工修理修配劳务（当前无数据）
        "row6_refund_goods_special_sales": 0.0, "row6_refund_goods_special_tax": 0.0,
        "row6_refund_goods_other_sales": 0.0, "row6_refund_goods_other_tax": 0.0,
        "row6_refund_goods_no_invoice_sales": 0.0, "row6_refund_goods_no_invoice_tax": 0.0,
        "row6_refund_goods_check_sales": 0.0, "row6_refund_goods_check_tax": 0.0,
        "row6_refund_goods_total_sales": 0.0, "row6_refund_goods_total_tax": 0.0,
        
        # 第7行：即征即退服务、不动产和无形资产（当前无数据）
        "row7_refund_service_special_sales": 0.0, "row7_refund_service_special_tax": 0.0,
        "row7_refund_service_other_sales": 0.0, "row7_refund_service_other_tax": 0.0,
        "row7_refund_service_no_invoice_sales": 0.0, "row7_refund_service_no_invoice_tax": 0.0,
        "row7_refund_service_check_sales": 0.0, "row7_refund_service_check_tax": 0.0,
        "row7_refund_service_total_sales": 0.0, "row7_refund_service_total_tax": 0.0,
        
        # 第8行：简易计税 6%征收率（当前无数据）
        "row8_6_simple_special_sales": 0.0, "row8_6_simple_special_tax": 0.0,
        "row8_6_simple_other_sales": 0.0, "row8_6_simple_other_tax": 0.0,
        "row8_6_simple_no_invoice_sales": 0.0, "row8_6_simple_no_invoice_tax": 0.0,
        "row8_6_simple_check_sales": 0.0, "row8_6_simple_check_tax": 0.0,
        "row8_6_simple_total_sales": 0.0, "row8_6_simple_total_tax": 0.0,
        
        # 第9a行：简易计税 5%征收率（货物）（当前无数据）
        "row9a_5_goods_simple_special_sales": 0.0, "row9a_5_goods_simple_special_tax": 0.0,
        "row9a_5_goods_simple_other_sales": 0.0, "row9a_5_goods_simple_other_tax": 0.0,
        "row9a_5_goods_simple_no_invoice_sales": 0.0, "row9a_5_goods_simple_no_invoice_tax": 0.0,
        "row9a_5_goods_simple_check_sales": 0.0, "row9a_5_goods_simple_check_tax": 0.0,
        "row9a_5_goods_simple_total_sales": 0.0, "row9a_5_goods_simple_total_tax": 0.0,
        
        # 第9b行：简易计税 5%征收率（服务）（当前无数据）
        "row9b_5_service_simple_special_sales": 0.0, "row9b_5_service_simple_special_tax": 0.0,
        "row9b_5_service_simple_other_sales": 0.0, "row9b_5_service_simple_other_tax": 0.0,
        "row9b_5_service_simple_no_invoice_sales": 0.0, "row9b_5_service_simple_no_invoice_tax": 0.0,
        "row9b_5_service_simple_check_sales": 0.0, "row9b_5_service_simple_check_tax": 0.0,
        "row9b_5_service_simple_total_sales": 0.0, "row9b_5_service_simple_total_tax": 0.0,
        
        # 第10行：简易计税 4%征收率（当前无数据）
        "row10_4_simple_special_sales": 0.0, "row10_4_simple_special_tax": 0.0,
        "row10_4_simple_other_sales": 0.0, "row10_4_simple_other_tax": 0.0,
        "row10_4_simple_no_invoice_sales": 0.0, "row10_4_simple_no_invoice_tax": 0.0,
        "row10_4_simple_check_sales": 0.0, "row10_4_simple_check_tax": 0.0,
        "row10_4_simple_total_sales": 0.0, "row10_4_simple_total_tax": 0.0,
        
        # 第11行：简易计税 3%征收率（货物）（当前无数据）
        "row11_3_goods_simple_special_sales": 0.0, "row11_3_goods_simple_special_tax": 0.0,
        "row11_3_goods_simple_other_sales": 0.0, "row11_3_goods_simple_other_tax": 0.0,
        "row11_3_goods_simple_no_invoice_sales": 0.0, "row11_3_goods_simple_no_invoice_tax": 0.0,
        "row11_3_goods_simple_check_sales": 0.0, "row11_3_goods_simple_check_tax": 0.0,
        "row11_3_goods_simple_total_sales": 0.0, "row11_3_goods_simple_total_tax": 0.0,
        
        # 第12行：简易计税 3%征收率（服务）（当前无数据）
        "row12_3_service_simple_special_sales": 0.0, "row12_3_service_simple_special_tax": 0.0,
        "row12_3_service_simple_other_sales": 0.0, "row12_3_service_simple_other_tax": 0.0,
        "row12_3_service_simple_no_invoice_sales": 0.0, "row12_3_service_simple_no_invoice_tax": 0.0,
        "row12_3_service_simple_check_sales": 0.0, "row12_3_service_simple_check_tax": 0.0,
        "row12_3_service_simple_total_sales": 0.0, "row12_3_service_simple_total_tax": 0.0,
    }
    # 附表一：计算所有税率行的合计
    # 第13行 合计 = 第1行+第2行+第3行+第4行+第5行+第6行+第7行+    }
    
    # 附表一：计算所有税率行的第9列（合计销售额）和第10列（合计销项税额）
    # 第9列 = 第1列 + 第3列 + 第5列 + 第7列
    # 第10列 = 第2列 + 第4列 + 第6列 + 第8列
    # 第11列 = 第9列 + 第10列（价税合计额）
    
    # 第1行：13%税率的货物及加工修理修配劳务
    form_sales["row1_13_goods_total_sales"] = round(
        form_sales["row1_13_goods_special_sales"] + form_sales["row1_13_goods_other_sales"] + 
        form_sales["row1_13_goods_no_invoice_sales"] + form_sales["row1_13_goods_check_sales"], 2)
    form_sales["row1_13_goods_total_tax"] = round(
        form_sales["row1_13_goods_special_tax"] + form_sales["row1_13_goods_other_tax"] + 
        form_sales["row1_13_goods_no_invoice_tax"] + form_sales["row1_13_goods_check_tax"], 2)
    
    # 第2行：13%税率的服务、不动产和无形资产
    form_sales["row2_13_service_total_sales"] = round(
        form_sales["row2_13_service_special_sales"] + form_sales["row2_13_service_other_sales"] + 
        form_sales["row2_13_service_no_invoice_sales"] + form_sales["row2_13_service_check_sales"], 2)
    form_sales["row2_13_service_total_tax"] = round(
        form_sales["row2_13_service_special_tax"] + form_sales["row2_13_service_other_tax"] + 
        form_sales["row2_13_service_no_invoice_tax"] + form_sales["row2_13_service_check_tax"], 2)
    
    # 第3行：9%税率的货物及加工修理修配劳务
    form_sales["row3_9_goods_total_sales"] = round(
        form_sales["row3_9_goods_special_sales"] + form_sales["row3_9_goods_other_sales"] + 
        form_sales["row3_9_goods_no_invoice_sales"] + form_sales["row3_9_goods_check_sales"], 2)
    form_sales["row3_9_goods_total_tax"] = round(
        form_sales["row3_9_goods_special_tax"] + form_sales["row3_9_goods_other_tax"] + 
        form_sales["row3_9_goods_no_invoice_tax"] + form_sales["row3_9_goods_check_tax"], 2)
    
    # 第4行：9%税率的服务、不动产和无形资产
    form_sales["row4_9_service_total_sales"] = round(
        form_sales["row4_9_service_special_sales"] + form_sales["row4_9_service_other_sales"] + 
        form_sales["row4_9_service_no_invoice_sales"] + form_sales["row4_9_service_check_sales"], 2)
    form_sales["row4_9_service_total_tax"] = round(
        form_sales["row4_9_service_special_tax"] + form_sales["row4_9_service_other_tax"] + 
        form_sales["row4_9_service_no_invoice_tax"] + form_sales["row4_9_service_check_tax"], 2)
    
    # 第5行：6%税率（服务、不动产和无形资产）
    form_sales["row5_6_service_total_sales"] = round(
        form_sales["row5_6_service_special_sales"] + form_sales["row5_6_service_other_sales"] + 
        form_sales["row5_6_service_no_invoice_sales"] + form_sales["row5_6_service_check_sales"], 2)
    form_sales["row5_6_service_total_tax"] = round(
        form_sales["row5_6_service_special_tax"] + form_sales["row5_6_service_other_tax"] + 
        form_sales["row5_6_service_no_invoice_tax"] + form_sales["row5_6_service_check_tax"], 2)
    
    # 计算第13行合计（一般计税方法计税 - 全部征税项目）
    # 第13行第9列 = 第1行 + 第2行 + 第3行 + 第4行 + 第5行
    row13_col9 = round(
        form_sales["row1_13_goods_total_sales"] + form_sales["row2_13_service_total_sales"] + 
        form_sales["row3_9_goods_total_sales"] + form_sales["row4_9_service_total_sales"] + 
        form_sales["row5_6_service_total_sales"], 2)
    row13_col10 = round(
        form_sales["row1_13_goods_total_tax"] + form_sales["row2_13_service_total_tax"] + 
        form_sales["row3_9_goods_total_tax"] + form_sales["row4_9_service_total_tax"] + 
        form_sales["row5_6_service_total_tax"], 2)
    
    form_sales["row13_total_special_sales"] = row13_col9
    form_sales["row13_total_special_tax"] = row13_col10
    
    # 计算第14行（即征即退项目）
    # 第14行第9列 = 第6行 + 第7行
    # 当前无数据，暂为0
    form_sales["row14_refund_special_sales"] = 0.0
    form_sales["row14_refund_special_tax"] = 0.0
    
    # 主表第1栏 = 第13行第9列 - 第14行第9列
    sales_total = round(row13_col9 - form_sales["row14_refund_special_sales"], 2)
    # 主表第11栏 = 第13行第10列 - 第14行第10列
    output_tax = round(row13_col10 - form_sales["row14_refund_special_tax"], 2)
    
    vd.form_sales = json.dumps(form_sales, ensure_ascii=False)
    
    # 主表第1栏「按适用税率计税销售额」= 附表一第13行第9列 - 第14行第9列
    # 覆盖之前临时计算的 sales_total
    # sales_total 和 output_tax 已经在上面计算了

    # ====== 附列资料（二）：本期进项税额明细 ======
    # 数据源：进项抵扣表 InputVATDeduction，按「抵扣所属期」取数
    # 这才是增值税申报附表二的正确取数规则，不是按开票日期从PurchaseInvoice取
    # 复用上方的 input_deductions 查询结果，保证主表和附表二数据源一致
    form2_deductions = input_deductions

    def _ded_sum(ded_list, cat_filter=None):
        """汇总进项抵扣记录：返回(份数, 不含税金额, 税额)
        税额取 deductible_tax_amount（有效抵扣税额）"""
        if cat_filter is not None:
            ded_list = [d for d in ded_list if (d.invoice_category or "") == cat_filter]
        return (
            len(ded_list),
            sum(float(d.amount or 0) for d in ded_list),
            sum(float(d.deductible_tax_amount or 0) for d in ded_list),
        )

    # 增值税专用发票类（默认票种）
    def _is_special_invoice(d):
        cat = (d.invoice_category or "") or (d.invoice_category_label or "")
        return "专用发票" in cat or "专用" in cat or (not cat)  # 无标签默认视为专票
    special_list = [d for d in form2_deductions if _is_special_invoice(d)]

    cert_count, cert_amount, cert_tax = _ded_sum(special_list)

    # 非专票（普票/电子票等）→ 旅客运输 / 其他分类
    non_special_list = [d for d in form2_deductions if not _is_special_invoice(d)]
    # - 铁路电子客票/航空电子客票/公路水路客票 → row10 旅客运输服务
    # - 其他非专票 → row4 其他扣税凭证
    def _is_travel_invoice(d):
        cat = (d.invoice_category or "") or (d.invoice_category_label or "")
        return any(kw in cat for kw in ["铁路", "航空", "公路", "水路", "旅客", "客票", "运输服务", "通行费"])
    travel_list = [d for d in non_special_list if _is_travel_invoice(d)]
    other_list  = [d for d in non_special_list if not _is_travel_invoice(d)]

    travel_count, travel_amount, travel_tax = _ded_sum(travel_list)
    other_count, other_amount, other_tax = _ded_sum(other_list)
    # 当期申报抵扣进项税额合计 = 第1栏+第4栏+第9栏+第10栏+第11栏
    # 按填表说明：第12栏 = 第1栏+第4栏+第9栏+第10栏+第11栏
    real_estate_count, real_estate_amount, real_estate_tax = 0, 0.0, 0.0  # 第9栏 本期用于购建不动产的扣税凭证（当前无数据）
    foreign_count, foreign_tax = 0, 0.0  # 第11栏 外贸企业进项税额抵扣证明（当前无数据）
    total_ded_count = cert_count + other_count + travel_count + real_estate_count + foreign_count
    total_ded_amount = round(cert_amount + other_amount + travel_amount + real_estate_amount, 2)
    total_ded_tax = round(cert_tax + other_tax + travel_tax + real_estate_tax + foreign_tax, 2)

    form_input = {
        "period": period,
        # 一、申报抵扣的进项税额
        "row1_certified_count": cert_count,              # 认证相符的增值税专用发票-份数
        "row1_certified_amount": round(cert_amount, 2),   # 金额（不含税）
        "row1_certified_tax": round(cert_tax, 2),         # 税额
        "row2_certified_curr_count": cert_count,
        "row2_certified_curr_amount": round(cert_amount, 2),
        "row2_certified_curr_tax": round(cert_tax, 2),
        "row3_certified_prior_count": 0, "row3_certified_prior_amount": 0.0, "row3_certified_prior_tax": 0.0,
        # 其他扣税凭证（不含旅客运输，旅客运输单独填入 row10）
        "row4_other_count": other_count, "row4_other_amount": round(other_amount, 2), "row4_other_tax": round(other_tax, 2),
        "row5_customs_count": 0, "row5_customs_amount": 0.0, "row5_customs_tax": 0.0,
        "row6_agri_count": 0, "row6_agri_amount": 0.0, "row6_agri_tax": 0.0,
        "row7_wht_count": 0, "row7_wht_tax": 0.0,
        "row8a_agri_extra": 0.0,
        "row8b_other_count": 0, "row8b_other_amount": 0.0, "row8b_other_tax": 0.0,
        # 不动产/旅客/外贸
        "row9_real_estate_count": 0, "row9_real_estate_amount": 0.0, "row9_real_estate_tax": 0.0,
        "row10_travel_count": travel_count, "row10_travel_amount": round(travel_amount, 2), "row10_travel_tax": round(travel_tax, 2),
        "row11_foreign_trade_count": 0, "row11_foreign_trade_tax": 0.0,
        # 二、进项税额转出额
        "row13_transfer_out_total": 0.0,
        "row14_exempt_transfer": 0.0, "row15_collective_transfer": 0.0,
        "row16_abnormal_loss": 0.0, "row17_simple_tax_transfer": 0.0,
        "row18_exempt_credit_transfer": 0.0, "row19_tax_check_transfer": 0.0,
        "row20_red_letter_transfer": 0.0, "row21_prior_credit_arrears": 0.0,
        "row22_prior_credit_refund": 0.0, "row23a_abnormal_transfer": 0.0, "row23b_other_transfer": 0.0,
        # 三、待抵扣进项税额
        "row25_pending_begin_count": 0, "row25_pending_begin_amount": 0.0, "row25_pending_begin_tax": 0.0,
        "row26_pending_curr_count": 0, "row26_pending_curr_amount": 0.0, "row26_pending_curr_tax": 0.0,
        "row27_pending_end_count": 0, "row27_pending_end_amount": 0.0, "row27_pending_end_tax": 0.0,
        "row28_not_allowed_count": 0, "row28_not_allowed_amount": 0.0, "row28_not_allowed_tax": 0.0,
        "row29_other_pending_count": 0, "row29_other_pending_amount": 0.0, "row29_other_pending_tax": 0.0,
        "row30_customs_pending_count": 0, "row30_customs_pending_amount": 0.0, "row30_customs_pending_tax": 0.0,
        "row31_agri_pending_count": 0, "row31_agri_pending_amount": 0.0, "row31_agri_pending_tax": 0.0,
        "row32_wht_pending_count": 0, "row32_wht_pending_tax": 0.0,
        "row33_other_pending_count": 0, "row33_other_pending_amount": 0.0, "row33_other_pending_tax": 0.0,
        # 四、其他
        "row35_cert_count": cert_count,
        "row35_cert_amount": round(cert_amount, 2),
        "row35_cert_tax": round(cert_tax, 2),
        "row36_wht_total_tax": 0.0,
        # 兼容旧字段
        "certified_count": cert_count,
        "certified_amount": round(cert_amount, 2),
        "certified_tax": round(cert_tax, 2),
        "total_deductible": round(total_ded_tax, 2),
    }
    vd.form_input = json.dumps(form_input, ensure_ascii=False)

    # ====== 附列资料（三）：服务、不动产和无形资产扣除项目明细 ======
    # 本表仅在发生服务/不动产/无形资产扣除项目时填写。
    # 系统当前不跟踪扣除项目明细，默认全部填0，由用户手动填写扣除数据。
    form_deduction = {
        "period": period,
        # 各项目：价税合计额/期初余额/本期发生额/本期应扣除金额/本期实际扣除金额/期末余额
        # 17%税率项目（已废止，保留占位）
        "row1_17_price_tax": 0.0, "row1_17_begin": 0.0, "row1_17_occur": 0.0,
        "row1_17_should": 0.0, "row1_17_actual": 0.0, "row1_17_end": 0.0,
        # 13%税率项目
        "row1_13_price_tax": 0.0, "row1_13_begin": 0.0, "row1_13_occur": 0.0,
        "row1_13_should": 0.0, "row1_13_actual": 0.0, "row1_13_end": 0.0,
        # 9%税率项目
        "row2_9_price_tax": 0.0, "row2_9_begin": 0.0, "row2_9_occur": 0.0,
        "row2_9_should": 0.0, "row2_9_actual": 0.0, "row2_9_end": 0.0,
        # 6%税率项目（不含金融商品转让）
        "row3_6_price_tax": 0.0, "row3_6_begin": 0.0, "row3_6_occur": 0.0,
        "row3_6_should": 0.0, "row3_6_actual": 0.0, "row3_6_end": 0.0,
        # 6%税率的金融商品转让项目
        "row4_6_fin_price_tax": 0.0, "row4_6_fin_begin": 0.0, "row4_6_fin_occur": 0.0,
        "row4_6_fin_should": 0.0, "row4_6_fin_actual": 0.0, "row4_6_fin_end": 0.0,
        # 5%征收率项目
        "row5_5_price_tax": 0.0, "row5_5_begin": 0.0, "row5_5_occur": 0.0,
        "row5_5_should": 0.0, "row5_5_actual": 0.0, "row5_5_end": 0.0,
        # 3%征收率
        "row6_3_price_tax": 0.0, "row6_3_begin": 0.0, "row6_3_occur": 0.0,
        "row6_3_should": 0.0, "row6_3_actual": 0.0, "row6_3_end": 0.0,
        # 免抵退税
        "row7_exempt_credit_price_tax": 0.0, "row7_exempt_credit_begin": 0.0, "row7_exempt_credit_occur": 0.0,
        "row7_exempt_credit_should": 0.0, "row7_exempt_credit_actual": 0.0, "row7_exempt_credit_end": 0.0,
        # 免税
        "row8_exempt_price_tax": 0.0, "row8_exempt_begin": 0.0, "row8_exempt_occur": 0.0,
        "row8_exempt_should": 0.0, "row8_exempt_actual": 0.0, "row8_exempt_end": 0.0,
        # 合计
        "row_total_price_tax": 0.0, "row_total_begin": 0.0,
        "row_total_occur": 0.0, "row_total_should": 0.0, "row_total_actual": 0.0, "row_total_end": 0.0,
        # 兼容旧字段
        "row1_total_price_tax": 0.0,
        "row1_deduction": 0.0,
        "row1_after_deduction": 0.0,
    }
    vd.form_deduction = json.dumps(form_deduction, ensure_ascii=False)

    # ====== 附列资料（四）：税额抵减情况表 ======
    form_credit = {
        "period": period,
        # 一、税额抵减情况（5项）
        "tax_control_device": 0.0,
        "tax_control_begin": 0.0, "tax_control_occur": 0.0, "tax_control_should": 0.0,
        "tax_control_actual": 0.0, "tax_control_end": 0.0,
        "subtotal_tax": 0.0,
        "branch_begin": 0.0, "branch_occur": 0.0, "branch_should": 0.0, "branch_actual": 0.0, "branch_end": 0.0,
        "construction_begin": 0.0, "construction_occur": 0.0, "construction_should": 0.0, "construction_actual": 0.0, "construction_end": 0.0,
        "real_estate_begin": 0.0, "real_estate_occur": 0.0, "real_estate_should": 0.0, "real_estate_actual": 0.0, "real_estate_end": 0.0,
        "rental_begin": 0.0, "rental_occur": 0.0, "rental_should": 0.0, "rental_actual": 0.0, "rental_end": 0.0,
        # 二、加计抵减情况（一般项目）
        "item1_begin": 0.0, "item1_occur": 0.0, "item1_adjust": 0.0,
        "item1_can_deduct": 0.0, "item1_should_deduct": 0.0,
        "item1_actual_deduct": 0.0, "item1_end": 0.0,
        # 即征即退项目
        "item2_begin": 0.0, "item2_occur": 0.0, "item2_adjust": 0.0,
        "item2_can_deduct": 0.0, "item2_should_deduct": 0.0,
        "item2_actual_deduct": 0.0, "item2_end": 0.0,
    }
    vd.form_credit = json.dumps(form_credit, ensure_ascii=False)

    # ====== 附列资料（五）：附加税费情况表 ======
    city_full = round(tax_payable * 0.07, 2)
    edu_full = round(tax_payable * 0.03, 2)
    local_edu_full = round(tax_payable * 0.02, 2)
    reduction_rate = 0.5 if vd.micro_enterprise and vd.six_tax_reduction else 0.0

    form_surcharge = {
        "period": period,
        "micro_enterprise": vd.micro_enterprise,
        "six_tax_reduction": vd.six_tax_reduction,
        # 城市维护建设税
        "city_base": round(tax_payable, 2),
        "city_rate": 0.07,
        "city_tax": city_tax,                                          # 减免后实际应纳
        "city_full": round(city_full, 2),                              # 减免前原值
        "city_reduction_code": "六税两费减征" if reduction_rate > 0 else "",
        "city_reduction_amount": round(city_full * reduction_rate, 2),
        "city_six_tax_amount": round(city_full * reduction_rate, 2),
        "city_final": city_tax,
        "city_reduction_rate": reduction_rate,
        "city_vat_exempt_credit": 0.0,
        "city_vat_refund_deduct": 0.0,
        "city_edu_pilot_code": "",
        "city_edu_pilot_amount": 0.0,
        "city_paid": 0.0,
        # 教育费附加
        "edu_base": round(tax_payable, 2),
        "edu_rate": 0.03,
        "edu_tax": edu_surcharge,                                      # 减免后实际应纳
        "edu_full": round(edu_full, 2),                                # 减免前原值
        "edu_reduction_code": "六税两费减征" if reduction_rate > 0 else "",
        "edu_reduction_amount": round(edu_full * reduction_rate, 2),
        "edu_six_tax_amount": round(edu_full * reduction_rate, 2),
        "edu_final": edu_surcharge,
        "edu_reduction_rate": reduction_rate,
        "edu_vat_exempt_credit": 0.0,
        "edu_vat_refund_deduct": 0.0,
        "edu_edu_pilot_code": "",
        "edu_edu_pilot_amount": 0.0,
        "edu_paid": 0.0,
        # 地方教育附加
        "local_edu_base": round(tax_payable, 2),
        "local_edu_rate": 0.02,
        "local_edu_tax": local_edu,                                    # 减免后实际应纳
        "local_edu_full": round(local_edu_full, 2),                    # 减免前原值
        "local_edu_reduction_code": "六税两费减征" if reduction_rate > 0 else "",
        "local_edu_reduction_amount": round(local_edu_full * reduction_rate, 2),
        "local_edu_six_tax_amount": round(local_edu_full * reduction_rate, 2),
        "local_edu_final": local_edu,
        "local_edu_reduction_rate": reduction_rate,
        "local_edu_vat_exempt_credit": 0.0,
        "local_edu_vat_refund_deduct": 0.0,
        "local_edu_edu_pilot_code": "",
        "local_edu_edu_pilot_amount": 0.0,
        "local_edu_paid": 0.0,
        # 合计
        "total_tax": round(city_full + edu_full + local_edu_full, 2),
        "total_reduction": round((city_full + edu_full + local_edu_full) * reduction_rate, 2),
        "total_six_tax_reduction": round((city_full + edu_full + local_edu_full) * reduction_rate, 2),
        "total_edu_pilot": 0.0,
        "total_paid": 0.0,
        "total_final": round(city_tax + edu_surcharge + local_edu, 2),
        # 兼容
        "vat_exempt_credit": 0.0, "vat_refund_deduct": 0.0,
        "city_reduction_type": "六税两费减征" if reduction_rate > 0 else "",
        "edu_reduction_type": "六税两费减征" if reduction_rate > 0 else "",
        "local_edu_reduction_type": "六税两费减征" if reduction_rate > 0 else "",
    }
    vd.form_surcharge = json.dumps(form_surcharge, ensure_ascii=False)

    # ====== 增值税减免税申报明细表 ======
    form_reduction = {
        "period": period,
        "micro_enterprise": vd.micro_enterprise,
        "six_tax_reduction": vd.six_tax_reduction,
        # 一、减税项目
        "tax_reduction_items": [],
        "tax_reduction_total_occur": 0.0,    # 本期发生额
        "tax_reduction_total_should": 0.0,    # 本期应抵减税额
        "tax_reduction_total_actual": 0.0,    # 本期实际抵减税额
        "tax_reduction_total_end": 0.0,       # 期末余额
        # 二、免税项目
        "exempt_items": [],
        "exempt_total_exempt_sales": 0.0,     # 免征增值税项目销售额
        "exempt_total_exempt_tax": 0.0,       # 免税销售额对应的进项税额
        "exempt_total_exempt_amount": 0.0,    # 免税额
        # 兼容旧字段
        "reduction_items": [],
    }
    vd.form_reduction = json.dumps(form_reduction, ensure_ascii=False)

    # ===== 主表计算（按官方填写说明公式）=====
    # 按正确顺序：先计算所有附表，再根据附表计算主表
    
    # 取上期留抵：从同公司上期申报表取期末留抵
    prev_year = period_date.year
    prev_month = period_date.month - 1
    if prev_month == 0:
        prev_year -= 1
        prev_month = 12
    prev_period = f"{prev_year}-{prev_month:02d}"
    prior_vd = db.query(VATDeclaration).filter(
        VATDeclaration.company_id == company_id,
        VATDeclaration.period == prev_period
    ).first()
    prior_credit = 0.0
    if prior_vd and prior_vd.form_main:
        prior_main = json.loads(prior_vd.form_main) if isinstance(prior_vd.form_main, str) else prior_vd.form_main
        prior_credit = prior_main.get("row20_end_credit", 0.0)
    
    # 读取附表一和附表二
    s1 = json.loads(vd.form_sales) if isinstance(vd.form_sales, str) else (vd.form_sales or {})
    s2 = json.loads(vd.form_input) if isinstance(vd.form_input, str) else (vd.form_input or {})
    
    # ===== 主表第1栏：按适用税率计税销售额 =====
    # 官方公式：第1栏 = 附表一第9列第1至5行之和 - 第9列第6、7行之和
    sales_total = round(
        s1.get("row1_13_goods_total_sales", 0) +
        s1.get("row2_13_service_total_sales", 0) +
        s1.get("row3_9_goods_total_sales", 0) +
        s1.get("row4_9_service_total_sales", 0) +
        s1.get("row5_6_service_total_sales", 0) -
        s1.get("row6_refund_goods_total_sales", 0) -
        s1.get("row7_refund_service_total_sales", 0),
        2
    )
    
    # ===== 主表第11栏：销项税额 =====
    # 官方公式：第11栏 = 附表一（第10列第1、3行之和 - 第10列第6行）+（第14列第2、4、5行之和 - 第14列第7行）
    # 注：当"扣除后销项税额"（第14列）未实现时，回退使用"合计销项税额"（第10列）
    def _tax_s1(row_key):
        """取第14列(扣除后)，若为0则回退第10列(合计)"""
        after = s1.get(f"{row_key}_after_deduction_tax", 0) or 0
        total = s1.get(f"{row_key}_total_tax", 0) or 0
        return after if after else total
    
    output_tax_part1 = round(
        s1.get("row1_13_goods_total_tax", 0) +
        s1.get("row3_9_goods_total_tax", 0) -
        s1.get("row6_refund_goods_total_tax", 0),
        2
    )
    output_tax_part2 = round(
        _tax_s1("row2_13_service") +
        _tax_s1("row4_9_service") +
        _tax_s1("row5_6_service") -
        _tax_s1("row7_refund_service"),
        2
    )
    output_tax = round(output_tax_part1 + output_tax_part2, 2)
    
    # ===== 主表第12栏：进项税额 =====
    input_tax = s2.get("total_deductible", 0)
    
    # ===== 主表第21栏：简易计税办法计算的应纳税额 =====
    # 官方公式：第21栏 = 附表一（第10列第8、9a、10、11行之和 - 第10列第14行）+（第14列第9b、12、13a、13b行之和 - 第14列第15行）
    # 注意：当前系统简化计算，仅计算第10列（合计销项税额），未实现第14列（扣除后销项税额）
    simple_tax = round(
        s1.get("row8_6_simple_total_tax", 0) +
        s1.get("row9a_5_goods_simple_total_tax", 0) +
        s1.get("row10_4_simple_total_tax", 0) +
        s1.get("row11_3_goods_simple_total_tax", 0) +
        s1.get("row12_3_service_simple_total_tax", 0) +
        s1.get("row13a_prepay_total_tax", 0) +
        s1.get("row13b_prepay_2_total_tax", 0) -
        s1.get("row14_refund_simple_total_tax", 0) -
        s1.get("row15_refund_simple_total_tax", 0),
        2
    )
    
    # 主表计算链（按填表说明公式）
    input_transfer_out = 0.0
    exempt_refund = 0.0
    tax_check = 0.0
    total_deduct = round(input_tax + prior_credit - input_transfer_out - exempt_refund + tax_check, 2)
    actual_deduct = min(total_deduct, output_tax)
    actual_reduction = 0.0
    tax_payable = max(0, round(output_tax - actual_deduct - actual_reduction, 2))
    end_credit = round(total_deduct - actual_deduct, 2)
    reduction = 0.0
    tax_payable_total = round(tax_payable + simple_tax - reduction, 2)
    
    year_str = period[:4]
    ytd_periods = [f"{year_str}-{m:02d}" for m in range(1, period_date.month + 1)]
    
    # 查询本年所有已保存申报表（不含当前正在计算的本条，因为 form_main 尚未入库）
    ytd_declarations = db.query(VATDeclaration).filter(
        VATDeclaration.company_id == company_id,
        VATDeclaration.period.in_(ytd_periods),
        VATDeclaration.id != vd.id,
    ).all()
    
    # 需要累计的字段列表（主表所有数字栏次）
    ytd_fields = [
        "row1_sales", "row2_other_invoice", "row3_no_invoice", "row4_tax_check",
        "row5_simple_method", "row6_exempt_sales", "row7_export_exempt",
        "row8_tax_free", "row9_exempt_goods", "row10_exempt_service",
        "row11_output_tax", "row12_input_tax", "row13_prior_credit",
        "row14_input_transfer_out", "row15_exempt_refund", "row16_actual_deduct_by_item",
        "row17_total_deductible", "row18_actual_deduct", "row19_tax_payable",
        "row20_end_credit", "row21_simple_tax", "row22_simple_tax_reduction",
        "row23_reduction", "row24_tax_payable_total",
        "row25_prior_unpaid", "row26_real_paid_during", "row27_installment_prepaid",
        "row28_export_tax_refund", "row29_remote_prepaid", "row30_already_paid_total",
        "row31_should_pay_refund", "row32_check_tax_should", "row33_check_prepaid",
        "row34_should_check", "row36_prior_unpaid_check", "row37_check_paid",
        "row38_end_check",
        "row39_city_maintenance_tax", "row40_education_surcharge", "row41_local_education_surcharge",
    ]
    
    # 累加历史各月数据
    ytd_sums = {f: 0.0 for f in ytd_fields}
    for d in ytd_declarations:
        m = json.loads(d.form_main) if isinstance(d.form_main, str) else (d.form_main or {})
        for f in ytd_fields:
            ytd_sums[f] += m.get(f, 0.0)
    
    # 构建主表数据
    form_main = {
        "period": period,
        "taxpayer_name": vd.taxpayer_name,
        # 一、销售额
        "row1_sales": round(sales_total, 2),         # 按适用税率计税销售额
        "row2_other_invoice": 0.0,                     # 其中：开具其他发票
        "row3_no_invoice": 0.0,                        # 未开具发票
        "row4_tax_check": round(s1.get("row4_check_total_sales", 0), 2),  # 纳税检查调整 = 附表一第7列第1至5行之和
        "row5_simple_method": max(0, round(
            s1.get("row8_simple_6_total_sales", 0) +
            s1.get("row9a_simple_5a_total_sales", 0) +
            s1.get("row10_simple_4_total_sales", 0) +
            s1.get("row11_simple_3a_total_sales", 0) +
            s1.get("row12_simple_3b_total_sales", 0) +
            s1.get("row13a_prepay_total_sales", 0) +
            s1.get("row13b_prepay_2_total_sales", 0) -
            s1.get("row14_refund_simple_total_sales", 0) -
            s1.get("row15_refund_simple_total_sales", 0),
            2
        )),  # 简易计税销售额
        "row6_exempt_sales": 0.0,                      # 免税销售额
        "row7_export_exempt": round(
            s1.get("row16_export_goods_total_sales", 0) +
            s1.get("row17_export_service_total_sales", 0),
            2
        ),  # 免抵退销售额 = 附表一第9列第16、17行之和
        "row8_tax_free": round(
            s1.get("row18_exempt_goods_total_sales", 0) +
            s1.get("row19_exempt_service_total_sales", 0),
            2
        ),  # 免税销售额 = 附表一第9列第18、19行之和
        "row9_exempt_goods": 0.0,                      # 免税货物销售额
        "row10_exempt_service": 0.0,                   # 免税劳务销售额
        # 二、税款计算
        "row11_output_tax": round(output_tax, 2),      # 销项税额
        "row12_input_tax": round(input_tax, 2),         # 进项税额
        "row13_prior_credit": round(prior_credit, 2),   # 上期留抵税额
        "row14_input_transfer_out": round(s2.get("row13_transfer_out_total", 0), 2),  # 进项税额转出 = 附表二第13栏
        "row15_exempt_refund": 0.0,                     # 免抵退应退税额
        "row16_actual_deduct_by_item": 0.0,             # 按适用税率计算的纳税检查应补缴税额
        "row17_total_deductible": total_deduct,         # 应抵扣税额合计 =12+13-14-15+16
        "row18_actual_deduct": round(actual_deduct, 2), # 实际抵扣税额
        "row19_tax_payable": round(tax_payable, 2),     # 应纳税额 =11-18
        "row20_end_credit": round(end_credit, 2),       # 期末留抵税额 =17-18
        "row21_simple_tax": simple_tax,                        # 简易计税办法计算的应纳税额
        "row22_simple_tax_reduction": 0.0,              # 按简易计税办法计算的纳税检查应补缴税额
        "row23_reduction": 0.0,                         # 应纳税额减征额
        "row24_tax_payable_total": tax_payable_total,   # 应纳税额合计 =19+21-23
        # 三、税款缴纳
        "row25_prior_unpaid": round(prior_main.get("row32_end_unpaid", 0.0) if prior_vd and prior_vd.form_main else 0.0, 2),  # 期初未缴税额 = 上期第32栏
        "row26_real_paid_during": 0.0,                  # 本期已缴税额
        "row27_installment_prepaid": 0.0,               # 分次预缴税额
        "row28_export_tax_refund": 0.0,                 # 出口开具专用缴款书预缴税额
        "row29_remote_prepaid": 0.0,                    # 本期缴纳上期应纳税额
        "row30_already_paid_total": 0.0,                # 本期缴纳欠缴税额
        "row31_should_pay_refund": 0.0,                 # 期末未缴税额
        "row32_check_tax_should": 0.0,                  # 其中：欠缴税额
        "row33_check_prepaid": 0.0,                     # 本期入库查补税额
        "row34_should_check": 0.0,                      # 期末未缴查补税额
        "row36_prior_unpaid_check": 0.0,                # 期初未缴查补税额
        "row37_check_paid": 0.0,                        # 本期入库查补税额
        "row38_end_check": 0.0,                         # 期末未缴查补税额
        # 四、附加税费
        "row39_city_maintenance_tax": city_tax,          # 城市维护建设税
        "row40_education_surcharge": edu_surcharge,      # 教育费附加
        "row41_local_education_surcharge": local_edu,    # 地方教育附加
        # 合计
        "city_maintenance_tax": city_tax,
        "education_surcharge": edu_surcharge,
        "local_education_surcharge": local_edu,
        "total_surcharge": round(city_tax + edu_surcharge + local_edu, 2),
    }
    
    # 本年累计(YTD) —— 注入 form_main
    for _k in ytd_fields:
        # 本年累计 = 历史各月累计 + 本月发生额
        this_month = form_main.get(_k, 0.0)
        form_main[_k + "_ytd"] = round(ytd_sums.get(_k, 0.0) + this_month, 2)
        # 即征即退项目（当前系统不涉及，默认0）
        form_main[_k + "_refund"] = 0.0
        form_main[_k + "_refund_ytd"] = 0.0
    
    vd.form_main = json.dumps(form_main, ensure_ascii=False)
