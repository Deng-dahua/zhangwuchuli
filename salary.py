"""
工资薪金所得模块 - 预扣预缴明细

支持：
1. 按期间管理工资表
2. 从税务模板Excel导入（综合所得月工资薪所得.xls）
3. 自动计算个税（累计预扣法）
4. 自动添加人员档案
"""
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, Form, Body
from fastapi.responses import JSONResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, or_
from typing import Optional, List
from datetime import date, datetime
import io
import openpyxl
import json

from database import (
    get_db, SessionLocal,
    SalaryRecord, Company, Employee,
)

router = APIRouter(prefix="/api/salary", tags=["工资薪金"])


# ========== Pydantic Schemas ==========
from pydantic import BaseModel


class SalaryRecordCreate(BaseModel):
    period: str
    employee_name: str
    id_number: str
    id_type: str = "居民身份证"
    tax_period_start: Optional[str] = None
    tax_period_end: Optional[str] = None
    income_type: str = "正常工资薪金"
    current_income: float = 0.0
    tax_free_income: float = 0.0
    basic_deduction: float = 5000.0
    # 专项扣除
    pension_insurance: float = 0.0
    medical_insurance: float = 0.0
    unemployment_insurance: float = 0.0
    housing_fund: float = 0.0
    enterprise_annuity: float = 0.0
    commercial_health: float = 0.0
    tax_deferred_pension: float = 0.0
    other_special_deduction: float = 0.0
    # 专项附加扣除
    child_education: float = 0.0
    continuing_education: float = 0.0
    housing_loan_interest: float = 0.0
    housing_rent: float = 0.0
    elderly_support: float = 0.0
    infant_care: float = 0.0
    major_medical: float = 0.0
    other_additional_deduction: float = 0.0
    # 累计
    cumulative_income: float = 0.0
    cumulative_tax_free: float = 0.0
    cumulative_deduction: float = 0.0
    cumulative_special: float = 0.0
    cumulative_additional: float = 0.0
    cumulative_other: float = 0.0
    cumulative_tax_withheld: float = 0.0
    other_deduction: float = 0.0
    # 税额
    taxable_income: float = 0.0
    tax_rate: float = 0.0
    quick_deduction: float = 0.0
    tax_payable: float = 0.0
    tax_already_withheld: float = 0.0
    tax_to_pay: float = 0.0
    tax_refund: float = 0.0
    net_salary: float = 0.0


class SalaryRecordUpdate(BaseModel):
    current_income: Optional[float] = None
    tax_free_income: Optional[float] = None
    basic_deduction: Optional[float] = None
    pension_insurance: Optional[float] = None
    medical_insurance: Optional[float] = None
    unemployment_insurance: Optional[float] = None
    housing_fund: Optional[float] = None
    enterprise_annuity: Optional[float] = None
    commercial_health: Optional[float] = None
    tax_deferred_pension: Optional[float] = None
    other_special_deduction: Optional[float] = None
    child_education: Optional[float] = None
    continuing_education: Optional[float] = None
    housing_loan_interest: Optional[float] = None
    housing_rent: Optional[float] = None
    elderly_support: Optional[float] = None
    infant_care: Optional[float] = None
    major_medical: Optional[float] = None
    other_additional_deduction: Optional[float] = None
    cumulative_income: Optional[float] = None
    cumulative_tax_free: Optional[float] = None
    cumulative_deduction: Optional[float] = None
    cumulative_special: Optional[float] = None
    cumulative_additional: Optional[float] = None
    cumulative_other: Optional[float] = None
    cumulative_tax_withheld: Optional[float] = None
    other_deduction: Optional[float] = None
    taxable_income: Optional[float] = None
    tax_rate: Optional[float] = None
    quick_deduction: Optional[float] = None
    tax_payable: Optional[float] = None
    tax_already_withheld: Optional[float] = None
    tax_to_pay: Optional[float] = None
    tax_refund: Optional[float] = None
    net_salary: Optional[float] = None


# ========== 工具函数 ==========


def _parse_id_number(id_number: str):
    """从身份证号提取出生日期和性别"""
    if not id_number or len(id_number) < 15:
        return None, None
    id_number = id_number.strip().upper()
    birthday = None
    gender = None
    try:
        if len(id_number) == 18:
            birthday = f"{id_number[6:10]}-{id_number[10:12]}-{id_number[12:14]}"
            gender = "女" if int(id_number[16]) % 2 == 0 else "男"
        elif len(id_number) == 15:
            birthday = f"19{id_number[6:8]}-{id_number[8:10]}-{id_number[10:12]}"
            gender = "女" if int(id_number[14]) % 2 == 0 else "男"
    except Exception:
        pass
    return birthday, gender


def _auto_create_employee(db: Session, company_id: int, name: str, id_number: str):
    """自动创建人员档案（如果不存在）"""
    if not name or not id_number:
        return None
    existing = db.query(Employee).filter(
        Employee.company_id == company_id,
        Employee.id_card == id_number
    ).first()
    if existing:
        # 如果姓名不一致，更新姓名
        if existing.name != name:
            existing.name = name
            db.flush()
        return existing

    # 生成人员编码 RY001, RY002...
    max_code = db.query(Employee.code).filter(
        Employee.company_id == company_id
    ).order_by(Employee.id.desc()).first()
    if max_code and max_code[0] and max_code[0].startswith('RY'):
        try:
            num = int(max_code[0][2:]) + 1
        except ValueError:
            num = 1
    else:
        num = 1
    code = f"RY{num:03d}"

    emp = Employee(
        company_id=company_id,
        code=code,
        name=name,
        id_card=id_number,
    )
    db.add(emp)
    db.flush()
    return emp


def _compute_tax_cumulative(income: float, prev_cumulative_income: float,
                            prev_cumulative_tax: float,
                            basic_deduction: float = 5000,
                            special_deduction: float = 0,
                            additional_deduction: float = 0,
                            prev_cumulative_deduction: float = 0,
                            prev_cumulative_special: float = 0,
                            prev_cumulative_additional: float = 0):
    """
    累计预扣法计算个税（P1-1: 使用上期累计值而非月份数×月扣除额，避免入职月份不同导致的误差）
    返回: (taxable_income, tax_rate, quick_deduction, cum_tax, tax_this_period,
           cum_deduction, cum_special, cum_additional)
    """
    # 累计收入 = 上期累计 + 本期
    cum_income = prev_cumulative_income + income
    # 累计减除费用 = 上期累计 + 本期
    cum_deduction = prev_cumulative_deduction + basic_deduction
    # 累计专项扣除 = 上期累计 + 本期
    cum_special = prev_cumulative_special + special_deduction
    # 累计专项附加扣除 = 上期累计 + 本期
    cum_additional = prev_cumulative_additional + additional_deduction

    # 应纳税所得额
    taxable = cum_income - cum_deduction - cum_special - cum_additional
    if taxable <= 0:
        taxable = 0

    # 七级超额累进税率表
    brackets = [
        (0, 36000, 0.03, 0),
        (36000, 144000, 0.10, 2520),
        (144000, 300000, 0.20, 16920),
        (300000, 420000, 0.25, 31920),
        (420000, 660000, 0.30, 52920),
        (660000, 960000, 0.35, 85920),
        (960000, float('inf'), 0.45, 181920),
    ]

    rate = 0
    qd = 0
    for lo, hi, r, q in brackets:
        if taxable > lo:
            rate = r
            qd = q
        else:
            break

    cum_tax = taxable * rate - qd
    tax_this_period = max(0, cum_tax - prev_cumulative_tax)

    return taxable, rate, qd, cum_tax, tax_this_period, cum_deduction, cum_special, cum_additional


# ========== API 端点 ==========


@router.get("/periods")
def list_salary_periods(company_id: int = Query(...), db: Session = Depends(get_db)):
    """列出有工资记录的期间"""
    periods = db.query(SalaryRecord.period).filter(
        SalaryRecord.company_id == company_id
    ).distinct().order_by(SalaryRecord.period.desc()).all()
    return [p[0] for p in periods]


@router.get("/records")
def list_salary_records(
    company_id: int = Query(...),
    period: Optional[str] = Query(None),
    keyword: Optional[str] = Query(None),
    db: Session = Depends(get_db)
):
    """查询工资记录列表"""
    q = db.query(SalaryRecord).filter(SalaryRecord.company_id == company_id)
    if period:
        q = q.filter(SalaryRecord.period == period)
    if keyword:
        q = q.filter(or_(
            SalaryRecord.employee_name.contains(keyword),
            SalaryRecord.id_number.contains(keyword)
        ))
    items = q.order_by(SalaryRecord.id.asc()).all()
    return [
        {
            "id": r.id,
            "period": r.period,
            "employee_name": r.employee_name,
            "id_number": r.id_number,
            "id_type": r.id_type,
            "tax_period_start": r.tax_period_start,
            "tax_period_end": r.tax_period_end,
            "income_type": r.income_type,
            "current_income": r.current_income,
            "basic_deduction": r.basic_deduction,
            "pension_insurance": r.pension_insurance,
            "medical_insurance": r.medical_insurance,
            "unemployment_insurance": r.unemployment_insurance,
            "housing_fund": r.housing_fund,
            "special_deduction_total": round(
                (r.pension_insurance or 0) + (r.medical_insurance or 0) +
                (r.unemployment_insurance or 0) + (r.housing_fund or 0) +
                (r.enterprise_annuity or 0) + (r.commercial_health or 0) +
                (r.tax_deferred_pension or 0) + (r.other_special_deduction or 0), 2
            ),
            "child_education": r.child_education,
            "continuing_education": r.continuing_education,
            "housing_loan_interest": r.housing_loan_interest,
            "housing_rent": r.housing_rent,
            "elderly_support": r.elderly_support,
            "infant_care": r.infant_care,
            "major_medical": r.major_medical,
            "additional_deduction_total": round(
                (r.child_education or 0) + (r.continuing_education or 0) +
                (r.housing_loan_interest or 0) + (r.housing_rent or 0) +
                (r.elderly_support or 0) + (r.infant_care or 0) +
                (r.major_medical or 0) + (r.other_additional_deduction or 0), 2
            ),
            "cumulative_income": r.cumulative_income,
            "cumulative_deduction": r.cumulative_deduction,
            "cumulative_special": r.cumulative_special,
            "cumulative_additional": r.cumulative_additional,
            "cumulative_tax_withheld": r.cumulative_tax_withheld,
            "tax_free_income": r.tax_free_income,
            "enterprise_annuity": r.enterprise_annuity,
            "commercial_health": r.commercial_health,
            "tax_deferred_pension": r.tax_deferred_pension,
            "other_special_deduction": r.other_special_deduction,
            "other_additional_deduction": r.other_additional_deduction,
            "other_deduction": r.other_deduction,
            "cumulative_tax_free": r.cumulative_tax_free,
            "cumulative_other": r.cumulative_other,
            "taxable_income": r.taxable_income,
            "tax_rate": r.tax_rate,
            "quick_deduction": r.quick_deduction,
            "tax_payable": r.tax_payable,
            "tax_already_withheld": r.tax_already_withheld,
            "tax_to_pay": r.tax_to_pay,
            "tax_refund": r.tax_refund,
            "net_salary": r.net_salary,
            "created_at": r.created_at.isoformat() if r.created_at else None,
            "updated_at": r.updated_at.isoformat() if r.updated_at else None,
        }
        for r in items
    ]


@router.post("/records")
def create_salary_record(data: SalaryRecordCreate, company_id: int = Query(...), db: Session = Depends(get_db)):
    """新增工资记录"""
    record = SalaryRecord(
        company_id=company_id,
        **data.model_dump()
    )
    db.add(record)
    db.flush()

    # 自动创建人员档案
    _auto_create_employee(db, company_id, data.employee_name, data.id_number)

    db.commit()
    db.refresh(record)
    return {"msg": "ok", "id": record.id}


@router.get("/records/{record_id}")
def get_salary_record(record_id: int, company_id: int = Query(...), db: Session = Depends(get_db)):
    """获取单条工资记录详情"""
    r = db.query(SalaryRecord).filter(
        SalaryRecord.id == record_id,
        SalaryRecord.company_id == company_id
    ).first()
    if not r:
        raise HTTPException(404, "记录不存在")
    return {
        "id": r.id,
        "period": r.period,
        "employee_name": r.employee_name,
        "id_type": r.id_type,
        "id_number": r.id_number,
        "tax_period_start": r.tax_period_start,
        "tax_period_end": r.tax_period_end,
        "income_type": r.income_type,
        "current_income": r.current_income,
        "tax_free_income": r.tax_free_income,
        "basic_deduction": r.basic_deduction,
        "pension_insurance": r.pension_insurance,
        "medical_insurance": r.medical_insurance,
        "unemployment_insurance": r.unemployment_insurance,
        "housing_fund": r.housing_fund,
        "enterprise_annuity": r.enterprise_annuity,
        "commercial_health": r.commercial_health,
        "tax_deferred_pension": r.tax_deferred_pension,
        "other_special_deduction": r.other_special_deduction,
        "child_education": r.child_education,
        "continuing_education": r.continuing_education,
        "housing_loan_interest": r.housing_loan_interest,
        "housing_rent": r.housing_rent,
        "elderly_support": r.elderly_support,
        "infant_care": r.infant_care,
        "major_medical": r.major_medical,
        "other_additional_deduction": r.other_additional_deduction,
        "cumulative_income": r.cumulative_income,
        "cumulative_tax_free": r.cumulative_tax_free,
        "cumulative_deduction": r.cumulative_deduction,
        "cumulative_special": r.cumulative_special,
        "cumulative_additional": r.cumulative_additional,
        "cumulative_other": r.cumulative_other,
        "cumulative_tax_withheld": r.cumulative_tax_withheld,
        "other_deduction": r.other_deduction,
        "taxable_income": r.taxable_income,
        "tax_rate": r.tax_rate,
        "quick_deduction": r.quick_deduction,
        "tax_payable": r.tax_payable,
        "tax_already_withheld": r.tax_already_withheld,
        "tax_to_pay": r.tax_to_pay,
        "tax_refund": r.tax_refund,
        "net_salary": r.net_salary,
        "raw_data": json.loads(r.raw_data) if r.raw_data else None,
        "created_at": r.created_at.isoformat() if r.created_at else None,
        "updated_at": r.updated_at.isoformat() if r.updated_at else None,
    }


@router.put("/records/{record_id}")
def update_salary_record(record_id: int, data: SalaryRecordUpdate, company_id: int = Query(...), db: Session = Depends(get_db)):
    """更新工资记录"""
    r = db.query(SalaryRecord).filter(
        SalaryRecord.id == record_id,
        SalaryRecord.company_id == company_id
    ).first()
    if not r:
        raise HTTPException(404, "记录不存在")
    for k, v in data.model_dump(exclude_none=True).items():
        setattr(r, k, v)
    r.updated_at = datetime.now()
    db.commit()
    return {"msg": "ok"}


@router.delete("/records/{record_id}")
def delete_salary_record(record_id: int, company_id: int = Query(...), db: Session = Depends(get_db)):
    """删除工资记录"""
    r = db.query(SalaryRecord).filter(
        SalaryRecord.id == record_id,
        SalaryRecord.company_id == company_id
    ).first()
    if not r:
        raise HTTPException(404, "记录不存在")
    db.delete(r)
    db.commit()
    return {"msg": "ok"}


@router.post("/records/batch-delete")
def batch_delete_salary_records(ids: List[int], company_id: int = Query(...), db: Session = Depends(get_db)):
    """批量删除"""
    deleted = db.query(SalaryRecord).filter(
        SalaryRecord.company_id == company_id,
        SalaryRecord.id.in_(ids)
    ).delete(synchronize_session=False)
    db.commit()
    return {"msg": f"已删除{deleted}条"}


@router.post("/import")
def import_salary_excel(
    file: UploadFile = File(...),
    company_id: int = Query(...),
    period: str = Query(...),
    db: Session = Depends(get_db)
):
    """
    从税务模板Excel导入工资表
    支持 .xls 和 .xlsx 格式
    自动根据人员信息创建人员档案
    """
    content = file.file.read()

    # 统一读取 Excel，支持 .xls 和 .xlsx
    rows = None
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    except Exception:
        try:
            import xlrd
            wb_xls = xlrd.open_workbook(file_contents=content, encoding_override='gb2312')
            ws_xls = wb_xls.sheet_by_index(0)
            rows = []
            for r in range(ws_xls.nrows):
                row_data = []
                for c in range(ws_xls.ncols):
                    cell = ws_xls.cell(r, c)
                    if cell.ctype == 2:  # number
                        row_data.append(cell.value)
                    elif cell.ctype == 0:  # empty
                        row_data.append(None)
                    else:
                        # text/date/bool/error → 转为字符串，保留空串不转None
                        val = str(cell.value).strip() if cell.value else ""
                        row_data.append(val if val else None)
                rows.append(tuple(row_data))
        except Exception as e:
            raise HTTPException(400, f"无法读取Excel文件: {str(e)}")

    if not rows:
        raise HTTPException(400, "Excel文件为空")

    # ====== .xls 模板显式位置映射（兼容 GBK 编码乱码）======
    POS_MAP = {
        1: "姓名", 2: "证件类型", 3: "证件号码",
        4: "税款所属期起", 5: "税款所属期止", 6: "所得项目",
        7: "本期收入", 8: "免税收入", 9: "基本减除费用",
        10: "基本养老保险", 11: "基本医疗保险", 12: "失业保险", 13: "住房公积金",
        14: "企业年金", 15: "商业健康保险", 16: "税延养老保险", 17: "其他专项扣除",
        18: "累计收入额", 19: "累计免税收入", 20: "累计减除费用",
        21: "累计专项扣除",
        22: "累计子女教育", 23: "累计继续教育",
        24: "累计住房贷款利息", 25: "累计住房租金",
        26: "累计赡养老人", 27: "累计3岁以下婴幼儿照护",
        28: "累计大病医疗", 29: "累计其他扣除",
        31: "其他扣除",
        34: "累计已预扣预缴税额",
        35: "应纳税所得额", 36: "税率", 37: "速算扣除数",
        38: "累计应预扣预缴税额",
        40: "本期应预扣预缴税额", 41: "已缴税额", 42: "应补退税额",
        43: "实发工资",
    }

    # 找到表头行：先按列名匹配，失败则按位置判断
    header_row_idx = None
    for i, row in enumerate(rows):
        if not row:
            continue
        # 尝试1：按中文列名匹配
        if any(cell and ("姓" in str(cell) or "名" in str(cell)) for cell in row):
            header_row_idx = i
            break
        # 尝试2：按位置判断（col1有文字 + col7是数字 = 可能是表头）
        name_col = row[1] if len(row) > 1 else None
        income_col = row[7] if len(row) > 7 else None
        if name_col and isinstance(name_col, str) and len(name_col) > 0 and \
           income_col is not None and isinstance(income_col, str) and any(
               kw in str(income_col) for kw in ["收入", "本期"]):
            header_row_idx = i
            break

    if header_row_idx is None:
        # 最后兜底：检查第一行是否有足够的列
        if len(rows) > 0 and len(rows[0]) >= 8:
            header_row_idx = 0
        else:
            raise HTTPException(400, "未找到表头行，请检查模板格式")

    headers = rows[header_row_idx]

    # 构建列映射：优先按名称匹配，名称匹配不到时用位置映射兜底
    col_map = {}
    for j, h in enumerate(headers):
        if h and isinstance(h, str) and h.strip():
            col_map[h.strip()] = j
    # 位置映射兜底（优先级低于名称匹配）
    for idx, name in POS_MAP.items():
        if name not in col_map and idx < len(headers):
            col_map[name] = idx

    # 数据行：用姓名列是否有值来过滤（不用 col0，因为序号列可能为空）
    name_col_idx = col_map.get("姓名", 1)
    data_rows = []
    for r in rows[header_row_idx + 1:]:
        if not r:
            continue
        # 姓名列有值且不是"合计"=数据行
        name_val = r[name_col_idx] if len(r) > name_col_idx else None
        if name_val and isinstance(name_val, str) and name_val.strip() and "合计" not in str(name_val):
            data_rows.append(r)

    created = 0
    for row in data_rows:
        # 跳过合计行
        name = row[col_map.get("姓名", 1)] if "姓名" in col_map else None
        if not name or "合计" in str(name):
            continue

        id_number = str(row[col_map.get("证件号码", 3)]).strip() if "证件号码" in col_map else ""

        # 读取各列（兼容不同模板列名）
        def _get(col_names, default=0.0):
            for cn in (col_names if isinstance(col_names, list) else [col_names]):
                if cn in col_map:
                    v = row[col_map[cn]]
                    if v is None:
                        return default
                    try:
                        return float(v)
                    except Exception:
                        return default
            return default

        current_income = _get(["本期收入", "收入额"])
        tax_period_start = str(row[col_map["税款所属期起"]]) if "税款所属期起" in col_map and row[col_map["税款所属期起"]] else None
        tax_period_end = str(row[col_map["税款所属期止"]]) if "税款所属期止" in col_map and row[col_map["税款所属期止"]] else None

        # 专项扣除
        pension = _get(["基本养老保险"])
        medical = _get(["基本医疗保险"])
        unemployment = _get(["失业保险"])
        housing = _get(["住房公积金"])

        # 专项附加扣除
        child = _get(["累计子女教育", "子女教育"])
        continuing = _get(["累计继续教育", "继续教育"])
        housing_loan = _get(["累计住房贷款利息", "住房贷款利息"])
        rent = _get(["累计住房租金", "住房租金"])
        elderly = _get(["累计赡养老人", "赡养老人"])
        infant = _get(["累计3岁以下婴幼儿照护", "3岁以下婴幼儿照护"])
        major = _get(["累计大病医疗", "大病医疗"])

        # 累计数据
        cum_income = _get(["累计收入额"])
        cum_deduction = _get(["累计减除费用"])
        cum_special = _get(["累计专项扣除"])
        cum_additional = _get(["累计专项附加扣除"])
        cum_tax_withheld = _get(["累计已预扣预缴税额"])

        # 税额
        taxable = _get(["应纳税所得额"])
        tax_rate = _get(["税率"])
        quick_ded = _get(["速算扣除数"])
        tax_payable = _get(["累计应预扣预缴税额"])
        tax_this = _get(["本期应预扣预缴税额"])
        tax_to_pay_refund = _get(["本期应补(退)税额", "应补(退)税额"])
        net_salary = _get(["实发工资"])

        record = SalaryRecord(
            company_id=company_id,
            period=period,
            employee_name=str(name).strip(),
            id_type=str(row[col_map.get("证件类型", 2)]).strip() if "证件类型" in col_map else "居民身份证",
            id_number=id_number,
            tax_period_start=tax_period_start,
            tax_period_end=tax_period_end,
            income_type=str(row[col_map.get("所得项目", 6)]).strip() if "所得项目" in col_map else "正常工资薪金",
            current_income=current_income,
            pension_insurance=pension,
            medical_insurance=medical,
            unemployment_insurance=unemployment,
            housing_fund=housing,
            child_education=child,
            continuing_education=continuing,
            housing_loan_interest=housing_loan,
            housing_rent=rent,
            elderly_support=elderly,
            infant_care=infant,
            major_medical=major,
            cumulative_income=cum_income,
            cumulative_deduction=cum_deduction,
            cumulative_special=cum_special,
            cumulative_additional=cum_additional,
            cumulative_tax_withheld=cum_tax_withheld,
            taxable_income=taxable,
            tax_rate=tax_rate,
            quick_deduction=quick_ded,
            tax_payable=tax_payable,
            tax_to_pay=tax_this,
            tax_refund=tax_to_pay_refund,
            net_salary=net_salary,
            raw_data=json.dumps([str(c) if c is not None else "" for c in row], ensure_ascii=False),
        )
        db.add(record)
        created += 1

        # 自动创建人员档案
        _auto_create_employee(db, company_id, str(name).strip(), id_number)

    db.commit()
    return {"msg": f"成功导入{created}条工资记录", "count": created}


@router.get("/stats")
def salary_stats(company_id: int = Query(...), period: Optional[str] = Query(None), db: Session = Depends(get_db)):
    """工资统计"""
    q = db.query(SalaryRecord).filter(SalaryRecord.company_id == company_id)
    if period:
        q = q.filter(SalaryRecord.period == period)

    records = q.all()
    total_income = sum(r.current_income or 0 for r in records)
    total_tax = sum(r.tax_to_pay or 0 for r in records)
    total_net = sum(r.net_salary or 0 for r in records)
    count = len(records)

    return {
        "period": period,
        "count": count,
        "total_income": round(total_income, 2),
        "total_tax": round(total_tax, 2),
        "total_net": round(total_net, 2),
        "avg_income": round(total_income / count, 2) if count else 0,
    }


@router.post("/compute")
def compute_salary_tax(
    company_id: int = Query(...),
    period: str = Query(...),
    db: Session = Depends(get_db)
):
    """
    为指定期间的工资记录重新计算个税（累计预扣法）
    需要按员工累计历史数据
    """
    # 获取本期所有记录
    records = db.query(SalaryRecord).filter(
        SalaryRecord.company_id == company_id,
        SalaryRecord.period == period
    ).all()

    if not records:
        return {"msg": "未找到工资记录"}

    updated = 0

    for r in records:
        # 获取该员工历史累计数据（本期之前）
        prev = db.query(SalaryRecord).filter(
            SalaryRecord.company_id == company_id,
            SalaryRecord.id_number == r.id_number,
            SalaryRecord.period < period
        ).order_by(SalaryRecord.period.desc()).first()

        prev_cum_income = prev.cumulative_income if prev else 0
        prev_cum_tax = prev.cumulative_tax_withheld if prev else 0
        prev_cum_deduction = prev.cumulative_deduction if prev else 0
        prev_cum_special = prev.cumulative_special if prev else 0
        prev_cum_additional = prev.cumulative_additional if prev else 0

        # 专项扣除合计（月）
        special_total = (r.pension_insurance or 0) + (r.medical_insurance or 0) + \
                        (r.unemployment_insurance or 0) + (r.housing_fund or 0)
        # 附加扣除合计（月）
        additional_total = (r.child_education or 0) + (r.continuing_education or 0) + \
                           (r.housing_loan_interest or 0) + (r.housing_rent or 0) + \
                           (r.elderly_support or 0) + (r.infant_care or 0) + \
                           (r.major_medical or 0)

        # 累计预扣法计算（使用上期累计扣除值，不再用 months × 月扣除额）
        (taxable, rate, qd, cum_tax, tax_this,
         cum_deduction, cum_special, cum_additional) = _compute_tax_cumulative(
            r.current_income or 0,
            prev_cum_income,
            prev_cum_tax,
            r.basic_deduction or 5000,
            special_total,
            additional_total,
            prev_cumulative_deduction=prev_cum_deduction,
            prev_cumulative_special=prev_cum_special,
            prev_cumulative_additional=prev_cum_additional,
        )

        r.taxable_income = round(taxable, 2)
        r.tax_rate = rate
        r.quick_deduction = qd
        r.tax_payable = round(cum_tax, 2)
        r.tax_to_pay = round(tax_this, 2)
        r.tax_already_withheld = round(tax_this, 2)
        r.cumulative_income = round(prev_cum_income + (r.current_income or 0), 2)
        r.cumulative_deduction = round(cum_deduction, 2)
        r.cumulative_special = round(cum_special, 2)
        r.cumulative_additional = round(cum_additional, 2)
        r.cumulative_tax_withheld = round(cum_tax, 2)
        r.net_salary = round((r.current_income or 0) - tax_this - special_total, 2)
        r.updated_at = datetime.now()
        updated += 1

    db.commit()
    return {"msg": f"已重新计算{updated}条记录的个税"}


@router.post("/auto-create-employees")
def auto_create_employees(company_id: int = Query(...), period: Optional[str] = Query(None), db: Session = Depends(get_db)):
    """根据工资表自动创建人员档案（全量或指定期间）"""
    q = db.query(SalaryRecord).filter(SalaryRecord.company_id == company_id)
    if period:
        q = q.filter(SalaryRecord.period == period)

    records = q.all()
    created = 0
    for r in records:
        emp = _auto_create_employee(db, company_id, r.employee_name, r.id_number)
        if emp:
            created += 1

    db.commit()
    return {"msg": f"已自动创建/更新{created}条人员档案"}
