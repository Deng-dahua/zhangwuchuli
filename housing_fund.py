"""
公积金缴存 API 路由
"""
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from pydantic import BaseModel
from sqlalchemy.orm import Session
from datetime import datetime
import openpyxl
import io
import logging
from typing import Optional, List

from database import HousingFundDetail, Company, get_db, _generate_hf_accrual_journals, _match_hf_payment_journals

router = APIRouter(prefix="/api/housing-fund", tags=["公积金缴存"])


# ============ CRUD ============

@router.get("/details")
def list_details(
    company_id: int,
    period: str = Query(None),
    db: Session = Depends(get_db),
):
    """列表查询"""
    q = db.query(HousingFundDetail).filter(
        HousingFundDetail.company_id == company_id
    )
    if period:
        q = q.filter(HousingFundDetail.period == period)
    items = q.order_by(HousingFundDetail.id).all()
    result = []
    for item in items:
        result.append({
            "id": item.id,
            "company_id": item.company_id,
            "period": item.period,
            "employee_id": item.employee_id,
            "employee_name": item.employee_name,
            "id_number": item.id_number,
            "deposit_base": item.deposit_base,
            "company_ratio": item.company_ratio,
            "personal_ratio": item.personal_ratio,
            "total_amount": item.total_amount,
            "company_amount": item.company_amount,
            "personal_amount": item.personal_amount,
            "status": item.status,
            "created_at": item.created_at.isoformat() if item.created_at else None,
            "updated_at": item.updated_at.isoformat() if item.updated_at else None,
        })
    return {"items": result, "total": len(result)}


@router.get("/details/{detail_id}")
def get_detail(detail_id: int, company_id: int, db: Session = Depends(get_db)):
    """获取单条详情"""
    item = db.query(HousingFundDetail).filter(
        HousingFundDetail.id == detail_id,
        HousingFundDetail.company_id == company_id
    ).first()
    if not item:
        raise HTTPException(404, "记录不存在")
    return {
        "id": item.id,
        "company_id": item.company_id,
        "period": item.period,
        "employee_id": item.employee_id,
        "employee_name": item.employee_name,
        "id_number": item.id_number,
        "deposit_base": item.deposit_base,
        "company_ratio": item.company_ratio,
        "personal_ratio": item.personal_ratio,
        "total_amount": item.total_amount,
        "company_amount": item.company_amount,
        "personal_amount": item.personal_amount,
        "status": item.status,
        "created_at": item.created_at.isoformat() if item.created_at else None,
    }


@router.post("/details")
def create_detail(
    company_id: int,
    period: str,
    employee_id: str = "",
    employee_name: str = "",
    id_number: str = "",
    deposit_base: float = 0,
    company_ratio: float = 0,
    personal_ratio: float = 0,
    status: str = "正常",
    db: Session = Depends(get_db),
):
    """新增一条缴存记录（自动计算缴存额）"""
    if 0 < company_ratio <= 1:
        logging.warning(f"[公积金] 员工{employee_name} 单位缴存比例自动从小数 {company_ratio} 转为百分比 {company_ratio*100}%")
        company_ratio *= 100
    if 0 < personal_ratio <= 1:
        logging.warning(f"[公积金] 员工{employee_name} 个人缴存比例自动从小数 {personal_ratio} 转为百分比 {personal_ratio*100}%")
        personal_ratio *= 100
    company_amount = round(deposit_base * company_ratio / 100, 2)
    personal_amount = round(deposit_base * personal_ratio / 100, 2)
    total_amount = round(company_amount + personal_amount, 2)

    detail = HousingFundDetail(
        company_id=company_id,
        period=period,
        employee_id=employee_id,
        employee_name=employee_name,
        id_number=id_number,
        deposit_base=deposit_base,
        company_ratio=company_ratio,
        personal_ratio=personal_ratio,
        total_amount=total_amount,
        company_amount=company_amount,
        personal_amount=personal_amount,
        status=status,
    )
    db.add(detail)
    db.commit()
    return {"id": detail.id, "message": "保存成功", "total_amount": total_amount}


@router.put("/details/{detail_id}")
def update_detail(
    detail_id: int,
    company_id: int,
    employee_id: str = None,
    employee_name: str = None,
    id_number: str = None,
    deposit_base: float = None,
    company_ratio: float = None,
    personal_ratio: float = None,
    status: str = None,
    db: Session = Depends(get_db),
):
    """更新一条缴存记录"""
    item = db.query(HousingFundDetail).filter(
        HousingFundDetail.id == detail_id,
        HousingFundDetail.company_id == company_id
    ).first()
    if not item:
        raise HTTPException(404, "记录不存在")

    if employee_id is not None:
        item.employee_id = employee_id
    if employee_name is not None:
        item.employee_name = employee_name
    if id_number is not None:
        item.id_number = id_number
    if deposit_base is not None:
        item.deposit_base = deposit_base
    if company_ratio is not None:
        item.company_ratio = company_ratio
    if personal_ratio is not None:
        item.personal_ratio = personal_ratio
    if status is not None:
        item.status = status

    # 重新计算缴存额
    cr = (item.company_ratio or 0)
    pr = (item.personal_ratio or 0)
    if 0 < cr <= 1:
        cr *= 100
        item.company_ratio = cr
    if 0 < pr <= 1:
        pr *= 100
        item.personal_ratio = pr
    item.company_amount = round((item.deposit_base or 0) * cr / 100, 2)
    item.personal_amount = round((item.deposit_base or 0) * pr / 100, 2)
    item.total_amount = round(item.company_amount + item.personal_amount, 2)
    item.updated_at = datetime.now()

    db.commit()
    return {"id": item.id, "message": "更新成功", "total_amount": item.total_amount}


@router.delete("/details/{detail_id}")
def delete_detail(detail_id: int, company_id: int, db: Session = Depends(get_db)):
    """删除一条缴存记录"""
    item = db.query(HousingFundDetail).filter(
        HousingFundDetail.id == detail_id,
        HousingFundDetail.company_id == company_id
    ).first()
    if not item:
        raise HTTPException(404, "记录不存在")
    db.delete(item)
    db.commit()
    return {"message": "删除成功"}


class BatchDeleteRequest(BaseModel):
    ids: List[int]


@router.post("/details/batch-delete")
def batch_delete(data: BatchDeleteRequest, company_id: int, db: Session = Depends(get_db)):
    """批量删除"""
    deleted = db.query(HousingFundDetail).filter(
        HousingFundDetail.id.in_(data.ids),
        HousingFundDetail.company_id == company_id
    ).delete(synchronize_session=False)
    db.commit()
    return {"deleted": deleted}


# ============ 统计 ============

@router.get("/stats")
def get_stats(company_id: int, period: str = Query(None), db: Session = Depends(get_db)):
    """获取统计信息"""
    q = db.query(HousingFundDetail).filter(
        HousingFundDetail.company_id == company_id
    )
    if period:
        q = q.filter(HousingFundDetail.period == period)

    items = q.all()
    person_count = len(items)
    total_company = sum(item.company_amount or 0 for item in items)
    total_personal = sum(item.personal_amount or 0 for item in items)
    total_all = sum(item.total_amount or 0 for item in items)

    return {
        "person_count": person_count,
        "total_company_amount": round(total_company, 2),
        "total_personal_amount": round(total_personal, 2),
        "total_amount": round(total_all, 2),
    }


# ============ Excel 导入 ============

@router.post("/import")
async def import_excel(
    company_id: int,
    period: str = Query(...),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    """从 Excel 导入公积金缴存明细"""
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(400, "仅支持 .xlsx / .xls 文件")

    try:
        contents = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
        ws = wb.active
    except Exception as e:
        raise HTTPException(400, f"文件读取失败: {str(e)}")

    rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))
    if len(rows) < 2:
        raise HTTPException(400, "文件格式不正确：行数不足")

    # 表头映射
    header = rows[0]
    col_map = {}
    field_names = {
        "工号": "employee_id", "姓名": "employee_name", "身份证号": "id_number",
        "缴存基数": "deposit_base", "单位缴存比例": "company_ratio",
        "个人缴存比例": "personal_ratio", "缴存额": "total_amount",
        "单位缴存额": "company_amount", "个人缴存额": "personal_amount",
    }
    for i, h in enumerate(header):
        h_str = str(h or "").strip()
        if h_str in field_names:
            col_map[field_names[h_str]] = i

    if "employee_name" not in col_map:
        raise HTTPException(400, f"未找到'姓名'列，表头: {[str(h) for h in header]}")

    imported = 0
    skipped = 0
    errors = []
    auto_corrected_ratios = 0

    for i, row in enumerate(rows[1:], start=2):
        name = str(row[col_map["employee_name"]] or "").strip() if "employee_name" in col_map else ""
        if not name:
            continue

        try:
            emp_id = str(row[col_map["employee_id"]] or "").strip() if "employee_id" in col_map else ""
            id_num = str(row[col_map["id_number"]] or "").strip() if "id_number" in col_map else ""
            deposit_base = float(row[col_map["deposit_base"]] or 0) if "deposit_base" in col_map else 0
            company_ratio = float(row[col_map["company_ratio"]] or 0) if "company_ratio" in col_map else 0
            personal_ratio = float(row[col_map["personal_ratio"]] or 0) if "personal_ratio" in col_map else 0

            # 自动检测小数比例（如 0.1=10%），转换为百分比
            if 0 < company_ratio <= 1:
                auto_corrected_ratios += 1
                logging.warning(f"[公积金导入] 第{i}行 {name}: 单位比例从小数 {company_ratio} 自动转为 {company_ratio*100}%")
                company_ratio *= 100
            if 0 < personal_ratio <= 1:
                auto_corrected_ratios += 1
                logging.warning(f"[公积金导入] 第{i}行 {name}: 个人比例从小数 {personal_ratio} 自动转为 {personal_ratio*100}%")
                personal_ratio *= 100

            company_amount = round(deposit_base * company_ratio / 100, 2)
            personal_amount = round(deposit_base * personal_ratio / 100, 2)
            total_amount = round(company_amount + personal_amount, 2)

            detail = HousingFundDetail(
                company_id=company_id,
                period=period,
                employee_id=emp_id,
                employee_name=name,
                id_number=id_num,
                deposit_base=deposit_base,
                company_ratio=company_ratio,
                personal_ratio=personal_ratio,
                total_amount=total_amount,
                company_amount=company_amount,
                personal_amount=personal_amount,
                status="正常",
            )
            db.add(detail)
            imported += 1
        except Exception as e:
            errors.append(f"第{i}行 {name}: {str(e)}")

    db.commit()

    msg = f"成功导入 {imported} 条记录"
    if auto_corrected_ratios > 0:
        msg += f"（已自动将 {auto_corrected_ratios} 个小数比例转为百分比）"

    return {
        "imported": imported,
        "skipped": skipped,
        "errors": errors[:10],
        "message": msg,
    }


# ========== 凭证生成 ==========

@router.post("/generate-accrual")
def generate_hf_accrual(
    company_id: int = Query(...),
    period: str = Query(...),
    db: Session = Depends(get_db),
):
    """
    为指定期间的公积金明细生成计提凭证。
    借：管理费用-住房公积金（单位部分）
    贷：应付职工薪酬-住房公积金（单位部分）
    """
    result = _generate_hf_accrual_journals(db, company_id, period)
    db.commit()
    return result


@router.post("/match-payment")
def match_hf_payment(
    company_id: int = Query(...),
    db: Session = Depends(get_db),
):
    """
    匹配银行流水与公积金明细，生成缴纳凭证。
    借：应付职工薪酬-住房公积金（单位+个人）
    贷：银行存款
    """
    result = _match_hf_payment_journals(db, company_id)
    db.commit()
    return result
