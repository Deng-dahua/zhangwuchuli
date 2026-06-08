"""
社保申报 API 路由
使用 FastAPI APIRouter，在 main.py 中 include_router 加载。
"""
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from sqlalchemy.orm import Session
from datetime import datetime
import json
import openpyxl
import io
from typing import Optional, List

from database import (
    SocialSecurityDeclaration, SocialSecurityDetail, Company, get_db,
    _generate_ss_accrual_journals, _match_ss_payment_journals,
)

router = APIRouter(prefix="/api/social-security", tags=["社保申报"])

# 保险项目名称映射（Excel表头 → 中文名）
INSURANCE_NAMES = [
    "基本养老保险（单位）",
    "基本养老保险（个人）",
    "基本医疗保险（单位）",
    "地方补充医疗（单位）",
    "基本医疗保险（个人）",
    "生育保险",
    "职业年金（单位）",
    "职业年金（个人）",
    "工伤保险（单位）",
    "公务员医疗补助",
    "家属统筹医疗（单位）",
    "家属统筹医疗（个人）",
    "失业保险（单位）",
    "失业保险（个人）",
    "地方补充养老（单位）",
]


# ============ CRUD ============

@router.get("/declarations")
def list_declarations(
    company_id: int,
    period: str = Query(None),
    period_from: str = Query(None),
    period_to: str = Query(None),
    db: Session = Depends(get_db),
):
    """列表查询"""
    q = db.query(SocialSecurityDeclaration).filter(
        SocialSecurityDeclaration.company_id == company_id
    )
    if period:
        q = q.filter(SocialSecurityDeclaration.period == period)
    if period_from:
        q = q.filter(SocialSecurityDeclaration.period >= period_from)
    if period_to:
        q = q.filter(SocialSecurityDeclaration.period <= period_to)
    items = q.order_by(SocialSecurityDeclaration.period.desc()).all()
    result = []
    for item in items:
        detail_count = db.query(SocialSecurityDetail).filter(
            SocialSecurityDetail.declaration_id == item.id
        ).count()
        result.append({
            "id": item.id,
            "company_id": item.company_id,
            "period": item.period,
            "status": item.status,
            "note": item.note,
            "detail_count": detail_count,
            "created_at": item.created_at.isoformat() if item.created_at else None,
            "updated_at": item.updated_at.isoformat() if item.updated_at else None,
        })
    return {"items": result, "total": len(result)}


@router.get("/declarations/{declaration_id}")
def get_declaration(declaration_id: int, company_id: int, db: Session = Depends(get_db)):
    """获取申报详情（含明细）"""
    decl = db.query(SocialSecurityDeclaration).filter(
        SocialSecurityDeclaration.id == declaration_id,
        SocialSecurityDeclaration.company_id == company_id
    ).first()
    if not decl:
        raise HTTPException(404, "申报记录不存在")

    details = db.query(SocialSecurityDetail).filter(
        SocialSecurityDetail.declaration_id == declaration_id
    ).order_by(SocialSecurityDetail.seq).all()

    return {
        "id": decl.id,
        "company_id": decl.company_id,
        "period": decl.period,
        "status": decl.status,
        "note": decl.note,
        "created_at": decl.created_at.isoformat() if decl.created_at else None,
        "details": [{
            "id": d.id,
            "seq": d.seq,
            "employee_name": d.employee_name,
            "id_number": d.id_number,
            "period_start": d.period_start,
            "period_end": d.period_end,
            "total_amount": d.total_amount,
            "personal_amount": d.personal_amount,
            "company_amount": d.company_amount,
            "salary_base": d.salary_base,
            "category": d.category,
            "insurance_items": json.loads(d.insurance_items) if d.insurance_items else [],
        } for d in details]
    }


@router.post("/declarations")
def create_declaration(
    company_id: int,
    period: str,
    details: list = None,
    note: str = None,
    status: str = "草稿",
    db: Session = Depends(get_db),
):
    """创建或更新申报（按期间去重：若已存在则更新）"""
    existing = db.query(SocialSecurityDeclaration).filter(
        SocialSecurityDeclaration.company_id == company_id,
        SocialSecurityDeclaration.period == period
    ).first()

    if existing:
        # 更新现有记录
        if note is not None:
            existing.note = note
        if status is not None:
            existing.status = status
        existing.updated_at = datetime.now()
        decl = existing
        # 删除旧明细
        db.query(SocialSecurityDetail).filter(
            SocialSecurityDetail.declaration_id == decl.id
        ).delete()
        db.flush()
    else:
        decl = SocialSecurityDeclaration(
            company_id=company_id,
            period=period,
            status=status or "草稿",
            note=note,
        )
        db.add(decl)
        db.flush()

    # 添加明细
    if details:
        for d in details:
            items_json = json.dumps(d.get("insurance_items", []), ensure_ascii=False)
            detail = SocialSecurityDetail(
                declaration_id=decl.id,
                seq=d.get("seq", 0),
                employee_name=d.get("employee_name", ""),
                id_number=d.get("id_number", ""),
                period_start=d.get("period_start", period),
                period_end=d.get("period_end", period),
                total_amount=d.get("total_amount", 0),
                personal_amount=d.get("personal_amount", 0),
                company_amount=d.get("company_amount", 0),
                salary_base=d.get("salary_base", 0),
                category=d.get("category", "在职人员"),
                insurance_items=items_json,
            )
            db.add(detail)

    db.commit()

    # 自动生成社保计提凭证（使用 savepoint 保证一致性）
    try:
        result = _generate_ss_accrual_journals(db, company_id, decl.id)
    except Exception:
        db.rollback()
        raise HTTPException(500, "社保计提凭证生成失败，申报记录已回滚")

    return {"id": decl.id, "message": "保存成功", "journal": result}


@router.delete("/declarations/{declaration_id}")
def delete_declaration(declaration_id: int, company_id: int, db: Session = Depends(get_db)):
    """删除申报记录及其明细"""
    decl = db.query(SocialSecurityDeclaration).filter(
        SocialSecurityDeclaration.id == declaration_id,
        SocialSecurityDeclaration.company_id == company_id
    ).first()
    if not decl:
        raise HTTPException(404, "申报记录不存在")
    # 级联删除明细
    db.query(SocialSecurityDetail).filter(
        SocialSecurityDetail.declaration_id == declaration_id
    ).delete()
    db.delete(decl)
    db.commit()
    return {"message": "删除成功"}


# ============ 统计 ============

@router.get("/stats")
def get_stats(company_id: int, period: str = Query(None), db: Session = Depends(get_db)):
    """获取统计信息"""
    q = db.query(SocialSecurityDeclaration).filter(
        SocialSecurityDeclaration.company_id == company_id
    )
    if period:
        q = q.filter(SocialSecurityDeclaration.period == period)

    declarations = q.all()
    total_company = 0.0
    total_personal = 0.0
    detail_count = 0
    for decl in declarations:
        details = db.query(SocialSecurityDetail).filter(
            SocialSecurityDetail.declaration_id == decl.id
        ).all()
        detail_count += len(details)
        for d in details:
            total_company += d.company_amount or 0
            total_personal += d.personal_amount or 0

    return {
        "total_declarations": len(declarations),
        "total_details": detail_count,
        "total_company_amount": round(total_company, 2),
        "total_personal_amount": round(total_personal, 2),
        "total_amount": round(total_company + total_personal, 2),
    }


# ============ Excel 导入 ============

@router.post("/import")
async def import_excel(
    company_id: int,
    period: str = Query(...),
    declaration_id: int = Query(None),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    """从日常申报明细表 Excel 导入数据"""
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(400, "仅支持 .xlsx / .xls 文件")

    try:
        contents = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
        ws = wb.active
    except Exception as e:
        raise HTTPException(400, f"文件读取失败: {str(e)}")

    rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))
    if len(rows) < 4:
        raise HTTPException(400, "文件格式不正确：行数不足")

    # 解析保险项目列索引（从第9列开始，每2列一组：费率+应缴费额）
    num_insurance = len(INSURANCE_NAMES)

    details = []
    current_category = "在职人员"
    seq = 0
    errors = []

    for i, row in enumerate(rows):
        if i < 3:
            continue  # 跳过前三行表头

        if row[0] is None and row[1] is None:
            continue  # 空行

        # 检测分类行（如"在职人员"、"退休人员"等）
        first_cell = str(row[0] or "").strip()
        second_cell = str(row[1] or "").strip()
        if first_cell in ("在职人员", "退休人员", "家属统筹人员"):
            current_category = first_cell
            continue

        # 检测小计/合计行
        if second_cell in ("小计", "合计") or first_cell in ("", "合计"):
            continue

        # 跳过空行
        emp_name = str(row[1] or "").strip()
        if not emp_name:
            continue

        try:
            seq += 1
            id_number = str(row[2] or "").strip()
            period_start = str(row[3] or "").strip()
            period_end = str(row[4] or "").strip()

            total_amount = float(row[5] or 0)
            personal_amount = float(row[6] or 0)
            company_amount = float(row[7] or 0)
            salary_base = float(row[8] or 0)

            # 解析保险项目（第9列开始，每2列：费率+金额）
            insurance_items = []
            for j in range(num_insurance):
                col_idx = 9 + j * 2
                rate = ""
                if col_idx < len(row) and row[col_idx] is not None:
                    rate = str(row[col_idx]).strip()
                amount = 0.0
                if col_idx + 1 < len(row) and row[col_idx + 1] is not None:
                    try:
                        amount = float(row[col_idx + 1])
                    except (ValueError, TypeError):
                        pass
                if amount > 0 or rate:
                    insurance_items.append({
                        "name": INSURANCE_NAMES[j] if j < len(INSURANCE_NAMES) else f"保险项{j+1}",
                        "rate": rate,
                        "amount": amount,
                    })

            details.append({
                "seq": seq,
                "employee_name": emp_name,
                "id_number": id_number,
                "period_start": period_start if period_start else period,
                "period_end": period_end if period_end else period,
                "total_amount": total_amount,
                "personal_amount": personal_amount,
                "company_amount": company_amount,
                "salary_base": salary_base,
                "category": current_category,
                "insurance_items": insurance_items,
            })
        except Exception as e:
            errors.append(f"第{i+1}行 {emp_name}: {str(e)}")

    if not details:
        return {"imported": 0, "errors": errors, "message": "未识别到有效数据"}

    # 创建/更新申报记录
    if declaration_id:
        decl = db.query(SocialSecurityDeclaration).filter(
            SocialSecurityDeclaration.id == declaration_id,
            SocialSecurityDeclaration.company_id == company_id
        ).first()
        if not decl:
            raise HTTPException(404, "申报记录不存在")
        # 追加模式：不清空旧明细
        existing_count = db.query(SocialSecurityDetail).filter(
            SocialSecurityDetail.declaration_id == decl.id
        ).count()
        seq = existing_count
        decl.updated_at = datetime.now()
    else:
        # 按公司+期间查找（不删除旧记录，改为追加）
        decl = db.query(SocialSecurityDeclaration).filter(
            SocialSecurityDeclaration.company_id == company_id,
            SocialSecurityDeclaration.period == period
        ).first()
        if decl:
            existing_count = db.query(SocialSecurityDetail).filter(
                SocialSecurityDetail.declaration_id == decl.id
            ).count()
            seq = existing_count
            decl.updated_at = datetime.now()
        else:
            decl = SocialSecurityDeclaration(
                company_id=company_id, period=period, status="已确认"
            )
            db.add(decl)
            db.flush()
            seq = 0

    for d in details:
        seq += 1
        d["seq"] = seq
        detail = SocialSecurityDetail(
            declaration_id=decl.id,
            seq=d["seq"],
            employee_name=d["employee_name"],
            id_number=d["id_number"],
            period_start=d["period_start"],
            period_end=d["period_end"],
            total_amount=d["total_amount"],
            personal_amount=d["personal_amount"],
            company_amount=d["company_amount"],
            salary_base=d["salary_base"],
            category=d["category"],
            insurance_items=json.dumps(d["insurance_items"], ensure_ascii=False),
        )
        db.add(detail)

    db.commit()

    # 自动生成社保计提凭证
    result = _generate_ss_accrual_journals(db, company_id, decl.id)

    return {
        "imported": len(details),
        "total": len(details),
        "declaration_id": decl.id,
        "errors": errors[:10],
        "message": f"成功导入 {len(details)} 条记录",
        "journal": result
    }


# ========== 生成凭证 ==========

@router.post("/declarations/{declaration_id}/details")
def add_detail(declaration_id: int, company_id: int, payload: dict, db: Session = Depends(get_db)):
    """新增参保人员明细"""
    decl = db.query(SocialSecurityDeclaration).filter(
        SocialSecurityDeclaration.id == declaration_id,
        SocialSecurityDeclaration.company_id == company_id
    ).first()
    if not decl:
        raise HTTPException(404, "申报记录不存在")

    # 计算总金额、个人合计、单位合计
    insurance_items = payload.get("insurance_items", [])
    personal_amount = sum(i.get("amount", 0) for i in insurance_items if i.get("type") == "personal")
    company_amount = sum(i.get("amount", 0) for i in insurance_items if i.get("type") == "unit")
    total_amount = personal_amount + company_amount

    # 自动设置 seq
    max_seq = db.query(SocialSecurityDetail).filter(
        SocialSecurityDetail.declaration_id == declaration_id
    ).count()

    detail = SocialSecurityDetail(
        declaration_id=declaration_id,
        seq=max_seq + 1,
        employee_name=payload.get("employee_name", ""),
        id_number=payload.get("id_number", ""),
        period_start=payload.get("period_start", decl.period),
        period_end=payload.get("period_end", decl.period),
        total_amount=total_amount,
        personal_amount=personal_amount,
        company_amount=company_amount,
        salary_base=payload.get("salary_base", 0),
        category=payload.get("category", "在职人员"),
        insurance_items=json.dumps(insurance_items, ensure_ascii=False),
    )
    db.add(detail)
    decl.updated_at = datetime.now()
    db.commit()
    db.refresh(detail)
    return {"id": detail.id, "message": "新增成功"}


@router.put("/details/{detail_id}")
def update_detail(detail_id: int, company_id: int, payload: dict, db: Session = Depends(get_db)):
    """更新参保人员明细"""
    detail = db.query(SocialSecurityDetail).filter(
        SocialSecurityDetail.id == detail_id
    ).first()
    if not detail:
        raise HTTPException(404, "明细记录不存在")

    # 校验公司归属
    decl = db.query(SocialSecurityDeclaration).filter(
        SocialSecurityDeclaration.id == detail.declaration_id,
        SocialSecurityDeclaration.company_id == company_id
    ).first()
    if not decl:
        raise HTTPException(403, "无权限操作此记录")

    insurance_items = payload.get("insurance_items", [])
    personal_amount = sum(i.get("amount", 0) for i in insurance_items if i.get("type") == "personal")
    company_amount = sum(i.get("amount", 0) for i in insurance_items if i.get("type") == "unit")
    total_amount = personal_amount + company_amount

    detail.employee_name = payload.get("employee_name", detail.employee_name)
    detail.id_number = payload.get("id_number", detail.id_number)
    detail.period_start = payload.get("period_start", detail.period_start)
    detail.period_end = payload.get("period_end", detail.period_end)
    detail.total_amount = total_amount
    detail.personal_amount = personal_amount
    detail.company_amount = company_amount
    detail.salary_base = payload.get("salary_base", detail.salary_base)
    detail.category = payload.get("category", detail.category)
    detail.insurance_items = json.dumps(insurance_items, ensure_ascii=False)
    decl.updated_at = datetime.now()
    db.commit()
    return {"message": "更新成功"}


@router.delete("/details/{detail_id}")
def delete_detail(detail_id: int, company_id: int, db: Session = Depends(get_db)):
    """删除参保人员明细"""
    detail = db.query(SocialSecurityDetail).filter(
        SocialSecurityDetail.id == detail_id
    ).first()
    if not detail:
        raise HTTPException(404, "明细记录不存在")

    decl = db.query(SocialSecurityDeclaration).filter(
        SocialSecurityDeclaration.id == detail.declaration_id,
        SocialSecurityDeclaration.company_id == company_id
    ).first()
    if not decl:
        raise HTTPException(403, "无权限操作此记录")

    db.delete(detail)
    decl.updated_at = datetime.now()
    db.commit()
    return {"message": "删除成功"}


@router.post("/generate-payment-journals")
def generate_ss_payment_journals(company_id: int = Query(...), db: Session = Depends(get_db)):
    """手动触发社保缴纳凭证匹配（银行流水 → 社保缴纳凭证）"""
    result = _match_ss_payment_journals(db, company_id)
    db.commit()
    return result


@router.post("/declarations/{declaration_id}/generate-accrual-journal")
def generate_ss_accrual_journal(declaration_id: int, company_id: int = Query(...), db: Session = Depends(get_db)):
    """手动触发社保计提凭证生成"""
    result = _generate_ss_accrual_journals(db, company_id, declaration_id)
    db.commit()
    return result
