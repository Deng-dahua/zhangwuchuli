"""
财税系统 - AI 智能助手对话引擎
支持：税务政策问答、账务处理指导、财务管理咨询、风险预警提示
LLM 驱动 + 本地规则引擎双模
"""
from fastapi import APIRouter, Depends, Query, UploadFile, File, Form
from fastapi.responses import HTMLResponse
from sqlalchemy.orm import Session
from pydantic import BaseModel
from typing import Optional, List, Dict, Any
from datetime import date, datetime
import os
import csv
import io
import re
import uuid
import json
import openpyxl
import httpx
from pypdf import PdfReader

from database import (
    get_db, Customer, Supplier, Employee, JournalEntry, Account,
    Company
)

router = APIRouter()

# ==================== LLM 配置 ====================

# OpenAI 兼容 API 配置（支持 OpenAI / DeepSeek / Ollama / 本地模型等）
LLM_CONFIG = {
    "api_base": os.environ.get("TAX_LLM_BASE", "https://api.deepseek.com/v1"),
    "api_key": os.environ.get("TAX_LLM_KEY", ""),
    "model": os.environ.get("TAX_LLM_MODEL", "deepseek-chat"),
    "max_tokens": 2000,
    "temperature": 0.3,
    "timeout": 30,
}

TAX_SYSTEM_PROMPT = """你是「存勤财税助手」，一家专注为中小制造企业提供财税服务的 AI 专家。

你的专业领域：
1. **财务管理**：成本核算、费用控制、资金管理、预算编制、财务分析
2. **账务处理**：会计分录编制、凭证处理、科目运用、期末结转、报表编制
3. **税务政策**：增值税、企业所得税、个人所得税、印花税、城建税及附加等
4. **风险提示**：税务稽查风险、发票合规风险、经营实质风险、税负异常预警
5. **实务操作**：纳税申报流程、发票开具规范、社保公积金处理、工商年报

回答要求：
- 回答简洁专业，先说结论再展开
- 涉及具体金额/税率时，注明依据和注意事项
- 遇到不确定的政策，建议用户咨询当地税务机关
- 用中文回答，适当使用表格和分点增强可读性
- 每轮回答控制在 500 字以内，复杂问题可分步解答

重要纪律：
- 不做税务筹划建议，只做政策解读和风险提示
- 不替用户做决策，只提供专业分析和选项
- 涉及具体企业数据的问题，先说明「根据你提供的信息分析」"""


def _build_context_prompt(company_id: int, db: Session) -> str:
    """构建当前系统上下文信息，注入到 LLM prompt 中"""
    ctx_parts = []
    try:
        company = db.query(Company).filter(Company.id == company_id).first()
        if company:
            name = company.name or "未设置"
            ctype = company.company_type or "未填写"
            scope = (company.business_scope or "")[:100]
            ctx_parts.append(f"当前企业：{name}（类型：{ctype}）")
            if scope.strip():
                ctx_parts.append(f"经营范围：{scope}")
        
        # 基本统计
        voucher_count = db.query(JournalEntry).filter(
            JournalEntry.company_id == company_id
        ).count()
        account_count = db.query(Account).filter(
            Account.company_id == company_id
        ).count()
        if voucher_count > 0:
            ctx_parts.append(f"系统已有 {voucher_count} 条凭证、{account_count} 个科目")
    except Exception:
        pass
    
    if ctx_parts:
        return "【当前系统信息】\n" + "\n".join(ctx_parts)
    return ""


def _call_llm(user_message: str, system_prompt: str = None, context: str = "") -> Optional[str]:
    """调用 LLM API 进行问答"""
    if not LLM_CONFIG["api_key"]:
        return None
    
    try:
        messages = []
        full_system = (system_prompt or TAX_SYSTEM_PROMPT)
        if context:
            full_system += "\n\n" + context
        messages.append({"role": "system", "content": full_system})
        messages.append({"role": "user", "content": user_message})
        
        headers = {
            "Authorization": f"Bearer {LLM_CONFIG['api_key']}",
            "Content-Type": "application/json",
        }
        payload = {
            "model": LLM_CONFIG["model"],
            "messages": messages,
            "max_tokens": LLM_CONFIG["max_tokens"],
            "temperature": LLM_CONFIG["temperature"],
        }
        
        client = httpx.Client(timeout=LLM_CONFIG["timeout"])
        resp = client.post(
            f"{LLM_CONFIG['api_base'].rstrip('/')}/chat/completions",
            headers=headers,
            json=payload,
        )
        resp.raise_for_status()
        data = resp.json()
        return data["choices"][0]["message"]["content"].strip()
    except Exception as e:
        return None


def _chat_tax_qa(user_message: str, company_id: int, db: Session) -> str:
    """财税问答：先尝试 LLM，失败则回退到规则引擎"""
    context = _build_context_prompt(company_id, db)
    reply = _call_llm(user_message, TAX_SYSTEM_PROMPT, context)
    
    if reply:
        return reply
    
    # LLM 不可用时的回退规则引擎
    msg_lower = user_message.strip().lower()
    
    # 税务政策常见问答
    tax_qa = {
        r"增值税.*税率|税率.*多少|几个点": (
            "📋 **增值税税率速查**\n\n"
            "| 税率 | 适用范围 |\n"
            "|------|----------|\n"
            "| 13% | 销售货物、加工修理修配、有形动产租赁 |\n"
            "| 9% | 农产品、自来水、暖气、图书、运输、邮政、建筑、不动产租赁 |\n"
            "| 6% | 现代服务业、金融服务、生活服务、无形资产转让 |\n"
            "| 3% | 小规模纳税人（2023-2027减按1%） |\n\n"
            "⚠️ 具体适用税率需根据你企业的纳税人身份和业务类型确定。"
        ),
        r"小规模.*免税|小规模.*起征点|小规模.*额度|小规模.*优惠|小规模.*政策": (
            "📋 **小规模纳税人增值税优惠政策**\n\n"
            "• 月销售额 10 万元以下（季度 30 万元以下）免征增值税\n"
            "• 适用 3% 征收率的应税销售收入，减按 1% 征收\n"
            "• 政策有效期至 2027 年 12 月 31 日\n\n"
            "💡 注意：专票部分不享受免税，需按适用税率缴税。"
        ),
        r"企业所得税.*税率|所得税.*多少|企业所得税.*怎么|所得税.*怎么算": (
            "📋 **企业所得税税率**\n\n"
            "• 基本税率：25%\n"
            "• 小型微利企业：年应纳税所得额≤300万\n"
            "  - 0-300万部分：减按 25% 计入，按 20% 税率（实际 5%）\n"
            "• 高新技术企业：15%\n"
            "• 西部大开发鼓励类：15%\n\n"
            "💡 小型微利企业需同时满足：资产≤5000万、人数≤300、所得≤300万。"
        ),
        r"印花税|印花税.*税率|合同.*印花": (
            "📋 **印花税主要税目税率**\n\n"
            "| 税目 | 税率 |\n"
            "|------|------|\n"
            "| 购销合同 | 0.3‰ |\n"
            "| 加工承揽合同 | 0.5‰ |\n"
            "| 建设工程勘察设计 | 0.5‰ |\n"
            "| 建筑安装工程承包 | 0.3‰ |\n"
            "| 财产租赁合同 | 1‰ |\n"
            "| 货物运输合同 | 0.5‰ |\n"
            "| 借款合同 | 0.05‰ |\n"
            "| 营业账簿（实收资本+资本公积） | 0.25‰ |\n\n"
            "💡 2022年7月起施行《印花税法》，部分税目有调整。"
        ),
        r"个税.*计算|个人所得税.*怎么算|工资.*个税": (
            "📋 **工资薪金个人所得税（累计预扣法）**\n\n"
            "应纳税所得额 = 累计收入 - 累计减除费用(5000/月) - 累计专项扣除 - 累计专项附加扣除\n\n"
            "| 级数 | 累计预扣应纳税所得额 | 税率 | 速算扣除 |\n"
            "|------|---------------------|------|----------|\n"
            "| 1 | ≤36,000 | 3% | 0 |\n"
            "| 2 | 36,000-144,000 | 10% | 2,520 |\n"
            "| 3 | 144,000-300,000 | 20% | 16,920 |\n"
            "| 4 | 300,000-420,000 | 25% | 31,920 |\n"
            "| 5 | 420,000-660,000 | 30% | 52,920 |\n"
            "| 6 | 660,000-960,000 | 35% | 85,920 |\n"
            "| 7 | >960,000 | 45% | 181,920 |\n\n"
            "💡 系统支持累计预扣法自动计算，可在「工资薪金」模块导入数据。"
        ),
        r"社保.*基数|社保.*比例|社保.*费率": (
            "📋 **社保缴费比例参考（以各地政策为准）**\n\n"
            "| 险种 | 单位比例 | 个人比例 |\n"
            "|------|---------|----------|\n"
            "| 养老保险 | 16% | 8% |\n"
            "| 医疗保险 | 8-10% | 2% |\n"
            "| 失业保险 | 0.5-1% | 0.5% |\n"
            "| 工伤保险 | 0.2-1.9% | 0 |\n"
            "| 生育保险 | 0.5-1% | 0 |\n\n"
            "💡 缴费基数一般为职工上年度月平均工资，上下限为当地社平工资的 60%-300%。各地政策有差异，请以当地社保局规定为准。"
        ),
        r"折旧.*年限|固定资产.*折旧|怎么折旧": (
            "📋 **固定资产折旧最低年限**\n\n"
            "| 资产类别 | 最低年限 |\n"
            "|----------|----------|\n"
            "| 房屋、建筑物 | 20年 |\n"
            "| 飞机、火车、轮船、机器设备 | 10年 |\n"
            "| 与生产经营有关的器具工具家具 | 5年 |\n"
            "| 飞机火车轮船以外的运输工具 | 4年 |\n"
            "| 电子设备 | 3年 |\n\n"
            "💡 企业可选用直线法或加速折旧法（双倍余额递减、年数总和），一经选定不得随意变更。"
        ),
        r"发票.*红冲|红字.*发票|开错.*发票": (
            "📋 **增值税发票红冲操作要点**\n\n"
            "1. 销方申请：在开票系统中填开《红字发票信息表》→ 上传审核 → 开具红字发票\n"
            "2. 购买方已抵扣：由购买方填开信息表 → 销方开具红字发票\n"
            "3. 红字发票金额可为负数，与原蓝字发票一一对应\n\n"
            "⚠️ 注意：红字发票必须在原发票开具当月或次月处理，跨年需在汇算清缴前完成。\n"
            "💡 系统支持红字发票处理，导入时金额为负数即可自动识别。"
        ),
        r"进项.*抵扣|抵扣.*进项|认证.*抵扣": (
            "📋 **增值税进项税额抵扣要点**\n\n"
            "✅ 可抵扣：\n"
            "• 增值税专用发票（含机动车销售统一发票）\n"
            "• 海关进口增值税专用缴款书\n"
            "• 农产品收购/销售发票（按买价×9%/10%计算抵扣）\n"
            "• 通行费电子普通发票\n"
            "• 旅客运输服务票据（飞机票、火车票等）\n\n"
            "❌ 不可抵扣：\n"
            "• 用于简易计税、免税、集体福利、个人消费的进项\n"
            "• 非正常损失的购进货物\n"
            "• 贷款服务、餐饮服务、居民日常服务、娱乐服务\n\n"
            "💡 认证期限已取消（2017.7.1后开具的专票无限期），系统可在「进项发票」模块管理。"
        ),
        r"小规模.*转.*一般|一般纳税人.*认定|怎么.*一般纳税人": (
            "📋 **一般纳税人认定标准**\n\n"
            "• 年应税销售额超过 500 万元 → 应当登记为一般纳税人\n"
            "• 年应税销售额未超过 500 万元 → 可选择登记（会计核算健全）\n"
            "• 登记后不得转回小规模纳税人（除特定政策窗口期）\n\n"
            "💡 一般纳税人适用税率抵扣制，小规模适用简易征收。选择哪种身份需综合考虑客户群体和进项来源。"
        ),
        r"零申报|没收入.*申报|怎么.*零申报": (
            "📋 **零申报注意事项**\n\n"
            "• 没有发生应税收入也需要按期申报\n"
            "• 增值税零申报 → 连续 6 个月可能触发税务预警\n"
            "• 企业所得税零申报 → 长期零申报是稽查重点\n"
            "• 即使无收入，如有成本费用也应如实申报亏损\n\n"
            "⚠️ 风险提示：长期零申报（超过6个月）会被税务机关列为风险纳税人，可能面临实地核查。\n"
            "💡 系统「涉税风险分析报告」可检测零申报风险。"
        ),
        r"业务招待费|招待费.*扣除|招待费.*标准": (
            "📋 **业务招待费税前扣除标准**\n\n"
            "• 按照发生额的 **60%** 扣除\n"
            "• 但最高不得超过当年销售（营业）收入的 **5‰**\n"
            "• 两者孰低原则\n"
            "• 超标部分需纳税调整，不得结转以后年度\n\n"
            "💡 示例：年收入1000万，实际招待费20万\n"
            "  - 20万×60% = 12万\n"
            "  - 1000万×5‰ = 5万\n"
            "  - 可扣除：5万（孰低），需调增15万"
        ),
        r"广告费|宣传费|广告.*宣传.*扣除": (
            "📋 **广告费和业务宣传费税前扣除**\n\n"
            "• 一般企业：不超过当年销售（营业）收入 **15%** 的部分准予扣除\n"
            "• 超过部分可结转以后年度扣除\n"
            "• 化妆品制造、医药制造、饮料制造（不含酒类）：**30%**\n"
            "• 烟草企业：不得扣除\n\n"
            "💡 系统可在「所得税申报」模块自动计算调整额。"
        ),
        r"职工教育|教育经费|培训费.*扣除": (
            "📋 **职工教育经费税前扣除**\n\n"
            "• 一般企业：工资薪金总额 **8%** 以内准予扣除\n"
            "• 超过部分可结转以后年度扣除\n"
            "• 软件企业、集成电路设计企业：全额扣除（职工培训费）\n\n"
            "💡 与「工会经费（2%）」「福利费（14%）」分别计算，不可混用。"
        ),
        r"工会经费|工会.*扣除|工会.*比例": (
            "📋 **工会经费税前扣除**\n\n"
            "• 扣除比例：工资薪金总额的 **2%**\n"
            "• 需实际拨缴取得合法票据方可扣除\n"
            "• 计提未拨缴部分不得扣除\n\n"
            "💡 「福利费（14%）」「工会经费（2%）」「教育经费（8%）」三项分别计算限额。"
        ),
        r"福利费|职工福利.*扣除|福利.*标准": (
            "📋 **职工福利费税前扣除标准**\n\n"
            "• 扣除比例：工资薪金总额的 **14%**\n"
            "• 超过部分不得扣除，也不得结转\n"
            "• 包括：职工食堂、浴室、医疗、交通补贴等\n"
            "•  Exclusion：□□费、□□费、□□费（应并入工资缴个税）\n\n"
            "💡 福利费超标的，汇算清缴时需做纳税调增。"
        ),
        r"捐赠|公益性捐赠|慈善.*扣除": (
            "📋 **公益性捐赠税前扣除**\n\n"
            "• 一般公益性捐赠：年度利润总额 **12%** 以内准予扣除\n"
            "• 超过部分可结转以后 **3 年** 扣除\n"
            "• 扶贫捐赠：全额扣除（政策延续至2025年底）\n"
            "• 疫情防控捐赠：全额扣除\n"
            "• 需通过公益性社会组织或县级以上人民政府捐赠\n\n"
            "⚠️ 直接捐赠不得扣除。"
        ),
        r"亏损.*弥补|亏损.*年限|以前年度亏损": (
            "📋 **亏损弥补年限**\n\n"
            "• 一般企业：最长 **5 年** 弥补亏损\n"
            "• 高新技术企业、科技型中小企业：**10 年**\n"
            "• 受疫情影响困难行业企业（2020-2021）：**8 年**\n"
            "• 亏损弥补期限从亏损年度次年起算\n\n"
            "💡 系统「利润表」可查看历年亏损情况，辅助亏损弥补测算。"
        ),
        r"高新.*技术|高新技术企业.*优惠|研发费用.*加计扣除": (
            "📋 **高新技术企业税收优惠**\n\n"
            "• 企业所得税税率：**15%**（一般企业为25%）\n"
            "• 认定条件：核心自主知识产权、研发费用占比、高新收入占比等\n"
            "• 研发费用加计扣除：\n"
            "  - 制造业/科技型中小企业：**100%** 加计扣除\n"
            "  - 其他企业：**75%** 加计扣除（政策延续至2027年底）\n\n"
            "💡 认定有效期3年，需每年报送年报，期满后需重新认定。"
        ),
        r"房产税|房产.*怎么算|房产.*税率": (
            "📋 **房产税计算方式**\n\n"
            "**从价计征**（自用）：\n"
            "• 计税依据：房产原值 × (1 - 扣除比例10%-30%）\n"
            "• 税率：**1.2%**\n"
            "• 年应纳税额 = 房产原值 × (1 - 扣除比例） × 1.2%\n\n"
            "**从租计征**（出租）：\n"
            "• 计税依据：租金收入\n"
            "• 税率：**12%**（个人出租住房减按 **4%**）\n\n"
            "💡 扣除比例由各省、自治区、直辖市人民政府规定，一般为30%。"
        ),
        r"土地使用税|城镇.*土地使用|土地.*税率": (
            "📋 **城镇土地使用税**\n\n"
            "• 计税依据：**实际占用的土地面积**（㎡）\n"
            "• 税率：**分级幅度税额**（各地不同）\n"
            "  - 大城市：1.5 ～ 30 元/㎡·年\n"
            "  - 中等城市：1.2 ～ 24 元/㎡·年\n"
            "  - 小城市：0.9 ～ 18 元/㎡·年\n"
            "  - 县城/建制镇/工矿区：0.6 ～ 12 元/㎡·年\n\n"
            "💡 由省、自治区、直辖市人民政府确定具体适用税额标准。"
        ),
    }
    
    for pattern, answer in tax_qa.items():
        if re.search(pattern, msg_lower):
            return answer
    
    # 通用财务问题
    finance_tips = {
        r"(怎么做|如何处理|账务处理|分录.*怎么|记什么.*科目)": (
            "💡 关于账务处理，建议参考以下思路：\n\n"
            "1. 确认业务性质（采购/销售/费用/资产等）\n"
            "2. 确定涉及科目（可查看系统「科目表」）\n"
            "3. 借方记增加/费用，贷方记减少/收入\n"
            "4. 确保借贷金额相等\n\n"
            "📝 如需具体分录建议，请描述：\n"
            "• 业务类型（如：采购原材料）\n"
            "• 金额\n"
            "• 发票类型（专票/普票）\n"
            "• 付款方式（银行/现金）"
        ),
        r"(合理|合规|规范|有没有.*问题|风险|会不会)": (
            "🔍 关于财务合规性，建议关注以下几点：\n\n"
            "1. **三流一致**：合同流、发票流、资金流保持一致\n"
            "2. **凭证完整**：每笔业务需有合法有效凭证\n"
            "3. **票据规范**：发票内容与实际业务相符\n"
            "4. **按时申报**：各税种按期申报缴纳\n\n"
            "💡 系统「涉税风险分析报告」可根据你的财务数据自动扫描 61 个风险维度。\n"
            "建议定期生成报告以排查风险。"
        ),
    }
    
    for pattern, answer in finance_tips.items():
        if re.search(pattern, msg_lower):
            return answer
    
    return None

# ==================== 数据模型 ====================

class ChatRequest(BaseModel):
    message: str
    session_id: Optional[str] = None

sessions: dict = {}  # { session_id: { "intent": str, "step": int, "data": dict, "updated": float } }

def get_session(sid: str):
    if sid not in sessions:
        sessions[sid] = {"intent": None, "step": 0, "data": {}, "updated": 0}
    return sessions[sid]


# ==================== 意图识别与文本提取 ====================

def intent_from_text(msg: str) -> str:
    """从消息中识别意图 — 基于关键词优先级得分匹配，返回置信度最高的意图"""
    msg_lower = msg.strip().lower()

    # 文件上传 — 最高优先级，先于所有其他意图
    if msg_lower.startswith("[上传文件]"):
        return "file_upload"

    # 意图关键词库：(intent_name, [(pattern_regex, score), ...])
    # score 越高越优先；正则基于 msg_lower 匹配
    INTENTS = [
        ("create_voucher", [
            (r"\b(录[入记]凭证|制[单证]|做[一]?[笔张]?(账|分录|凭证)|填制凭证|记[账一]笔|录入分录)\b", 20),
            (r"(添加|新增|录入).{0,4}(凭证|分录)", 15),
            (r"nova.*voucher", 10),
        ]),
        ("list_vouchers", [
            (r"\b(查[看询]?凭证|凭证列[表出]|序时账|日记账|会计分录)\b", 20),
            (r"voucher.*list", 10),
            (r"(查|看|显示).{0,4}(凭证|分录)", 10),
        ]),
        ("create_customer", [
            (r"\b(新[增添加][客]户|录入客户|添加客户|建档.*客户|客户.*建档)\b", 20),
            (r"new.*customer|create.*customer", 10),
            (r"(添加|新增|录入).{0,4}客户", 15),
        ]),
        ("list_customers", [
            (r"\b(查[看询]?客户|客户列[表出]|客户管理)\b", 20),
            (r"(查|看).{0,4}客户", 10),
        ]),
        ("create_supplier", [
            (r"\b(新[增添加]供应商|录入供应商|添加供应商|建档.*供应商|供应商.*建档)\b", 20),
            (r"new.*supplier|create.*supplier", 10),
            (r"(添加|新增|录入).{0,4}供应商", 15),
        ]),
        ("list_suppliers", [
            (r"\b(查[看询]?供应商|供应商列[表出]|供应商管理)\b", 20),
            (r"(查|看).{0,4}供应商", 10),
        ]),
        ("create_employee", [
            (r"\b(新[增添加](员工|人员|职员)|录入(员工|人员)|添加(员工|人员))\b", 20),
            (r"new.*employee|create.*employee", 10),
            (r"(添加|新增|录入|建档).{0,4}(员工|人员|职员)", 15),
        ]),
        ("list_employees", [
            (r"\b(查[看询]?(员工|人员|职员)|(员工|人员|职员)列[表出])\b", 20),
            (r"(查|看).{0,4}(员工|人员|职员)", 10),
        ]),
        ("query_profit_loss", [
            (r"\b(利润表|损益表|利润报表|损益报表)\b", 20),
            (r"\bprofit.*loss|loss.*profit\b", 10),
            (r"(查|看|生成).{0,4}(利润|损益)", 12),
        ]),
        ("query_balance_sheet", [
            (r"\b(资产负债表|资产负债|balance\s*sheet)\b", 20),
            (r"(查|看|生成).{0,4}(资产|负债).{0,4}表", 12),
        ]),
        ("query_general_ledger", [
            (r"\b(总账|总分类账|general\s*ledger)\b", 20),
            (r"(查|看).{0,4}总账", 15),
        ]),
        ("query_detail_ledger", [
            (r"\b(明细账|明细分类账|detail\s*ledger)\b", 20),
            (r"(查|看).{0,4}明细", 12),
        ]),
        ("company_info", [
            (r"\b((公司|企业)信息|设置公司|录入公司|公司设置)\b", 20),
            (r"(修改|编辑|查看).{0,4}(公司|企业)", 12),
        ]),
        ("list_accounts", [
            (r"\b(科目表|科目列表|会计科目|chart\s*of\s*account)\b", 20),
            (r"(查|看|浏览).{0,4}科目", 12),
        ]),
        ("dashboard", [
            (r"\b(看板|dashboard|数据概览|统计看板|首页)\b", 20),
            (r"(看|打开).{0,4}(看板|概览|面板)", 12),
        ]),
        ("help", [
            (r"\b(帮助|help|能做什么|会什么|功能|指令|命令|怎么用)\b", 20),
        ]),
        ("cancel", [
            (r"\b(取消|退出|算了|不要了|返回|撤销)\b", 20),
        ]),
    ]

    # 得分制匹配：计算每个意图的总分，选最高分
    best_intent = None
    best_score = 0
    for intent_name, patterns in INTENTS:
        score = 0
        for pattern_regex, pattern_score in patterns:
            if re.search(pattern_regex, msg_lower):
                score += pattern_score
        if score > best_score:
            best_score = score
            best_intent = intent_name

    # 智能问询：如果没有任何匹配，尝试自然语言判断
    if best_score == 0:
        # 优先判断是否为财税政策问答（避免被下面的泛化规则误判）
        tax_keywords = r"(增值税|企业所得税|个人所得税|印花税|城建税|消费税|关税|房产税|土地使用税|"
        tax_keywords += r"契税|资源税|环保税|社保|公积金|个税|小规模|一般纳税人|发票|红冲|抵扣|"
        tax_keywords += r"进项|销项|零申报|申报|纳税|税金|税率|征收率|起征点|免税|退税|留抵|"
        tax_keywords += r"折旧|摊销|存货|坏账|预提|待摊|成本核算|利润分配|盈余公积|"
        tax_keywords += r"分录|记账|科目|凭证|借贷|借方|贷方|账务处理|做账|会计准则|"
        tax_keywords += r"优惠政策|税收优惠|减免税|稽查|合规|风险|预警|"
        tax_keywords += r"招待费|工会|教育经费|福利费|捐赠|亏损弥补|房产|土地使用|"
        tax_keywords += r"高新技术|研发|加计扣除|广告费|宣传费|印花税|城建税|教育费附加|"
        tax_keywords += r"车船税|契税|关税|消费税|环保税|文化事业|建设费)"
        if re.search(tax_keywords, msg_lower):
            return None  # 返回 None 表示进入 LLM 财税问答
        
        # 问报表
        if re.search(r"(看|查|显示).{0,4}(报表|报告)", msg_lower):
            return "query_general_ledger"
        # 问数据
        if re.search(r"(多少|几个|哪些|什么)", msg_lower):
            return "query_general_ledger"

    return best_intent


def extract_date(text: str) -> Optional[str]:
    """从自由文本中提取日期"""
    m = re.search(r"(\d{4}[-/年]\d{1,2}[-/月]\d{1,2})[日号]?", text)
    if m:
        raw = m.group(1)
        raw = raw.replace("年", "-").replace("月", "-").replace("/", "-")
        parts = raw.split("-")
        if len(parts) == 3:
            return f"{parts[0]}-{parts[1].zfill(2)}-{parts[2].zfill(2)}"
    # 仅年月
    m = re.search(r"(\d{4})[-/年](\d{1,2})[月]?", text)
    if m:
        return f"{m.group(1)}-{m.group(2).zfill(2)}-01"
    return None


def extract_amount(text: str) -> Optional[float]:
    """从文本中提取金额"""
    m = re.search(r"(\d+\.?\d*)\s*(万|万元|元|块|块钱)?", text)
    if m:
        amt = float(m.group(1))
        if m.group(2) in ("万", "万元"):
            amt *= 10000
        return round(amt, 2)
    return None


def extract_number(text: str) -> Optional[int]:
    """从文本中提取数字"""
    m = re.search(r"(\d+)", text)
    return int(m.group(1)) if m else None


# ==================== 文件读取与上传 ====================

def read_excel_content(file_bytes: bytes, filename: str) -> str:
    """读取 Excel 文件内容"""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    lines = [f"[Excel] {filename}"]
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        lines.append(f"\n--- 工作表: {sheet_name} ---")
        headers = []
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)
            headers.append(str(cell.value) if cell.value is not None else "")
        lines.append(" | ".join(headers))
        lines.append("-" * 60)
        row_count = 0
        for row in range(2, ws.max_row + 1):
            vals = []
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                vals.append(str(cell.value) if cell.value is not None else "")
            if any(v.strip() for v in vals):
                lines.append(" | ".join(vals))
                row_count += 1
                if row_count >= 200:
                    lines.append(f"... (共 {ws.max_row - 1} 行，仅显示前 200 行)")
                    break
        lines.append(f"→ 共 {ws.max_row - 1} 行数据\n")
    return "\n".join(lines)


def read_csv_content(file_bytes: bytes, filename: str) -> str:
    """读取 CSV 文件内容"""
    text = file_bytes.decode("utf-8-sig")
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        return f"[CSV] {filename}\n(空文件)"
    lines = [f"[CSV] {filename}"]
    lines.append(" | ".join(rows[0]))
    lines.append("-" * 60)
    for i, row in enumerate(rows[1:], 1):
        lines.append(" | ".join(row))
        if i >= 200:
            lines.append(f"... (共 {len(rows) - 1} 行，仅显示前 200 行)")
            break
    return "\n".join(lines)


def read_pdf_content(file_bytes: bytes, filename: str) -> str:
    """读取 PDF 文件文本内容"""
    reader = PdfReader(io.BytesIO(file_bytes))
    lines = [f"[PDF] {filename}"]
    total_text = ""
    for i, page in enumerate(reader.pages):
        text = page.extract_text()
        if text:
            total_text += text + "\n"
    if not total_text.strip():
        return f"[PDF] {filename}\n(无法提取文本内容，可能是扫描件或图片型 PDF)"
    if len(total_text) > 5000:
        total_text = total_text[:5000] + f"\n...(共 {len(total_text)} 字符，仅显示前 5000)"
    lines.append(total_text.strip())
    return "\n".join(lines)


@router.post("/api/chat/upload")
async def upload_file(
    file: UploadFile = File(...),
    session_id: str = Form("")
):
    """上传文件并识别内容"""
    try:
        content_bytes = await file.read()
        fname = file.filename or "unknown"
        ext = os.path.splitext(fname)[1].lower()

        if ext in (".xlsx", ".xls"):
            content = read_excel_content(content_bytes, fname)
        elif ext == ".csv":
            content = read_csv_content(content_bytes, fname)
        elif ext == ".pdf":
            content = read_pdf_content(content_bytes, fname)
        elif ext in (".txt", ".md", ".log"):
            text = content_bytes.decode("utf-8")
            if len(text) > 5000:
                text = text[:5000] + f"\n...(共 {len(text)} 字符，仅显示前 5000)"
            content = f"[文本] {fname}\n{text}"
        elif ext in (".png", ".jpg", ".jpeg", ".gif", ".bmp", ".webp"):
            content = f"[图片] {fname}\n(不支持图片文字识别，请直接描述需求或将数据整理为 Excel/CSV 格式上传)"
        else:
            return {"error": f"不支持的文件格式：{ext}。支持格式：xlsx, csv, pdf, txt, md, log", "session_id": session_id}

        return {
            "file_name": fname,
            "file_type": ext,
            "content": content,
            "session_id": session_id
        }
    except Exception as e:
        return {"error": f"文件处理失败：{str(e)}", "session_id": session_id}


# ==================== 主对话接口 ====================

@router.post("/api/chat")
def chat_endpoint(payload: ChatRequest, company_id: int = Query(...), db: Session = Depends(get_db)):
    """AI 助手对话接口"""
    message = payload.message.strip()
    sid = payload.session_id or str(uuid.uuid4())
    sess = get_session(sid)

    if not message:
        return {"reply": "请说点什么吧 😊", "session_id": sid, "action": None}

    # 取消当前流程
    if re.search(r"^(取消|退出|算了|不要了|返回)$", message):
        sess["intent"] = None
        sess["step"] = 0
        sess["data"] = {}
        return {"reply": "好的，已取消当前操作。\n\n💡 试试问我财税问题：\n• 增值税税率、个税计算、印花税\n• 账务处理建议\n• 风险合规提示\n\n或说「帮助」查看系统功能。", "session_id": sid, "action": None}

    intent = intent_from_text(message)

    # 如果在流程中
    if sess["intent"]:
        intent = sess["intent"]
    elif intent == "cancel":
        sess["intent"] = None
        return {"reply": "当前没有进行中的任务。有什么我可以帮你的？", "session_id": sid, "action": None}

    # ────────────── 录入凭证 ──────────────
    if intent == "create_voucher":
        return handle_create_voucher(sess, message, db, sid, company_id)

    # ────────────── 新增客户 ──────────────
    if intent == "create_customer":
        return handle_create_customer(sess, message, db, sid, company_id)

    # ────────────── 新增供应商 ──────────────
    if intent == "create_supplier":
        return handle_create_supplier(sess, message, db, sid, company_id)

    # ────────────── 新增员工 ──────────────
    if intent == "create_employee":
        return handle_create_employee(sess, message, db, sid, company_id)

    # ────────────── 查询类 ──────────────
    if intent == "query_profit_loss":
        sess["intent"] = None
        today = date.today().strftime("%Y-%m")
        return {"reply": f"📅 请告诉我你要查询的期间范围，例如：\n• `{today}` → 查 {today} 月\n• `2026-01 到 2026-05` → 查1-5月", "session_id": sid, "action": None}

    if intent == "query_balance_sheet":
        sess["intent"] = None
        today = date.today().strftime("%Y-%m")
        return {"reply": f"📅 请告诉我你要查询的截止期间，例如：\n• `{today}` → 截止 {today} 月", "session_id": sid, "action": None}

    if intent == "query_general_ledger":
        return handle_query_general(sess, message, db, sid)

    if intent == "query_detail_ledger":
        return handle_query_detail(sess, message, db, sid)

    if intent == "list_vouchers":
        sess["intent"] = None
        return handle_list_vouchers(message, db, sid)

    if intent in ("list_customers", "list_suppliers", "list_employees", "list_accounts"):
        sess["intent"] = None
        return {"reply": f"✅ 请打开侧边栏的对应页面查看。\n💡 提示：你可以说「新增客户」来快捷录入。", "session_id": sid, "action": {"type": "navigate", "page": intent.replace("list_", "")}}

    if intent == "company_info":
        sess["intent"] = None
        return {"reply": "🏢 请打开侧边栏「公司信息」页面填写。\n你也可以直接告诉我：\n• 公司名称\n• 统一社会信用代码\n• 法定代表人\n• 地址电话等\n\n我会帮你一次性填好！", "session_id": sid, "action": None}

    if intent == "dashboard":
        sess["intent"] = None
        return {"reply": "📊 请打开侧边栏「数据看板」查看统计。", "session_id": sid, "action": {"type": "navigate", "page": "dashboard"}}

    if intent == "help":
        sess["intent"] = None
        return {
            "reply": "📋 **财税问答助手 — 功能指引**\n\n"
                     "**💬 财税问答**\n"
                     "直接输入你的财税问题，我会基于专业知识和政策法规为你解答：\n"
                     "• 税务政策：增值税、企业所得税、个税、印花税等\n"
                     "• 账务处理：会计分录、科目运用、凭证编制\n"
                     "• 风险提示：稽查风险、发票合规、经营实质\n"
                     "• 财务管理：成本控制、费用管理、资金规划\n\n"
                     "**📎 文件分析**\n"
                     "上传 Excel/CSV/PDF 文件，我帮你识别数据类型并引导录入。\n\n"
                     "**🖥️ 系统操作**\n"
                     "• 「录入凭证」— 引导填制记账凭证\n"
                     "• 「新增客户 / 供应商 / 员工」— 快速建档\n"
                     "• 「查看利润表 / 资产负债表」— 查询报表\n"
                     "• 「生成风险报告」— 61 维度涉税风险扫描\n\n"
                     "输入「取消」可随时退出当前流程。",
            "session_id": sid,
            "action": None
        }

    # ────────────── 文件上传 ──────────────
    if intent == "file_upload":
        return handle_file_upload(sess, message, db, sid)

    # ────────────── 文件确认后续 ──────────────
    if intent == "file_confirm":
        return handle_file_confirm(sess, message, db, sid)

    # 未识别意图 → LLM 财税问答（带本地规则回退）
    sess["intent"] = None
    ai_reply = _chat_tax_qa(message, company_id, db)
    if ai_reply:
        return {
            "reply": ai_reply,
            "session_id": sid,
            "action": None
        }
    # 本地规则也无法回答 → 友好引导
    return {
        "reply": "抱歉，这个问题我暂时无法给出精确答案 🤔\n\n"
                 "你可以试试问我这些：\n"
                 "• **税务政策**：增值税税率、个税计算、印花税等\n"
                 "• **账务处理**：某笔业务怎么做分录、记什么科目\n"
                 "• **财务管理**：成本控制、费用管理的建议\n"
                 "• **风险提示**：税务稽查风险、合规性检查\n\n"
                 "💡 也可以说「帮助」查看系统操作功能，或上传文件让我帮你分析。",
        "session_id": sid,
        "action": None
    }


# ==================== 文件上传智能分析 ====================

DATA_PATTERNS = {
    "voucher": {
        "name": "记账凭证",
        "keywords": ["日期", "摘要", "科目", "借方", "贷方", "金额", "凭证", "voucher", "date", "account", "debit", "credit"],
        "min_match": 3,
        "action_hint": "我可以帮你**逐笔录入凭证**，你只需要确认每一笔即可。"
    },
    "customer": {
        "name": "客户档案",
        "keywords": ["客户", "名称", "联系人", "电话", "手机", "地址", "税号", "customer", "phone", "contact", "address"],
        "min_match": 2,
        "action_hint": "我可以帮你**批量导入客户档案**，请确认以下信息无误。"
    },
    "supplier": {
        "name": "供应商档案",
        "keywords": ["供应商", "厂家", "供货", "supplier", "采购"],
        "min_match": 2,
        "action_hint": "我可以帮你**批量导入供应商档案**，请确认以下信息无误。"
    },
    "employee": {
        "name": "员工花名册",
        "keywords": ["员工", "人员", "职员", "部门", "职位", "入职", "身份证", "employee", "department", "position"],
        "min_match": 2,
        "action_hint": "我可以帮你**批量导入员工信息**，请确认以下信息无误。"
    },
    "account": {
        "name": "会计科目",
        "keywords": ["科目编码", "科目名称", "科目类别", "余额方向", "account_code", "account_name"],
        "min_match": 2,
        "action_hint": "我可以帮你**批量导入会计科目**，请确认后处理。"
    },
}


def analyze_file_columns(headers: List[str]) -> dict:
    """分析文件列名，猜测数据类型"""
    headers_lower = [h.strip().lower() for h in headers]
    best_type = None
    best_score = 0

    for ptype, pattern in DATA_PATTERNS.items():
        score = 0
        for kw in pattern["keywords"]:
            for h in headers_lower:
                if kw in h:
                    score += 1
                    break
        if score >= pattern["min_match"] and score > best_score:
            best_score = score
            best_type = ptype

    return {
        "detected_type": best_type,
        "confidence": best_score,
        "pattern": DATA_PATTERNS.get(best_type) if best_type else None
    }


def handle_file_upload(sess, message: str, db, sid: str):
    """处理文件上传：分析内容 → 提问确认 → 等待用户指令"""
    lines = message.split("\n")
    file_name = ""
    content_start = 0

    m = re.match(r"\[上传文件\]\s*(.+)", lines[0])
    if m:
        file_name = m.group(1).strip()

    for i, line in enumerate(lines):
        if line.strip().startswith("[Excel]") or line.strip().startswith("[CSV]") or \
           line.strip().startswith("[PDF]") or line.strip().startswith("[文本]") or \
           line.strip().startswith("[图片]"):
            content_start = i
            break

    format_label = ""
    if content_start < len(lines):
        format_label = lines[content_start].strip()
        fm = re.match(r"\[(\w+)\]\s*(.+)", format_label)
        if fm and not file_name:
            file_name = fm.group(2).strip()

    headers = []
    data_rows = 0
    for line in lines[content_start:]:
        stripped = line.strip()
        if " | " in stripped and not stripped.startswith("---"):
            parts = [p.strip() for p in stripped.split(" | ")]
            if not headers:
                headers = parts
            else:
                data_rows += 1
        rm = re.search(r"共\s*(\d+)\s*行", stripped)
        if rm:
            data_rows = max(data_rows, int(rm.group(1)))

    if not headers:
        sess["intent"] = None
        sess["step"] = 0
        sess["data"] = {}
        return {
            "reply": f"📎 已收到文件 **{file_name}**。\n\n"
                     f"我暂时无法自动识别这个文件的列结构。\n\n"
                     f"🤔 **请告诉我：**\n"
                     f"• 这个文件包含什么数据？（凭证 / 客户 / 供应商 / 员工 / 科目 / 其他）\n"
                     f"• 你希望我怎么处理？（录入系统 / 仅查看 / 导入到对应模块）\n\n"
                     f"也可以点击左侧菜单进入对应页面手动操作。",
            "session_id": sid,
            "action": None
        }

    analysis = analyze_file_columns(headers)
    detected = analysis["detected_type"]
    confidence = analysis["confidence"]
    pattern = analysis["pattern"]

    cols_display = "、".join(headers[:8])
    if len(headers) > 8:
        cols_display += f" ... 共 {len(headers)} 列"

    if not detected or confidence < 2:
        sess["intent"] = None
        sess["step"] = 0
        sess["data"] = {}
        return {
            "reply": f"📎 已收到文件 **{file_name}**（约 {data_rows} 行数据）\n\n"
                     f"📋 识别到的列：{cols_display}\n\n"
                     f"⚠️ 我无法确定这是什么类型的数据。\n\n"
                     f"🤔 **请确认：**\n"
                     f"1️⃣ 这是**凭证数据**？→ 回复「凭证」\n"
                     f"2️⃣ 这是**客户名单**？→ 回复「客户」\n"
                     f"3️⃣ 这是**供应商名单**？→ 回复「供应商」\n"
                     f"4️⃣ 这是**员工信息**？→ 回复「员工」\n"
                     f"5️⃣ 这是**会计科目**？→ 回复「科目」\n"
                     f"6️⃣ 只是给你看看，不做处理？→ 回复「不用」\n\n"
                     f"💡 也可以直接描述，例如：「把这些客户信息录入系统」",
            "session_id": sid,
            "action": None
        }

    sess["intent"] = "file_confirm"
    sess["step"] = 0
    sess["data"] = {
        "file_name": file_name,
        "detected_type": detected,
        "headers": headers,
        "data_rows": data_rows,
        "raw_content": message
    }

    preview_lines = []
    preview_count = 0
    for line in lines[content_start:]:
        stripped = line.strip()
        if " | " in stripped and not stripped.startswith("---") and not stripped.startswith("[") and preview_count < 3:
            parts = [p.strip() for p in stripped.split(" | ")]
            if parts != headers:
                preview_lines.append(" | ".join(parts))
                preview_count += 1

    preview_text = ""
    if preview_lines:
        preview_text = "\n\n📋 **数据预览（前3行）：**\n" + "\n".join(f"  `{p}`" for p in preview_lines)

    return {
        "reply": f"📎 已收到文件 **{file_name}**（约 {data_rows} 行数据）\n\n"
                 f"🔍 我识别到这可能是一份 **{pattern['name']}** 数据。\n"
                 f"📋 包含列：{cols_display}"
                 f"{preview_text}\n\n"
                 f"⚠️ **请确认：**\n"
                 f"• 回复「**是**」或「**确认**」→ {pattern['action_hint']}\n"
                 f"• 回复「**不是**」→ 告诉我这是什么数据\n"
                 f"• 回复「**不用**」→ 仅查看，不做处理\n"
                 f"• 回复「**取消**」→ 放弃本次上传",
        "session_id": sid,
        "action": None
    }


def handle_file_confirm(sess, message: str, db, sid: str):
    """处理文件确认后的用户回应"""
    msg = message.strip()
    msg_lower = msg.lower()
    data = sess.get("data", {})
    detected_type = data.get("detected_type", "")
    file_name = data.get("file_name", "未知文件")

    if re.search(r"^(是|对|确认|好|可以|行|ok|yes|没错|是的|对的)$", msg_lower):
        sess["intent"] = None
        sess["step"] = 0
        sess["data"] = {}
        if detected_type == "voucher":
            return {
                "reply": f"✅ 好的！我将根据 **{file_name}** 的数据逐笔录入凭证。\n\n"
                         f"📝 请回复「**录入凭证**」开始，我会引导你一步步操作。\n\n"
                         f"💡 提示：你可以把文件中每一行的数据逐条告诉我，我来填。",
                "session_id": sid,
                "action": None
            }
        elif detected_type == "customer":
            return {
                "reply": f"✅ 好的！我将把 **{file_name}** 中的客户信息录入系统。\n\n"
                         f"请回复「**新增客户**」开始逐条录入。\n\n"
                         f"💡 提示：也可以告诉我「批量导入所有客户」，我会遍历每一行。",
                "session_id": sid,
                "action": None
            }
        elif detected_type == "supplier":
            return {
                "reply": f"✅ 好的！我将把 **{file_name}** 中的供应商信息录入系统。\n\n"
                         f"请回复「**新增供应商**」开始逐条录入。\n\n"
                         f"💡 提示：也可以告诉我「批量导入所有供应商」，我会遍历每一行。",
                "session_id": sid,
                "action": None
            }
        elif detected_type == "employee":
            return {
                "reply": f"✅ 好的！我将把 **{file_name}** 中的员工信息录入系统。\n\n"
                         f"请回复「**新增员工**」开始逐条录入。\n\n"
                         f"💡 提示：也可以告诉我「批量导入所有员工」，我会遍历每一行。",
                "session_id": sid,
                "action": None
            }
        else:
            return {
                "reply": f"✅ 好的！请告诉我具体要怎么处理 **{file_name}** 的数据？",
                "session_id": sid,
                "action": None
            }

    if re.search(r"^(不是|不对|错了|不对的|no|不对哦)$", msg_lower):
        sess["intent"] = None
        sess["step"] = 0
        sess["data"] = {}
        return {
            "reply": f"🤔 抱歉识别错了！\n\n"
                     f"请告诉我 **{file_name}** 里是什么数据：\n"
                     f"• 回复「**凭证**」→ 记账凭证\n"
                     f"• 回复「**客户**」→ 客户档案\n"
                     f"• 回复「**供应商**」→ 供应商档案\n"
                     f"• 回复「**员工**」→ 员工信息\n"
                     f"• 回复「**科目**」→ 会计科目\n"
                     f"• 或直接描述你要做什么",
            "session_id": sid,
            "action": None
        }

    if re.search(r"^(不用|算了|不需要|看看就行|仅查看|只是看看)$", msg_lower):
        sess["intent"] = None
        sess["step"] = 0
        sess["data"] = {}
        return {
            "reply": f"👌 好的，**{file_name}** 的数据仅供查看，不做录入处理。\n\n"
                     f"如果后续需要处理，随时告诉我！\n\n"
                     f"💡 你可以继续上传其他文件，或告诉我其他需求。",
            "session_id": sid,
            "action": None
        }

    sess["intent"] = None
    sess["step"] = 0
    sess["data"] = {}

    type_hint = None
    if re.search(r"凭证|voucher|记账", msg_lower):
        type_hint = "voucher"
    elif re.search(r"客户|customer", msg_lower):
        type_hint = "customer"
    elif re.search(r"供应商|supplier", msg_lower):
        type_hint = "supplier"
    elif re.search(r"员工|人员|职员|employee", msg_lower):
        type_hint = "employee"
    elif re.search(r"科目|account", msg_lower):
        type_hint = "account"

    if type_hint:
        pattern = DATA_PATTERNS.get(type_hint, {})
        sess["intent"] = "file_confirm"
        sess["step"] = 0
        sess["data"] = {**data, "detected_type": type_hint}
        return {
            "reply": f"🔍 收到，你指定为 **{pattern.get('name', type_hint)}** 数据。\n\n"
                     f"⚠️ **再次确认：**\n"
                     f"• 回复「**是**」→ {pattern.get('action_hint', '开始处理')}\n"
                     f"• 回复「**不是**」→ 重新指定",
            "session_id": sid,
            "action": None
        }

    return {
        "reply": f"🤔 收到你的回复：「{msg}」\n\n"
                 f"关于 **{file_name}** 的数据处理，请明确告诉我：\n"
                 f"• 回复「**是**」→ 确认按识别类型处理\n"
                 f"• 回复「**不是**」→ 重新指定数据类型\n"
                 f"• 回复「**不用**」→ 仅查看\n"
                 f"• 回复「**取消**」→ 放弃\n\n"
                 f"💡 也可以直接说数据类型如「凭证」「客户」等。",
        "session_id": sid,
        "action": None
    }


# ==================== CRUD 处理器 ====================

def handle_create_voucher(sess: Dict[str, Any], msg: str, db: Session, sid: str, company_id: int) -> Dict[str, Any]:
    """AI 录入凭证：引导用户去序时账页面操作"""
    sess["intent"] = None
    return {
        "reply": "📝 录入凭证请打开侧边栏「序时账」页面操作。\n\n在序时账页面可以：\n• 新增凭证\n• 编辑凭证\n• 从销项/进项/银行流水自动生成凭证\n\n💡 提示：你也可以上传文件，我会帮你识别并引导导入。",
        "session_id": sid,
        "action": {"type": "navigate", "page": "journal"}
    }


def handle_create_customer(sess: Dict[str, Any], msg: str, db: Session, sid: str, company_id: int) -> Dict[str, Any]:
    """AI 新增客户：多步骤引导录入客户档案"""
    data = sess["data"]

    if step == 0:
        sess["intent"] = "create_customer"
        sess["step"] = 1
        parts = msg.strip()
        code_match = re.search(r"编码[：:]*\s*(\S+)", msg)
        name_match = re.sub(r"新[增添加]客户|录入客户|添加客户", "", msg).strip("，。,.").strip()
        if name_match:
            data["name"] = name_match
        if code_match:
            data["code"] = code_match.group(1)

        if data.get("name"):
            if not data.get("code"):
                data["code"] = f"KH{db.query(Customer).filter(Customer.company_id == company_id).count() + 1:03d}"
            sess["data"] = data
            sess["step"] = 2
            return {"reply": f"👤 客户名称：**{data['name']}**\n📋 编码：**{data['code']}**\n\n还需要添加其他信息吗？可以直接告诉我：\n• 联系人\n• 电话\n• 信用额度\n\n或说「**完成**」直接保存。", "session_id": sid, "action": None}

        return {"reply": "👤 好的，新增客户。\n\n请告诉我**客户名称**和**编码**（可选），例如：\n• 「广州钢材贸易有限公司」\n• 「编码 KH001 广州钢材贸易有限公司」", "session_id": sid, "action": None}

    elif step == 1:
        data["name"] = msg.strip()
        if not data.get("code"):
            data["code"] = f"KH{db.query(Customer).filter(Customer.company_id == company_id).count() + 1:03d}"
        sess["data"] = data
        sess["step"] = 2
        return {"reply": f"👤 客户名称：**{data['name']}**\n📋 编码：**{data['code']}**\n\n还需要添加其他信息吗？可以告诉我联系人、电话、地址等。\n或说「**完成**」直接保存。", "session_id": sid, "action": None}

    elif step >= 2:
        if re.search(r"^(完成|好了|结束|确认|提交|保存|ok|done)$", msg.strip(), re.IGNORECASE):
            return save_customer(data, db, sess, sid, company_id)
        contact_m = re.search(r"联系人[：:]*\s*(\S+)", msg)
        phone_m = re.search(r"电话[：:]*\s*(\S+)", msg)
        addr_m = re.search(r"地址[：:]*\s*(.+?)(?:$|电话|联系人)", msg)
        credit_m = re.search(r"额度[：:]*\s*(\d+\.?\d*)", msg)
        if contact_m: data["contact"] = contact_m.group(1)
        if phone_m: data["phone"] = phone_m.group(1)
        if addr_m: data["address"] = addr_m.group(1).strip()
        if credit_m: data["credit_limit"] = float(credit_m.group(1))
        if not contact_m and not phone_m and not addr_m and not credit_m:
            pt = msg.strip().split()
            if len(pt) >= 1 and not data.get("contact"):
                data["contact"] = pt[0]
            if len(pt) >= 2 and not data.get("phone"):
                data["phone"] = pt[1]

        sess["data"] = data
        info_lines = []
        for k, label in [("name", "名称"), ("code", "编码"), ("contact", "联系人"), ("phone", "电话"), ("address", "地址"), ("credit_limit", "信用额度")]:
            if data.get(k):
                info_lines.append(f"  {label}：{data[k]}")
        return {"reply": f"已更新客户信息：\n" + "\n".join(info_lines) + "\n\n说「**完成**」保存，或继续补充信息。", "session_id": sid, "action": None}

    return {"reply": "请继续...", "session_id": sid, "action": None}


def save_customer(data: Dict[str, Any], db: Session, sess: Dict[str, Any], sid: str, company_id: int) -> Dict[str, Any]:
    """保存客户档案到数据库，处理异常回滚"""
    try:
        name = data.get("name", "")
        code = data.get("code", "")
        c = Customer(
            company_id=company_id,
            code=data.get("code", ""),
            name=name or "未命名客户",
            uscc=data.get("uscc"),
            contact=data.get("contact"),
            phone=data.get("phone"),
            address=data.get("address"),
            credit_limit=data.get("credit_limit", 0.0)
        )
        db.add(c)
        db.commit()
        sess["intent"] = None
        sess["step"] = 0
        sess["data"] = {}
        return {"reply": f"🎉 客户 **{c.name}**（{c.code}）添加成功！\n\n💡 接下来可以「新增客户」继续添加，或「查看利润表」查询报表。", "session_id": sid, "action": {"type": "reload", "page": "customers"}}
    except Exception as e:
        return {"reply": f"❌ 保存失败：{str(e)}", "session_id": sid, "action": None}


def handle_create_supplier(sess: Dict[str, Any], msg: str, db: Session, sid: str, company_id: int) -> Dict[str, Any]:
    """AI 新增供应商：多步骤引导录入供应商档案"""
    data = sess["data"]

    if step == 0:
        sess["intent"] = "create_supplier"
        sess["step"] = 1
        name_match = re.sub(r"新[增添加]供应商|录入供应商|添加供应商", "", msg).strip("，。,.").strip()
        code_match = re.search(r"编码[：:]*\s*(\S+)", msg)
        if name_match: data["name"] = name_match
        if code_match: data["code"] = code_match.group(1)
        if data.get("name"):
            if not data.get("code"): data["code"] = f"GYS{db.query(Supplier).filter(Supplier.company_id == company_id).count() + 1:03d}"
            sess["data"] = data; sess["step"] = 2
            return {"reply": f"📦 供应商：**{data['name']}**（{data['code']}）\n\n需要补充其他信息吗？或说「**完成**」直接保存。", "session_id": sid, "action": None}
        return {"reply": "📦 新增供应商。请告诉我**供应商名称**，例如：\n「广州钢铁供应链有限公司」", "session_id": sid, "action": None}

    elif step == 1:
        data["name"] = msg.strip()
        if not data.get("code"): data["code"] = f"GYS{db.query(Supplier).filter(Supplier.company_id == company_id).count() + 1:03d}"
        sess["data"] = data; sess["step"] = 2
        return {"reply": f"📦 供应商：**{data['name']}**（{data['code']}）\n\n需要补充其他信息吗？或说「**完成**」保存。", "session_id": sid, "action": None}

    elif step >= 2:
        if re.search(r"^(完成|好了|结束|确认|提交|保存|ok|done)$", msg.strip(), re.IGNORECASE):
            return save_supplier(data, db, sess, sid, company_id)
        sess["data"] = data
        lines = [f"  {k}：{v}" for k, v in data.items() if v and k in ("name", "code")]
        return {"reply": "已更新：\n" + "\n".join(lines) + "\n\n说「**完成**」保存。", "session_id": sid, "action": None}

    return {"reply": "请继续...", "session_id": sid, "action": None}


def save_supplier(data: Dict[str, Any], db: Session, sess: Dict[str, Any], sid: str, company_id: int) -> Dict[str, Any]:
    """保存供应商档案到数据库"""
    try:
        name = data.get("name", "")
        code = data.get("code", "")
        s = Supplier(company_id=company_id, code=data.get("code", ""), name=name, uscc=data.get("uscc", ""))
        db.add(s); db.commit()
        sess["intent"] = None; sess["step"] = 0; sess["data"] = {}
        return {"reply": f"🎉 供应商 **{s.name}**（{s.code}）添加成功！", "session_id": sid, "action": {"type": "reload", "page": "suppliers"}}
    except Exception as e:
        return {"reply": f"❌ {e}", "session_id": sid, "action": None}


def handle_create_employee(sess: Dict[str, Any], msg: str, db: Session, sid: str, company_id: int) -> Dict[str, Any]:
    """AI 新增员工：多步骤引导录入员工档案"""
    data = sess["data"]

    if step == 0:
        sess["intent"] = "create_employee"
        sess["step"] = 1
        name_match = re.sub(r"新[增添加](员工|人员|职员)|录入(员工|人员)|添加(员工|人员)", "", msg).strip("，。,.").strip()
        dept_m = re.search(r"部门[：:]*\s*(\S+)", msg)
        if name_match: data["name"] = name_match
        if dept_m: data["department_name"] = dept_m.group(1)
        if data.get("name"):
            if not data.get("code"): data["code"] = f"RY{db.query(Employee).filter(Employee.company_id == company_id).count() + 1:03d}"
            sess["data"] = data; sess["step"] = 2
            return {"reply": f"👤 员工：**{data['name']}**（{data['code']}）\n\n还需要补充部门、职位、电话吗？或说「**完成**」保存。", "session_id": sid, "action": None}
        return {"reply": "👤 新增员工。请告诉我**姓名**，例如：「张三」", "session_id": sid, "action": None}

    elif step == 1:
        data["name"] = msg.strip()
        if not data.get("code"): data["code"] = f"RY{db.query(Employee).filter(Employee.company_id == company_id).count() + 1:03d}"
        sess["data"] = data; sess["step"] = 2
        return {"reply": f"👤 员工：**{data['name']}**\n\n需要补充部门、职位、电话吗？或说「**完成**」保存。", "session_id": sid, "action": None}

    elif step >= 2:
        if re.search(r"^(完成|好了|结束|确认|ok|done)$", msg.strip(), re.IGNORECASE):
            return save_employee(data, db, sess, sid, company_id)
        dept_m = re.search(r"部门[：:]*\s*(\S+)", msg)
        pos_m = re.search(r"职位[：:]*\s*(\S+)", msg)
        phone_m = re.search(r"电话[：:]*\s*(\S+)", msg)
        if dept_m: data["department_name"] = dept_m.group(1)
        if pos_m: data["position"] = pos_m.group(1)
        if phone_m: data["phone"] = phone_m.group(1)
        sess["data"] = data
        lines = [f"  {k}：{v}" for k, v in data.items() if v and k in ("name", "code", "department_name", "position", "phone")]
        return {"reply": "已更新：\n" + "\n".join(lines) + "\n\n说「**完成**」保存。", "session_id": sid, "action": None}

    return {"reply": "请继续...", "session_id": sid, "action": None}


def save_employee(data: Dict[str, Any], db: Session, sess: Dict[str, Any], sid: str, company_id: int) -> Dict[str, Any]:
    """保存员工档案到数据库"""
    try:
        e = Employee(company_id=company_id, code=data.get("code", ""), name=data.get("name", ""), id_card=data.get("id_card"), email=data.get("email"))
        db.add(e); db.commit()
        sess["intent"] = None; sess["step"] = 0; sess["data"] = {}
        return {"reply": f"🎉 员工 **{e.name}** 添加成功！", "session_id": sid, "action": {"type": "reload", "page": "employees"}}
    except Exception as e:
        return {"reply": f"❌ {e}", "session_id": sid, "action": None}


# ==================== 查询处理器（存根） ====================

def handle_list_vouchers(message, db, sid):
    """查询凭证列表 — 引导用户去序时账页面"""
    return {
        "reply": "📋 请打开侧边栏「序时账」页面查看凭证列表。\n\n💡 你也可以告诉我具体的查询条件（如「5月的凭证」），我会帮你筛选。",
        "session_id": sid,
        "action": {"type": "navigate", "page": "journal"}
    }


def handle_query_general(sess, message, db, sid):
    """查询总账 — 引导用户去报表页面"""
    sess["intent"] = None
    return {
        "reply": "📊 请打开侧边栏「科目余额表」查看总账。\n\n💡 你可以告诉我具体的科目编码或名称来查询明细。",
        "session_id": sid,
        "action": {"type": "navigate", "page": "trial-balance"}
    }


def handle_query_detail(sess, message, db, sid):
    """查询明细账 — 引导用户去序时账页面"""
    sess["intent"] = None
    return {
        "reply": "📋 请打开侧边栏「序时账」页面查看明细账。\n\n💡 你可以告诉我具体的科目编码来筛选。",
        "session_id": sid,
        "action": {"type": "navigate", "page": "journal"}
    }
