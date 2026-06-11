"""
文化事业建设费申报 API 路由
使用 FastAPI APIRouter，在 main.py 中 include_router 加载。
完全参照 vat.py 的架构复刻。
"""
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, Form, Request
from pydantic import BaseModel
from sqlalchemy.orm import Session
from sqlalchemy import func, or_
from datetime import datetime
from typing import Optional, List
import json
import calendar

from database import (
    CulturalConstructionFeeDeclaration,
    CulturalConstructionFeeDeduction,
    Company, get_db,
    SalesInvoice, PurchaseInvoice, BookkeepingInvoice,
    JournalEntry, Account, Supplier,
)

router = APIRouter(prefix="/api/cultural-construction-fee", tags=["文化事业建设费"])


# ==================== 工具函数 ====================


# ==================== 外部导入申报表 ====================

@router.post("/declarations/import")
async def import_ccf_declaration(
    company_id: int = Form(...),
    file: UploadFile = File(...),
    declaration_id: Optional[str] = Form(None),
    db: Session = Depends(get_db)
):
    """
    外部导入文化事业建设费申报表（Excel文件）
    接收用户上传的已填好的申报表Excel文件，解析并保存到数据库。
    """
    import tempfile, os, re
    from openpyxl import load_workbook
    from datetime import datetime
    
    # 保存上传的文件到临时目录
    suffix = file.filename.rsplit('.', 1)[-1] if '.' in file.filename else 'xlsx'
    with tempfile.NamedTemporaryFile(suffix='.' + suffix, delete=False) as tmp:
        content = await file.read()
        tmp.write(content)
        tmp_path = tmp.name
    
    try:
        wb = load_workbook(tmp_path, data_only=True)
        
        # 获取或创建申报表记录
        decl = None
        if declaration_id:
            try:
                decl_id = int(declaration_id)
                decl = db.query(CulturalConstructionFeeDeclaration).filter(
                    CulturalConstructionFeeDeclaration.id == decl_id,
                    CulturalConstructionFeeDeclaration.company_id == company_id
                ).first()
            except:
                pass
        
        if not decl:
            # 从文件名或Excel内容推断期间
            period = None
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(values_only=True))
                for i, row in enumerate(rows[:30]):
                    for j, cell in enumerate(row):
                        if cell and isinstance(cell, str) and ('所属期' in cell or '申报期' in cell):
                            for k in range(j + 1, min(j + 4, len(row))):
                                val = row[k]
                                if val:
                                    m = re.search(r'(\d{4})[年\-](\d{1,2})', str(val))
                                    if m:
                                        period = f"{m.group(1)}-{m.group(2).zfill(2)}"
                                        break
                            if not period and i + 1 < len(rows):
                                next_row = rows[i + 1]
                                if j < len(next_row) and next_row[j]:
                                    m = re.search(r'(\d{4})[年\-](\d{1,2})', str(next_row[j]))
                                    if m:
                                        period = f"{m.group(1)}-{m.group(2).zfill(2)}"
                        if period:
                            break
                    if period:
                        break
                if period:
                    break
            
            if not period:
                match = re.search(r'(\d{4})[年\-](\d{1,2})', file.filename)
                if match:
                    period = f"{match.group(1)}-{match.group(2).zfill(2)}"
                else:
                    now = datetime.now()
                    period = f"{now.year}-{now.month:02d}"
            
            # 创建新申报表
            max_id = db.query(func.max(CulturalConstructionFeeDeclaration.id)).filter(
                CulturalConstructionFeeDeclaration.company_id == company_id
            ).scalar() or 0
            
            decl = CulturalConstructionFeeDeclaration(
                id=max_id + 1,
                company_id=company_id,
                period=period,
                status="已导入",
                created_at=datetime.now(),
                updated_at=datetime.now()
            )
            db.add(decl)
            db.flush()
        
        # 解析Excel文件，提取申报表数据
        form_main = None
        form_deduction = None
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # 将工作表转换为2D数组
            data = []
            for row in ws.iter_rows(values_only=True):
                data.append(list(row))
            
            # 根据工作表名称判断类型（精确匹配：扣除清单优先，主表最后匹配）
            sn = sheet_name
            if '扣除' in sn or 'deduction' in sn.lower() or '清单' in sn:
                form_deduction = data
            elif '主表' in sn or '申报表' in sn or '文化' in sn:
                form_main = data
        
        # 保存到数据库
        if form_main:
            decl.form_main = json.dumps(form_main, ensure_ascii=False, default=str)
        if form_deduction:
            decl.form_deduction = json.dumps(form_deduction, ensure_ascii=False, default=str)
        
        decl.status = "已导入"
        decl.updated_at = datetime.now()
        
        db.commit()
        
        return {
            "success": True,
            "message": "申报表导入成功",
            "declaration_id": decl.id,
            "period": decl.period
        }
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=400, detail=f"文件解析失败: {str(e)}")
    finally:
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)


def _ccf_build_form_main_from_db(decl):
    """从 ORM 列构建 form_main dict（JSON 为空时的 fallback）"""
    row_keys = [
        "row1_taxable_income", "row2_tax_exempt_income",
        "row3_deduction_beginning", "row4_deduction_current_period",
        "row5_taxable_income_deduction", "row6_tax_exempt_deduction",
        "row7_deduction_ending_balance", "row8_taxable_sales",
        "row9_fee_rate",
        "row10_payable_fee", "row10a_fee_reduction",
        "row11_unpaid_beginning",
        "row12_paid_current_period", "row13_prepaid",
        "row14_paid_last_period", "row15_paid_arrears",
        "row16_unpaid_ending", "row17_arrears",
        "row18_fill_refund", "row19_inspected_supplement",
    ]
    result = {}
    for key in row_keys:
        if key == "row9_fee_rate":
            result["row9_fee_rate"] = float(getattr(decl, "row9_fee_rate", 0.03) or 0.03)
        elif key == "row10a_fee_reduction":
            result["row10a_fee_reduction_current"] = float(getattr(decl, "row10a_fee_reduction_current", 0) or 0)
            result["row10a_fee_reduction_ytd"] = float(getattr(decl, "row10a_fee_reduction_ytd", 0) or 0)
        else:
            result[key + "_current"] = float(getattr(decl, key + "_current", 0) or 0)
            result[key + "_ytd"] = float(getattr(decl, key + "_ytd", 0) or 0)
    result["fee_reduction_rate"] = float(getattr(decl, "fee_reduction_rate", 0.5) or 0.5)
    return result


# ==================== Pydantic 模型 ====================

class DeductionItem(BaseModel):
    seq: int = 0
    invoice_supplier_tax_no: str = ""
    invoice_supplier_name: str = ""
    service_item_name: str = ""
    voucher_type: str = ""
    voucher_no: str = ""
    amount: float = 0.0


class DeclarationCreate(BaseModel):
    period: str = ""
    taxpayer_name: str = ""
    note: str = ""
    # 主表栏次（本月数 / 本年累计）
    row1_taxable_income_current: float = 0.0
    row1_taxable_income_ytd: float = 0.0
    row2_tax_exempt_income_current: float = 0.0
    row2_tax_exempt_income_ytd: float = 0.0
    row3_deduction_beginning_current: float = 0.0
    row3_deduction_beginning_ytd: float = 0.0
    row4_deduction_current_period_current: float = 0.0
    row4_deduction_current_period_ytd: float = 0.0
    row5_taxable_income_deduction_current: float = 0.0
    row5_taxable_income_deduction_ytd: float = 0.0
    row6_tax_exempt_deduction_current: float = 0.0
    row6_tax_exempt_deduction_ytd: float = 0.0
    row7_deduction_ending_balance_current: float = 0.0
    row7_deduction_ending_balance_ytd: float = 0.0
    row8_taxable_sales_current: float = 0.0
    row8_taxable_sales_ytd: float = 0.0
    row9_fee_rate: float = 0.03
    row10_payable_fee_current: float = 0.0
    row10_payable_fee_ytd: float = 0.0
    row10a_fee_reduction_current: float = 0.0
    row10a_fee_reduction_ytd: float = 0.0
    fee_reduction_rate: float = 0.5
    row11_unpaid_beginning_current: float = 0.0
    row11_unpaid_beginning_ytd: float = 0.0
    row12_paid_current_period_current: float = 0.0
    row12_paid_current_period_ytd: float = 0.0
    row13_prepaid_current: float = 0.0
    row13_prepaid_ytd: float = 0.0
    row14_paid_last_period_current: float = 0.0
    row14_paid_last_period_ytd: float = 0.0
    row15_paid_arrears_current: float = 0.0
    row15_paid_arrears_ytd: float = 0.0
    row16_unpaid_ending_current: float = 0.0
    row16_unpaid_ending_ytd: float = 0.0
    row17_arrears_current: float = 0.0
    row17_arrears_ytd: float = 0.0
    row18_fill_refund_current: float = 0.0
    row18_fill_refund_ytd: float = 0.0
    row19_inspected_supplement_current: float = 0.0
    row19_inspected_supplement_ytd: float = 0.0
    # JSON 表单数据
    form_main: dict = {}
    form_deduction: dict = {}
    # 扣除项目清单
    deductions: List[DeductionItem] = []


class DeclarationUpdate(DeclarationCreate):
    pass


# ==================== 统计卡片 ====================

@router.get("/stats")
def get_stats(company_id: int = Query(), db: Session = Depends(get_db)):
    """统计卡片数据"""
    declarations = db.query(CulturalConstructionFeeDeclaration).filter(
        CulturalConstructionFeeDeclaration.company_id == company_id
    ).all()
    count = len(declarations)
    total_taxable_income = sum(float(d.row1_taxable_income_current or 0) for d in declarations)
    total_fee = sum(float(d.row10_payable_fee_current or 0) for d in declarations)
    total_fill_refund = sum(float(d.row18_fill_refund_current or 0) for d in declarations)
    return {
        "count": count,
        "total_taxable_income": round(total_taxable_income, 2),
        "total_fee": round(total_fee, 2),
        "total_fill_refund": round(total_fill_refund, 2),
    }


# ==================== CRUD ====================

@router.get("/declarations")
def list_declarations(
    company_id: int = Query(),
    period: str = Query(None),
    period_from: str = Query(None),
    period_to: str = Query(None),
    db: Session = Depends(get_db),
):
    q = db.query(CulturalConstructionFeeDeclaration).filter(
        CulturalConstructionFeeDeclaration.company_id == company_id
    )
    if period:
        q = q.filter(CulturalConstructionFeeDeclaration.period == period)
    if period_from:
        q = q.filter(CulturalConstructionFeeDeclaration.period >= period_from)
    if period_to:
        q = q.filter(CulturalConstructionFeeDeclaration.period <= period_to)
    items = q.order_by(CulturalConstructionFeeDeclaration.period.desc()).all()
    result = []
    for item in items:
        deduction_count = db.query(CulturalConstructionFeeDeduction).filter(
            CulturalConstructionFeeDeduction.declaration_id == item.id
        ).count()
        result.append({
            "id": item.id,
            "company_id": item.company_id,
            "period": item.period,
            "status": item.status,
            "note": item.note,
            "taxpayer_name": item.taxpayer_name,
            "taxpayer_id": item.taxpayer_id,
            "fill_date": str(item.fill_date) if item.fill_date else None,
            "deduction_count": deduction_count,
            "payable_fee_current": float(item.row10_payable_fee_current or 0),
            "payable_fee_ytd": float(item.row10_payable_fee_ytd or 0),
            "fill_refund_current": float(item.row18_fill_refund_current or 0),
            "fill_refund_ytd": float(item.row18_fill_refund_ytd or 0),
            "created_at": item.created_at.isoformat() if item.created_at else None,
            "updated_at": item.updated_at.isoformat() if item.updated_at else None,
        })
    return {"items": result, "total": len(result)}


@router.get("/declarations/{declaration_id}")
def get_declaration(declaration_id: int, company_id: int = Query(), db: Session = Depends(get_db)):
    decl = db.query(CulturalConstructionFeeDeclaration).filter(
        CulturalConstructionFeeDeclaration.id == declaration_id,
        CulturalConstructionFeeDeclaration.company_id == company_id,
    ).first()
    if not decl:
        raise HTTPException(404, "申报记录不存在")

    deductions = db.query(CulturalConstructionFeeDeduction).filter(
        CulturalConstructionFeeDeduction.declaration_id == declaration_id
    ).order_by(CulturalConstructionFeeDeduction.seq).all()

    # 从 JSON 读取，若为空则从 DB 列合成
    form_main = {}
    form_deduction = {}
    try:
        if decl.form_main:
            form_main = json.loads(decl.form_main) if isinstance(decl.form_main, str) else decl.form_main
    except Exception:
        form_main = {}
    if not form_main:
        form_main = _ccf_build_form_main_from_db(decl)
    try:
        if decl.form_deduction:
            form_deduction = json.loads(decl.form_deduction) if isinstance(decl.form_deduction, str) else decl.form_deduction
    except Exception:
        form_deduction = {}

    return {
        "id": decl.id,
        "company_id": decl.company_id,
        "period": decl.period,
        "status": decl.status,
        "note": decl.note,
        "taxpayer_name": decl.taxpayer_name,
        "taxpayer_id": decl.taxpayer_id,
        "fill_date": str(decl.fill_date) if decl.fill_date else None,
        "form_main": form_main,
        "form_deduction": form_deduction,
        "deductions": [{
            "id": d.id,
            "seq": d.seq,
            "invoice_supplier_tax_no": d.invoice_supplier_tax_no,
            "invoice_supplier_name": d.invoice_supplier_name,
            "service_item_name": d.service_item_name,
            "voucher_type": d.voucher_type,
            "voucher_no": d.voucher_no,
            "amount": float(d.amount or 0),
        } for d in deductions],
        "created_at": decl.created_at.isoformat() if decl.created_at else None,
        "updated_at": decl.updated_at.isoformat() if decl.updated_at else None,
    }


@router.post("/declarations")
def create_declaration(payload: DeclarationCreate, company_id: int = Query(), db: Session = Depends(get_db)):
    period = payload.period
    if not period:
        raise HTTPException(400, "税款所属期不能为空")
    company = db.query(Company).filter(Company.id == company_id).first()
    if not company:
        raise HTTPException(404, "公司不存在")

    # 按期间去重
    existing = db.query(CulturalConstructionFeeDeclaration).filter(
        CulturalConstructionFeeDeclaration.company_id == company_id,
        CulturalConstructionFeeDeclaration.period == period,
    ).first()
    if existing:
        raise HTTPException(400, f"期间 {period} 已有申报记录")

    decl = CulturalConstructionFeeDeclaration(
        company_id=company_id,
        period=period,
        taxpayer_name=payload.taxpayer_name or company.name,
        taxpayer_id=company.uscc or "",
        status="草稿",
        note=payload.note or "",
        fill_date=datetime.now().date(),
    )
    # 主表栏次（通过 setattr 批量赋值）
    for field in [
        "row1_taxable_income_current", "row1_taxable_income_ytd",
        "row2_tax_exempt_income_current", "row2_tax_exempt_income_ytd",
        "row3_deduction_beginning_current", "row3_deduction_beginning_ytd",
        "row4_deduction_current_period_current", "row4_deduction_current_period_ytd",
        "row5_taxable_income_deduction_current", "row5_taxable_income_deduction_ytd",
        "row6_tax_exempt_deduction_current", "row6_tax_exempt_deduction_ytd",
        "row7_deduction_ending_balance_current", "row7_deduction_ending_balance_ytd",
        "row8_taxable_sales_current", "row8_taxable_sales_ytd",
        "row10_payable_fee_current", "row10_payable_fee_ytd",
        "row11_unpaid_beginning_current", "row11_unpaid_beginning_ytd",
        "row12_paid_current_period_current", "row12_paid_current_period_ytd",
        "row13_prepaid_current", "row13_prepaid_ytd",
        "row14_paid_last_period_current", "row14_paid_last_period_ytd",
        "row15_paid_arrears_current", "row15_paid_arrears_ytd",
        "row16_unpaid_ending_current", "row16_unpaid_ending_ytd",
        "row17_arrears_current", "row17_arrears_ytd",
        "row18_fill_refund_current", "row18_fill_refund_ytd",
        "row19_inspected_supplement_current", "row19_inspected_supplement_ytd",
    ]:
        if hasattr(payload, field):
            setattr(decl, field, getattr(payload, field, 0.0))

    if hasattr(payload, "row9_fee_rate"):
        decl.row9_fee_rate = payload.row9_fee_rate or 0.03

    # 保存 JSON 表单
    if payload.form_main:
        decl.form_main = json.dumps(payload.form_main, ensure_ascii=False)
    if payload.form_deduction:
        decl.form_deduction = json.dumps(payload.form_deduction, ensure_ascii=False)

    db.add(decl)
    db.flush()

    # 添加扣除项目清单
    for i, d in enumerate(payload.deductions):
        deduction = CulturalConstructionFeeDeduction(
            declaration_id=decl.id,
            seq=i + 1,
            invoice_supplier_tax_no=d.invoice_supplier_tax_no,
            invoice_supplier_name=d.invoice_supplier_name,
            service_item_name=d.service_item_name,
            voucher_type=d.voucher_type,
            voucher_no=d.voucher_no,
            amount=d.amount,
        )
        db.add(deduction)

    db.commit()
    db.refresh(decl)
    return {"id": decl.id, "message": "创建成功"}


@router.put("/declarations/{declaration_id}")
def update_declaration(declaration_id: int, payload: dict, company_id: int = Query(), db: Session = Depends(get_db)):
    decl = db.query(CulturalConstructionFeeDeclaration).filter(
        CulturalConstructionFeeDeclaration.id == declaration_id,
        CulturalConstructionFeeDeclaration.company_id == company_id,
    ).first()
    if not decl:
        raise HTTPException(404, "申报记录不存在")

    # 更新基础字段
    if "period" in payload:
        decl.period = payload["period"]
    if "taxpayer_name" in payload:
        decl.taxpayer_name = payload["taxpayer_name"]
    if "status" in payload:
        decl.status = payload["status"]
    if "note" in payload:
        decl.note = payload["note"]
    decl.fill_date = datetime.now().date()

    # 更新主表栏次
    field_map = {
        "row1_taxable_income_current": "row1_taxable_income_current",
        "row1_taxable_income_ytd": "row1_taxable_income_ytd",
        "row2_tax_exempt_income_current": "row2_tax_exempt_income_current",
        "row2_tax_exempt_income_ytd": "row2_tax_exempt_income_ytd",
        "row3_deduction_beginning_current": "row3_deduction_beginning_current",
        "row3_deduction_beginning_ytd": "row3_deduction_beginning_ytd",
        "row4_deduction_current_period_current": "row4_deduction_current_period_current",
        "row4_deduction_current_period_ytd": "row4_deduction_current_period_ytd",
        "row5_taxable_income_deduction_current": "row5_taxable_income_deduction_current",
        "row5_taxable_income_deduction_ytd": "row5_taxable_income_deduction_ytd",
        "row6_tax_exempt_deduction_current": "row6_tax_exempt_deduction_current",
        "row6_tax_exempt_deduction_ytd": "row6_tax_exempt_deduction_ytd",
        "row7_deduction_ending_balance_current": "row7_deduction_ending_balance_current",
        "row7_deduction_ending_balance_ytd": "row7_deduction_ending_balance_ytd",
        "row8_taxable_sales_current": "row8_taxable_sales_current",
        "row8_taxable_sales_ytd": "row8_taxable_sales_ytd",
        "row9_fee_rate": "row9_fee_rate",
        "row10_payable_fee_current": "row10_payable_fee_current",
        "row10_payable_fee_ytd": "row10_payable_fee_ytd",
        "row11_unpaid_beginning_current": "row11_unpaid_beginning_current",
        "row11_unpaid_beginning_ytd": "row11_unpaid_beginning_ytd",
        "row12_paid_current_period_current": "row12_paid_current_period_current",
        "row12_paid_current_period_ytd": "row12_paid_current_period_ytd",
        "row13_prepaid_current": "row13_prepaid_current",
        "row13_prepaid_ytd": "row13_prepaid_ytd",
        "row14_paid_last_period_current": "row14_paid_last_period_current",
        "row14_paid_last_period_ytd": "row14_paid_last_period_ytd",
        "row15_paid_arrears_current": "row15_paid_arrears_current",
        "row15_paid_arrears_ytd": "row15_paid_arrears_ytd",
        "row16_unpaid_ending_current": "row16_unpaid_ending_current",
        "row16_unpaid_ending_ytd": "row16_unpaid_ending_ytd",
        "row17_arrears_current": "row17_arrears_current",
        "row17_arrears_ytd": "row17_arrears_ytd",
        "row18_fill_refund_current": "row18_fill_refund_current",
        "row18_fill_refund_ytd": "row18_fill_refund_ytd",
        "row19_inspected_supplement_current": "row19_inspected_supplement_current",
        "row19_inspected_supplement_ytd": "row19_inspected_supplement_ytd",
    }
    for key in field_map:
        if key in payload:
            setattr(decl, key, payload[key])

    # 更新 JSON 表单
    if "form_main" in payload:
        decl.form_main = json.dumps(payload["form_main"], ensure_ascii=False)
    if "form_deduction" in payload:
        decl.form_deduction = json.dumps(payload["form_deduction"], ensure_ascii=False)

    # 替换扣除项目清单
    db.query(CulturalConstructionFeeDeduction).filter(
        CulturalConstructionFeeDeduction.declaration_id == declaration_id
    ).delete()
    for i, d in enumerate(payload.get("deductions", [])):
        deduction = CulturalConstructionFeeDeduction(
            declaration_id=declaration_id,
            seq=i + 1,
            invoice_supplier_tax_no=d.get("invoice_supplier_tax_no", ""),
            invoice_supplier_name=d.get("invoice_supplier_name", ""),
            service_item_name=d.get("service_item_name", ""),
            voucher_type=d.get("voucher_type", ""),
            voucher_no=d.get("voucher_no", ""),
            amount=d.get("amount", 0.0),
        )
        db.add(deduction)

    decl.updated_at = datetime.now()
    db.commit()
    return {"message": "更新成功"}


@router.delete("/declarations/{declaration_id}")
def delete_declaration(declaration_id: int, company_id: int = Query(), db: Session = Depends(get_db)):
    decl = db.query(CulturalConstructionFeeDeclaration).filter(
        CulturalConstructionFeeDeclaration.id == declaration_id,
        CulturalConstructionFeeDeclaration.company_id == company_id,
    ).first()
    if not decl:
        raise HTTPException(404, "申报记录不存在")
    db.query(CulturalConstructionFeeDeduction).filter(
        CulturalConstructionFeeDeduction.declaration_id == declaration_id
    ).delete()
    db.delete(decl)
    db.commit()
    return {"message": "删除成功"}


# ==================== 自动计算 ====================

@router.post("/declarations/{declaration_id}/auto-calculate")
def auto_calculate(declaration_id: int, company_id: int = Query(), db: Session = Depends(get_db)):
    """自动计算文化事业建设费申报表各栏次"""
    decl = db.query(CulturalConstructionFeeDeclaration).filter(
        CulturalConstructionFeeDeclaration.id == declaration_id,
        CulturalConstructionFeeDeclaration.company_id == company_id,
    ).first()
    if not decl:
        raise HTTPException(404, "申报记录不存在")

    # 栏次7 = 3 + 4 - 5 - 6
    decl.row7_deduction_ending_balance_current = round(
        (decl.row3_deduction_beginning_current or 0)
        + (decl.row4_deduction_current_period_current or 0)
        - (decl.row5_taxable_income_deduction_current or 0)
        - (decl.row6_tax_exempt_deduction_current or 0),
        2
    )

    # 栏次8 = 1 - 5
    decl.row8_taxable_sales_current = round(
        (decl.row1_taxable_income_current or 0)
        - (decl.row5_taxable_income_deduction_current or 0),
        2
    )

    # 栏次10 = 8 × 9
    decl.row10_payable_fee_current = round(
        (decl.row8_taxable_sales_current or 0) * (decl.row9_fee_rate or 0.03),
        2
    )

    # 栏次12 = 13 + 14 + 15
    decl.row12_paid_current_period_current = round(
        (decl.row13_prepaid_current or 0)
        + (decl.row14_paid_last_period_current or 0)
        + (decl.row15_paid_arrears_current or 0),
        2
    )

    # 栏次16 = 10 + 11 - 12
    decl.row16_unpaid_ending_current = round(
        (decl.row10_payable_fee_current or 0)
        + (decl.row11_unpaid_beginning_current or 0)
        - (decl.row12_paid_current_period_current or 0),
        2
    )

    # 栏次17 = 11 - 14 - 15
    decl.row17_arrears_current = round(
        (decl.row11_unpaid_beginning_current or 0)
        - (decl.row14_paid_last_period_current or 0)
        - (decl.row15_paid_arrears_current or 0),
        2
    )

    # 栏次18 = 10 - 13
    decl.row18_fill_refund_current = round(
        (decl.row10_payable_fee_current or 0)
        - (decl.row13_prepaid_current or 0),
        2
    )

    # 本年累计 = 本月数 + 历史期间数据
    # 简化：暂不实现 YTD 自动计算

    decl.updated_at = datetime.now()
    db.commit()
    return {"message": "自动计算完成"}


# ==================== AI 自动填列 ====================

# 文化事业建设费 — 广告服务关键词（用于匹配销项/进项发票的货物/服务名称）
CCF_AD_KEYWORDS = ["广告服务", "广告发布费", "广告制作费", "广告代理", "广告策划", "广告设计"]

# 文化事业建设费 — 非广告服务关键词（有广告字眼但实际不是广告服务）
CCF_AD_EXCLUDE = ["非广告", "非广告服务"]


def _get_next_voucher_no(db, company_id: int) -> int:
    """获取下一个凭证号"""
    max_no = db.query(func.max(JournalEntry.voucher_no)).filter(
        JournalEntry.company_id == company_id
    ).scalar()
    return (max_no or 0) + 1


def _get_or_create_account(db, company_id: int, code: str, name: str,
                           parent_code: str = None, category: str = "负债",
                           balance_direction: str = "贷") -> Account:
    """查找科目，不存在则自动创建"""
    acct = db.query(Account).filter(
        Account.company_id == company_id,
        Account.code == code,
    ).first()
    if not acct:
        acct = Account(
            company_id=company_id,
            code=code,
            name=name,
            parent_code=parent_code,
            category=category,
            balance_direction=balance_direction,
        )
        db.add(acct)
        db.flush()
    return acct


def _generate_ccf_voucher(db, company_id: int, decl, fee_amount: float,
                          entry_date, period: str) -> dict:
    """生成文化事业建设费计提凭证（借：税金及附加 / 贷：应交文化事业建设费）"""
    # 确保科目存在
    tax_account = _get_or_create_account(
        db, company_id, "6403", "税金及附加",
        category="损益", balance_direction="借"
    )
    ccf_account = _get_or_create_account(
        db, company_id, "221016", "应交文化事业建设费",
        parent_code="2210", category="负债", balance_direction="贷"
    )

    # 检查是否已存在同日同金额的 CCF 凭证
    existing = db.query(JournalEntry).filter(
        JournalEntry.company_id == company_id,
        JournalEntry.source == "文化事业建设费",
        JournalEntry.entry_date == entry_date,
        JournalEntry.debit_amount == fee_amount,
    ).first()
    if existing:
        vno = existing.voucher_no
        return {"voucher_no": vno, "exists": True, "message": f"凭证 记-{vno} 已存在，跳过生成"}

    vno = _get_next_voucher_no(db, company_id)
    summary = f"计提文化事业建设费-{period}"

    # 借方：税金及附加
    db.add(JournalEntry(
        company_id=company_id,
        entry_date=entry_date,
        period=period,
        voucher_word="记",
        voucher_no=vno,
        summary=summary,
        account_code=tax_account.code,
        account_name=tax_account.name,
        debit_amount=fee_amount,
        credit_amount=0,
        source="文化事业建设费",
        ref_id=decl.id,
        attach_count=0,
    ))

    # 贷方：应交文化事业建设费
    db.add(JournalEntry(
        company_id=company_id,
        entry_date=entry_date,
        period=period,
        voucher_word="记",
        voucher_no=vno,
        summary=summary,
        account_code=ccf_account.code,
        account_name=ccf_account.name,
        debit_amount=0,
        credit_amount=fee_amount,
        source="文化事业建设费",
        ref_id=decl.id,
        attach_count=0,
    ))

    db.flush()
    return {"voucher_no": vno, "exists": False, "message": f"已生成凭证 记-{vno}"}


def _match_ccf_payment_voucher(db, company_id: int, decl, fee_amount: float,
                               entry_date, period: str) -> dict:
    """匹配银行流水中的文化事业建设费支付记录"""
    from database import BankTransaction

    # 查找银行流水中对方户名含"国家金库"且金额匹配的支付
    bank_txs = db.query(BankTransaction).filter(
        BankTransaction.company_id == company_id,
        BankTransaction.transaction_date >= entry_date,
        BankTransaction.credit_amount == fee_amount,
        BankTransaction.counterparty_name.contains("国家金库"),
    ).all()

    if not bank_txs:
        return {"matched": False, "message": "未找到匹配的银行支付流水"}

    results = []
    for tx in bank_txs:
        # 检查是否已有支付凭证
        existing_je = db.query(JournalEntry).filter(
            JournalEntry.company_id == company_id,
            JournalEntry.source == "文化事业建设费缴纳",
            JournalEntry.entry_date == tx.transaction_date,
        ).first()
        if existing_je:
            results.append({"voucher_no": existing_je.voucher_no, "matched": True,
                          "message": f"支付凭证 记-{existing_je.voucher_no} 已存在",
                          "bank_tx_id": tx.id})
            continue

        vno = _get_next_voucher_no(db, company_id)
        summary = f"缴纳文化事业建设费-{period}"

        # 确保科目存在
        ccf_account = _get_or_create_account(
            db, company_id, "221016", "应交文化事业建设费",
            parent_code="2210", category="负债", balance_direction="贷"
        )
        bank_account = db.query(Account).filter(
            Account.company_id == company_id,
            Account.code == "1002",
        ).first()
        if not bank_account:
            bank_account = _get_or_create_account(
                db, company_id, "1002", "银行存款",
                category="资产", balance_direction="借"
            )

        # 借方：应交文化事业建设费（冲减）
        db.add(JournalEntry(
            company_id=company_id,
            entry_date=tx.transaction_date,
            period=period,
            voucher_word="记",
            voucher_no=vno,
            summary=summary,
            account_code=ccf_account.code,
            account_name=ccf_account.name,
            debit_amount=fee_amount,
            credit_amount=0,
            source="文化事业建设费缴纳",
            ref_id=decl.id,
        ))

        # 贷方：银行存款
        db.add(JournalEntry(
            company_id=company_id,
            entry_date=tx.transaction_date,
            period=period,
            voucher_word="记",
            voucher_no=vno,
            summary=summary,
            account_code=bank_account.code,
            account_name=bank_account.name,
            debit_amount=0,
            credit_amount=fee_amount,
            source="文化事业建设费缴纳",
            ref_id=decl.id,
        ))

        db.flush()
        # 回写银行流水的凭证号
        tx.journal_voucher_no = f"记-{vno}"
        decl.row15_paid_arrears_current = fee_amount
        results.append({"voucher_no": vno, "matched": True,
                       "message": f"已生成支付凭证 记-{vno}",
                       "bank_tx_id": tx.id})

    return {"matched": True, "results": results}


@router.post("/declarations/{declaration_id}/ai-auto-fill")
def ai_auto_fill(declaration_id: int, company_id: int = Query(),
                 db: Session = Depends(get_db)):
    """
    🤖 AI 自动填列文化事业建设费申报表
    1. 从销项发票扫描广告服务含税收入 → 应征收入
    2. 从进项/记账发票扫描可扣除的广告相关费用 → 扣除项目清单
    3. 自动计算主表各栏次
    4. 自动生成计提凭证到序时账
    5. 自动匹配银行流水生成缴纳凭证
    6. 自动建档供应商
    """
    decl = db.query(CulturalConstructionFeeDeclaration).filter(
        CulturalConstructionFeeDeclaration.id == declaration_id,
        CulturalConstructionFeeDeclaration.company_id == company_id,
    ).first()
    if not decl:
        raise HTTPException(404, "申报记录不存在")

    period = decl.period  # "2025-10"
    try:
        year, month = period.split("-")
        year, month = int(year), int(month)
    except:
        raise HTTPException(400, f"期间格式错误：{period}")

    # 计算期间起止日期
    from calendar import monthrange
    _, last_day = monthrange(year, month)
    period_start = f"{year}-{month:02d}-01"
    period_end = f"{year}-{month:02d}-{last_day:02d}"
    entry_date = datetime(year, month, last_day).date()

    log = []  # 操作日志

    # ========== 1. 扫描销项发票 → 应征收入 ==========
    # 筛选含"广告服务"关键词的正数发票
    sales_invoices = db.query(SalesInvoice).filter(
        SalesInvoice.company_id == company_id,
        SalesInvoice.invoice_date.between(period_start, period_end),
        SalesInvoice.total_amount > 0,
    ).all()

    taxable_income = 0.0
    exempt_income = 0.0
    taxable_invoices = []
    for si in sales_invoices:
        goods = (si.goods_name or "").lower()
        # 匹配广告服务关键词
        is_ad = any(kw.lower() in goods for kw in CCF_AD_KEYWORDS)
        is_exclude = any(kw.lower() in goods for kw in CCF_AD_EXCLUDE)
        if is_ad and not is_exclude:
            taxable_income += float(si.total_amount or 0)
            taxable_invoices.append(si)
            log.append(f"  销项发票 #{si.id} 广告服务 含税{si.total_amount:,.2f}")
        elif is_ad:
            exempt_income += float(si.total_amount or 0)

    taxable_income = round(taxable_income, 2)
    log.insert(0, f"📊 应征收入（广告服务含税）= {taxable_income:,.2f}（{len(taxable_invoices)} 张发票）")

    # 更新主表
    decl.row1_taxable_income_current = taxable_income
    decl.row1_taxable_income_ytd = taxable_income
    decl.row2_tax_exempt_income_current = 0  # 有应征收入时免征为0
    decl.row2_tax_exempt_income_ytd = 0

    # ========== 2. 扫描进项/记账发票 → 扣除项目清单 ==========
    # 清空旧扣除项目
    db.query(CulturalConstructionFeeDeduction).filter(
        CulturalConstructionFeeDeduction.declaration_id == declaration_id
    ).delete()

    purchase_invoices = db.query(PurchaseInvoice).filter(
        PurchaseInvoice.company_id == company_id,
        PurchaseInvoice.invoice_date.between(period_start, period_end),
        PurchaseInvoice.total_amount > 0,
    ).all()

    bookkeeping_invoices = db.query(BookkeepingInvoice).filter(
        BookkeepingInvoice.company_id == company_id,
        BookkeepingInvoice.invoice_date.between(period_start, period_end),
        BookkeepingInvoice.total_amount > 0,
    ).all()

    deduction_total = 0.0
    deduction_count = 0
    all_purchase = list(purchase_invoices) + list(bookkeeping_invoices)

    for inv in all_purchase:
        goods = (getattr(inv, 'goods_name', '') or '').lower()
        is_ad = any(kw.lower() in goods for kw in CCF_AD_KEYWORDS)
        is_exclude = any(kw.lower() in goods for kw in CCF_AD_EXCLUDE)
        if is_ad and not is_exclude:
            amount = float(getattr(inv, 'total_amount', 0) or 0)
            seller_name = getattr(inv, 'seller_name', '') or getattr(inv, 'party_name', '') or ''
            deduction_total += amount
            deduction_count += 1
            # 判断发票来源类型
            inv_type = "purchase" if isinstance(inv, PurchaseInvoice) else "bookkeeping"
            # 取纳税人识别号和发票号
            tax_no = ""
            inv_no = ""
            if inv_type == "purchase":
                tax_no = getattr(inv, 'seller_tax_no', '') or ''
                # 优先数电发票号码，其次传统发票代码+号码
                digital = getattr(inv, 'digital_invoice_no', '') or ''
                inv_code = getattr(inv, 'invoice_code', '') or ''
                inv_no_raw = getattr(inv, 'invoice_no', '') or ''
                if digital:
                    inv_no = digital
                elif inv_code or inv_no_raw:
                    inv_no = (inv_code + inv_no_raw).strip()
                else:
                    inv_no = digital  # 都为空时保持空字符串
            else:
                # 记账发票用 party_tax_no
                tax_no = getattr(inv, 'party_tax_no', '') or ''
                digital = getattr(inv, 'digital_invoice_no', '') or ''
                inv_code = getattr(inv, 'invoice_code', '') or ''
                inv_no_raw = getattr(inv, 'invoice_no', '') or ''
                if digital:
                    inv_no = digital
                elif inv_code or inv_no_raw:
                    inv_no = (inv_code + inv_no_raw).strip()
                else:
                    inv_no = ''
            deduction = CulturalConstructionFeeDeduction(
                declaration_id=declaration_id,
                seq=deduction_count,
                invoice_supplier_tax_no=tax_no[:50] if tax_no else "",
                invoice_supplier_name=seller_name[:100] if seller_name else "",
                service_item_name=(getattr(inv, 'goods_name', '') or '')[:100],
                voucher_type="增值税专用发票" if inv_type == "purchase" else "增值税普通发票",
                voucher_no=inv_no[:50] if inv_no else "",
                amount=amount,
            )
            db.add(deduction)
            log.append(f"  进项发票 #{inv.id} 广告扣除 {seller_name[:15]} {amount:,.2f}")

    # 如无扣除项目则保持空
    if deduction_count == 0:
        log.append("📋 扣除项目清单：本期无广告服务进项发票，清单为空")

    # ========== 2.5 自动建档供应商（只建 CCF 扣除相关的，且去重） ==========
    # 只对扣除项目中的供应商建档
    seen_suppliers = set()
    for d in db.query(CulturalConstructionFeeDeduction).filter(
        CulturalConstructionFeeDeduction.declaration_id == declaration_id
    ).all():
        seller_name = d.invoice_supplier_name
        if not seller_name or seller_name in seen_suppliers:
            continue
        seen_suppliers.add(seller_name)
        existing_supplier = db.query(Supplier).filter(
            Supplier.company_id == company_id,
            Supplier.name == seller_name,
        ).first()
        if not existing_supplier:
            # 获取最大编码
            max_code = db.query(func.max(Supplier.code)).filter(
                Supplier.company_id == company_id
            ).scalar()
            try:
                next_num = int(max_code.replace('GYS', '')) + 1 if max_code else 1
            except:
                next_num = 1
            next_code = f"GYS{next_num:03d}"
            db.add(Supplier(
                company_id=company_id,
                code=next_code,
                name=seller_name,
            ))
            db.flush()  # 立即 flush 确保下个供应商编码不重复
            log.append(f"  🏷️ 自动建档供应商：{seller_name}（{next_code}）")

    # 扣除项目相关栏次
    # 取前期期末余额
    prior_period = f"{year}-{month-1:02d}" if month > 1 else f"{year-1}-12"
    prior_decl = db.query(CulturalConstructionFeeDeclaration).filter(
        CulturalConstructionFeeDeclaration.company_id == company_id,
        CulturalConstructionFeeDeclaration.period == prior_period,
    ).first()
    r3_beginning = float(prior_decl.row7_deduction_ending_balance_current or 0) if prior_decl else 0

    decl.row3_deduction_beginning_current = r3_beginning
    decl.row3_deduction_beginning_ytd = r3_beginning
    decl.row4_deduction_current_period_current = round(deduction_total, 2)
    decl.row4_deduction_current_period_ytd = round(deduction_total, 2)

    # 第5栏：应征收入减除额 = min(期初+本期, 应征收入)
    available_deduction = r3_beginning + deduction_total
    decl.row5_taxable_income_deduction_current = round(min(available_deduction, taxable_income), 2)
    decl.row5_taxable_income_deduction_ytd = round(min(available_deduction, taxable_income), 2)

    # 第6栏：免征收入减除额（暂无免征收入，为0）
    decl.row6_tax_exempt_deduction_current = 0
    decl.row6_tax_exempt_deduction_ytd = 0

    # ========== 3. 自动计算 ==========
    # 栏次7 = 3+4-5-6
    decl.row7_deduction_ending_balance_current = round(
        r3_beginning + deduction_total
        - float(decl.row5_taxable_income_deduction_current or 0)
        - float(decl.row6_tax_exempt_deduction_current or 0), 2
    )
    decl.row7_deduction_ending_balance_ytd = decl.row7_deduction_ending_balance_current

    # 栏次8 = 1-5
    decl.row8_taxable_sales_current = round(
        taxable_income - float(decl.row5_taxable_income_deduction_current or 0), 2
    )
    decl.row8_taxable_sales_ytd = decl.row8_taxable_sales_current

    # 费率
    decl.row9_fee_rate = 0.03

    # 栏次10 = 8×9
    decl.row10_payable_fee_current = round(
        float(decl.row8_taxable_sales_current or 0) * 0.03, 2
    )
    decl.row10_payable_fee_ytd = decl.row10_payable_fee_current

    # 栏次11 = 前期期末未缴费
    decl.row11_unpaid_beginning_current = float(prior_decl.row16_unpaid_ending_current or 0) if prior_decl else 0
    decl.row11_unpaid_beginning_ytd = decl.row11_unpaid_beginning_current

    # 栏次12 = 13+14+15
    decl.row12_paid_current_period_current = round(
        (decl.row13_prepaid_current or 0)
        + (decl.row14_paid_last_period_current or 0)
        + (decl.row15_paid_arrears_current or 0), 2
    )
    decl.row12_paid_current_period_ytd = decl.row12_paid_current_period_current

    # 栏次16 = 10+11-12
    decl.row16_unpaid_ending_current = round(
        (decl.row10_payable_fee_current or 0)
        + (decl.row11_unpaid_beginning_current or 0)
        - (decl.row12_paid_current_period_current or 0), 2
    )
    decl.row16_unpaid_ending_ytd = decl.row16_unpaid_ending_current

    # 栏次17 = 11-14-15
    decl.row17_arrears_current = round(
        (decl.row11_unpaid_beginning_current or 0)
        - (decl.row14_paid_last_period_current or 0)
        - (decl.row15_paid_arrears_current or 0), 2
    )
    decl.row17_arrears_ytd = decl.row17_arrears_current

    # 栏次18 = 10-13
    decl.row18_fill_refund_current = round(
        (decl.row10_payable_fee_current or 0)
        - (decl.row13_prepaid_current or 0), 2
    )
    decl.row18_fill_refund_ytd = decl.row18_fill_refund_current

    # 栏次19 = 0（默认）
    decl.row19_inspected_supplement_current = 0
    decl.row19_inspected_supplement_ytd = 0

    # 更新状态和时间
    decl.status = "已申报"
    decl.updated_at = datetime.now()

    fee_amount = float(decl.row10_payable_fee_current or 0)
    log.append(f"🧮 应缴费额 = {fee_amount:,.2f}（计费销售额 × 3%）")

    # ========== 4. 生成计提凭证到序时账 ==========
    if fee_amount > 0:
        voucher_result = _generate_ccf_voucher(db, company_id, decl, fee_amount, entry_date, period)
        log.append(f"📝 计提凭证：{voucher_result['message']}")
    else:
        log.append("📝 应缴费额为 0，跳过凭证生成")

    # ========== 5. 自动匹配银行流水生成缴纳凭证 ==========
    if fee_amount > 0:
        payment_result = _match_ccf_payment_voucher(db, company_id, decl, fee_amount, entry_date, period)
        if payment_result.get("matched") and payment_result.get("results"):
            for r in payment_result["results"]:
                log.append(f"💰 缴纳凭证：{r['message']}")
        else:
            log.append(f"💰 {payment_result.get('message', '未找到银行支付流水')}")

    # 状态更新
    decl.status = "已申报"
    decl.updated_at = datetime.now()

    db.commit()

    return {
        "success": True,
        "message": "AI 自动填列完成",
        "declaration_id": declaration_id,
        "period": period,
        "summary": {
            "taxable_income": taxable_income,
            "deduction_total": round(deduction_total, 2),
            "deduction_count": deduction_count,
            "taxable_sales": float(decl.row8_taxable_sales_current or 0),
            "fee_rate": 0.03,
            "payable_fee": fee_amount,
            "unpaid_ending": float(decl.row16_unpaid_ending_current or 0),
            "fill_refund": float(decl.row18_fill_refund_current or 0),
        },
        "log": log,
    }
