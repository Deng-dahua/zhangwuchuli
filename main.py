"""
中小制造业账务处理系统 - 后端 API
"""
from fastapi import FastAPI, Depends, HTTPException, Query, UploadFile, File, Form, Body
from fastapi.staticfiles import StaticFiles
from fastapi.responses import HTMLResponse, FileResponse
from sqlalchemy.orm import Session
from sqlalchemy.sql import text
from sqlalchemy import func, and_, or_
from pydantic import BaseModel
from typing import Optional, List
from datetime import date, datetime
from contextlib import asynccontextmanager
import os
import csv
import io
import re
import logging
import hashlib
import uuid
import openpyxl
import json
from pypdf import PdfReader

from database import (
    get_db, init_db, init_company_data,
    Company, Department, Employee, Customer, Supplier,
    Account, Period,
    FixedAsset, FixedAssetDepreciation,
    IntangibleAsset, IntangibleAssetAmortization,
    InventoryItem, InventoryTransaction, InventoryBalance,
    Contract, ContractPayment,
    Payment,
    SalesInvoice, PurchaseInvoice, BookkeepingInvoice,
    BankConfig, BankTransaction, BankRule,
    InputVATDeduction, ColumnTemplate, JournalEntry,
    SalaryRecord, VATDeclaration,
    CompanyShareholder, CompanyDirector, CompanySupervisor, CompanyFinanceContact,
    auto_generate_single_invoice,
    auto_generate_input_vat_for_period, auto_generate_input_vat_journals,
    _normalize_customer_name, _match_customer, _generate_bank_journals, _classify_bank_tx, _build_entity_index, _ensure_account,
    _generate_salary_journals, _generate_hf_accrual_journals, _match_hf_payment_journals,
    _match_ss_payment_journals, _match_tax_payment_journals,
    auto_generate_purchase_journal, auto_generate_bookkeeping_journal, _next_voucher_no, _classify_purchase_debit,
)

from vat import router as vat_router
from salary import router as salary_router
from social_security import router as social_security_router
from housing_fund import router as housing_fund_router
from cultural_construction_fee import router as cultural_construction_fee_router
from tax_risk import router as tax_risk_router


@asynccontextmanager
async def lifespan(app: FastAPI):
    """应用生命周期：启动时初始化数据库"""
    init_db()
    # 启动时不自动处理——单条/批量导入时已自动触发供应商建档+凭证生成+科目创建
    yield

app = FastAPI(title="账务处理系统", description="中小制造业账务管理系统", version="1.0.0", lifespan=lifespan)
# ==================== 开发模式：强制无缓存 ====================
@app.middleware("http")
async def add_cache_headers(request, call_next):
    """给所有响应加 no-cache 头，强制浏览器不用本地缓存。"""
    response = await call_next(request)
    response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    return response
app.include_router(vat_router)
app.include_router(salary_router)
app.include_router(social_security_router)
app.include_router(housing_fund_router)
app.include_router(cultural_construction_fee_router)
app.include_router(tax_risk_router)

# 挂载静态文件
app.mount("/static", StaticFiles(directory="static"), name="static")

# ==================== 文件上传安全常数 (P2-4/5) ====================
MAX_UPLOAD_SIZE = 10 * 1024 * 1024  # 10MB
ALLOWED_EXTENSIONS = {'.xlsx', '.xls', '.csv', '.pdf', '.txt'}

def _validate_upload(file: UploadFile):
    """验证上传文件大小和扩展名"""
    ext = os.path.splitext(file.filename or '')[1].lower()
    if ext not in ALLOWED_EXTENSIONS:
        raise HTTPException(400, f"不支持的文件类型: {ext}，仅接受 {'/'.join(sorted(ALLOWED_EXTENSIONS))}")
    # 检查文件大小 — 先读入内存再判断
    content = file.file.read()
    file.file.seek(0)  # 重置让后续代码正常读取
    if len(content) > MAX_UPLOAD_SIZE:
        raise HTTPException(400, f"文件过大（{len(content)/1024/1024:.1f}MB），上限10MB")
    return content


# ==================== Pydantic 模型 ====================

# 公司信息
class CompanyUpdate(BaseModel):
    company_name: Optional[str] = None
    uscc: Optional[str] = None
    registered_capital: Optional[float] = None
    established_date: Optional[date] = None
    legal_representative: Optional[str] = None
    legal_representative_id: Optional[str] = None
    address: Optional[str] = None
    business_scope: Optional[str] = None
    company_type: Optional[str] = None
    shareholders: Optional[List[dict]] = None
    directors: Optional[List[dict]] = None
    supervisors: Optional[List[dict]] = None
    finance_contacts: Optional[List[dict]] = None

# 部门
class DepartmentCreate(BaseModel):
    code: str
    name: str
    parent_code: Optional[str] = None
    manager: Optional[str] = None
    description: Optional[str] = None

class DepartmentUpdate(BaseModel):
    name: Optional[str] = None
    manager: Optional[str] = None
    description: Optional[str] = None
    is_active: Optional[bool] = None

# 人员
class EmployeeCreate(BaseModel):
    code: str
    name: str
    id_card: Optional[str] = None
    email: Optional[str] = None
    salary: Optional[float] = 0.0

class EmployeeUpdate(BaseModel):
    name: Optional[str] = None
    id_card: Optional[str] = None
    email: Optional[str] = None
    salary: Optional[float] = None
    leave_date: Optional[date] = None

# 客户
class BatchDelete(BaseModel):
    ids: list[int]

class CustomerCreate(BaseModel):
    code: str
    name: str
    uscc: Optional[str] = None
    tax_no: Optional[str] = None
    address: Optional[str] = None
    bank_name: Optional[str] = None
    bank_account: Optional[str] = None
    remark: Optional[str] = None

class CustomerUpdate(BaseModel):
    code: Optional[str] = None
    name: Optional[str] = None
    uscc: Optional[str] = None
    tax_no: Optional[str] = None
    address: Optional[str] = None
    bank_name: Optional[str] = None
    bank_account: Optional[str] = None
    is_active: Optional[bool] = None
    remark: Optional[str] = None

# 供应商
class SupplierCreate(BaseModel):
    code: str
    name: str
    uscc: Optional[str] = None
    tax_no: Optional[str] = None
    bank_name: Optional[str] = None
    bank_account: Optional[str] = None
    remark: Optional[str] = None

class SupplierUpdate(BaseModel):
    name: Optional[str] = None
    uscc: Optional[str] = None
    tax_no: Optional[str] = None
    bank_name: Optional[str] = None
    bank_account: Optional[str] = None
    is_active: Optional[bool] = None
    remark: Optional[str] = None



# ==================== 首页 ====================

@app.get("/", response_class=HTMLResponse)
async def index():
    with open("static/index.html", "r", encoding="utf-8") as f:
        return f.read()


# ==================== 公司信息 ====================

@app.get("/api/company")
def get_company(company_id: int = Query(...), db: Session = Depends(get_db)):
    info = db.query(Company).filter(Company.id == company_id).first()
    if not info:
        return {"company_name": "", "uscc": ""}
    return {
        "id": info.id,
        "company_name": info.name,
        "uscc": info.uscc or "",
        "registered_capital": info.registered_capital,
        "established_date": str(info.established_date) if info.established_date else "",
        "legal_representative": info.legal_representative or "",
        "legal_representative_id": info.legal_representative_id or "",
        "address": info.address or "",
        "business_scope": info.business_scope or "",
        "company_type": info.company_type or "",
        "shareholders": [{"name": s.name, "id_number": s.id_number or "", "ratio": s.ratio, "contribution_amount": s.contribution_amount} for s in info.shareholders],
        "directors": [{"name": d.name, "id_number": d.id_number or ""} for d in info.directors],
        "supervisors": [{"name": s.name, "id_number": s.id_number or ""} for s in info.supervisors],
        "finance_contacts": [{"name": f.name, "id_number": f.id_number or "", "phone": f.phone or ""} for f in info.finance_contacts],
    }

@app.put("/api/company")
def update_company(data: CompanyUpdate, company_id: int = Query(...), db: Session = Depends(get_db)):
    info = db.query(Company).filter(Company.id == company_id).first()
    if not info:
        info = Company(id=company_id, name=data.company_name or "")
        db.add(info)
        db.flush()
    if data.uscc:
        ok, msg = validate_uscc(data.uscc)
        if not ok:
            raise HTTPException(400, detail=f"公司统一社会信用代码：{msg}")
    # 更新主表字段
    main_fields = {k: v for k, v in data.model_dump(exclude_unset=True).items()
                   if k not in ("shareholders", "directors", "supervisors", "finance_contacts")}
    for k, v in main_fields.items():
        if k == 'company_name':
            info.name = v
        else:
            setattr(info, k, v)
    # 更新子表
    _update_company_subtable(db, info, CompanyShareholder, data.shareholders)
    _update_company_subtable(db, info, CompanyDirector, data.directors)
    _update_company_subtable(db, info, CompanySupervisor, data.supervisors)
    _update_company_subtable(db, info, CompanyFinanceContact, data.finance_contacts)
    db.commit()
    return {"message": "保存成功"}


def _update_company_subtable(db, company, model, items):
    """更新公司子表：清空旧数据，写入新数据"""
    if items is None:
        return
    db.query(model).filter(model.company_id == company.id).delete()
    for item in items:
        db.add(model(company_id=company.id, **item))


# ==================== 部门档案 ====================

@app.get("/api/departments")
def list_departments(
    keyword: Optional[str] = None,
    company_id: int = Query(...),
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=500),
    db: Session = Depends(get_db)
):
    q = db.query(Department).filter(Department.company_id == company_id)
    if keyword:
        q = q.filter(or_(
            Department.code.contains(keyword),
            Department.name.contains(keyword)
        ))
    depts = q.order_by(Department.code).offset(skip).limit(limit).all()
    return [
        {
            "id": d.id, "code": d.code, "name": d.name,
            "parent_code": d.parent_code, "manager": d.manager,
            "description": d.description, "is_active": d.is_active,
            "has_journal": False
        } for d in depts
    ]

# ── 档案锁定检查（被序时账引用时禁止编辑/删除）──

def _check_archive_lock(db, company_id, archive_type, archive_id) -> bool:
    """检查档案是否被序时账引用。返回 True=已锁定"""
    if archive_type == "department":
        return False
    elif archive_type == "employee":
        emp = db.query(Employee).filter(Employee.company_id == company_id, Employee.id == archive_id).first()
        if emp and emp.name:
            return db.query(JournalEntry).filter(
                JournalEntry.company_id == company_id,
                JournalEntry.contact_project == emp.name
            ).first() is not None
    elif archive_type == "customer":
        cust = db.query(Customer).filter(Customer.company_id == company_id, Customer.id == archive_id).first()
        if cust and cust.name:
            return db.query(JournalEntry).filter(
                JournalEntry.company_id == company_id,
                JournalEntry.contact_project == cust.name
            ).first() is not None
    elif archive_type == "supplier":
        supp = db.query(Supplier).filter(Supplier.company_id == company_id, Supplier.id == archive_id).first()
        if supp and supp.name:
            return db.query(JournalEntry).filter(
                JournalEntry.company_id == company_id,
                JournalEntry.contact_project == supp.name
            ).first() is not None
    return False

@app.post("/api/departments")
def create_department(data: DepartmentCreate, company_id: int = Query(...), db: Session = Depends(get_db)):
    d = Department(company_id=company_id, **data.model_dump())
    db.add(d)
    db.commit()
    return {"message": "新增成功"}

@app.put("/api/departments/{dept_id}")
def update_department(dept_id: int, data: DepartmentUpdate, company_id: int = Query(...), db: Session = Depends(get_db)):
    if _check_archive_lock(db, company_id, "department", dept_id):
        raise HTTPException(403, detail="该部门已被序时账引用，不可编辑")
    d = db.query(Department).filter(Department.company_id == company_id, Department.id == dept_id).first()
    if not d:
        raise HTTPException(404, detail="部门不存在")
    for k, v in data.model_dump(exclude_unset=True).items():
        setattr(d, k, v)
    db.commit()
    return {"message": "更新成功"}

@app.delete("/api/departments/{dept_id}")
def delete_department(dept_id: int, company_id: int = Query(...), db: Session = Depends(get_db)):
    if _check_archive_lock(db, company_id, "department", dept_id):
        raise HTTPException(403, detail="该部门已被序时账引用，不可删除")
    d = db.query(Department).filter(Department.company_id == company_id, Department.id == dept_id).first()
    if not d:
        raise HTTPException(404, detail="部门不存在")
    db.delete(d)
    db.flush()
    _renumber_archive(db, company_id, Department, 'BM')
    db.commit()
    return {"message": "删除成功"}

class DeptBatchDelete(BaseModel):
    ids: list[int]

@app.post("/api/departments/batch-delete")
def batch_delete_departments(req: DeptBatchDelete, company_id: int = Query(...), db: Session = Depends(get_db)):
    # 过滤掉被序时账引用的部门
    locked_ids = [did for did in req.ids if _check_archive_lock(db, company_id, "department", did)]
    deletable_ids = [did for did in req.ids if did not in locked_ids]
    if not deletable_ids:
        raise HTTPException(403, detail="所选部门均已被序时账引用，不可删除")
    deleted = db.query(Department).filter(
        Department.company_id == company_id,
        Department.id.in_(deletable_ids)
    ).delete(synchronize_session=False)
    db.flush()
    _renumber_archive(db, company_id, Department, 'BM')
    db.commit()
    return {"message": f"成功删除 {deleted} 个部门", "count": deleted}

@app.post("/api/departments/import")
async def import_departments(
    file: UploadFile = File(...),
    company_id: int = Query(...),
    db: Session = Depends(get_db)
):
    """从 CSV/XLSX 导入部门（编码+名称），编码为空时自动生成 BM001 格式"""
    ext = os.path.splitext(file.filename or "unknown")[1].lower()
    content_bytes = await file.read()

    rows = []
    if ext in (".xlsx", ".xls"):
        try:
            wb = openpyxl.load_workbook(io.BytesIO(content_bytes), data_only=True)
            ws = wb.active
            for r in ws.iter_rows(values_only=True):
                rows.append([str(c) if c is not None else "" for c in r])
        except Exception as e:
            raise HTTPException(400, f"无法解析 Excel 文件: {e}")
    elif ext == ".csv":
        try:
            text = content_bytes.decode("utf-8-sig")
            rows = list(csv.reader(io.StringIO(text)))
        except UnicodeDecodeError as e:
            raise HTTPException(400, f"文件编码错误，请使用 UTF-8 编码的 CSV: {e}")
    else:
        raise HTTPException(400, f"不支持的文件格式: {ext}，请上传 .csv 或 .xlsx")

    if not rows:
        raise HTTPException(400, "文件为空")
    headers = [h.strip() for h in rows[0]]
    ci = next((i for i, h in enumerate(headers) if h in ("编码", "code", "部门编码")), None)
    ni = next((i for i, h in enumerate(headers) if h in ("名称", "name", "部门名称", "部门")), 1)


    # 获取当前最大编码（仅匹配 BM 前缀，提取数字部分）
    existing_codes = db.query(Department.code).filter(
        Department.company_id == company_id,
        Department.code.like('BM%')
    ).all()
    code_counter = 0
    for c in existing_codes:
        try:
            num = int(c[0][2:])
            if num > code_counter:
                code_counter = num
        except Exception: pass

    imported = 0
    skipped = 0
    for row in rows[1:]:
        # 跳过完全空行
        if not any(str(c).strip() for c in row):
            continue
        code = row[ci].strip() if (ci is not None and ci < len(row)) else ""
        name = row[ni].strip() if ni < len(row) else ""
        if not name:
            continue


        # 编码为空时自动生成 BM001 格式
        if not code:
            code_counter += 1
            code = f"BM{code_counter:03d}"

        existing = db.query(Department).filter(
            Department.company_id == company_id, Department.code == code
        ).first()
        if existing:
            existing.name = name
        else:
            db.add(Department(
                company_id=company_id, code=code, name=name
            ))
        imported += 1
    db.commit()
    msg = f"成功导入 {imported} 条部门"
    if skipped > 0:
        msg += f"，跳过 {skipped} 条重复"
    return {"message": msg, "count": imported, "skipped": skipped}


# ==================== 人员档案 ====================

@app.get("/api/employees")
def list_employees(
    keyword: Optional[str] = None,
    company_id: int = Query(...),
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=500),
    db: Session = Depends(get_db)
):
    q = db.query(Employee).filter(Employee.company_id == company_id)
    if keyword:
        q = q.filter(or_(
            Employee.code.contains(keyword),
            Employee.name.contains(keyword)
        ))
    emps = q.order_by(Employee.code).offset(skip).limit(limit).all()

    # 检测哪些人员在序时账往来项目中出现过
    emp_names = [e.name for e in emps if e.name]
    names_with_entries = set()
    if emp_names:
        hits = db.query(JournalEntry.contact_project).filter(
            JournalEntry.company_id == company_id,
            JournalEntry.contact_project.in_(emp_names)
        ).distinct().all()
        names_with_entries.update(r[0] for r in hits if r[0])

    return [
        {
            "id": e.id, "code": e.code, "name": e.name,
            "id_card": e.id_card or "",
            "email": e.email or "", "salary": e.salary or 0,
            "has_journal": e.name in names_with_entries if e.name else False,
        } for e in emps
    ]

@app.post("/api/employees")
def create_employee(data: EmployeeCreate, company_id: int = Query(...), db: Session = Depends(get_db)):
    e = Employee(company_id=company_id, **data.model_dump())
    db.add(e)
    db.commit()
    return {"message": "新增成功"}

@app.put("/api/employees/{emp_id}")
def update_employee(emp_id: int, data: EmployeeUpdate, company_id: int = Query(...), db: Session = Depends(get_db)):
    e = db.query(Employee).filter(Employee.company_id == company_id, Employee.id == emp_id).first()
    if not e:
        raise HTTPException(404, detail="员工不存在")
    # 检查该人员是否已被序时账往来项目引用
    if e.name:
        ref = db.query(JournalEntry.id).filter(
            JournalEntry.company_id == company_id,
            JournalEntry.contact_project == e.name
        ).first()
        if ref:
            raise HTTPException(400, detail=f"人员「{e.name}」已被序时账往来项目引用，不可编辑。")
    for k, v in data.model_dump(exclude_unset=True).items():
        setattr(e, k, v)
    db.commit()
    return {"message": "更新成功"}

@app.delete("/api/employees/{emp_id}")
def delete_employee(emp_id: int, company_id: int = Query(...), db: Session = Depends(get_db)):
    e = db.query(Employee).filter(Employee.company_id == company_id, Employee.id == emp_id).first()
    if not e:
        raise HTTPException(404, detail="员工不存在")
    # 检查该人员是否已被序时账往来项目引用
    if e.name:
        ref = db.query(JournalEntry.id).filter(
            JournalEntry.company_id == company_id,
            JournalEntry.contact_project == e.name
        ).first()
        if ref:
            raise HTTPException(400, detail=f"人员「{e.name}」已被序时账往来项目引用，不可删除。")
    db.delete(e)
    db.flush()
    _renumber_archive(db, company_id, Employee, 'RY')
    db.commit()
    return {"message": "删除成功"}

@app.post("/api/employees/batch-delete")
def batch_delete_employees(data: dict, company_id: int = Query(...), db: Session = Depends(get_db)):
    ids = data.get("ids", [])
    if not ids:
        raise HTTPException(400, detail="请选择要删除的记录")

    # 查询被序时账往来项目引用的人员名称
    locked_names = set()
    emp_names = db.query(Employee.name).filter(
        Employee.company_id == company_id,
        Employee.id.in_(ids)
    ).all()
    all_names = [n[0] for n in emp_names if n[0]]
    if all_names:
        hits = db.query(JournalEntry.contact_project).filter(
            JournalEntry.company_id == company_id,
            JournalEntry.contact_project.in_(all_names)
        ).distinct().all()
        locked_names.update(r[0] for r in hits if r[0])

    # 过滤掉被锁定的人员
    if locked_names:
        safe_ids = [
            eid for eid in ids
            if db.query(Employee).filter(Employee.company_id == company_id, Employee.id == eid).first().name not in locked_names
        ]
    else:
        safe_ids = ids

    if not safe_ids:
        raise HTTPException(400, detail=f"所选人员均已被序时账往来项目引用，不可删除。")

    deleted = db.query(Employee).filter(Employee.company_id == company_id, Employee.id.in_(safe_ids)).delete(synchronize_session=False)
    db.flush()
    _renumber_archive(db, company_id, Employee, 'RY')
    db.commit()

    skipped = len(ids) - len(safe_ids)
    msg = f"成功删除 {deleted} 条人员记录"
    if skipped > 0:
        msg += f"，{skipped} 条因被序时账引用已跳过"
    return {"message": msg}


# ==================== 客户档案 ====================


@app.get("/api/customers")
def list_customers(
    keyword: Optional[str] = None,
    is_active: Optional[bool] = None,
    company_id: int = Query(...),
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=500),
    db: Session = Depends(get_db)
):
    q = db.query(Customer).filter(Customer.company_id == company_id)
    if keyword:
        q = q.filter(or_(
            Customer.code.contains(keyword),
            Customer.name.contains(keyword),
            Customer.contact.contains(keyword)
        ))
    if is_active is not None:
        q = q.filter(Customer.is_active == is_active)
    items = q.order_by(Customer.code).offset(skip).limit(limit).all()

    # 检测哪些客户名称存在于序时账中（contact_project 或 summary）
    cust_names = [c.name for c in items if c.name]
    names_with_entries = set()
    if cust_names:
        # 精确匹配 contact_project
        contact_hits = db.query(JournalEntry.contact_project).filter(
            JournalEntry.company_id == company_id,
            JournalEntry.contact_project.in_(cust_names)
        ).distinct().all()
        names_with_entries.update(r[0] for r in contact_hits if r[0])

        # 模糊匹配 summary（仅检查尚未匹配的客户）
        remaining = [n for n in cust_names if n not in names_with_entries]
        if remaining:
            conds = [JournalEntry.summary.contains(name) for name in remaining]
            summary_rows = db.query(JournalEntry.summary).filter(
                JournalEntry.company_id == company_id,
                or_(*conds)
            ).all()
            for row in summary_rows:
                if row[0]:
                    for name in remaining:
                        if name in row[0]:
                            names_with_entries.add(name)

    return [
        {
            "id": c.id, "code": c.code, "name": c.name,
            "uscc": c.uscc or "",
            "tax_no": c.tax_no,
            "bank_name": c.bank_name,
            "bank_account": c.bank_account,
            "is_active": c.is_active,
            "remark": c.remark,
            "has_journal": c.name in names_with_entries if c.name else False
        } for c in items
    ]

@app.post("/api/customers")
def create_customer(data: CustomerCreate, company_id: int = Query(...), db: Session = Depends(get_db)):
    if data.uscc:
        ok, msg = validate_uscc(data.uscc)
        if not ok:
            raise HTTPException(400, detail=f"客户统一社会信用代码：{msg}")
    # 计算全行指纹
    fp_values = (
        str(company_id),
        str(data.code or ""),
        str(data.name or ""),
        str(data.uscc or ""),
        str(data.tax_no or ""),
        str(data.contact or ""),
        str(data.phone or ""),
        str(data.address or ""),
        str(data.credit_limit if data.credit_limit is not None else ""),
        str(data.payment_terms if data.payment_terms is not None else ""),
        str(data.bank_name or ""),
        str(data.bank_account or ""),
        str(data.is_active if data.is_active is not None else ""),
        str(data.remark or "")
    )
    fp_raw = "|".join(fp_values)
    fp = hashlib.sha256(fp_raw.encode("utf-8")).hexdigest()
    # 去重检查
    existing = db.query(Customer).filter(
        Customer.company_id == company_id,
        Customer._fingerprint == fp
    ).first()
    if existing:
        raise HTTPException(400, detail="该客户数据已存在（全行比对重复），请勿重复录入")
    c = Customer(company_id=company_id, _fingerprint=fp, **data.model_dump())
    db.add(c)
    db.commit()
    return {"message": "新增成功"}

@app.put("/api/customers/{cust_id}")
def update_customer(cust_id: int, data: CustomerUpdate, company_id: int = Query(...), db: Session = Depends(get_db)):
    if _check_archive_lock(db, company_id, "customer", cust_id):
        raise HTTPException(403, detail="该客户已被序时账引用，不可编辑")
    c = db.query(Customer).filter(Customer.company_id == company_id, Customer.id == cust_id).first()
    if not c:
        raise HTTPException(404, detail="客户不存在")
    if data.uscc:
        ok, msg = validate_uscc(data.uscc)
        if not ok:
            raise HTTPException(400, detail=f"客户统一社会信用代码：{msg}")
    for k, v in data.model_dump(exclude_unset=True).items():
        setattr(c, k, v)
    # 重新计算全行指纹
    fp_values = (
        str(company_id),
        str(c.code or ""),
        str(c.name or ""),
        str(c.uscc or ""),
        str(c.tax_no or ""),
        str(c.contact or ""),
        str(c.phone or ""),
        str(c.address or ""),
        str(c.credit_limit if c.credit_limit is not None else ""),
        str(c.payment_terms if c.payment_terms is not None else ""),
        str(c.bank_name or ""),
        str(c.bank_account or ""),
        str(c.is_active if c.is_active is not None else ""),
        str(c.remark or "")
    )
    fp_raw = "|".join(fp_values)
    fp = hashlib.sha256(fp_raw.encode("utf-8")).hexdigest()
    c._fingerprint = fp
    db.commit()
    return {"message": "更新成功"}

@app.post("/api/customers/batch-delete")
def batch_delete_customers(
    body: BatchDelete,
    company_id: int = Query(...),
    db: Session = Depends(get_db)
):
    # 过滤被序时账引用的客户
    locked_ids = [cid for cid in body.ids if _check_archive_lock(db, company_id, "customer", cid)]
    deletable_ids = [cid for cid in body.ids if cid not in locked_ids]
    if not deletable_ids:
        raise HTTPException(403, detail="所选客户均已被序时账引用，不可删除")
    deleted = db.query(Customer).filter(
        Customer.company_id == company_id,
        Customer.id.in_(deletable_ids)
    ).delete(synchronize_session=False)
    db.flush()
    _renumber_archive(db, company_id, Customer, 'KH')
    db.commit()
    return {"message": f"成功删除 {deleted} 条客户"}

@app.delete("/api/customers/{cust_id}")
def delete_customer(cust_id: int, company_id: int = Query(...), db: Session = Depends(get_db)):
    if _check_archive_lock(db, company_id, "customer", cust_id):
        raise HTTPException(403, detail="该客户已被序时账引用，不可删除")
    c = db.query(Customer).filter(Customer.company_id == company_id, Customer.id == cust_id).first()
    if not c:
        raise HTTPException(404, detail="客户不存在")
    db.delete(c)
    db.flush()
    _renumber_archive(db, company_id, Customer, 'KH')
    db.commit()
    return {"message": "删除成功"}


# ==================== 客户智能建档（新规则 2026-06-06） ====================

@app.post("/api/customers/auto-create")
def auto_create_customers(company_id: int = Query(...), db: Session = Depends(get_db)):
    """智能客户建档：
    1. 唯一来源：销项发票购方名称
    2. 只要开具发票模块有信息，就一定是客户
    3. 排除人员档案和公司内部人（存量清理）
    """
    created = 0
    updated = 0
    skipped = 0
    infos = []

    # 构建实体索引
    idx = _build_entity_index(db, company_id)
    existing_cust_map = {}  # norm -> Customer obj
    for cust in db.query(Customer).filter(Customer.company_id == company_id).all():
        if cust._fingerprint:
            existing_cust_map[cust._fingerprint] = cust
        elif cust.name:
            existing_cust_map[_normalize_customer_name(cust.name)] = cust
    insider_norms = idx['insiders'] | idx['shareholders']

    # 清理：将已在人员档案/内部人的客户从客户档案中移除
    removed_names = []
    for norm, cust in list(existing_cust_map.items()):
        if norm in insider_norms:
            db.delete(cust)
            removed_names.append(cust.name)
    if removed_names:
        db.flush()
        infos.append(f"已从客户档案移除{len(removed_names)}条内部人员：{', '.join(removed_names)}")
        # 从 existing_cust_map 中移除已删除的条目
        for norm in list(existing_cust_map.keys()):
            if norm in insider_norms:
                del existing_cust_map[norm]

    sources = []

    # 1. 销项发票购方名称（主要来源）
    invoices = db.query(SalesInvoice).filter(
        SalesInvoice.company_id == company_id,
        SalesInvoice.buyer_name.isnot(None)
    ).all()
    for inv in invoices:
        name = inv.buyer_name.strip() if inv.buyer_name else ""
        if name:
            sources.append({
                'name': name,
                'tax_no': inv.buyer_tax_no.strip() if inv.buyer_tax_no else None,
                'source': f'销项发票:{inv.invoice_no or inv.id}',
            })

    # 去重 & 过滤
    seen = {}
    for s in sources:
        norm = _normalize_customer_name(s['name'])
        # 跳过公司内部人
        if norm in insider_norms:
            skipped += 1
            continue
        # 跳过已存在的客户
        if norm in existing_cust_map:
            skipped += 1
            # 如果有税号且现有记录没有，更新税号
            cust = existing_cust_map[norm]
            if s['tax_no'] and not cust.uscc:
                cust.uscc = s['tax_no']
                cust.tax_no = s['tax_no']
                db.flush()
                updated += 1
            continue
        if norm not in seen:
            seen[norm] = s
        elif s['tax_no'] and not seen[norm]['tax_no']:
            seen[norm] = s

    # 逐个创建
    for norm, s in seen.items():
        # 生成编码
        max_cust = db.query(Customer.code).filter(
            Customer.company_id == company_id,
            Customer.code.like('KH%')
        ).order_by(Customer.code.desc()).first()
        if max_cust and max_cust[0] and max_cust[0].startswith('KH'):
            try:
                num = int(max_cust[0][2:]) + 1
            except ValueError:
                num = 1
        else:
            num = 1
        code = f"KH{num:03d}"

        cust = Customer(
            company_id=company_id,
            code=code,
            name=s['name'],
            tax_no=s['tax_no'] or '',
            uscc=s['tax_no'] or '',
            is_active=True,
            _fingerprint=norm,
        )
        db.add(cust)
        db.flush()
        created += 1
        infos.append(f"已创建客户：{s['name']}（来源：{s['source']}）")

    db.commit()
    return {
        "message": f"智能建档完成：新建{created}条，跳过{skipped}条",
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "infos": infos,
    }


# ==================== 供应商档案 ====================

@app.get("/api/suppliers")
def list_suppliers(
    keyword: Optional[str] = None,
    is_active: Optional[bool] = None,
    company_id: int = Query(...),
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=500),
    db: Session = Depends(get_db)
):
    # 查询序时账中出现的供应商名称
    names_with_entries = set()
    try:
        entries = db.query(JournalEntry).filter(
            JournalEntry.company_id == company_id,
            JournalEntry.counterparty.isnot(None)
        ).all()
        names_with_entries = {e.counterparty for e in entries if e.counterparty}
    except Exception as e:
        logging.warning(f"供应商往来查询失败: {e}")
    q = db.query(Supplier).filter(Supplier.company_id == company_id)
    if keyword:
        q = q.filter(or_(
            Supplier.code.contains(keyword),
            Supplier.name.contains(keyword)
        ))
    if is_active is not None:
        q = q.filter(Supplier.is_active == is_active)
    items = q.order_by(Supplier.code).offset(skip).limit(limit).all()
    return [
        {
            "id": s.id, "code": s.code, "name": s.name,
            "uscc": s.uscc or "",
            "tax_no": s.tax_no,
            "bank_name": s.bank_name,
            "bank_account": s.bank_account,
            "is_active": s.is_active,
            "remark": s.remark,
            "has_journal": s.name in names_with_entries if s.name else False
        } for s in items
    ]

@app.post("/api/suppliers")
def create_supplier(data: SupplierCreate, company_id: int = Query(...), db: Session = Depends(get_db)):
    if data.uscc:
        ok, msg = validate_uscc(data.uscc)
        if not ok:
            raise HTTPException(400, detail=f"供应商统一社会信用代码：{msg}")
    s = Supplier(company_id=company_id, **data.model_dump())
    db.add(s)
    db.commit()
    return {"message": "新增成功"}

@app.put("/api/suppliers/{supp_id}")
def update_supplier(supp_id: int, data: SupplierUpdate, company_id: int = Query(...), db: Session = Depends(get_db)):
    if _check_archive_lock(db, company_id, "supplier", supp_id):
        raise HTTPException(403, detail="该供应商已被序时账引用，不可编辑")
    s = db.query(Supplier).filter(Supplier.company_id == company_id, Supplier.id == supp_id).first()
    if not s:
        raise HTTPException(404, detail="供应商不存在")
    if data.uscc:
        ok, msg = validate_uscc(data.uscc)
        if not ok:
            raise HTTPException(400, detail=f"供应商统一社会信用代码：{msg}")
    for k, v in data.model_dump(exclude_unset=True).items():
        setattr(s, k, v)
    db.commit()
    return {"message": "更新成功"}

@app.post("/api/suppliers/batch-delete")
def batch_delete_suppliers(
    body: BatchDelete,
    company_id: int = Query(...),
    db: Session = Depends(get_db)
):
    locked_ids = [sid for sid in body.ids if _check_archive_lock(db, company_id, "supplier", sid)]
    deletable_ids = [sid for sid in body.ids if sid not in locked_ids]
    if not deletable_ids:
        raise HTTPException(403, detail="所选供应商均已被序时账引用，不可删除")
    deleted = db.query(Supplier).filter(
        Supplier.company_id == company_id,
        Supplier.id.in_(deletable_ids)
    ).delete(synchronize_session=False)
    db.flush()
    _renumber_archive(db, company_id, Supplier, 'GYS')
    db.commit()
    return {"message": f"成功删除 {deleted} 条供应商"}

@app.delete("/api/suppliers/{supp_id}")
def delete_supplier(supp_id: int, company_id: int = Query(...), db: Session = Depends(get_db)):
    if _check_archive_lock(db, company_id, "supplier", supp_id):
        raise HTTPException(403, detail="该供应商已被序时账引用，不可删除")
    s = db.query(Supplier).filter(Supplier.company_id == company_id, Supplier.id == supp_id).first()
    if not s:
        raise HTTPException(404, detail="供应商不存在")
    db.delete(s)
    db.flush()
    _renumber_archive(db, company_id, Supplier, 'GYS')
    db.commit()
    return {"message": "删除成功"}


# ==================== 供应商智能建档（新规则 2026-06-06） ====================

def _extract_company_names(text: str) -> set:
    """从自由文本中提取企业名称（用于摘要/交易附言等字段）"""
    if not text:
        return set()
    # 常见前缀动词/介词（提取后需剥离）
    _LEADING_PREFIXES = [
        '待支付', '支付给', '转账给', '汇给', '转给', '退还', '退回',
        '支付', '转账', '付款给', '付款至', '付给', '付至', '付款',
        '预付', '预付给', '归还', '汇入', '汇出', '转付', '代付',
        '付', '转', '如', '给', '向',
    ]
    # 伪名称关键词：提取结果包含这些的视为非公司名（避免误提取）
    _NON_NAME_KEYWORDS = [
        '项目款', '保证金', '投标', '服务费', '货款', '租费',
        '顾问费', '咨询费', '代理费', '赞助费', '劳务费',
    ]
    names = set()
    # 企业后缀，按长度降序优先匹配长的
    _NAME_SUFFIXES = sorted([
        '有限责任公司', '股份有限公司', '集团有限公司', '有限公司',
        '总公司', '分公司', '公司', '厂', '中心', '机构', '店', '行',
        '协会', '所', '部', '网', '工作室', '事务所', '经营部'
    ], key=len, reverse=True)
    for suffix in _NAME_SUFFIXES:
        idx = 0
        while True:
            idx = text.find(suffix, idx)
            if idx < 0:
                break
            end = idx + len(suffix)
            start = idx
            # 向前扩展到标点/空格/换行
            while start > 0 and text[start - 1] not in '，,。. \t;；:：、（）()\n\r【】《》""\'\'!！?？':
                start -= 1
            name = text[start:end].strip()
            # 剥离常见前缀
            for prefix in sorted(_LEADING_PREFIXES, key=len, reverse=True):
                if name.startswith(prefix) and len(name) > len(prefix) + 2:
                    name = name[len(prefix):]
                    break
            # 合理长度范围（至少4个字符，不超过80字符）
            if 4 <= len(name) <= 80:
                # 排除包含业务关键词的伪名称
                if not any(kw in name for kw in _NON_NAME_KEYWORDS):
                    names.add(name)
            idx = end
    return names


def _enrich_archive_info(db: Session, company_id: int) -> dict:
    """档案信息补全：从发票/银行流水等数据源提取缺失字段，更新客户/供应商档案。
    触发时机：每次文件导入后自动运行，第一时间填补新信息。
    """
    from database import _normalize_customer_name
    enriched_cust = 0
    enriched_supp = 0
    fields_filled = []

    # ── 客户档案补全 ──
    custs = db.query(Customer).filter(Customer.company_id == company_id).all()
    cust_norm_map = {}
    for c in custs:
        # 用归一化名称做键（_fingerprint是SHA256，不能直接匹配银行流水的归一化名）
        norm = _normalize_customer_name(c.name or "")
        if norm:
            cust_norm_map[norm] = c

    # 来源1：销项发票购方信息
    for inv in db.query(SalesInvoice).filter(
        SalesInvoice.company_id == company_id,
        SalesInvoice.buyer_name.isnot(None)
    ).all():
        norm = _normalize_customer_name(inv.buyer_name.strip())
        c = cust_norm_map.get(norm)
        if not c:
            continue
        changed = False
        # 税号=统一社会信用代码（双向同步）
        inv_tax = (inv.buyer_tax_no or "").strip()
        if inv_tax:
            if not c.tax_no:
                c.tax_no = inv_tax
                fields_filled.append(f"客户[{c.name}]税号←销项发票")
                changed = True
            if not c.uscc:
                c.uscc = inv_tax
                changed = True
        if inv.buyer_address and not c.address:
            c.address = inv.buyer_address.strip()
            fields_filled.append(f"客户[{c.name}]地址←销项发票")
            changed = True
        if inv.buyer_bank_name and not c.bank_name:
            c.bank_name = inv.buyer_bank_name.strip()
            changed = True
        if inv.buyer_bank_account and not c.bank_account:
            c.bank_account = inv.buyer_bank_account.strip()
            changed = True
        # uscc→tax_no 反向
        if c.uscc and not c.tax_no:
            c.tax_no = c.uscc
            changed = True
        if c.tax_no and not c.uscc:
            c.uscc = c.tax_no
            changed = True
        if changed:
            enriched_cust += 1

    # 来源2：银行流水对方信息
    for tx in db.query(BankTransaction).filter(
        BankTransaction.company_id == company_id,
        BankTransaction.counterparty_name.isnot(None)
    ).all():
        norm = _normalize_customer_name(tx.counterparty_name.strip())
        c = cust_norm_map.get(norm)
        if not c:
            continue
        changed = False
        if tx.counterparty_bank and not c.bank_name:
            c.bank_name = tx.counterparty_bank.strip()
            changed = True
        if tx.counterparty_account and not c.bank_account:
            c.bank_account = tx.counterparty_account.strip()
            changed = True
        if changed:
            enriched_cust += 1

    # ── 供应商档案补全 ──
    supps = db.query(Supplier).filter(Supplier.company_id == company_id).all()
    supp_norm_map = {}
    for s in supps:
        norm = _normalize_customer_name(s.name or "")
        if norm:
            supp_norm_map[norm] = s

    # 来源1：取得发票销方信息
    for inv in db.query(PurchaseInvoice).filter(
        PurchaseInvoice.company_id == company_id,
        PurchaseInvoice.seller_name.isnot(None)
    ).all():
        norm = _normalize_customer_name(inv.seller_name.strip())
        s = supp_norm_map.get(norm)
        if not s:
            continue
        changed = False
        # 税号=统一社会信用代码（双向同步）
        inv_tax = (inv.seller_tax_no or "").strip()
        if inv_tax:
            if not s.tax_no:
                s.tax_no = inv_tax
                fields_filled.append(f"供应商[{s.name}]税号←取得发票")
                changed = True
            if not s.uscc:
                s.uscc = inv_tax
                changed = True
        # uscc→tax_no 反向
        if s.uscc and not s.tax_no:
            s.tax_no = s.uscc
            changed = True
        if s.tax_no and not s.uscc:
            s.uscc = s.tax_no
            changed = True
        if changed:
            enriched_supp += 1

    # 来源2：银行流水对方信息
    for tx in db.query(BankTransaction).filter(
        BankTransaction.company_id == company_id,
        BankTransaction.counterparty_name.isnot(None)
    ).all():
        norm = _normalize_customer_name(tx.counterparty_name.strip())
        s = supp_norm_map.get(norm)
        if not s:
            continue
        changed = False
        if tx.counterparty_bank and not s.bank_name:
            s.bank_name = tx.counterparty_bank.strip()
            changed = True
        if tx.counterparty_account and not s.bank_account:
            s.bank_account = tx.counterparty_account.strip()
            changed = True
        if changed:
            enriched_supp += 1

    if enriched_cust or enriched_supp:
        db.flush()

    return {
        "customer_enriched": enriched_cust,
        "supplier_enriched": enriched_supp,
        "fields_filled": fields_filled,
    }


def _close_archive_gap(db: Session, company_id: int) -> dict:
    """档案缺失自动补齐：序时账往来科目中有contact_project但档案中不存在的实体 → 自动建档
    修复范围：1122应收账款 → 客户 / 2202应付账款 → 供应商 / 1123预付账款 → 供应商
    这是确定性修复——有明细账就必需有档案，不允许gap存在。
    """
    from database import _normalize_customer_name
    created_cust = 0
    created_supp = 0

    # 现有档案归一化集合
    cust_norms = set()
    for c in db.query(Customer).filter(Customer.company_id == company_id).all():
        fp = c._fingerprint or _normalize_customer_name(c.name or "")
        if fp:
            cust_norms.add(fp)
    supp_norms = set()
    for s in db.query(Supplier).filter(Supplier.company_id == company_id).all():
        fp = s._fingerprint or _normalize_customer_name(s.name or "")
        if fp:
            supp_norms.add(fp)

    # 扫描序时账往来科目：1122→客户, 2202/1123→供应商
    for code, entity_type in [("1122", "customer"), ("2202", "supplier"), ("1123", "supplier")]:
        entries = db.query(JournalEntry.contact_project).filter(
            JournalEntry.company_id == company_id,
            JournalEntry.account_code == code,
            JournalEntry.contact_project.isnot(None),
            JournalEntry.contact_project != "",
        ).distinct().all()

        for (cp,) in entries:
            name = cp.strip()
            if not name or len(name) < 4:
                continue
            norm = _normalize_customer_name(name)

            # 排除内部人员（1221的才是人员，1122/2202是企业往来）
            # 排除明显非企业名称
            _NON_ENTITY = ("手续费", "金库", "公积金", "待处理", "出售凭证", "业务收入",
                          "国家金库", "税务", "国库", "工资", "社保", "个税")
            if any(kw in name for kw in _NON_ENTITY):
                continue
            # 排除个人名（3字以下纯中文名）
            if len(name) <= 3 and all('\u4e00' <= c <= '\u9fff' for c in name):
                continue

            if entity_type == "customer":
                if norm in cust_norms:
                    continue
                max_c = db.query(Customer.code).filter(
                    Customer.company_id == company_id, Customer.code.like('KH%')
                ).order_by(Customer.code.desc()).first()
                num = int(max_c[0][2:]) + 1 if max_c and max_c[0] and max_c[0].startswith('KH') else 1
                db.add(Customer(company_id=company_id, code=f"KH{num:03d}", name=name,
                               _fingerprint=norm, is_active=True))
                db.flush()
                cust_norms.add(norm)
                created_cust += 1

            elif entity_type == "supplier":
                if norm in supp_norms:
                    continue
                max_s = db.query(Supplier.code).filter(
                    Supplier.company_id == company_id, Supplier.code.like('GYS%')
                ).order_by(Supplier.code.desc()).first()
                num = int(max_s[0][3:]) + 1 if max_s and max_s[0] and max_s[0].startswith('GYS') else 1
                db.add(Supplier(company_id=company_id, code=f"GYS{num:03d}", name=name,
                               _fingerprint=norm, is_active=True))
                db.flush()
                supp_norms.add(norm)
                created_supp += 1

    return {"customer_created": created_cust, "supplier_created": created_supp}


def _do_auto_create_suppliers(db: Session, company_id: int) -> dict:
    """供应商智能建档核心逻辑（可被API和导入流程复用）"""
    created = 0
    updated = 0
    skipped = 0
    infos = []

    # 构建实体索引
    idx = _build_entity_index(db, company_id)
    shareholder_norms = idx['shareholders']
    insider_norms = idx['insiders'] | idx['shareholders']

    # 清理：将已在人员档案/内部人的供应商移除
    removed_names = []
    for supp in db.query(Supplier).filter(Supplier.company_id == company_id).all():
        if supp.name:
            norm = _normalize_customer_name(supp.name)
            if norm in insider_norms:
                db.delete(supp)
                removed_names.append(supp.name)
    if removed_names:
        db.flush()
        infos.append(f"已从供应商档案移除{len(removed_names)}条内部人员：{', '.join(removed_names)}")

    # 1. 取得发票销方名称集合
    pi_names = set()
    pi_sources = {}
    invoices = db.query(PurchaseInvoice).filter(
        PurchaseInvoice.company_id == company_id,
        PurchaseInvoice.seller_name.isnot(None)
    ).all()
    for inv in invoices:
        name = inv.seller_name.strip() if inv.seller_name else ""
        if not name:
            continue
        norm = _normalize_customer_name(name)
        pi_names.add(norm)
        pi_sources[norm] = {
            'name': name,
            'tax_no': inv.seller_tax_no.strip() if inv.seller_tax_no else None,
            'source': f'进项发票:{inv.invoice_no or inv.id}',
        }

    # 2. 银行流水付款方（借方=付款，即 debit_amount > 0）
    #   取数来源：对方户名 + 摘要 + 交易附言（老邓 2026-06-10 三源综合）
    # 手续费/税费/政府机构等非供应商关键词
    _NON_SUPPLIER_KEYWORDS = ('手续费', '金库', '公积金', '待处理', '出售凭证', '业务收入', '国家金库', '税务', '国库', '工资', '社保', '个税', '薪金', '薪酬')
    _BIZ_SUFFIXES = ('公司', '厂', '中心', '机构', '店', '行', '协会', '所', '部', '网')
    bt_names = set()
    bt_sources = {}
    txs = db.query(BankTransaction).filter(
        BankTransaction.company_id == company_id,
        BankTransaction.debit_amount > 0
    ).all()
    for tx in txs:
        # 从三个来源收集名称
        raw_names = set()
        # ① 对方户名
        if tx.counterparty_name and tx.counterparty_name.strip():
            raw_names.add(tx.counterparty_name.strip())
        # ② 摘要
        if tx.summary:
            raw_names |= _extract_company_names(tx.summary)
        # ③ 交易附言
        if tx.transaction_remark:
            raw_names |= _extract_company_names(tx.transaction_remark)

        for name in raw_names:
            if not name:
                continue
            # 跳过股东/内部人员
            norm = _normalize_customer_name(name)
            if norm in insider_norms:
                continue
            # 跳过手续费/税费/政府机构关键词
            if any(kw in name for kw in _NON_SUPPLIER_KEYWORDS):
                continue
            # 跳过明显非企业名称（长度<6且不含公司等后缀）
            if len(name) < 6 and not any(s in name for s in _BIZ_SUFFIXES):
                continue
            bt_names.add(norm)
            if norm not in bt_sources:
                bt_sources[norm] = {
                    'name': name,
                    'tax_no': None,
                    'source': f'银行流水:#{tx.id}',
                }

    # 3. 候选供应商：银行流水付款方 ∩ 取得发票销方（双源信号，老邓 2026-06-10 回归铁律）
    candidate_names = pi_names & bt_names
    # ⚠️ 用数据库直接查判重，不能用 idx['suppliers']——它被 _build_entity_index 污染了（含所有发票销方）
    existing_supp_norms = set()
    for s in db.query(Supplier).filter(Supplier.company_id == company_id).all():
        fp = s._fingerprint or _normalize_customer_name(s.name or "")
        if fp:
            existing_supp_norms.add(fp)

    for norm in candidate_names:
        if norm in shareholder_norms:
            skipped += 1
            continue
        if norm in existing_supp_norms:
            skipped += 1
            continue

        s = pi_sources.get(norm) or bt_sources.get(norm)
        if not s:
            continue

        source_tag = []
        if norm in pi_names:
            source_tag.append('取得发票')
        if norm in bt_names:
            source_tag.append('银行流水')
        full_source = '+'.join(source_tag)

        max_supp = db.query(Supplier.code).filter(
            Supplier.company_id == company_id,
            Supplier.code.like('GYS%')
        ).order_by(Supplier.code.desc()).first()
        if max_supp and max_supp[0] and max_supp[0].startswith('GYS'):
            try:
                num = int(max_supp[0][3:]) + 1
            except ValueError:
                num = 1
        else:
            num = 1
        code = f"GYS{num:03d}"

        supp = Supplier(
            company_id=company_id,
            code=code,
            name=s['name'],
            uscc=s['tax_no'] if s['tax_no'] and s['tax_no'].strip() else None,
            _fingerprint=norm,
            is_active=True,
        )
        db.add(supp)
        db.flush()
        created += 1
        infos.append(f"已创建供应商：{s['name']}（来源：{full_source}）")

    db.commit()
    return {
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "infos": infos,
    }


@app.post("/api/suppliers/auto-create")
def auto_create_suppliers(company_id: int = Query(...), db: Session = Depends(get_db)):
    """智能供应商建档：
    1. 以银行流水付款方为注来源
    2. 银行流水 + 取得发票双源出现 → 强信号 → 创建
    3. 单源（仅银行或仅发票）→ 不创建
    4. 排除股东（投资款归实收资本，付款归分红）
    """
    result = _do_auto_create_suppliers(db, company_id)
    return {
        "message": f"智能建档完成：新建{result['created']}条，跳过{result['skipped']}条",
        "created": result['created'],
        "updated": result['updated'],
        "skipped": result['skipped'],
        "infos": result['infos'],
    }


@app.get("/api/suppliers/diagnose")
def diagnose_suppliers(company_id: int = Query(...), db: Session = Depends(get_db)):
    """双源供应商诊断：展示取得发票销方 ∩ 银行流水付款方的匹配过程"""
    from database import _normalize_customer_name, _build_entity_index
    idx = _build_entity_index(db, company_id)
    insider_norms = idx['insiders']

    # 排除关键词
    _NON_SUPPLIER_KEYWORDS = ('手续费', '金库', '公积金', '待处理', '出售凭证', '业务收入', '国家金库', '税务', '国库')
    _BIZ_SUFFIXES = ('公司', '厂', '中心', '机构', '店', '行', '协会', '所', '部', '网')

    # 1. 取得发票销方
    pi_items = []
    pi_norms = set()
    for inv in db.query(PurchaseInvoice).filter(
        PurchaseInvoice.company_id == company_id,
        PurchaseInvoice.seller_name.isnot(None)
    ).all():
        name = inv.seller_name.strip()
        if not name: continue
        norm = _normalize_customer_name(name)
        pi_norms.add(norm)
        pi_items.append({
            "name": name, "norm": norm,
            "tax_no": inv.seller_tax_no.strip() if inv.seller_tax_no else "",
            "invoice_no": inv.invoice_no or inv.digital_invoice_no or str(inv.id),
        })

    # 2. 银行流水付款方
    bt_items = []
    bt_norms = set()
    for tx in db.query(BankTransaction).filter(
        BankTransaction.company_id == company_id,
        BankTransaction.debit_amount > 0,
        BankTransaction.counterparty_name.isnot(None)
    ).all():
        name = tx.counterparty_name.strip()
        if not name: continue
        norm = _normalize_customer_name(name)
        if norm in insider_norms: continue
        if any(kw in name for kw in _NON_SUPPLIER_KEYWORDS): continue
        if len(name) < 6 and not any(s in name for s in _BIZ_SUFFIXES): continue
        bt_norms.add(norm)
        bt_items.append({
            "name": name, "norm": norm,
            "tx_id": tx.id, "amount": float(tx.debit_amount or 0),
        })

    # 3. 交集分析
    both = pi_norms & bt_norms
    pi_only = pi_norms - bt_norms
    bt_only = bt_norms - pi_norms

    # 4. 已有供应商（直接查数据库，不用 idx['suppliers']——已被发票销方污染）
    existing_list = [{"code": s.code, "name": s.name} for s in db.query(Supplier).filter(
        Supplier.company_id == company_id
    ).order_by(Supplier.code).all()]
    existing_norms = set()
    for s in db.query(Supplier).filter(Supplier.company_id == company_id).all():
        fp = s._fingerprint or _normalize_customer_name(s.name or "")
        if fp:
            existing_norms.add(fp)

    return {
        "summary": {
            "purchase_invoice_sellers": len(pi_norms),
            "bank_transaction_payers": len(bt_norms),
            "dual_source_match": len(both),
            "pi_only": len(pi_only),
            "bt_only": len(bt_only),
            "existing_suppliers": len(existing_list),
            "new_to_create": len(both - existing_norms),
        },
        "dual_source": sorted([
            {
                "name": next((bi["name"] for bi in bt_items if bi["norm"] == n), 
                              next((pi["name"] for pi in pi_items if pi["norm"] == n), n)),
                "norm": n,
                "from_purchase_invoice": bool(n in pi_norms),
                "from_bank_transaction": bool(n in bt_norms),
                "already_exists": n in existing_norms,
            }
            for n in both
        ], key=lambda x: x["name"]),
        "pi_only_sample": sorted(list(pi_only))[:10],
        "bt_only_sample": sorted(list(bt_only))[:10],
        "existing_suppliers": existing_list,
    }


@app.post("/api/process-all")
def process_all(company_id: int = Query(...), db: Session = Depends(get_db)):
    """三步统一处理流程（老邓 2026-06-10 铁律）：
    ① 确定供应商档案（双源：发票销方 ∩ 银行付款方）
    ② 取得发票序时账登记（根据供应商档案）
    ③ 银行流水序时账登记（根据供应商档案）
    """
    import logging
    log = logging.getLogger("process-all")

    # ── 第零步：档案缺失补齐（有明细账但档案缺失 → 自动建档，零容忍）──
    gap_result = {"customer_created": 0, "supplier_created": 0}
    try:
        gap_result = _close_archive_gap(db, company_id)
        db.commit()
    except Exception:
        pass

    # ── 第零步半：档案信息补全（第一时间填补缺失字段）──
    enrich_result = {"customer_enriched": 0, "supplier_enriched": 0}
    try:
        enrich_result = _enrich_archive_info(db, company_id)
        db.commit()
    except Exception:
        pass

    # ── 第一步：确定供应商档案 ──
    supp_result = _do_auto_create_suppliers(db, company_id)
    db.commit()

    # ── 第二步：未记账发票生成凭证 ──
    from database import BookkeepingInvoice
    pi_count = auto_generate_bookkeeping_journal(db, company_id)
    db.commit()

    db.commit()

    # ── 第三步：银行流水生成凭证 ──
    bank_result = _generate_bank_journals(db, company_id, None)
    db.commit()

    # ── 第四步：社保缴纳匹配 ──
    ss_result = {"generated": 0}
    try:
        ss_result = _match_ss_payment_journals(db, company_id)
        db.commit()
    except Exception:
        pass

    # ── 第五步：税费组合缴纳匹配（国家金库）──
    tax_result = {"generated": 0}
    try:
        tax_result = _match_tax_payment_journals(db, company_id)
        db.commit()
    except Exception:
        pass

    # ── 第六步：公积金缴纳匹配 ──
    hf_result = {"generated": 0}
    try:
        hf_result = _match_hf_payment_journals(db, company_id)
        db.commit()
    except Exception:
        pass

    return {
        "step0_archive_gap": gap_result,
        "step0_5_enrich": enrich_result,
        "step1_suppliers": supp_result,
        "step2_bookkeeping": {
            "generated": pi_count,
        },
        "step3_bank_transactions": bank_result,
        "step4_social_security": ss_result,
        "step5_tax_payment": tax_result,
        "step6_housing_fund": hf_result,
    }


@app.post("/api/generate-sample-archives")
def generate_sample_archives(company_id: int = Query(...), db: Session = Depends(get_db)):
    """为部门、人员、客户、供应商各生成25条样本数据"""
    results = {"departments": 0, "employees": 0, "customers": 0, "suppliers": 0}

    # --- 部门：25个常用部门 ---
    dept_names = [
        "总经理办公室", "财务部", "人力资源部", "市场部", "销售一部",
        "销售二部", "研发一部", "研发二部", "采购部", "质量管理部",
        "物流部", "行政部", "法务合规部", "信息技术部", "客户服务部",
        "公关部", "审计部", "战略发展部", "工程部", "设计部",
        "培训部", "安全环保部", "后勤保障部", "国际业务部", "投资管理部"
    ]
    # 先查现有最大编码
    max_dept = db.query(Department.code).filter(
        Department.company_id == company_id, Department.code.like('BM%')
    ).order_by(Department.code.desc()).first()
    dept_idx = int(max_dept[0][2:]) + 1 if max_dept else 1
    for name in dept_names:
        existing = db.query(Department).filter(
            Department.company_id == company_id, Department.name == name
        ).first()
        if not existing:
            db.add(Department(company_id=company_id, code=f"BM{dept_idx:03d}", name=name))
            dept_idx += 1
            results["departments"] += 1
    db.flush()

    # --- 人员：25个员工 ---
    emp_data = [
        ("张伟", "440101199001011234"), ("李娜", "440102199103152345"), ("王磊", "440103198807203456"),
        ("陈静", "440104199206184567"), ("刘洋", "440105199311255678"), ("杨帆", "440106198912106789"),
        ("赵敏", "440107199507157890"), ("黄超", "440108199008168901"), ("周婷", "440109199409179012"),
        ("吴强", "440110199110181123"), ("郑芳", "440111199211192234"), ("冯涛", "440112199312203345"),
        ("何丽", "440113199401214456"), ("韩明", "440114199502225567"), ("曹雪", "440115199603236678"),
        ("许杰", "440116199704247789"), ("邓辉", "440117199805258890"), ("萧琳", "440118199906269901"),
        ("唐波", "440119198701270112"), ("彭悦", "440120198802282223"), ("曾强", "440121198903013334"),
        ("董洁", "440122199004024445"), ("袁浩", "440123199105035556"), ("蒋霞", "440124199206046667"),
        ("沈飞", "440125199307057778")
    ]
    max_emp = db.query(Employee.code).filter(
        Employee.company_id == company_id, Employee.code.like('RY%')
    ).order_by(Employee.code.desc()).first()
    emp_idx = int(max_emp[0][2:]) + 1 if max_emp else 1
    for i, (name, id_card) in enumerate(emp_data):
        existing = db.query(Employee).filter(
            Employee.company_id == company_id, Employee.name == name, Employee.id_card == id_card
        ).first()
        if not existing:
            db.add(Employee(
                company_id=company_id, code=f"RY{emp_idx:03d}", name=name,
                id_card=id_card,
                email=f"{name.lower()}{emp_idx}@cunqin.com",
                salary=round(5000 + i * 800 + (hash(name) % 3000), -2)
            ))
            emp_idx += 1
            results["employees"] += 1
    db.flush()

    # --- 客户：25个企业客户 ---
    cust_data = [
        ("广州天宏科技有限公司", "91440101MA5ABCD123"), ("深圳鹏程实业有限公司", "91440300MA5EFGH456"),
        ("东莞华耀电子有限公司", "91441900MA5IJKL789"), ("佛山顺达建材有限公司", "91440600MA5MNOP012"),
        ("中山明辉灯饰有限公司", "91442000MA5QRST345"), ("珠海海天贸易有限公司", "91440400MA5UVWX678"),
        ("惠州鑫源五金有限公司", "91441300MA5YZAB901"), ("江门益丰食品有限公司", "91440700MA5CDEF234"),
        ("肇庆鼎湖旅游开发有限公司", "91441200MA5GHIJ567"), ("汕头潮阳纺织有限公司", "91440500MA5KLMN890"),
        ("北京中科创新科技有限公司", "91110108MA5OPQR123"), ("上海浦江物流有限公司", "91310115MA5STUV456"),
        ("杭州西湖软件有限公司", "91330108MA5WXYZ789"), ("南京金陵机械有限公司", "91320105MA5ABCD012"),
        ("武汉江城建设集团有限公司", "91420102MA5EFGH345"), ("成都天府餐饮管理有限公司", "91510104MA5IJKL678"),
        ("重庆山城商贸有限公司", "91500103MA5MNOP901"), ("长沙星城文化传媒有限公司", "91430102MA5QRST234"),
        ("厦门海西进出口有限公司", "91350203MA5UVWX567"), ("青岛海尔智能科技有限公司", "91370281MA5YZAB890"),
        ("大连滨海渔业有限公司", "91210202MA5CDEF123"), ("苏州园林设计院有限公司", "91320505MA5GHIJ456"),
        ("无锡太湖环保科技有限公司", "91320213MA5KLMN789"), ("合肥高新投资管理有限公司", "91340104MA5OPQR012"),
        ("福州闽江房地产开发有限公司", "91350102MA5STUV345")
    ]
    max_cust = db.query(Customer.code).filter(
        Customer.company_id == company_id, Customer.code.like('KH%')
    ).order_by(Customer.code.desc()).first()
    cust_idx = int(max_cust[0][2:]) + 1 if max_cust else 1
    banks = ["中国工商银行", "中国建设银行", "中国农业银行", "中国银行", "招商银行"]
    for name, uscc in cust_data:
        existing = db.query(Customer).filter(
            Customer.company_id == company_id, Customer.uscc == uscc
        ).first()
        if not existing:
            bi = hash(name) % len(banks)
            db.add(Customer(
                company_id=company_id, code=f"KH{cust_idx:03d}", name=name,
                uscc=uscc, tax_no=uscc[2:20],
                bank_name=banks[bi],
                bank_account=f"{62220000 + cust_idx * 137:020d}",
                remark="样本数据"
            ))
            cust_idx += 1
            results["customers"] += 1
    db.flush()

    # --- 供应商：25个供应商 ---
    supp_data = [
        ("广州龙腾电子科技有限公司", "91440101MA5AAAA111"), ("深圳星辰照明有限公司", "91440300MA5BBBB222"),
        ("东莞万丰模具制品有限公司", "91441900MA5CCCC333"), ("佛山新力包装材料有限公司", "91440600MA5DDDD444"),
        ("中山瑞安五金机电有限公司", "91442000MA5EEEE555"), ("珠海格力精密模具有限公司", "91440400MA5FFFF666"),
        ("惠州德盛化工有限公司", "91441300MA5GGGG777"), ("江门华盛纺织原料有限公司", "91440700MA5HHHH888"),
        ("肇庆大发木业有限公司", "91441200MA5IIII999"), ("汕头阳光印务有限公司", "91440500MA5JJJJ000"),
        ("北京云帆信息技术有限公司", "91110108MA5KKKK111"), ("上海博达广告传媒有限公司", "91310115MA5LLLL222"),
        ("杭州网联通信设备有限公司", "91330108MA5MMMM333"), ("南京翔宇机械设备有限公司", "91320105MA5NNNN444"),
        ("武汉盛丰粮油贸易有限公司", "91420102MA5OOOO555"), ("成都锦程物流有限公司", "91510104MA5PPPP666"),
        ("重庆利群商贸有限公司", "91500103MA5QQQQ777"), ("长沙恒达仪器仪表有限公司", "91430102MA5RRRR888"),
        ("厦门伟业建筑工程有限公司", "91350203MA5SSSS999"), ("青岛远洋渔业有限公司", "91370281MA5TTTT000"),
        ("大连宏发水产品有限公司", "91210202MA5UUUU111"), ("苏州鼎丰纺织有限公司", "91320505MA5VVVV222"),
        ("无锡大明金属材料有限公司", "91320213MA5WWWW333"), ("合肥利安医疗器材有限公司", "91340104MA5XXXX444"),
        ("福州东南汽车配件有限公司", "91350102MA5YYYY555")
    ]
    max_supp = db.query(Supplier.code).filter(
        Supplier.company_id == company_id, Supplier.code.like('GYS%')
    ).order_by(Supplier.code.desc()).first()
    supp_idx = int(max_supp[0][3:]) + 1 if max_supp else 1
    for name, uscc in supp_data:
        existing = db.query(Supplier).filter(
            Supplier.company_id == company_id, Supplier.uscc == uscc
        ).first()
        if not existing:
            bi = hash(name + "_s") % len(banks)
            db.add(Supplier(
                company_id=company_id, code=f"GYS{supp_idx:03d}", name=name,
                uscc=uscc, tax_no=uscc[2:20],
                bank_name=banks[bi],
                bank_account=f"{62280000 + supp_idx * 211:020d}",
                remark="样本数据"
            ))
            supp_idx += 1
            results["suppliers"] += 1

    db.commit()
    return {"message": "样本数据生成完成", "results": results}


# ==================== 会计科目（原有，保留）====================

def _build_account_hierarchy(db: Session, company_id: int) -> dict:
    """构建科目编码→全级次名称的映射"""
    all_accounts = db.query(Account).filter(Account.company_id == company_id).all()
    code_map = {a.code: a for a in all_accounts}

    def get_full_name(acct):
        parts = []
        current = acct
        visited = set()
        while current and current.code not in visited:
            visited.add(current.code)
            parts.append(f"{current.code} {current.name}")
            current = code_map.get(current.parent_code) if current.parent_code else None
        parts.reverse()
        return " / ".join(parts)

    return {a.code: get_full_name(a) for a in all_accounts}


@app.get("/api/accounts")
def list_accounts(
    category: Optional[str] = None,
    keyword: Optional[str] = None,
    level: Optional[int] = None,
    leaf_only: Optional[str] = None,
    company_id: int = Query(...),
    skip: int = Query(0, ge=0),
    limit: int = Query(200, ge=1, le=500),
    db: Session = Depends(get_db)
):
    q = db.query(Account).filter(Account.company_id == company_id)
    if category:
        q = q.filter(Account.category == category)
    if keyword:
        q = q.filter(or_(
            Account.code.contains(keyword),
            Account.name.contains(keyword)
        ))
    if level:
        q = q.filter(Account.level == level)
    accounts = q.order_by(Account.code).offset(skip).limit(limit).all()

    # 末级科目过滤：排除那些是其他科目parent_code的科目
    if leaf_only and leaf_only.lower() in ("1", "true", "yes"):
        all_codes = {a.code for a in accounts}
        parent_codes = {a.parent_code for a in accounts if a.parent_code}
        accounts = [a for a in accounts if a.code not in parent_codes]

    # 构建全级次名称映射
    hierarchy = _build_account_hierarchy(db, company_id)

    # 检查哪些科目有下级
    all_codes = {a.code for a in accounts}
    parent_codes = {a.parent_code for a in accounts if a.parent_code}
    has_children_codes = parent_codes & all_codes

    # 检查哪些科目被序时账使用
    journal_codes = set()
    try:
        journal_codes = {r[0] for r in db.query(JournalEntry.account_code).filter(
            JournalEntry.company_id == company_id
        ).distinct().all()}
    except Exception as e:
        logging.warning(f"科目序时账使用检查失败: {e}")

    return [
        {
            "id": a.id, "code": a.code, "name": a.name,
            "full_name": hierarchy.get(a.code, f"{a.code} {a.name}"),
            "category": a.category, "balance_direction": a.balance_direction,
            "level": a.level, "parent_code": a.parent_code,
            "is_active": a.is_active,
            "opening_balance": a.opening_balance or 0.0,
            "has_children": a.code in has_children_codes,
            "has_journal": a.code in journal_codes,
        } for a in accounts
    ]


class AccountCreate(BaseModel):
    code: str
    name: str
    category: str = ""
    balance_direction: str = ""
    level: int = 1
    parent_code: str = ""
    opening_balance: float = 0.0


class AccountUpdate(BaseModel):
    code: Optional[str] = None
    name: Optional[str] = None
    category: Optional[str] = None
    balance_direction: Optional[str] = None
    level: Optional[int] = None
    parent_code: Optional[str] = None
    opening_balance: Optional[float] = None
    password: str = ""
    is_active: Optional[bool] = None


@app.post("/api/accounts")
def create_account(data: AccountCreate, company_id: int = Query(...), db: Session = Depends(get_db)):
    code = data.code
    name = (data.name or "").strip()
    category = data.category
    balance_direction = data.balance_direction
    level = data.level
    parent_code = data.parent_code
    if not code or not name:
        raise HTTPException(400, detail="科目编码和名称不能为空")
    # 一级科目限制：仅6个往来科目可作为一级科目，其他必须设二级
    ALLOWED_L1_CODES = {'1122', '2202', '2203', '1123', '1221', '2241'}
    if level == 1:
        code_root = code[:4]
        if code_root not in ALLOWED_L1_CODES:
            raise HTTPException(400,
                detail="该科目不可作为一级科目使用。仅以下6个往来科目允许设置一级科目："
                       "应收账款(1122)、应付账款(2202)、预收账款(2203)、"
                       "预付账款(1123)、其他应收款(1221)、其他应付款(2241)。"
                       "请选择2级（含）以上级次。")
    # 1221其他应收款不允许设二级科目，应使用往来项目（人员/供应商档案）
    if code[:4] == '1221' and level and level >= 2:
        raise HTTPException(400, detail="1221其他应收款不需要二级科目，请直接使用往来项目（人员档案/供应商档案）")
    # 去重检查：同一公司内科目编码不能重复
    dup_code = db.query(Account).filter(Account.company_id == company_id, Account.code == code).first()
    if dup_code:
        raise HTTPException(400, detail=f"科目编码【{code}】已存在（{dup_code.name}），请更换编码")
    # 去重检查：同一公司内科目名称不能重复（本级名称）
    dup_name = db.query(Account).filter(Account.company_id == company_id, Account.name == name).first()
    if dup_name:
        raise HTTPException(400, detail=f"科目名称【{name}】已存在（{dup_name.code}），请更换名称")
    acc = Account(company_id=company_id, code=code, name=name, category=category,
                  balance_direction=balance_direction, level=level, parent_code=parent_code,
                  opening_balance=data.opening_balance)
    db.add(acc)
    db.commit()
    db.refresh(acc)
    return {"id": acc.id, "code": acc.code, "name": acc.name, "message": "创建成功"}


ACCOUNT_PWD = "123456"

def _account_needs_password(db, company_id, account_id):
    """检查科目是否有下级或被序时账使用，需要密码才能修改/删除"""
    acc = db.query(Account).filter(Account.company_id == company_id, Account.id == account_id).first()
    if not acc:
        return False, None, "科目不存在"
    # 检查是否有下级科目
    has_child = db.query(Account).filter(Account.company_id == company_id, Account.parent_code == acc.code).first()
    if has_child:
        return True, acc, "该科目下有下级科目，修改/删除需要密码"
    # 检查是否被序时账使用
    has_journal = db.query(JournalEntry).filter(
        JournalEntry.company_id == company_id,
        JournalEntry.account_code == acc.code
    ).first()
    if has_journal:
        return True, acc, "该科目已被序时账使用，修改/删除需要密码"
    return False, acc, ""

@app.put("/api/accounts/{account_id}")
def update_account(account_id: int, data: AccountUpdate, company_id: int = Query(...), db: Session = Depends(get_db)):
    needs_pwd, acc, msg = _account_needs_password(db, company_id, account_id)
    if needs_pwd:
        pwd = data.password
        if pwd != ACCOUNT_PWD:
            raise HTTPException(403, detail=f"{msg}，请输入正确密码")
    if not acc:
        raise HTTPException(404, detail="科目不存在")
    if data.name is not None:
        new_name = data.name.strip()
        # 去重检查：同一公司内科目名称不能重复（排除自身）
        dup_name = db.query(Account).filter(
            Account.company_id == company_id,
            Account.name == new_name,
            Account.id != account_id
        ).first()
        if dup_name:
            raise HTTPException(400, detail=f"科目名称【{new_name}】已存在（{dup_name.code}），请更换名称")
        acc.name = new_name
    if data.is_active is not None:
        acc.is_active = data.is_active
    if data.opening_balance is not None:
        acc.opening_balance = data.opening_balance
    db.commit()
    return {"message": "更新成功"}


@app.delete("/api/accounts/{account_id}")
def delete_account(account_id: int, company_id: int = Query(...), db: Session = Depends(get_db), password: str = Body("")):
    pwd = password
    needs_pwd, acc, msg = _account_needs_password(db, company_id, account_id)
    if needs_pwd:
        if pwd != ACCOUNT_PWD:
            raise HTTPException(403, detail=f"{msg}，请输入正确密码")
    if not acc:
        raise HTTPException(404, detail="科目不存在")
    db.delete(acc)
    db.commit()
    return {"message": "删除成功"}


# ==================== 期间管理 ====================

@app.get("/api/periods")
def list_periods(company_id: int = Query(...), db: Session = Depends(get_db)):
    periods = db.query(Period).filter(Period.company_id == company_id).order_by(Period.period.desc()).all()
    return [{"period": p.period, "status": p.status} for p in periods]


@app.post("/api/periods/{period}/close")
def close_period(period: str, company_id: int = Query(...), db: Session = Depends(get_db)):
    p = db.query(Period).filter(Period.company_id == company_id, Period.period == period).first()
    if not p:
        raise HTTPException(404, detail="期间不存在")
    if p.status == "已结账":
        raise HTTPException(400, detail="该期间已结账")
    p.status = "已结账"
    p.closed_at = datetime.now()

    # 自动创建下期
    year, month = int(period[:4]), int(period[5:])
    if month == 12:
        next_period = f"{year + 1}-01"
    else:
        next_period = f"{year}-{str(month + 1).zfill(2)}"
    existing = db.query(Period).filter(Period.company_id == company_id, Period.period == next_period).first()
    if not existing:
        db.add(Period(company_id=company_id, period=next_period))

    db.commit()
    return {"message": f"{period} 结账成功，已自动创建 {next_period} 期间"}


# ==================== 统计看板（原有，保留）====================

@app.get("/api/dashboard")
def dashboard(company_id: int = Query(...), db: Session = Depends(get_db)):
    """统计看板 - 基础档案统计"""
    period = date.today().strftime("%Y-%m")

    customer_count = db.query(Customer).filter(Customer.company_id == company_id).count()
    supplier_count = db.query(Supplier).filter(Supplier.company_id == company_id).count()
    employee_count = db.query(Employee).filter(Employee.company_id == company_id).count()
    account_count = db.query(Account).filter(Account.company_id == company_id, Account.is_active == True).count()

    # 本月发票数量
    si_count = db.query(SalesInvoice).filter(SalesInvoice.company_id == company_id).count()
    pi_count = db.query(PurchaseInvoice).filter(PurchaseInvoice.company_id == company_id).count()
    bi_count = db.query(BookkeepingInvoice).filter(BookkeepingInvoice.company_id == company_id).count()

    return {
        "period": period,
        "customer_count": customer_count,
        "supplier_count": supplier_count,
        "employee_count": employee_count,
        "account_count": account_count,
        "sales_invoice_count": si_count,
        "purchase_invoice_count": pi_count,
        "bookkeeping_invoice_count": bi_count,
    }


# ==================== 公司账套管理 ====================

class CompanyCreate(BaseModel):
    name: str
    uscc: Optional[str] = None
    registered_capital: Optional[float] = None
    established_date: Optional[date] = None
    legal_representative: Optional[str] = None
    legal_representative_id: Optional[str] = None
    address: Optional[str] = None
    business_scope: Optional[str] = None
    company_type: Optional[str] = None
    shareholders: Optional[List[dict]] = None
    directors: Optional[List[dict]] = None
    supervisors: Optional[List[dict]] = None
    finance_contacts: Optional[List[dict]] = None

class CompanyUpdateModel(BaseModel):
    name: Optional[str] = None
    uscc: Optional[str] = None
    registered_capital: Optional[float] = None
    established_date: Optional[date] = None
    legal_representative: Optional[str] = None
    legal_representative_id: Optional[str] = None
    address: Optional[str] = None
    business_scope: Optional[str] = None
    company_type: Optional[str] = None
    shareholders: Optional[List[dict]] = None
    directors: Optional[List[dict]] = None
    supervisors: Optional[List[dict]] = None
    finance_contacts: Optional[List[dict]] = None


@app.get("/api/companies")
def list_companies(db: Session = Depends(get_db)):
    """获取公司列表（账套选择）"""
    companies = db.query(Company).order_by(Company.id).all()
    return [{
        "id": c.id, "name": c.name, "uscc": c.uscc or "",
        "registered_capital": c.registered_capital,
        "established_date": str(c.established_date) if c.established_date else "",
        "legal_representative": c.legal_representative or "",
        "legal_representative_id": c.legal_representative_id or "",
        "address": c.address or "",
        "business_scope": c.business_scope or "",
        "created_at": str(c.created_at.date()) if c.created_at else "",
        "shareholders": [{"name": s.name, "id_number": s.id_number or "", "ratio": s.ratio, "contribution_amount": s.contribution_amount} for s in c.shareholders],
        "directors": [{"name": d.name, "id_number": d.id_number or ""} for d in c.directors],
        "supervisors": [{"name": s.name, "id_number": s.id_number or ""} for s in c.supervisors],
        "finance_contacts": [{"name": f.name, "id_number": f.id_number or "", "phone": f.phone or ""} for f in c.finance_contacts],
    } for c in companies]


@app.post("/api/companies")
def create_company(data: CompanyCreate, db: Session = Depends(get_db)):
    """创建新公司/账套"""
    if data.uscc:
        ok, msg = validate_uscc(data.uscc)
        if not ok:
            raise HTTPException(400, detail=f"统一社会信用代码：{msg}")

    main_fields = {k: v for k, v in data.model_dump(exclude_unset=True).items()
                   if k not in ("shareholders", "directors", "supervisors", "finance_contacts")}
    company = Company(**main_fields)
    db.add(company)
    db.flush()

    # 子表
    _update_company_subtable(db, company, CompanyShareholder, data.shareholders)
    _update_company_subtable(db, company, CompanyDirector, data.directors)
    _update_company_subtable(db, company, CompanySupervisor, data.supervisors)
    _update_company_subtable(db, company, CompanyFinanceContact, data.finance_contacts)

    # 初始化公司基础数据（科目表、部门、期间）
    init_company_data(db, company.id)
    db.commit()
    db.refresh(company)

    return {"id": company.id, "name": company.name, "message": f"公司 '{company.name}' 创建成功，已初始化科目表和基础档案"}


@app.put("/api/companies/{company_id}")
def update_company_detail(company_id: int, data: CompanyUpdateModel, db: Session = Depends(get_db)):
    """更新公司信息"""
    company = db.query(Company).filter(Company.id == company_id).first()
    if not company:
        raise HTTPException(404, detail="公司不存在")
    if data.uscc:
        ok, msg = validate_uscc(data.uscc)
        if not ok:
            raise HTTPException(400, detail=f"统一社会信用代码：{msg}")
    main_fields = {k: v for k, v in data.model_dump(exclude_unset=True).items()
                   if k not in ("shareholders", "directors", "supervisors", "finance_contacts")}
    for k, v in main_fields.items():
        setattr(company, k, v)
    _update_company_subtable(db, company, CompanyShareholder, data.shareholders)
    _update_company_subtable(db, company, CompanyDirector, data.directors)
    _update_company_subtable(db, company, CompanySupervisor, data.supervisors)
    _update_company_subtable(db, company, CompanyFinanceContact, data.finance_contacts)
    db.commit()
    return {"message": "更新成功"}


@app.delete("/api/companies/{company_id}")
def delete_company(company_id: int, db: Session = Depends(get_db)):
    """删除公司及其所有关联数据"""
    company = db.query(Company).filter(Company.id == company_id).first()
    if not company:
        raise HTTPException(404, detail="公司不存在")

    # 级联删除顺序：先删子表（有外键的表），再删中间表，最后删主表
    # 1. 公司治理层子表
    db.query(CompanyShareholder).filter(CompanyShareholder.company_id == company_id).delete()
    db.query(CompanyDirector).filter(CompanyDirector.company_id == company_id).delete()
    db.query(CompanySupervisor).filter(CompanySupervisor.company_id == company_id).delete()
    db.query(CompanyFinanceContact).filter(CompanyFinanceContact.company_id == company_id).delete()
    # 2. 档案类
    db.query(Department).filter(Department.company_id == company_id).delete()
    db.query(Employee).filter(Employee.company_id == company_id).delete()
    db.query(Customer).filter(Customer.company_id == company_id).delete()
    db.query(Supplier).filter(Supplier.company_id == company_id).delete()
    db.query(Account).filter(Account.company_id == company_id).delete()
    db.query(Period).filter(Period.company_id == company_id).delete()
    # 3. 资产/库存
    db.query(FixedAssetDepreciation).filter(FixedAssetDepreciation.company_id == company_id).delete()
    db.query(FixedAsset).filter(FixedAsset.company_id == company_id).delete()
    db.query(IntangibleAssetAmortization).filter(IntangibleAssetAmortization.company_id == company_id).delete()
    db.query(IntangibleAsset).filter(IntangibleAsset.company_id == company_id).delete()
    db.query(InventoryTransaction).filter(InventoryTransaction.company_id == company_id).delete()
    db.query(InventoryBalance).filter(InventoryBalance.company_id == company_id).delete()
    db.query(InventoryItem).filter(InventoryItem.company_id == company_id).delete()
    # 4. 合同/付款
    db.query(ContractPayment).filter(ContractPayment.company_id == company_id).delete()
    db.query(Contract).filter(Contract.company_id == company_id).delete()
    db.query(Payment).filter(Payment.company_id == company_id).delete()
    # 5. 业务核心
    db.query(SalesInvoice).filter(SalesInvoice.company_id == company_id).delete()
    db.query(PurchaseInvoice).filter(PurchaseInvoice.company_id == company_id).delete()
    db.query(BookkeepingInvoice).filter(BookkeepingInvoice.company_id == company_id).delete()
    db.query(InputVATDeduction).filter(InputVATDeduction.company_id == company_id).delete()
    db.query(BankTransaction).filter(BankTransaction.company_id == company_id).delete()
    db.query(BankConfig).filter(BankConfig.company_id == company_id).delete()
    db.query(JournalEntry).filter(JournalEntry.company_id == company_id).delete()
    db.query(ColumnTemplate).filter(ColumnTemplate.company_id == company_id).delete()
    # 6. 子模块表（salary_records / vat_declarations 通过 raw SQL 确保兼容）
    import importlib
    try:
        salary_mod = importlib.import_module('salary')
        vat_mod = importlib.import_module('vat')
    except Exception:
        salary_mod = None; vat_mod = None
    if salary_mod:
        from database import SalaryRecord
        db.query(SalaryRecord).filter(SalaryRecord.company_id == company_id).delete()
    if vat_mod:
        from database import VATDeclaration
        db.query(VATDeclaration).filter(VATDeclaration.company_id == company_id).delete()
    # 7. 字典表（不按company_id隔离，跳过）
    # 8. 终删公司
    db.delete(company)
    db.commit()
    return {"message": "公司及全部关联数据已删除"}


# ==================== 固定资产 ====================

class FixedAssetCreate(BaseModel):
    code: str
    name: str
    category: str = "机器设备"
    spec: Optional[str] = None
    unit: Optional[str] = "台"
    dept_code: Optional[str] = None
    location: Optional[str] = None
    purchase_date: Optional[date] = None
    original_value: float = 0.0
    residual_value: float = 0.0
    useful_life_months: int = 60
    depreciation_method: str = "直线法"
    supplier: Optional[str] = None
    warranty_expiry: Optional[date] = None
    remark: Optional[str] = None


class FixedAssetUpdate(BaseModel):
    code: Optional[str] = None
    name: Optional[str] = None
    category: Optional[str] = None
    spec: Optional[str] = None
    unit: Optional[str] = None
    dept_code: Optional[str] = None
    location: Optional[str] = None
    purchase_date: Optional[date] = None
    original_value: Optional[float] = None
    residual_value: Optional[float] = None
    useful_life_months: Optional[int] = None
    depreciation_method: Optional[str] = None
    supplier: Optional[str] = None
    status: Optional[str] = None
    remark: Optional[str] = None


@app.get("/api/fixed-assets")
def list_fixed_assets(
    company_id: int = Query(...),
    category: Optional[str] = None,
    status: Optional[str] = None,
    keyword: Optional[str] = None,
    db: Session = Depends(get_db)
):
    q = db.query(FixedAsset).filter(FixedAsset.company_id == company_id)
    if category:
        q = q.filter(FixedAsset.category == category)
    if status:
        q = q.filter(FixedAsset.status == status)
    if keyword:
        q = q.filter(or_(FixedAsset.code.contains(keyword), FixedAsset.name.contains(keyword)))
    assets = q.order_by(FixedAsset.code).all()
    return [{
        "id": a.id, "code": a.code, "name": a.name, "category": a.category,
        "spec": a.spec, "unit": a.unit, "dept_code": a.dept_code,
        "location": a.location, "purchase_date": str(a.purchase_date) if a.purchase_date else "",
        "original_value": a.original_value, "residual_value": a.residual_value,
        "useful_life_months": a.useful_life_months,
        "accumulated_depreciation": a.accumulated_depreciation,
        "monthly_depreciation": a.monthly_depreciation,
        "depreciation_method": a.depreciation_method,
        "status": a.status, "supplier": a.supplier,
        "net_value": round(a.original_value - a.accumulated_depreciation, 2),
        "net_rate": round((a.original_value - a.accumulated_depreciation) / a.original_value * 100, 1) if a.original_value > 0 else 0,
        "remark": a.remark
    } for a in assets]


@app.get("/api/fixed-assets/{fa_id}")
def get_fixed_asset(fa_id: int, company_id: int = Query(...), db: Session = Depends(get_db)):
    fa = db.query(FixedAsset).filter(FixedAsset.company_id == company_id, FixedAsset.id == fa_id).first()
    if not fa:
        raise HTTPException(404, detail="资产不存在")
    return {
        "id": fa.id, "code": fa.code, "name": fa.name, "category": fa.category,
        "spec": fa.spec, "unit": fa.unit, "dept_code": fa.dept_code,
        "location": fa.location, "purchase_date": str(fa.purchase_date) if fa.purchase_date else "",
        "original_value": fa.original_value, "residual_value": fa.residual_value,
        "useful_life_months": fa.useful_life_months,
        "accumulated_depreciation": fa.accumulated_depreciation,
        "monthly_depreciation": fa.monthly_depreciation,
        "depreciation_method": fa.depreciation_method,
        "status": fa.status, "supplier": fa.supplier,
        "net_value": round(fa.original_value - fa.accumulated_depreciation, 2),
        "net_rate": round((fa.original_value - fa.accumulated_depreciation) / fa.original_value * 100, 1) if fa.original_value > 0 else 0,
        "remark": fa.remark
    }


@app.post("/api/fixed-assets")
def create_fixed_asset(data: FixedAssetCreate, company_id: int = Query(...), db: Session = Depends(get_db)):
    # 计算月折旧额（直线法）
    monthly = 0.0
    if data.useful_life_months > 0:
        monthly = round((data.original_value - data.residual_value) / data.useful_life_months, 2)
    fa = FixedAsset(
        company_id=company_id, code=data.code, name=data.name,
        category=data.category, spec=data.spec, unit=data.unit,
        dept_code=data.dept_code, location=data.location,
        purchase_date=data.purchase_date, original_value=data.original_value,
        residual_value=data.residual_value, useful_life_months=data.useful_life_months,
        monthly_depreciation=monthly, depreciation_method=data.depreciation_method,
        supplier=data.supplier, warranty_expiry=data.warranty_expiry,
        remark=data.remark
    )
    db.add(fa)
    db.commit()
    db.refresh(fa)
    return {"id": fa.id, "code": fa.code, "message": "固定资产新增成功"}


@app.put("/api/fixed-assets/{fa_id}")
def update_fixed_asset(fa_id: int, data: FixedAssetUpdate, company_id: int = Query(...), db: Session = Depends(get_db)):
    fa = db.query(FixedAsset).filter(FixedAsset.company_id == company_id, FixedAsset.id == fa_id).first()
    if not fa:
        raise HTTPException(404, detail="资产不存在")
    for k, v in data.model_dump(exclude_unset=True).items():
        setattr(fa, k, v)
    # 重新计算月折旧额
    if data.original_value is not None or data.residual_value is not None or data.useful_life_months is not None:
        if fa.useful_life_months and fa.useful_life_months > 0:
            orig = float(fa.original_value or 0)
            resid = float(fa.residual_value or 0)
            fa.monthly_depreciation = round((orig - resid) / fa.useful_life_months, 2)
    fa.updated_at = datetime.now()
    db.commit()
    return {"message": "更新成功"}


@app.post("/api/fixed-assets/{fa_id}/depreciate")
def depreciate_asset(fa_id: int, period: str = Query(...), company_id: int = Query(...), db: Session = Depends(get_db)):
    """计提单月折旧"""
    fa = db.query(FixedAsset).filter(FixedAsset.company_id == company_id, FixedAsset.id == fa_id).first()
    if not fa:
        raise HTTPException(404, detail="资产不存在")
    if fa.status != "在用":
        raise HTTPException(400, detail=f"资产状态为'{fa.status}'，不能计提折旧")
    # 检查是否已折旧
    existing = db.query(FixedAssetDepreciation).filter(
        FixedAssetDepreciation.company_id == company_id,
        FixedAssetDepreciation.asset_id == fa_id,
        FixedAssetDepreciation.period == period
    ).first()
    if existing:
        raise HTTPException(400, detail=f"该资产在 {period} 期间已计提折旧")
    # 累计折旧不能超过（原值-残值）
    orig = float(fa.original_value or 0)
    resid = float(fa.residual_value or 0)
    accum = float(fa.accumulated_depreciation or 0)
    monthly = float(fa.monthly_depreciation or 0)
    if accum + monthly > orig - resid:
        dep_amount = orig - resid - accum
    else:
        dep_amount = monthly
    if dep_amount <= 0:
        raise HTTPException(400, detail="该资产已提足折旧")
    acc_before = accum
    fa.accumulated_depreciation = acc_before + dep_amount
    fa.updated_at = datetime.now()
    rec = FixedAssetDepreciation(
        company_id=company_id, asset_id=fa_id, period=period,
        depreciation_amount=dep_amount, accumulated_before=acc_before,
        accumulated_after=round(orig - resid, 2),
        net_value=round(orig - resid, 2)
    )
    db.add(rec)
    db.commit()
    return {"message": f"计提折旧 ¥{dep_amount:.2f}，累计折旧 ¥{fa.accumulated_depreciation:.2f}"}


@app.get("/api/fixed-assets/{fa_id}/depreciations")
def get_asset_depreciations(fa_id: int, company_id: int = Query(...), db: Session = Depends(get_db)):
    recs = db.query(FixedAssetDepreciation).filter(
        FixedAssetDepreciation.company_id == company_id,
        FixedAssetDepreciation.asset_id == fa_id
    ).order_by(FixedAssetDepreciation.period).all()
    return [{
        "id": r.id, "period": r.period, "depreciation_amount": r.depreciation_amount,
        "accumulated_before": r.accumulated_before, "accumulated_after": r.accumulated_after,
        "net_value": r.net_value
    } for r in recs]


@app.delete("/api/fixed-assets/{fa_id}")
def delete_fixed_asset(fa_id: int, company_id: int = Query(...), db: Session = Depends(get_db)):
    fa = db.query(FixedAsset).filter(FixedAsset.company_id == company_id, FixedAsset.id == fa_id).first()
    if not fa:
        raise HTTPException(404, detail="资产不存在")
    if fa.status == "在用":
        raise HTTPException(400, detail="在用资产不能删除，请先变更为闲置或报废")
    db.delete(fa)
    db.commit()
    return {"message": "删除成功"}


@app.post("/api/fixed-assets/depreciate")
def batch_depreciate(company_id: int = Query(...), period: Optional[str] = None,
                     db: Session = Depends(get_db)):
    """批量计提折旧 + 自动生成凭证"""
    if not period:
        period = datetime.now().strftime("%Y-%m")
    assets = db.query(FixedAsset).filter(
        FixedAsset.company_id == company_id, FixedAsset.status == "在用"
    ).all()
    if not assets:
        return {"depreciated_count": 0, "message": "无在用资产"}
    
    depreciated = []
    total_amount = 0.0
    for fa in assets:
        existing = db.query(FixedAssetDepreciation).filter(
            FixedAssetDepreciation.company_id == company_id,
            FixedAssetDepreciation.asset_id == fa.id,
            FixedAssetDepreciation.period == period
        ).first()
        if existing:
            continue
        monthly = float(fa.monthly_depreciation or 0)
        if monthly <= 0:
            continue
        orig = float(fa.original_value or 0)
        resid = float(fa.residual_value or 0)
        accum = float(fa.accumulated_depreciation or 0)
        max_dep = orig - resid - accum
        if max_dep <= 0:
            continue
        dep_amount = min(monthly, max_dep)
        acc_before = float(fa.accumulated_depreciation or 0)
        fa.accumulated_depreciation = acc_before + dep_amount
        fa.updated_at = datetime.now()
        rec = FixedAssetDepreciation(
            company_id=company_id, asset_id=fa.id, period=period,
            depreciation_amount=dep_amount, accumulated_before=acc_before,
            accumulated_after=round(orig - resid, 2),
            net_value=round(orig - resid, 2)
        )
        db.add(rec)
        depreciated.append((fa, dep_amount))
        total_amount += dep_amount
    
    if not depreciated:
        db.commit()
        return {"depreciated_count": 0, "total_amount": 0, "message": "所有资产已计提或无需折旧"}
    
    # 生成折旧凭证
    _ensure_account(db, company_id, "1602", "累计折旧", "资产", "贷")
    _ensure_account(db, company_id, "660203", "折旧费", "损益", "借", parent_code="6602")
    _ensure_account(db, company_id, "6602", "管理费用", "损益", "借")
    
    next_vno = _next_voucher_no(db, company_id, period)
    summary = f"计提{period}固定资产折旧（{len(depreciated)}项）"
    # 借方：管理费用-折旧费
    je_debit = JournalEntry(
        company_id=company_id, period=period, voucher_word="记", voucher_no=next_vno,
        entry_date=datetime.now().date(), summary=summary, account_code="660203", account_name="折旧费",
        debit_amount=round(total_amount, 2), credit_amount=0,
        source="折旧计提"
    )
    db.add(je_debit)
    # 贷方：累计折旧
    je_credit = JournalEntry(
        company_id=company_id, period=period, voucher_word="记", voucher_no=next_vno,
        entry_date=datetime.now().date(), summary=summary, account_code="1602", account_name="累计折旧",
        debit_amount=0, credit_amount=round(total_amount, 2),
        source="折旧计提"
    )
    db.add(je_credit)
    
    db.commit()
    return {
        "depreciated_count": len(depreciated),
        "total_amount": round(total_amount, 2),
        "voucher_no": f"记-{next_vno}",
        "message": f"计提{len(depreciated)}项资产折旧 ¥{total_amount:,.2f}"
    }


@app.get("/api/fixed-assets/stats")
def fixed_assets_stats(company_id: int = Query(...), db: Session = Depends(get_db)):
    """固定资产统计概览"""
    assets = db.query(FixedAsset).filter(FixedAsset.company_id == company_id).all()
    active = [a for a in assets if a.status == "在用"]
    return {
        "total_count": len(assets),
        "active_count": len(active),
        "total_original": round(sum(a.original_value for a in assets), 2),
        "total_depreciation": round(sum(a.accumulated_depreciation for a in assets), 2),
        "total_net_value": round(sum(a.original_value - a.accumulated_depreciation for a in assets), 2),
        "monthly_depreciation": round(sum(a.monthly_depreciation for a in active), 2),
    }


@app.post("/api/fixed-assets/batch-delete")
def batch_delete_fixed_assets(ids: List[int], company_id: int = Query(...), db: Session = Depends(get_db)):
    deleted = 0
    for fa_id in ids:
        fa = db.query(FixedAsset).filter(FixedAsset.company_id == company_id, FixedAsset.id == fa_id).first()
        if fa and fa.status != "在用":
            db.delete(fa)
            deleted += 1
    db.commit()
    return {"deleted": deleted, "message": f"删除 {deleted} 项"}


# ==================== 无形资产 ====================

class IntangibleAssetCreate(BaseModel):
    code: str
    name: str
    category: str = "专利权"
    purchase_date: Optional[date] = None
    original_value: float = 0.0
    useful_life_months: int = 120
    residual_value: float = 0.0
    remark: Optional[str] = None


class IntangibleAssetUpdate(BaseModel):
    code: Optional[str] = None
    name: Optional[str] = None
    category: Optional[str] = None
    purchase_date: Optional[date] = None
    original_value: Optional[float] = None
    residual_value: Optional[float] = None
    useful_life_months: Optional[int] = None
    status: Optional[str] = None
    remark: Optional[str] = None


@app.get("/api/intangible-assets")
def list_intangible_assets(
    company_id: int = Query(...),
    category: Optional[str] = None,
    keyword: Optional[str] = None,
    db: Session = Depends(get_db)
):
    q = db.query(IntangibleAsset).filter(IntangibleAsset.company_id == company_id)
    if category:
        q = q.filter(IntangibleAsset.category == category)
    if keyword:
        q = q.filter(or_(IntangibleAsset.code.contains(keyword), IntangibleAsset.name.contains(keyword)))
    assets = q.order_by(IntangibleAsset.code).all()
    return [{
        "id": a.id, "code": a.code, "name": a.name, "category": a.category,
        "purchase_date": str(a.purchase_date) if a.purchase_date else "",
        "original_value": a.original_value, "residual_value": a.residual_value,
        "useful_life_months": a.useful_life_months,
        "accumulated_amortization": a.accumulated_amortization,
        "monthly_amortization": a.monthly_amortization,
        "status": a.status,
        "net_value": round(a.original_value - a.accumulated_amortization, 2),
        "remark": a.remark
    } for a in assets]


@app.get("/api/intangible-assets/{ia_id}")
def get_intangible_asset(ia_id: int, company_id: int = Query(...), db: Session = Depends(get_db)):
    a = db.query(IntangibleAsset).filter(IntangibleAsset.company_id == company_id, IntangibleAsset.id == ia_id).first()
    if not a:
        raise HTTPException(404, detail="无形资产不存在")
    return {
        "id": a.id, "code": a.code, "name": a.name, "category": a.category,
        "purchase_date": str(a.purchase_date) if a.purchase_date else "",
        "original_value": a.original_value, "residual_value": a.residual_value,
        "useful_life_months": a.useful_life_months,
        "accumulated_amortization": a.accumulated_amortization,
        "monthly_amortization": a.monthly_amortization,
        "status": a.status,
        "net_value": round(a.original_value - a.accumulated_amortization, 2),
        "remark": a.remark
    }


@app.post("/api/intangible-assets")
def create_intangible_asset(data: IntangibleAssetCreate, company_id: int = Query(...), db: Session = Depends(get_db)):
    monthly = round((data.original_value - data.residual_value) / data.useful_life_months, 2) if data.useful_life_months > 0 else 0
    ia = IntangibleAsset(
        company_id=company_id, code=data.code, name=data.name,
        category=data.category, purchase_date=data.purchase_date,
        original_value=data.original_value, useful_life_months=data.useful_life_months,
        residual_value=data.residual_value, monthly_amortization=monthly,
        remark=data.remark
    )
    db.add(ia)
    db.commit()
    db.refresh(ia)
    return {"id": ia.id, "code": ia.code, "message": "无形资产新增成功"}


@app.put("/api/intangible-assets/{ia_id}")
def update_intangible_asset(ia_id: int, data: IntangibleAssetUpdate, company_id: int = Query(...), db: Session = Depends(get_db)):
    ia = db.query(IntangibleAsset).filter(IntangibleAsset.company_id == company_id, IntangibleAsset.id == ia_id).first()
    if not ia:
        raise HTTPException(404, detail="资产不存在")
    for k, v in data.model_dump(exclude_unset=True).items():
        setattr(ia, k, v)
    if ia.useful_life_months and ia.useful_life_months > 0:
        orig = float(ia.original_value or 0)
        resid = float(ia.residual_value or 0)
        ia.monthly_amortization = round((orig - resid) / ia.useful_life_months, 2)
    ia.updated_at = datetime.now()
    db.commit()
    return {"message": "更新成功"}


@app.post("/api/intangible-assets/{ia_id}/amortize")
def amortize_asset(ia_id: int, period: str = Query(...), company_id: int = Query(...), db: Session = Depends(get_db)):
    """计提单月摊销"""
    ia = db.query(IntangibleAsset).filter(IntangibleAsset.company_id == company_id, IntangibleAsset.id == ia_id).first()
    if not ia:
        raise HTTPException(404, detail="资产不存在")
    if ia.status != "在用":
        raise HTTPException(400, detail=f"资产状态为'{ia.status}'，不能摊销")
    existing = db.query(IntangibleAssetAmortization).filter(
        IntangibleAssetAmortization.company_id == company_id,
        IntangibleAssetAmortization.asset_id == ia_id,
        IntangibleAssetAmortization.period == period
    ).first()
    if existing:
        raise HTTPException(400, detail=f"该资产在 {period} 期间已摊销")
    orig = float(ia.original_value or 0)
    resid = float(ia.residual_value or 0)
    accum = float(ia.accumulated_amortization or 0)
    monthly = float(ia.monthly_amortization or 0)
    if accum + monthly > orig - resid:
        amt = orig - resid - accum
    else:
        amt = monthly
    if amt <= 0:
        raise HTTPException(400, detail="该资产已摊销完毕")
    acc_before = accum
    ia.accumulated_amortization = acc_before + amt
    ia.updated_at = datetime.now()
    rec = IntangibleAssetAmortization(
        company_id=company_id, asset_id=ia_id, period=period,
        amortization_amount=amt, accumulated_before=acc_before,
        accumulated_after=round(orig - resid, 2),
        net_value=round(orig - resid, 2)
    )
    db.add(rec)
    db.commit()
    return {"message": f"摊销 ¥{amt:.2f}，累计摊销 ¥{ia.accumulated_amortization:.2f}"}


@app.delete("/api/intangible-assets/{ia_id}")
def delete_intangible_asset(ia_id: int, company_id: int = Query(...), db: Session = Depends(get_db)):
    ia = db.query(IntangibleAsset).filter(IntangibleAsset.company_id == company_id, IntangibleAsset.id == ia_id).first()
    if not ia:
        raise HTTPException(404, detail="资产不存在")
    db.delete(ia)
    db.commit()
    return {"message": "删除成功"}


@app.get("/api/intangible-assets/{ia_id}/amortizations")
def get_asset_amortizations(ia_id: int, company_id: int = Query(...), db: Session = Depends(get_db)):
    """查询某项无形资产的摊销明细"""
    recs = db.query(IntangibleAssetAmortization).filter(
        IntangibleAssetAmortization.company_id == company_id,
        IntangibleAssetAmortization.asset_id == ia_id
    ).order_by(IntangibleAssetAmortization.period).all()
    return [{
        "id": r.id, "period": r.period, "amortization_amount": r.amortization_amount,
        "accumulated_before": r.accumulated_before, "accumulated_after": r.accumulated_after,
        "net_value": r.net_value
    } for r in recs]


@app.post("/api/intangible-assets/amortize")
def batch_amortize(company_id: int = Query(...), period: Optional[str] = None,
                   db: Session = Depends(get_db)):
    """批量摊销 + 自动生成凭证"""
    if not period:
        period = datetime.now().strftime("%Y-%m")
    assets = db.query(IntangibleAsset).filter(
        IntangibleAsset.company_id == company_id, IntangibleAsset.status == "在用"
    ).all()
    if not assets:
        return {"amortized_count": 0, "message": "无在用资产"}
    
    amortized = []
    total_amount = 0.0
    for ia in assets:
        existing = db.query(IntangibleAssetAmortization).filter(
            IntangibleAssetAmortization.company_id == company_id,
            IntangibleAssetAmortization.asset_id == ia.id,
            IntangibleAssetAmortization.period == period
        ).first()
        if existing:
            continue
        monthly = float(ia.monthly_amortization or 0)
        if monthly <= 0:
            continue
        orig = float(ia.original_value or 0)
        resid = float(ia.residual_value or 0)
        accum = float(ia.accumulated_amortization or 0)
        max_amt = orig - resid - accum
        if max_amt <= 0:
            continue
        amt = min(monthly, max_amt)
        acc_before = float(ia.accumulated_amortization or 0)
        ia.accumulated_amortization = acc_before + amt
        ia.updated_at = datetime.now()
        rec = IntangibleAssetAmortization(
            company_id=company_id, asset_id=ia.id, period=period,
            amortization_amount=amt, accumulated_before=acc_before,
            accumulated_after=round(orig - resid, 2),
            net_value=round(orig - resid, 2)
        )
        db.add(rec)
        amortized.append((ia, amt))
        total_amount += amt
    
    if not amortized:
        db.commit()
        return {"amortized_count": 0, "total_amount": 0, "message": "所有资产已摊销或无需摊销"}
    
    # 生成摊销凭证
    _ensure_account(db, company_id, "1702", "累计摊销", "资产", "贷")
    _ensure_account(db, company_id, "660208", "摊销费", "损益", "借", parent_code="6602")
    _ensure_account(db, company_id, "6602", "管理费用", "损益", "借")
    
    next_vno = _next_voucher_no(db, company_id, period)
    summary = f"计提{period}无形资产摊销（{len(amortized)}项）"
    je_debit = JournalEntry(
        company_id=company_id, period=period, voucher_word="记", voucher_no=next_vno,
        entry_date=datetime.now().date(), summary=summary, account_code="660208", account_name="摊销费",
        debit_amount=round(total_amount, 2), credit_amount=0,
        source="摊销计提"
    )
    db.add(je_debit)
    je_credit = JournalEntry(
        company_id=company_id, period=period, voucher_word="记", voucher_no=next_vno,
        entry_date=datetime.now().date(), summary=summary, account_code="1702", account_name="累计摊销",
        debit_amount=0, credit_amount=round(total_amount, 2),
        source="摊销计提"
    )
    db.add(je_credit)
    
    db.commit()
    return {
        "amortized_count": len(amortized),
        "total_amount": round(total_amount, 2),
        "voucher_no": f"记-{next_vno}",
        "message": f"摊销{len(amortized)}项资产 ¥{total_amount:,.2f}"
    }


@app.get("/api/intangible-assets/stats")
def intangible_assets_stats(company_id: int = Query(...), db: Session = Depends(get_db)):
    assets = db.query(IntangibleAsset).filter(IntangibleAsset.company_id == company_id).all()
    active = [a for a in assets if a.status == "在用"]
    return {
        "total_count": len(assets),
        "active_count": len(active),
        "total_original": round(sum(a.original_value for a in assets), 2),
        "total_amortization": round(sum(a.accumulated_amortization for a in assets), 2),
        "total_net_value": round(sum(a.original_value - a.accumulated_amortization for a in assets), 2),
        "monthly_amortization": round(sum(a.monthly_amortization for a in active), 2),
    }


@app.post("/api/intangible-assets/batch-delete")
def batch_delete_intangible_assets(ids: List[int], company_id: int = Query(...), db: Session = Depends(get_db)):
    deleted = 0
    for ia_id in ids:
        ia = db.query(IntangibleAsset).filter(IntangibleAsset.company_id == company_id, IntangibleAsset.id == ia_id).first()
        if ia:
            db.delete(ia)
            deleted += 1
    db.commit()
    return {"deleted": deleted, "message": f"删除 {deleted} 项"}


# ==================== 库存管理 ====================

class InventoryItemCreate(BaseModel):
    code: str
    name: str
    spec: Optional[str] = None
    unit: Optional[str] = "个"
    category: Optional[str] = "原材料"
    warehouse: Optional[str] = None
    safety_stock: float = 0.0
    cost_price: float = 0.0
    sale_price: float = 0.0
    account_code: Optional[str] = None
    remark: Optional[str] = None

class InventoryItemUpdate(BaseModel):
    name: Optional[str] = None
    spec: Optional[str] = None
    category: Optional[str] = None
    warehouse: Optional[str] = None
    safety_stock: Optional[float] = None
    cost_price: Optional[float] = None
    sale_price: Optional[float] = None
    is_active: Optional[bool] = None
    remark: Optional[str] = None

class InventoryTransactionCreate(BaseModel):
    transaction_date: date
    trans_type: str  # 入库/出库/调拨入/调拨出/盘盈/盘亏/其他
    item_code: str
    quantity: float
    unit_price: float = 0.0
    warehouse: Optional[str] = None
    warehouse_to: Optional[str] = None
    voucher_no: Optional[str] = None
    reference_no: Optional[str] = None
    operator: Optional[str] = "管理员"
    remark: Optional[str] = None


@app.get("/api/inventory-items")
def list_inventory_items(
    company_id: int = Query(...),
    category: Optional[str] = None,
    keyword: Optional[str] = None,
    db: Session = Depends(get_db)
):
    q = db.query(InventoryItem).filter(InventoryItem.company_id == company_id, InventoryItem.is_active == True)
    if category:
        q = q.filter(InventoryItem.category == category)
    if keyword:
        q = q.filter(or_(InventoryItem.code.contains(keyword), InventoryItem.name.contains(keyword)))
    items = q.order_by(InventoryItem.code).all()
    return [{
        "id": i.id, "code": i.code, "name": i.name, "spec": i.spec,
        "unit": i.unit, "category": i.category, "warehouse": i.warehouse,
        "safety_stock": i.safety_stock, "current_stock": i.current_stock,
        "cost_price": i.cost_price, "sale_price": i.sale_price,
        "stock_value": round(i.current_stock * i.cost_price, 2),
        "account_code": i.account_code, "remark": i.remark
    } for i in items]


@app.get("/api/inventory-items/{item_id}")
def get_inventory_item(item_id: int, company_id: int = Query(...), db: Session = Depends(get_db)):
    i = db.query(InventoryItem).filter(InventoryItem.company_id == company_id, InventoryItem.id == item_id).first()
    if not i:
        raise HTTPException(404, detail="存货不存在")
    return {
        "id": i.id, "code": i.code, "name": i.name, "spec": i.spec,
        "unit": i.unit, "category": i.category, "warehouse": i.warehouse,
        "safety_stock": i.safety_stock, "current_stock": i.current_stock,
        "cost_price": i.cost_price, "sale_price": i.sale_price,
        "stock_value": round(i.current_stock * i.cost_price, 2),
        "account_code": i.account_code, "remark": i.remark
    }


@app.post("/api/inventory-items")
def create_inventory_item(data: InventoryItemCreate, company_id: int = Query(...), db: Session = Depends(get_db)):
    item = InventoryItem(company_id=company_id, **data.model_dump())
    db.add(item)
    db.commit()
    db.refresh(item)
    return {"id": item.id, "code": item.code, "message": "新增成功"}


@app.put("/api/inventory-items/{item_id}")
def update_inventory_item(item_id: int, data: InventoryItemUpdate, company_id: int = Query(...), db: Session = Depends(get_db)):
    item = db.query(InventoryItem).filter(InventoryItem.company_id == company_id, InventoryItem.id == item_id).first()
    if not item:
        raise HTTPException(404, detail="商品不存在")
    for k, v in data.model_dump(exclude_unset=True).items():
        setattr(item, k, v)
    item.updated_at = datetime.now()
    db.commit()
    return {"message": "更新成功"}


@app.delete("/api/inventory-items/{item_id}")
def delete_inventory_item(item_id: int, company_id: int = Query(...), db: Session = Depends(get_db)):
    item = db.query(InventoryItem).filter(InventoryItem.company_id == company_id, InventoryItem.id == item_id).first()
    if not item:
        raise HTTPException(404, detail="存货不存在")
    item.is_active = False
    item.updated_at = datetime.now()
    db.commit()
    return {"message": "删除成功"}


@app.get("/api/inventory-transactions")
def list_inventory_transactions(
    company_id: int = Query(...),
    item_code: Optional[str] = None,
    trans_type: Optional[str] = None,
    limit: int = 100,
    db: Session = Depends(get_db)
):
    q = db.query(InventoryTransaction).filter(InventoryTransaction.company_id == company_id)
    if item_code:
        q = q.filter(InventoryTransaction.item_code == item_code)
    if trans_type:
        q = q.filter(InventoryTransaction.trans_type == trans_type)
    items = q.order_by(InventoryTransaction.transaction_date.desc(), InventoryTransaction.id.desc()).limit(limit).all()
    return [{
        "id": t.id, "item_code": t.item_code, "transaction_date": str(t.transaction_date),
        "trans_type": t.trans_type, "quantity": t.quantity, "unit_price": t.unit_price,
        "total_amount": t.total_amount, "warehouse": t.warehouse, "warehouse_to": t.warehouse_to,
        "voucher_no": t.voucher_no, "reference_no": t.reference_no,
        "operator": t.operator, "remark": t.remark, "created_at": str(t.created_at.date()) if t.created_at else ""
    } for t in items]


@app.post("/api/inventory-transactions")
def create_inventory_transaction(data: InventoryTransactionCreate, company_id: int = Query(...), db: Session = Depends(get_db)):
    # 校验商品存在
    item = db.query(InventoryItem).filter(InventoryItem.company_id == company_id, InventoryItem.code == data.item_code).first()
    if not item:
        raise HTTPException(400, detail=f"商品 {data.item_code} 不存在")
    qty = data.quantity
    total = round(abs(qty) * data.unit_price, 2)
    trans = InventoryTransaction(
        company_id=company_id, item_code=data.item_code,
        transaction_date=data.transaction_date, trans_type=data.trans_type,
        quantity=qty, unit_price=data.unit_price, total_amount=total,
        warehouse=data.warehouse, warehouse_to=data.warehouse_to,
        voucher_no=data.voucher_no, reference_no=data.reference_no,
        operator=data.operator, remark=data.remark
    )
    db.add(trans)
    # 更新库存
    if data.trans_type in ("入库", "调拨入", "盘盈", "其他"):
        item.current_stock += qty
    elif data.trans_type in ("出库", "调拨出", "盘亏"):
        item.current_stock -= qty
        if item.current_stock < 0:
            item.current_stock += qty
            raise HTTPException(400, detail=f"库存不足，当前库存: {item.current_stock}")
    item.updated_at = datetime.now()
    db.commit()
    db.refresh(trans)
    return {"id": trans.id, "message": f"{data.trans_type}成功，当前库存: {item.current_stock}"}


@app.post("/api/inventory-transactions/transfer")
def create_inventory_transfer(company_id: int = Query(...), db: Session = Depends(get_db),
    item_code: str = Form(...), transaction_date: str = Form(...),
    quantity: float = Form(...), warehouse_from: str = Form(...),
    warehouse_to: str = Form(...), unit_price: float = Form(0.0),
    reference_no: str = Form(""), operator: str = Form("管理员"),
    remark: str = Form("")):
    """仓库间调拨"""
    item = db.query(InventoryItem).filter(InventoryItem.company_id == company_id, InventoryItem.code == item_code).first()
    if not item:
        raise HTTPException(400, detail=f"商品 {item_code} 不存在")
    if quantity <= 0:
        raise HTTPException(400, detail="数量必须大于0")
    total = round(quantity * unit_price, 2)
    tx_date = datetime.strptime(transaction_date, "%Y-%m-%d").date() if transaction_date else datetime.now().date()
    
    # 调拨出
    out_tx = InventoryTransaction(
        company_id=company_id, item_code=item_code, transaction_date=tx_date,
        trans_type="调拨出", quantity=quantity, unit_price=unit_price, total_amount=total,
        warehouse=warehouse_from, warehouse_to=warehouse_to,
        reference_no=reference_no, operator=operator, remark=remark
    )
    db.add(out_tx)
    
    # 调拨入
    in_tx = InventoryTransaction(
        company_id=company_id, item_code=item_code, transaction_date=tx_date,
        trans_type="调拨入", quantity=quantity, unit_price=unit_price, total_amount=total,
        warehouse=warehouse_to, warehouse_to=warehouse_from,
        reference_no=reference_no, operator=operator, remark=remark
    )
    db.add(in_tx)
    
    db.commit()
    return {"message": f"调拨 {quantity} {item.unit}从{warehouse_from}到{warehouse_to}", "out_tx_id": out_tx.id, "in_tx_id": in_tx.id}


@app.post("/api/inventory-transactions/count")
def create_inventory_count(company_id: int = Query(...), db: Session = Depends(get_db),
    item_code: str = Form(...), transaction_date: str = Form(...),
    actual_quantity: float = Form(...), unit_price: float = Form(0.0),
    warehouse: str = Form(""), reference_no: str = Form(""),
    operator: str = Form("管理员"), remark: str = Form("")):
    """盘点（自动生成盘盈/盘亏）"""
    item = db.query(InventoryItem).filter(InventoryItem.company_id == company_id, InventoryItem.code == item_code).first()
    if not item:
        raise HTTPException(400, detail=f"商品 {item_code} 不存在")
    diff = actual_quantity - item.current_stock
    if abs(diff) < 0.001:
        return {"message": "库存账实相符，无需调整", "current": item.current_stock, "actual": actual_quantity}
    
    tx_date = datetime.strptime(transaction_date, "%Y-%m-%d").date() if transaction_date else datetime.now().date()
    trans_type = "盘盈" if diff > 0 else "盘亏"
    total = round(abs(diff) * unit_price, 2)
    
    trans = InventoryTransaction(
        company_id=company_id, item_code=item_code, transaction_date=tx_date,
        trans_type=trans_type, quantity=abs(diff), unit_price=unit_price, total_amount=total,
        warehouse=warehouse, reference_no=reference_no, operator=operator,
        remark=f"盘点调整：账存{item.current_stock} 实盘{actual_quantity} 差异{diff} {remark}"
    )
    db.add(trans)
    
    # 更新库存
    if diff > 0:
        item.current_stock += diff
    else:
        item.current_stock += diff  # diff is negative
    item.updated_at = datetime.now()
    db.commit()
    db.refresh(trans)
    return {
        "id": trans.id, "trans_type": trans_type,
        "difference": round(diff, 2), "total_amount": total,
        "message": f"{trans_type}已记录，差异: {diff:+.2f} {item.unit}，当前库存: {item.current_stock}"
    }


@app.get("/api/inventory-balances")
def list_inventory_balances(company_id: int = Query(...), period: str = Query(...),
    db: Session = Depends(get_db)):
    """库存余额表（按期核算）"""
    items = db.query(InventoryItem).filter(InventoryItem.company_id == company_id, InventoryItem.is_active == True).all()
    
    # 期初：取上月余额表或从0开始
    prev_period = _prev_period(period)
    prev_map = {}
    if prev_period:
        prev_balances = db.query(InventoryBalance).filter(
            InventoryBalance.company_id == company_id, InventoryBalance.period == prev_period
        ).all()
        prev_map = {b.item_code: b.end_quantity for b in prev_balances}
    
    # 本期收发汇总
    tx_start = datetime.strptime(period + "-01", "%Y-%m-%d").date()
    if len(period) == 7:
        y, m = int(period[:4]), int(period[5:7])
        if m == 12:
            tx_end = datetime(y + 1, 1, 1).date()
        else:
            tx_end = datetime(y, m + 1, 1).date()
    else:
        tx_end = datetime.now().date()
    
    txs = db.query(InventoryTransaction).filter(
        InventoryTransaction.company_id == company_id,
        InventoryTransaction.transaction_date >= tx_start,
        InventoryTransaction.transaction_date < tx_end
    ).all()
    
    # 按商品汇总
    from collections import defaultdict
    in_map = defaultdict(float)
    out_map = defaultdict(float)
    for t in txs:
        if t.trans_type in ("入库", "调拨入", "盘盈"):
            in_map[t.item_code] += t.quantity
        elif t.trans_type in ("出库", "调拨出", "盘亏"):
            out_map[t.item_code] += t.quantity
    
    results = []
    for item in items:
        begin = prev_map.get(item.code, 0.0)
        in_qty = round(in_map.get(item.code, 0), 2)
        out_qty = round(out_map.get(item.code, 0), 2)
        end_qty = round(begin + in_qty - out_qty, 2)
        end_amount = round(end_qty * item.cost_price, 2)
        results.append({
            "item_code": item.code, "item_name": item.name,
            "spec": item.spec, "unit": item.unit, "warehouse": item.warehouse,
            "begin_quantity": begin, "in_quantity": in_qty,
            "out_quantity": out_qty, "end_quantity": end_qty,
            "cost_price": item.cost_price, "end_amount": end_amount
        })
    
    return {"period": period, "items": results}


def _prev_period(period: str) -> Optional[str]:
    """计算上月期间"""
    if len(period) != 7:
        return None
    y, m = int(period[:4]), int(period[5:7])
    if m == 1:
        return f"{y-1}-12"
    return f"{y}-{m-1:02d}"


@app.post("/api/inventory-items/batch-delete")
def batch_delete_inventory_items(ids: List[int], company_id: int = Query(...), db: Session = Depends(get_db)):
    deleted = 0
    for item_id in ids:
        item = db.query(InventoryItem).filter(InventoryItem.company_id == company_id, InventoryItem.id == item_id).first()
        if item:
            item.is_active = False
            deleted += 1
    db.commit()
    return {"deleted": deleted, "message": f"停用 {deleted} 项"}


# ==================== 合同管理 ====================

class ContractCreate(BaseModel):
    contract_no: str
    name: str
    contract_type: str
    party_a: Optional[str] = None
    party_b: Optional[str] = None
    amount: float = 0.0
    signing_date: Optional[date] = None
    effective_date: Optional[date] = None
    expiry_date: Optional[date] = None
    status: str = "起草中"
    responsible_person: Optional[str] = None
    dept_code: Optional[str] = None
    content_summary: Optional[str] = None
    remark: Optional[str] = None


class ContractUpdate(BaseModel):
    name: Optional[str] = None
    contract_type: Optional[str] = None
    party_a: Optional[str] = None
    party_b: Optional[str] = None
    amount: Optional[float] = None
    signing_date: Optional[date] = None
    effective_date: Optional[date] = None
    expiry_date: Optional[date] = None
    status: Optional[str] = None
    responsible_person: Optional[str] = None
    content_summary: Optional[str] = None
    remark: Optional[str] = None


class ContractPaymentCreate(BaseModel):
    payment_no: int = 1
    payment_type: str
    amount: float
    due_date: Optional[date] = None
    remark: Optional[str] = None


@app.get("/api/contracts")
def list_contracts(
    company_id: int = Query(...),
    contract_type: Optional[str] = None,
    status: Optional[str] = None,
    keyword: Optional[str] = None,
    db: Session = Depends(get_db)
):
    q = db.query(Contract).filter(Contract.company_id == company_id)
    if contract_type:
        q = q.filter(Contract.contract_type == contract_type)
    if status:
        q = q.filter(Contract.status == status)
    if keyword:
        q = q.filter(or_(Contract.contract_no.contains(keyword), Contract.name.contains(keyword)))
    contracts = q.order_by(Contract.signing_date.desc()).all()
    return [{
        "id": c.id, "contract_no": c.contract_no, "name": c.name,
        "contract_type": c.contract_type, "party_a": c.party_a, "party_b": c.party_b,
        "amount": c.amount, "signing_date": str(c.signing_date) if c.signing_date else "",
        "effective_date": str(c.effective_date) if c.effective_date else "",
        "expiry_date": str(c.expiry_date) if c.expiry_date else "",
        "status": c.status, "responsible_person": c.responsible_person,
        "dept_code": c.dept_code, "content_summary": c.content_summary, "remark": c.remark
    } for c in contracts]


@app.post("/api/contracts")
def create_contract(data: ContractCreate, company_id: int = Query(...), db: Session = Depends(get_db)):
    contract = Contract(company_id=company_id, **data.model_dump())
    db.add(contract)
    db.commit()
    db.refresh(contract)
    return {"id": contract.id, "contract_no": contract.contract_no, "message": "合同创建成功"}


@app.put("/api/contracts/{contract_id}")
def update_contract(contract_id: int, data: ContractUpdate, company_id: int = Query(...), db: Session = Depends(get_db)):
    c = db.query(Contract).filter(Contract.company_id == company_id, Contract.id == contract_id).first()
    if not c:
        raise HTTPException(404, detail="合同不存在")
    for k, v in data.model_dump(exclude_unset=True).items():
        setattr(c, k, v)
    c.updated_at = datetime.now()
    db.commit()
    return {"message": "更新成功"}


@app.delete("/api/contracts/{contract_id}")
def delete_contract(contract_id: int, company_id: int = Query(...), db: Session = Depends(get_db)):
    c = db.query(Contract).filter(Contract.company_id == company_id, Contract.id == contract_id).first()
    if not c:
        raise HTTPException(404, detail="合同不存在")
    if c.status in ("履行中", "已签署"):
        raise HTTPException(400, detail=f"合同状态为'{c.status}'，不能删除")
    db.delete(c)
    db.commit()
    return {"message": "删除成功"}


@app.get("/api/contracts/{contract_id}/payments")
def get_contract_payments(contract_id: int, company_id: int = Query(...), db: Session = Depends(get_db)):
    payments = db.query(ContractPayment).filter(
        ContractPayment.company_id == company_id,
        ContractPayment.contract_id == contract_id
    ).order_by(ContractPayment.payment_no).all()
    return [{
        "id": p.id, "payment_no": p.payment_no, "payment_type": p.payment_type,
        "amount": p.amount, "due_date": str(p.due_date) if p.due_date else "",
        "paid_date": str(p.paid_date) if p.paid_date else "",
        "paid_amount": p.paid_amount, "status": p.status, "remark": p.remark
    } for p in payments]


@app.post("/api/contracts/{contract_id}/payments")
def add_contract_payment(contract_id: int, data: ContractPaymentCreate, company_id: int = Query(...), db: Session = Depends(get_db)):
    c = db.query(Contract).filter(Contract.company_id == company_id, Contract.id == contract_id).first()
    if not c:
        raise HTTPException(404, detail="合同不存在")
    payment = ContractPayment(
        company_id=company_id, contract_id=contract_id,
        payment_no=data.payment_no, payment_type=data.payment_type,
        amount=data.amount, due_date=data.due_date, remark=data.remark
    )
    db.add(payment)
    db.commit()
    db.refresh(payment)
    return {"id": payment.id, "message": "付款计划添加成功"}


# ==================== 付款管理 ====================

class PaymentCreate(BaseModel):
    payment_type: str = "外部单位"
    scenario: Optional[str] = None
    payment_no: str
    payment_date: date
    employee_id: Optional[int] = None
    employee_name: Optional[str] = None
    supplier_id: Optional[int] = None
    supplier_name: Optional[str] = None
    contract_id: Optional[int] = None
    contract_no: Optional[str] = None
    amount: float
    payment_method: str = "银行转账"
    payee: Optional[str] = None
    payee_account: Optional[str] = None
    payee_bank: Optional[str] = None
    department: Optional[str] = None
    purpose: Optional[str] = None
    remark: Optional[str] = None


class PaymentUpdate(BaseModel):
    payment_type: Optional[str] = None
    scenario: Optional[str] = None
    employee_id: Optional[int] = None
    employee_name: Optional[str] = None
    supplier_id: Optional[int] = None
    supplier_name: Optional[str] = None
    contract_id: Optional[int] = None
    contract_no: Optional[str] = None
    amount: Optional[float] = None
    payment_method: Optional[str] = None
    payee: Optional[str] = None
    payee_account: Optional[str] = None
    payee_bank: Optional[str] = None
    status: Optional[str] = None
    approved_by: Optional[str] = None
    department: Optional[str] = None
    purpose: Optional[str] = None
    remark: Optional[str] = None


@app.get("/api/payments")
def list_payments(
    company_id: int = Query(...),
    payment_type: Optional[str] = None,
    status: Optional[str] = None,
    supplier_id: Optional[int] = None,
    keyword: Optional[str] = None,
    db: Session = Depends(get_db)
):
    q = db.query(Payment).filter(Payment.company_id == company_id)
    if payment_type:
        q = q.filter(Payment.payment_type == payment_type)
    if status:
        q = q.filter(Payment.status == status)
    if supplier_id:
        q = q.filter(Payment.supplier_id == supplier_id)
    if keyword:
        q = q.filter(or_(
            Payment.payment_no.contains(keyword),
            Payment.supplier_name.contains(keyword),
            Payment.employee_name.contains(keyword),
            Payment.payee.contains(keyword),
            Payment.purpose.contains(keyword)
        ))
    payments = q.order_by(Payment.payment_date.desc()).all()
    return [{
        "id": p.id, "payment_type": p.payment_type, "scenario": p.scenario or "",
        "payment_no": p.payment_no,
        "payment_date": str(p.payment_date) if p.payment_date else "",
        "employee_id": p.employee_id, "employee_name": p.employee_name or "",
        "supplier_id": p.supplier_id, "supplier_name": p.supplier_name or "",
        "contract_id": p.contract_id, "contract_no": p.contract_no or "",
        "amount": p.amount, "payment_method": p.payment_method,
        "payee": p.payee or "", "payee_account": p.payee_account or "",
        "payee_bank": p.payee_bank or "", "status": p.status,
        "approved_by": p.approved_by or "",
        "approved_at": str(p.approved_at) if p.approved_at else "",
        "paid_at": str(p.paid_at) if p.paid_at else "",
        "department": p.department or "", "purpose": p.purpose or "",
        "remark": p.remark or "",
        "created_at": str(p.created_at) if p.created_at else ""
    } for p in payments]


@app.post("/api/payments")
def create_payment(data: PaymentCreate, company_id: int = Query(...), db: Session = Depends(get_db)):
    payment = Payment(company_id=company_id, **data.model_dump())
    db.add(payment)
    db.commit()
    db.refresh(payment)
    return {"id": payment.id, "payment_no": payment.payment_no, "message": "付款单创建成功"}


@app.get("/api/payments/stats")
def payment_stats(company_id: int = Query(...), db: Session = Depends(get_db)):
    """付款统计"""
    base = db.query(Payment).filter(Payment.company_id == company_id)
    total_count = base.count()
    total_amount = base.with_entities(func.sum(Payment.amount)).scalar() or 0
    
    # 按类型统计
    internal_base = base.filter(Payment.payment_type == "内部人员")
    internal_count = internal_base.count()
    internal_amount = internal_base.with_entities(func.sum(Payment.amount)).scalar() or 0
    
    external_base = base.filter(Payment.payment_type == "外部单位")
    external_count = external_base.count()
    external_amount = external_base.with_entities(func.sum(Payment.amount)).scalar() or 0
    
    pending_count = base.filter(Payment.status == "待审批").count()
    pending_amount = base.filter(Payment.status == "待审批").with_entities(func.sum(Payment.amount)).scalar() or 0
    approved_count = base.filter(Payment.status == "已审批").count()
    paid_count = base.filter(Payment.status == "已付款").count()
    paid_amount = base.filter(Payment.status == "已付款").with_entities(func.sum(Payment.amount)).scalar() or 0
    return {
        "total_count": total_count, "total_amount": total_amount,
        "internal_count": internal_count, "internal_amount": internal_amount,
        "external_count": external_count, "external_amount": external_amount,
        "pending_count": pending_count, "pending_amount": pending_amount,
        "approved_count": approved_count, "paid_count": paid_count,
        "paid_amount": paid_amount
    }


@app.put("/api/payments/{payment_id}")
def update_payment(payment_id: int, data: PaymentUpdate, company_id: int = Query(...), db: Session = Depends(get_db)):
    p = db.query(Payment).filter(Payment.company_id == company_id, Payment.id == payment_id).first()
    if not p:
        raise HTTPException(404, detail="付款单不存在")
    for k, v in data.model_dump(exclude_unset=True).items():
        setattr(p, k, v)
    p.updated_at = datetime.now()
    # 如果状态改为"已付款"，记录付款时间
    if data.status == "已付款":
        p.paid_at = datetime.now()
    db.commit()
    return {"message": "更新成功"}


@app.delete("/api/payments/{payment_id}")
def delete_payment(payment_id: int, company_id: int = Query(...), db: Session = Depends(get_db)):
    p = db.query(Payment).filter(Payment.company_id == company_id, Payment.id == payment_id).first()
    if not p:
        raise HTTPException(404, detail="付款单不存在")
    if p.status in ("已审批", "已付款"):
        raise HTTPException(400, detail=f"付款单状态为'{p.status}'，不能删除")
    db.delete(p)
    db.commit()
    return {"message": "删除成功"}


# ==================== 开具发票（销售发票）====================

class SalesInvoiceCreate(BaseModel):
    invoice_code: Optional[str] = None
    invoice_no: Optional[str] = None
    digital_invoice_no: Optional[str] = None
    seller_tax_no: Optional[str] = None
    seller_name: Optional[str] = None
    buyer_tax_no: Optional[str] = None
    buyer_name: Optional[str] = None
    invoice_date: Optional[date] = None
    tax_category_code: Optional[str] = None
    specific_business_type: Optional[str] = None
    goods_name: Optional[str] = None
    spec: Optional[str] = None
    unit: Optional[str] = None
    quantity: Optional[float] = 0
    unit_price: Optional[float] = 0
    amount: float = 0.0
    tax_rate: Optional[float] = 0.0
    tax_amount: Optional[float] = 0.0
    total_amount: Optional[float] = 0.0
    invoice_source: Optional[str] = None
    invoice_category: str = "增值税专用发票"
    status: str = "正常"
    is_positive: Optional[bool] = True
    invoice_risk_level: Optional[str] = None
    issuer: Optional[str] = None
    remark: Optional[str] = None


class SalesInvoiceUpdate(BaseModel):
    invoice_code: Optional[str] = None
    digital_invoice_no: Optional[str] = None
    seller_tax_no: Optional[str] = None
    seller_name: Optional[str] = None
    buyer_tax_no: Optional[str] = None
    buyer_name: Optional[str] = None
    invoice_date: Optional[date] = None
    tax_category_code: Optional[str] = None
    specific_business_type: Optional[str] = None
    goods_name: Optional[str] = None
    spec: Optional[str] = None
    unit: Optional[str] = None
    quantity: Optional[float] = None
    unit_price: Optional[float] = None
    amount: Optional[float] = None
    tax_rate: Optional[float] = None
    tax_amount: Optional[float] = None
    total_amount: Optional[float] = None
    invoice_source: Optional[str] = None
    invoice_category: Optional[str] = None
    status: Optional[str] = None
    is_positive: Optional[bool] = None
    invoice_risk_level: Optional[str] = None
    issuer: Optional[str] = None
    remark: Optional[str] = None


@app.get("/api/sales-invoices")
def list_sales_invoices(
    company_id: int = Query(...),
    invoice_category: Optional[str] = None,
    status: Optional[str] = None,
    keyword: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=500),
    db: Session = Depends(get_db)
):
    q = db.query(SalesInvoice).filter(SalesInvoice.company_id == company_id)
    if invoice_category:
        q = q.filter(SalesInvoice.invoice_category == invoice_category)
    if status:
        q = q.filter(SalesInvoice.status == status)
    if date_from:
        q = q.filter(SalesInvoice.invoice_date >= date_from)
    if date_to:
        q = q.filter(SalesInvoice.invoice_date <= date_to)
    if keyword:
        q = q.filter(or_(
            SalesInvoice.invoice_no.contains(keyword),
            SalesInvoice.invoice_code.contains(keyword),
            SalesInvoice.digital_invoice_no.contains(keyword),
            SalesInvoice.buyer_name.contains(keyword),
            SalesInvoice.goods_name.contains(keyword)
        ))
    invoices = q.order_by(SalesInvoice.invoice_date.desc()).offset(skip).limit(limit).all()
    # 构建凭证号映射（销项发票 → 序时账，通过摘要+借方金额+科目1122判重）
    voucher_map = {}
    for inv in invoices:
        buyer = inv.buyer_name or "客户"
        goods = inv.goods_name or ""
        summary = f"销售{goods or '货物'}给{buyer}"
        je = db.query(JournalEntry).filter(
            JournalEntry.company_id == company_id,
            JournalEntry.summary == summary,
            JournalEntry.debit_amount == inv.total_amount,
            JournalEntry.account_code == "1122"
        ).first()
        if je:
            voucher_map[inv.id] = f"{je.voucher_word}-{je.voucher_no}"
    return [{
        "id": inv.id,
        "invoice_code": inv.invoice_code or "",
        "invoice_no": inv.invoice_no,
        "digital_invoice_no": inv.digital_invoice_no or "",
        "seller_tax_no": inv.seller_tax_no or "",
        "seller_name": inv.seller_name or "",
        "buyer_tax_no": inv.buyer_tax_no or "",
        "buyer_name": inv.buyer_name or "",
        "invoice_date": str(inv.invoice_date) if inv.invoice_date else "",
        "tax_category_code": inv.tax_category_code or "",
        "specific_business_type": inv.specific_business_type or "",
        "goods_name": inv.goods_name or "",
        "spec": inv.spec or "",
        "unit": inv.unit or "",
        "quantity": inv.quantity or 0,
        "unit_price": inv.unit_price or 0,
        "amount": inv.amount or 0,
        "tax_rate": inv.tax_rate or 0,
        "tax_amount": inv.tax_amount or 0,
        "total_amount": inv.total_amount or 0,
        "invoice_source": inv.invoice_source or "",
        "invoice_category": inv.invoice_category or "增值税专用发票",
        "status": inv.status,
        "is_positive": inv.is_positive if inv.is_positive is not None else True,
        "invoice_risk_level": inv.invoice_risk_level or "",
        "issuer": inv.issuer or "",
        "remark": inv.remark or "",
        "journal_voucher_no": voucher_map.get(inv.id, ""),
        "created_at": str(inv.created_at) if inv.created_at else ""
    } for inv in invoices]


@app.post("/api/sales-invoices")
def create_sales_invoice(data: SalesInvoiceCreate, company_id: int = Query(...), db: Session = Depends(get_db)):
    # ── 按票号唯一去重 ──
    digital_no = (data.digital_invoice_no or "").strip()
    inv_code = (data.invoice_code or "").strip()
    inv_no = (data.invoice_no or "").strip()

    if digital_no:
        existing = db.query(SalesInvoice).filter(
            SalesInvoice.company_id == company_id,
            SalesInvoice.digital_invoice_no == digital_no
        ).first()
        if existing:
            raise HTTPException(400, detail=f"数电发票 {digital_no} 已存在，请勿重复录入")
    elif inv_no:
        q = db.query(SalesInvoice).filter(
            SalesInvoice.company_id == company_id,
            SalesInvoice.invoice_no == inv_no
        )
        if inv_code:
            q = q.filter(SalesInvoice.invoice_code == inv_code)
        existing = q.first()
        if existing:
            raise HTTPException(400, detail=f"发票 {inv_code}+{inv_no} 已存在，请勿重复录入")

    # 全行指纹去重
    fp_values = (
        str(company_id),
        str(data.invoice_no or ""),
        str(data.invoice_code or ""),
        str(data.digital_invoice_no or ""),
        str(data.seller_tax_no or ""),
        str(data.seller_name or ""),
        str(data.buyer_tax_no or ""),
        str(data.buyer_name or ""),
        str(data.invoice_date) if data.invoice_date else "",
        str(data.tax_category_code or ""),
        str(data.specific_business_type or ""),
        str(data.goods_name or ""),
        str(data.spec or ""),
        str(data.unit or ""),
        str(data.quantity),
        str(data.unit_price),
        str(data.amount),
        str(data.tax_rate),
        str(data.tax_amount),
        str(data.total_amount),
        str(data.invoice_source or ""),
        str(data.invoice_category or ""),
        str(data.status or ""),
        str(data.is_positive),
        str(data.invoice_risk_level or ""),
        str(data.issuer or ""),
        str(data.remark or ""),
    )
    fp_raw = "|".join(fp_values)
    fp = hashlib.sha256(fp_raw.encode("utf-8")).hexdigest()
    existing = db.query(SalesInvoice).filter(
        SalesInvoice.company_id == company_id,
        SalesInvoice._fingerprint == fp
    ).first()
    if existing:
        raise HTTPException(400, detail="该发票数据已存在（全行比对重复），请勿重复录入")
    inv = SalesInvoice(
        company_id=company_id,
        _fingerprint=fp,
        **data.model_dump()
    )
    db.add(inv)
    db.commit()
    db.refresh(inv)
    return {"id": inv.id, "invoice_no": inv.invoice_no, "message": "开具发票创建成功"}


@app.get("/api/sales-invoices/stats")
def sales_invoice_stats(company_id: int = Query(...), status: str = Query(None), db: Session = Depends(get_db)):
    base = db.query(SalesInvoice).filter(SalesInvoice.company_id == company_id)
    if status:
        base = base.filter(SalesInvoice.status.like(f"%{status}%"))
    total_count = base.count()
    total_amt = base.with_entities(func.sum(func.coalesce(SalesInvoice.amount, 0))).scalar() or 0
    total_amount = base.with_entities(func.sum(func.coalesce(SalesInvoice.total_amount, 0))).scalar() or 0
    total_tax = base.with_entities(func.sum(func.coalesce(SalesInvoice.tax_amount, 0))).scalar() or 0
    normal_count = base.filter(SalesInvoice.status == "正常").count() if not status else 0
    void_count = base.filter(SalesInvoice.status.like("%作废%")).count() if not status else 0
    red_count = base.filter(SalesInvoice.status.like("%红冲%")).count() if not status else 0
    return {
        "total_count": total_count, "total_amt": round(total_amt, 2),
        "total_amount": round(total_amount, 2),
        "total_tax": round(total_tax, 2),
        "normal_count": normal_count, "void_count": void_count,
        "red_count": red_count
    }


@app.get("/api/sales-invoices/{invoice_id}")
def get_sales_invoice(invoice_id: int, company_id: int = Query(...), db: Session = Depends(get_db)):
    inv = db.query(SalesInvoice).filter(SalesInvoice.company_id == company_id, SalesInvoice.id == invoice_id).first()
    if not inv:
        raise HTTPException(404, detail="发票不存在")
    return {
        "id": inv.id,
        "invoice_code": inv.invoice_code or "",
        "invoice_no": inv.invoice_no,
        "digital_invoice_no": inv.digital_invoice_no or "",
        "seller_tax_no": inv.seller_tax_no or "",
        "seller_name": inv.seller_name or "",
        "buyer_tax_no": inv.buyer_tax_no or "",
        "buyer_name": inv.buyer_name or "",
        "invoice_date": str(inv.invoice_date) if inv.invoice_date else "",
        "tax_category_code": inv.tax_category_code or "",
        "specific_business_type": inv.specific_business_type or "",
        "goods_name": inv.goods_name or "",
        "spec": inv.spec or "",
        "unit": inv.unit or "",
        "quantity": inv.quantity or 0,
        "unit_price": inv.unit_price or 0,
        "amount": inv.amount or 0,
        "tax_rate": inv.tax_rate or 0,
        "tax_amount": inv.tax_amount or 0,
        "total_amount": inv.total_amount or 0,
        "invoice_source": inv.invoice_source or "",
        "invoice_category": inv.invoice_category or "增值税专用发票",
        "status": inv.status,
        "is_positive": inv.is_positive if inv.is_positive is not None else True,
        "invoice_risk_level": inv.invoice_risk_level or "",
        "issuer": inv.issuer or "",
        "remark": inv.remark or "",
        "created_at": str(inv.created_at) if inv.created_at else "",
        "updated_at": str(inv.updated_at) if inv.updated_at else ""
    }


@app.put("/api/sales-invoices/{invoice_id}")
def update_sales_invoice(invoice_id: int, data: SalesInvoiceUpdate, company_id: int = Query(...), db: Session = Depends(get_db)):
    inv = db.query(SalesInvoice).filter(SalesInvoice.company_id == company_id, SalesInvoice.id == invoice_id).first()
    if not inv:
        raise HTTPException(404, detail="发票不存在")
    for k, v in data.model_dump(exclude_unset=True).items():
        setattr(inv, k, v)
    inv.updated_at = datetime.now()
    # 重新计算全行指纹
    fp_values = (
        str(company_id),
        str(inv.invoice_no or ""),
        str(inv.invoice_code or ""),
        str(inv.digital_invoice_no or ""),
        str(inv.seller_tax_no or ""),
        str(inv.seller_name or ""),
        str(inv.buyer_tax_no or ""),
        str(inv.buyer_name or ""),
        str(inv.invoice_date) if inv.invoice_date else "",
        str(inv.tax_category_code or ""),
        str(inv.specific_business_type or ""),
        str(inv.goods_name or ""),
        str(inv.spec or ""),
        str(inv.unit or ""),
        str(inv.quantity or 0),
        str(inv.unit_price or 0),
        str(inv.amount or 0),
        str(inv.tax_rate or 0),
        str(inv.tax_amount or 0),
        str(inv.total_amount or 0),
        str(inv.invoice_source or ""),
        str(inv.invoice_category or ""),
        str(inv.status or ""),
        str(inv.is_positive if inv.is_positive is not None else True),
        str(inv.invoice_risk_level or ""),
        str(inv.issuer or ""),
        str(inv.remark or ""),
    )
    fp_raw = "|".join(fp_values)
    fp = hashlib.sha256(fp_raw.encode("utf-8")).hexdigest()
    inv._fingerprint = fp
    db.commit()
    db.refresh(inv)
    return {"message": "更新成功"}



@app.delete("/api/sales-invoices/{invoice_id}")
def delete_sales_invoice(invoice_id: int, company_id: int = Query(...), db: Session = Depends(get_db)):
    inv = db.query(SalesInvoice).filter(SalesInvoice.company_id == company_id, SalesInvoice.id == invoice_id).first()
    if not inv:
        raise HTTPException(404, detail="发票不存在")
    db.delete(inv)
    db.commit()
    return {"message": "删除成功"}


@app.post("/api/sales-invoices/batch-delete")
def batch_delete_sales_invoices(ids: list[int], company_id: int = Query(...), db: Session = Depends(get_db)):
    deleted = db.query(SalesInvoice).filter(
        SalesInvoice.company_id == company_id,
        SalesInvoice.id.in_(ids)
    ).delete(synchronize_session=False)
    db.commit()
    return {"message": f"成功删除 {deleted} 条记录", "deleted": deleted}




# ==================== 取得发票（采购发票）====================

class PurchaseInvoiceCreate(BaseModel):
    invoice_code: Optional[str] = None
    invoice_no: Optional[str] = None
    digital_invoice_no: Optional[str] = None
    seller_tax_no: Optional[str] = None
    seller_name: Optional[str] = None
    buyer_tax_no: Optional[str] = None
    buyer_name: Optional[str] = None
    invoice_date: Optional[date] = None
    tax_category_code: Optional[str] = None
    specific_business_type: Optional[str] = None
    goods_name: Optional[str] = None
    spec: Optional[str] = None
    unit: Optional[str] = None
    quantity: Optional[float] = 0
    unit_price: Optional[float] = 0
    amount: float = 0.0
    tax_rate: Optional[float] = 0.0
    tax_amount: Optional[float] = 0.0
    total_amount: Optional[float] = 0.0
    invoice_source: Optional[str] = None
    invoice_category: str = "增值税专用发票"
    status: str = "正常"
    is_positive: Optional[bool] = True
    invoice_risk_level: Optional[str] = None
    issuer: Optional[str] = None
    remark: Optional[str] = None


class PurchaseInvoiceUpdate(BaseModel):
    invoice_code: Optional[str] = None
    digital_invoice_no: Optional[str] = None
    seller_tax_no: Optional[str] = None
    seller_name: Optional[str] = None
    buyer_tax_no: Optional[str] = None
    buyer_name: Optional[str] = None
    invoice_date: Optional[date] = None
    tax_category_code: Optional[str] = None
    specific_business_type: Optional[str] = None
    goods_name: Optional[str] = None
    spec: Optional[str] = None
    unit: Optional[str] = None
    quantity: Optional[float] = None
    unit_price: Optional[float] = None
    amount: Optional[float] = None
    tax_rate: Optional[float] = None
    tax_amount: Optional[float] = None
    total_amount: Optional[float] = None
    invoice_source: Optional[str] = None
    invoice_category: Optional[str] = None
    status: Optional[str] = None
    is_positive: Optional[bool] = None
    invoice_risk_level: Optional[str] = None
    issuer: Optional[str] = None
    remark: Optional[str] = None
@app.get("/api/purchase-invoices")
def list_purchase_invoices(
    company_id: int = Query(...),
    invoice_category: Optional[str] = None,
    status: Optional[str] = None,
    keyword: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=500),
    db: Session = Depends(get_db)
):
    q = db.query(PurchaseInvoice).filter(PurchaseInvoice.company_id == company_id)
    if invoice_category:
        q = q.filter(PurchaseInvoice.invoice_category == invoice_category)
    if status:
        q = q.filter(PurchaseInvoice.status == status)
    if date_from:
        q = q.filter(PurchaseInvoice.invoice_date >= date_from)
    if date_to:
        q = q.filter(PurchaseInvoice.invoice_date <= date_to)
    if keyword:
        q = q.filter(or_(
            PurchaseInvoice.invoice_no.contains(keyword),
            PurchaseInvoice.invoice_code.contains(keyword),
            PurchaseInvoice.digital_invoice_no.contains(keyword),
            PurchaseInvoice.seller_name.contains(keyword),
            PurchaseInvoice.goods_name.contains(keyword)
        ))
    invoices = q.order_by(PurchaseInvoice.invoice_date.desc()).offset(skip).limit(limit).all()
    # 构建凭证号映射（取得发票 → 序时账，通过 ref_id 精确匹配）
    voucher_map = {}
    if invoices:
        inv_ids = [inv.id for inv in invoices]
        entries = db.query(JournalEntry).filter(
            JournalEntry.company_id == company_id,
            JournalEntry.source == "未记账发票",
            JournalEntry.ref_id.in_(inv_ids)
        ).all()
        for je in entries:
            if je.ref_id not in voucher_map:
                voucher_map[je.ref_id] = f"{je.voucher_word}-{je.voucher_no}"
    return [{
        "id": inv.id,
        "invoice_code": inv.invoice_code or "",
        "invoice_no": inv.invoice_no,
        "digital_invoice_no": inv.digital_invoice_no or "",
        "seller_tax_no": inv.seller_tax_no or "",
        "seller_name": inv.seller_name or "",
        "buyer_tax_no": inv.buyer_tax_no or "",
        "buyer_name": inv.buyer_name or "",
        "invoice_date": str(inv.invoice_date) if inv.invoice_date else "",
        "tax_category_code": inv.tax_category_code or "",
        "specific_business_type": inv.specific_business_type or "",
        "goods_name": inv.goods_name or "",
        "spec": inv.spec or "",
        "unit": inv.unit or "",
        "quantity": inv.quantity or 0,
        "unit_price": inv.unit_price or 0,
        "amount": inv.amount or 0,
        "tax_rate": inv.tax_rate or 0,
        "tax_amount": inv.tax_amount or 0,
        "total_amount": inv.total_amount or 0,
        "invoice_source": inv.invoice_source or "",
        "invoice_category": inv.invoice_category or "增值税专用发票",
        "status": inv.status,
        "is_positive": inv.is_positive if inv.is_positive is not None else True,
        "invoice_risk_level": inv.invoice_risk_level or "",
        "issuer": inv.issuer or "",
        "remark": inv.remark or "",
        "journal_voucher_no": voucher_map.get(inv.id, ""),
        "skip_accounting": bool(inv.skip_accounting) if inv.skip_accounting is not None else False,
        "created_at": str(inv.created_at) if inv.created_at else ""
    } for inv in invoices]


@app.post("/api/purchase-invoices")
def create_purchase_invoice(data: PurchaseInvoiceCreate, company_id: int = Query(...), db: Session = Depends(get_db)):
    # ── 取得发票全指纹去重 ──
    import hashlib
    fp_values = (
        str(company_id), str(data.invoice_no or ""), str(data.invoice_code or ""),
        str(data.digital_invoice_no or ""),
        str(data.seller_tax_no or ""), str(data.seller_name or ""),
        str(data.buyer_tax_no or ""), str(data.buyer_name or ""),
        str(data.invoice_date) if data.invoice_date else "",
        str(data.tax_category_code or ""), str(data.specific_business_type or ""),
        str(data.goods_name or ""), str(data.spec or ""),
        str(data.unit or ""), str(data.quantity or 0), str(data.unit_price or 0),
        str(data.amount or 0), str(data.tax_rate or 0), str(data.tax_amount or 0), str(data.total_amount or 0),
        str(data.invoice_source or ""),
        str(data.invoice_category or "增值税专用发票"),
        str(data.status or "正常"),
        str("是" if data.is_positive else "否"),
        str(data.invoice_risk_level or ""),
        str(data.issuer or ""),
        str(data.remark or ""),
    )
    fp_raw = "|".join(fp_values)
    pi_fp = hashlib.sha256(fp_raw.encode("utf-8")).hexdigest()
    existing = db.query(PurchaseInvoice).filter(
        PurchaseInvoice.company_id == company_id,
        PurchaseInvoice._fingerprint == pi_fp
    ).first()
    if existing:
        raise HTTPException(400, detail="全指纹重复，发票已存在，请勿重复录入")
    inv = PurchaseInvoice(company_id=company_id, _fingerprint=pi_fp, **data.model_dump())
    db.add(inv)
    db.flush()
    db.commit()
    db.refresh(inv)
    return {"id": inv.id, "invoice_no": inv.invoice_no, "message": "取得发票创建成功"}


@app.get("/api/purchase-invoices/stats")
def purchase_invoice_stats(company_id: int = Query(...), tab: str = Query("all"), db: Session = Depends(get_db)):
    base = db.query(PurchaseInvoice).filter(PurchaseInvoice.company_id == company_id)
    # 按票种筛选
    if tab == "zpt":
        base = base.filter(PurchaseInvoice.invoice_category.contains("专用发票"))
    elif tab == "ppt":
        base = base.filter(PurchaseInvoice.invoice_category.contains("普通发票"))
    elif tab == "tlp":
        base = base.filter(PurchaseInvoice.invoice_category.contains("铁路"))
    total_count = base.count()
    total_amt = base.with_entities(func.sum(func.coalesce(PurchaseInvoice.amount, 0))).scalar() or 0
    total_amount = base.with_entities(func.sum(func.coalesce(PurchaseInvoice.total_amount, 0))).scalar() or 0
    total_raw_tax = base.with_entities(func.sum(func.coalesce(PurchaseInvoice.tax_amount, 0))).scalar() or 0
    # 可抵扣税额：专票/铁路票 = 税额合计，普票 = 0
    if tab == "ppt":
        total_tax = 0.0
    else:
        deduct_q = db.query(PurchaseInvoice).filter(
            PurchaseInvoice.company_id == company_id,
            PurchaseInvoice.tax_amount != 0,  # 含红字发票（负税额应扣减）
        )
        if tab == "zpt":
            deduct_q = deduct_q.filter(PurchaseInvoice.invoice_category.contains("专用发票"))
        elif tab == "tlp":
            deduct_q = deduct_q.filter(PurchaseInvoice.invoice_category.contains("铁路"))
        else:  # all：专票 + 铁路票
            deduct_q = deduct_q.filter(
                or_(PurchaseInvoice.invoice_category.contains("专用发票"),
                     PurchaseInvoice.invoice_category.contains("铁路")))
        total_tax = round(deduct_q.with_entities(func.sum(func.coalesce(PurchaseInvoice.tax_amount, 0))).scalar() or 0, 2)
    normal_count = base.filter(PurchaseInvoice.status == "正常").count()
    void_count = base.filter(PurchaseInvoice.status.like("%作废%")).count()
    red_count = base.filter(PurchaseInvoice.status.like("%红冲%")).count()
    return {
        "total_count": total_count, "total_amt": round(total_amt, 2),
        "total_amount": round(total_amount, 2),
        "total_raw_tax": round(total_raw_tax, 2),
        "normal_count": normal_count, "void_count": void_count,
        "red_count": red_count,
    }


@app.get("/api/purchase-invoices/{invoice_id}")
def get_purchase_invoice(invoice_id: int, company_id: int = Query(...), db: Session = Depends(get_db)):
    inv = db.query(PurchaseInvoice).filter(PurchaseInvoice.company_id == company_id, PurchaseInvoice.id == invoice_id).first()
    if not inv:
        raise HTTPException(404, detail="发票不存在")
    # 查询凭证号
    voucher_no = ""
    je = db.query(JournalEntry).filter(
        JournalEntry.company_id == company_id,
        JournalEntry.source == "取得发票",
        JournalEntry.ref_id == inv.id
    ).first()
    if je:
        voucher_no = f"{je.voucher_word}-{je.voucher_no}"
    return {
        "id": inv.id,
        "invoice_code": inv.invoice_code or "",
        "invoice_no": inv.invoice_no,
        "digital_invoice_no": inv.digital_invoice_no or "",
        "seller_tax_no": inv.seller_tax_no or "",
        "seller_name": inv.seller_name or "",
        "buyer_tax_no": inv.buyer_tax_no or "",
        "buyer_name": inv.buyer_name or "",
        "invoice_date": str(inv.invoice_date) if inv.invoice_date else "",
        "tax_category_code": inv.tax_category_code or "",
        "specific_business_type": inv.specific_business_type or "",
        "goods_name": inv.goods_name or "",
        "spec": inv.spec or "",
        "unit": inv.unit or "",
        "quantity": inv.quantity or 0,
        "unit_price": inv.unit_price or 0,
        "amount": inv.amount or 0,
        "tax_rate": inv.tax_rate or 0,
        "tax_amount": inv.tax_amount or 0,
        "total_amount": inv.total_amount or 0,
        "invoice_source": inv.invoice_source or "",
        "invoice_category": inv.invoice_category or "增值税专用发票",
        "status": inv.status,
        "is_positive": inv.is_positive if inv.is_positive is not None else True,
        "invoice_risk_level": inv.invoice_risk_level or "",
        "issuer": inv.issuer or "",
        "remark": inv.remark or "",
        "journal_voucher_no": voucher_no,
        "skip_accounting": bool(inv.skip_accounting) if inv.skip_accounting is not None else False,
        "created_at": str(inv.created_at) if inv.created_at else "",
        "updated_at": str(inv.updated_at) if inv.updated_at else ""
    }


def _sync_pi_update_to_bi(db, company_id, pi):
    """编辑取得发票后同步到未记账凭证（按三号匹配更新）"""
    bis = db.query(BookkeepingInvoice).filter(
        BookkeepingInvoice.company_id == company_id,
        BookkeepingInvoice.invoice_code == pi.invoice_code,
        BookkeepingInvoice.invoice_no == pi.invoice_no,
        BookkeepingInvoice.digital_invoice_no == pi.digital_invoice_no,
    ).all()
    for bi in bis:
        bi.invoice_date = pi.invoice_date
        bi.seller_tax_no = pi.seller_tax_no
        bi.seller_name = pi.seller_name
        bi.buyer_tax_no = pi.buyer_tax_no
        bi.buyer_name = pi.buyer_name
        bi.tax_category_code = pi.tax_category_code
        bi.specific_business_type = pi.specific_business_type
        bi.goods_name = pi.goods_name
        bi.spec = pi.spec
        bi.unit = pi.unit
        bi.quantity = pi.quantity
        bi.unit_price = pi.unit_price
        bi.amount = pi.amount
        bi.tax_rate = pi.tax_rate
        bi.tax_amount = pi.tax_amount
        bi.total_amount = pi.total_amount
        bi.invoice_source = pi.invoice_source
        bi.invoice_category = pi.invoice_category
        bi.status = pi.status
        bi.is_positive = pi.is_positive
        bi.invoice_risk_level = pi.invoice_risk_level
        bi.issuer = pi.issuer
        bi.remark = pi.remark

def _sync_pi_delete_to_bi(db, company_id, pi):
    """删除取得发票后同步删除未记账凭证（按三号匹配，仅删未记账的）
    注意：三号可能为 None/空字符串，需用 IS NULL 处理，因为 SQL 中 NULL != NULL"""
    conditions = [BookkeepingInvoice.company_id == company_id]
    # 三号：按实际值匹配，NULL/空字符串用 IS NULL
    for bi_field, pi_val in [
        (BookkeepingInvoice.invoice_code, pi.invoice_code),
        (BookkeepingInvoice.invoice_no, pi.invoice_no),
        (BookkeepingInvoice.digital_invoice_no, pi.digital_invoice_no),
    ]:
        if pi_val and pi_val.strip():
            conditions.append(bi_field == pi_val)
        else:
            conditions.append(or_(bi_field.is_(None), bi_field == ""))
    # 仅删未记账的
    conditions.append(or_(BookkeepingInvoice.voucher_no.is_(None), BookkeepingInvoice.voucher_no == ""))
    db.query(BookkeepingInvoice).filter(and_(*conditions)).delete(synchronize_session=False)


@app.put("/api/purchase-invoices/{invoice_id}")
def update_purchase_invoice(invoice_id: int, data: PurchaseInvoiceUpdate, company_id: int = Query(...), db: Session = Depends(get_db)):
    inv = db.query(PurchaseInvoice).filter(PurchaseInvoice.company_id == company_id, PurchaseInvoice.id == invoice_id).first()
    if not inv:
        raise HTTPException(404, detail="发票不存在")
    for k, v in data.model_dump(exclude_unset=True).items():
        setattr(inv, k, v)
    inv.updated_at = datetime.now()
    db.flush()
    # 同步到未记账凭证（按三号匹配）
    _sync_pi_update_to_bi(db, company_id, inv)
    db.commit()
    return {"message": "更新成功"}


@app.delete("/api/purchase-invoices/{invoice_id}")
def delete_purchase_invoice(invoice_id: int, company_id: int = Query(...), db: Session = Depends(get_db)):
    inv = db.query(PurchaseInvoice).filter(PurchaseInvoice.company_id == company_id, PurchaseInvoice.id == invoice_id).first()
    if not inv:
        raise HTTPException(404, detail="发票不存在")
    # 同步删除到未记账凭证
    _sync_pi_delete_to_bi(db, company_id, inv)
    db.delete(inv)
    db.commit()
    return {"message": "删除成功"}


@app.post("/api/purchase-invoices/batch-delete")
def batch_delete_purchase_invoices(ids: list[int], company_id: int = Query(...), db: Session = Depends(get_db)):
    invoices = db.query(PurchaseInvoice).filter(
        PurchaseInvoice.company_id == company_id,
        PurchaseInvoice.id.in_(ids)
    ).all()
    # 同步删除到未记账凭证
    for inv in invoices:
        _sync_pi_delete_to_bi(db, company_id, inv)
    db.flush()
    deleted = db.query(PurchaseInvoice).filter(
        PurchaseInvoice.company_id == company_id,
        PurchaseInvoice.id.in_(ids)
    ).delete(synchronize_session=False)
    db.commit()
    return {"message": f"成功删除 {deleted} 条记录", "deleted": deleted}


@app.post("/api/purchase-invoices/transfer-to-bookkeeping")
def transfer_purchase_to_bookkeeping(ids: list[int], company_id: int = Query(...), db: Session = Depends(get_db)):
    """取得发票 → 转入记账发票（同时生成凭证入账）"""
    invoices = db.query(PurchaseInvoice).filter(
        PurchaseInvoice.company_id == company_id,
        PurchaseInvoice.id.in_(ids)
    ).all()
    if not invoices:
        raise HTTPException(400, "未找到匹配的发票")
    period_groups = {}
    for inv in invoices:
        if not inv.invoice_date: continue
        p = str(inv.invoice_date)[:7]
        period_groups.setdefault(p, []).append(inv)
    if not period_groups:
        raise HTTPException(400, "所选发票均无开票日期")
    transferred = 0
    for period, invs in period_groups.items():
        vno = _next_voucher_no(db, company_id, period, "记")
        for inv in invs:
            bi = BookkeepingInvoice(
                company_id=company_id,
                invoice_code=inv.invoice_code, invoice_no=inv.invoice_no,
                digital_invoice_no=inv.digital_invoice_no,
                seller_tax_no=inv.seller_tax_no, seller_name=inv.seller_name,
                buyer_tax_no=inv.buyer_tax_no, buyer_name=inv.buyer_name,
                invoice_date=inv.invoice_date,
                tax_category_code=inv.tax_category_code,
                specific_business_type=inv.specific_business_type,
                goods_name=inv.goods_name, spec=inv.spec, unit=inv.unit,
                quantity=inv.quantity, unit_price=inv.unit_price,
                amount=inv.amount, tax_rate=inv.tax_rate, tax_amount=inv.tax_amount,
                total_amount=inv.total_amount,
                invoice_source=inv.invoice_source, invoice_category=inv.invoice_category,
                status=inv.status, is_positive=inv.is_positive,
                invoice_risk_level=inv.invoice_risk_level,
                issuer=inv.issuer, remark=inv.remark,
                voucher_no=f"记-{vno}"
            )
            db.add(bi); db.flush()
            debit_account, _ = _classify_purchase_debit(db, company_id, inv)
            amt = float(inv.amount or 0); tax = float(inv.tax_amount or 0)
            is_special = inv.invoice_category and "专用发票" in str(inv.invoice_category)
            db.add(JournalEntry(company_id=company_id, entry_date=inv.invoice_date, period=period, voucher_word="记", voucher_no=vno,
                account_code=debit_account, debit_amount=amt, credit_amount=0,
                summary=f"{inv.invoice_date} {inv.seller_name or ''} {inv.goods_name or '发票'} 入账"))
            if is_special and tax > 0:
                db.add(JournalEntry(company_id=company_id, entry_date=inv.invoice_date, period=period, voucher_word="记", voucher_no=vno,
                    account_code="221001002", debit_amount=tax, credit_amount=0,
                    summary=f"{inv.invoice_date} {inv.seller_name or ''} 进项税额"))
            db.add(JournalEntry(company_id=company_id, entry_date=inv.invoice_date, period=period, voucher_word="记", voucher_no=vno,
                account_code="2202", debit_amount=0, credit_amount=amt + (tax if is_special else 0),
                summary=f"{inv.invoice_date} {inv.seller_name or ''} {inv.goods_name or '发票'}"))
            transferred += 1
        vno += 1
    for inv in invoices: db.delete(inv)
    db.commit()
    return {"message": f"成功转入 {transferred} 条发票到记账发票并生成凭证", "transferred": transferred}


@app.post("/api/purchase-invoices/transfer-to-unbookkept")
def transfer_purchase_to_unbookkept(ids: list[int], company_id: int = Query(...), db: Session = Depends(get_db)):
    """取得发票 → 转入未记账发票（不生成凭证，voucher_no为空）"""
    invoices = db.query(PurchaseInvoice).filter(
        PurchaseInvoice.company_id == company_id,
        PurchaseInvoice.id.in_(ids)
    ).all()
    if not invoices:
        raise HTTPException(400, "未找到匹配的发票")
    transferred = 0
    for inv in invoices:
        db.add(BookkeepingInvoice(
            company_id=company_id,
            invoice_code=inv.invoice_code, invoice_no=inv.invoice_no,
            digital_invoice_no=inv.digital_invoice_no,
            seller_tax_no=inv.seller_tax_no, seller_name=inv.seller_name,
            buyer_tax_no=inv.buyer_tax_no, buyer_name=inv.buyer_name,
            invoice_date=inv.invoice_date,
            tax_category_code=inv.tax_category_code,
            specific_business_type=inv.specific_business_type,
            goods_name=inv.goods_name, spec=inv.spec, unit=inv.unit,
            quantity=inv.quantity, unit_price=inv.unit_price,
            amount=inv.amount, tax_rate=inv.tax_rate, tax_amount=inv.tax_amount,
            total_amount=inv.total_amount,
            invoice_source=inv.invoice_source, invoice_category=inv.invoice_category,
            status=inv.status, is_positive=inv.is_positive,
            invoice_risk_level=inv.invoice_risk_level,
            issuer=inv.issuer, remark=inv.remark,
        ))
        db.delete(inv)
        transferred += 1
    db.commit()
    return {"message": f"成功转入 {transferred} 条发票到未记账发票", "transferred": transferred}


@app.post("/api/purchase-invoices/sync-to-unbookkept")
def sync_purchase_to_unbookkept(company_id: int = Query(...), db: Session = Depends(get_db)):
    """同步：取得发票 → 未记账发票（逐条创建BookkeepingInvoice，不去重）"""
    all_pi = db.query(PurchaseInvoice).filter(PurchaseInvoice.company_id == company_id).all()
    if not all_pi:
        return {"message": "无待同步发票", "synced": 0}
    
    synced = 0
    for inv in all_pi:
        db.add(BookkeepingInvoice(
            company_id=company_id,
            invoice_code=inv.invoice_code, invoice_no=inv.invoice_no,
            digital_invoice_no=inv.digital_invoice_no,
            seller_tax_no=inv.seller_tax_no, seller_name=inv.seller_name,
            buyer_tax_no=inv.buyer_tax_no, buyer_name=inv.buyer_name,
            invoice_date=inv.invoice_date,
            tax_category_code=inv.tax_category_code,
            specific_business_type=inv.specific_business_type,
            goods_name=inv.goods_name, spec=inv.spec, unit=inv.unit,
            quantity=inv.quantity, unit_price=inv.unit_price,
            amount=inv.amount, tax_rate=inv.tax_rate, tax_amount=inv.tax_amount,
            total_amount=inv.total_amount,
            invoice_source=inv.invoice_source, invoice_category=inv.invoice_category,
            status=inv.status, is_positive=inv.is_positive,
            invoice_risk_level=inv.invoice_risk_level,
            issuer=inv.issuer, remark=inv.remark,
        ))
        synced += 1
    
    db.commit()
    return {"message": f"成功同步 {synced} 条发票到未记账发票", "synced": synced}


@app.post("/api/purchase-invoices/generate-voucher-only")
def purchase_invoice_generate_voucher_only(body: dict, company_id: int = Query(...), db: Session = Depends(get_db)):
    """取得发票 → 仅生成序时账凭证（不入进项认证模块）"""
    ids = body.get("ids", [])
    if not ids:
        raise HTTPException(400, "请提供发票ID列表")
    invoices = db.query(PurchaseInvoice).filter(
        PurchaseInvoice.company_id == company_id,
        PurchaseInvoice.id.in_(ids)
    ).all()
    if not invoices:
        raise HTTPException(400, "未找到匹配的发票")
    
    # 按期分组
    period_groups = {}
    for inv in invoices:
        if not inv.invoice_date: continue
        p = str(inv.invoice_date)[:7]
        period_groups.setdefault(p, []).append(inv)
    if not period_groups:
        raise HTTPException(400, "所选发票均无开票日期")
    
    count = 0
    for period, invs in period_groups.items():
        vno = _next_voucher_no(db, company_id, period, "记")
        for inv in invs:
            debit_account, _ = _classify_purchase_debit(db, company_id, inv)
            amt = float(inv.amount or 0); tax = float(inv.tax_amount or 0)
            is_special = inv.invoice_category and "专用发票" in str(inv.invoice_category)
            # 费用/成本（借方）
            db.add(JournalEntry(company_id=company_id, entry_date=inv.invoice_date, period=period, voucher_word="记", voucher_no=vno,
                account_code=debit_account, debit_amount=amt, credit_amount=0,
                summary=f"{inv.invoice_date} {inv.seller_name or ''} {inv.goods_name or '发票'}"))
            # 进项税额（专票）
            if is_special and tax > 0:
                db.add(JournalEntry(company_id=company_id, entry_date=inv.invoice_date, period=period, voucher_word="记", voucher_no=vno,
                    account_code="221001002", debit_amount=tax, credit_amount=0,
                    summary=f"{inv.invoice_date} {inv.seller_name or ''} 进项税额"))
            # 应付账款（贷方）
            db.add(JournalEntry(company_id=company_id, entry_date=inv.invoice_date, period=period, voucher_word="记", voucher_no=vno,
                account_code="2202", debit_amount=0, credit_amount=amt + (tax if is_special else 0),
                summary=f"{inv.invoice_date} {inv.seller_name or ''} {inv.goods_name or '发票'}"))
            count += 1
        vno += 1
    db.commit()
    return {"message": f"成功生成 {count} 张凭证（不入进项认证）", "count": count}


@app.post("/api/purchase-invoices/{invoice_id}/to-journal")
def purchase_invoice_to_journal(invoice_id: int, company_id: int = Query(...), db: Session = Depends(get_db)):
    """将单张取得发票生成进项抵扣记录并生成凭证"""
    inv = db.query(PurchaseInvoice).filter(
        PurchaseInvoice.id == invoice_id,
        PurchaseInvoice.company_id == company_id
    ).first()
    if not inv:
        raise HTTPException(404, "发票不存在")

    period = inv.invoice_date.strftime("%Y-%m") if inv.invoice_date else datetime.now().strftime("%Y-%m")

    # 查找或创建进项抵扣记录
    ded = db.query(InputVATDeduction).filter(
        InputVATDeduction.company_id == company_id,
        InputVATDeduction.invoice_no == inv.invoice_no
    ).first()

    if not ded:
        ded = InputVATDeduction(
            company_id=company_id,
            purchase_invoice_id=inv.id,
            invoice_code=inv.invoice_code or "",
            invoice_no=inv.invoice_no or "",
            invoice_date=inv.invoice_date,
            seller_name=inv.seller_name or "",
            seller_tax_id=inv.seller_tax_no or "",
            amount=inv.amount or 0,
            tax_amount=inv.tax_amount or 0,
            deductible_tax_amount=inv.tax_amount or 0,
            total_amount=inv.total_amount or 0,
            invoice_category=inv.invoice_category or "增值税专用发票",
            goods_name=inv.goods_name or "",
            deduction_period=period,
            deduction_status="待抵扣",
        )
        db.add(ded)
    else:
        if not ded.deduction_period:
            ded.deduction_period = period
        if not ded.deduction_status:
            ded.deduction_status = "待抵扣"

    db.flush()

    # 供应商建档由 /api/process-all 统一处理，此处直接生成凭证（2026-06-10 铁律）

    # 生成采购入账凭证（借：库存商品 / 贷：应付账款）
    purchase_count = auto_generate_purchase_journal(db, company_id, invoice_id)

    # 按月汇总生成进项抵扣凭证
    vat_count = auto_generate_input_vat_for_period(db, company_id, period)
    db.commit()
    return {"message": f"已为 {period} 生成采购凭证 {purchase_count} 张、进项抵扣凭证 {vat_count} 条", "period": period}


@app.post("/api/purchase-invoices/batch-to-journal")
def purchase_invoice_batch_to_journal(
    body: dict = Body(default=None),
    company_id: int = Query(...),
    db=Depends(get_db)
):
    """一键将勾选的取得发票生成进项抵扣凭证（按月汇总）"""
    # body=None 表示前端传了 null，即为所有发票生成凭证
    if body is None:
        ids = None
    else:
        ids = body.get("ids")
    if ids is None:
        # 为所有未取得凭证的发票生成
        invoices = db.query(PurchaseInvoice).filter(
            PurchaseInvoice.company_id == company_id
        ).order_by(PurchaseInvoice.invoice_date, PurchaseInvoice.id).all()
    else:
        if not ids:
            return {"message": "未选择任何发票", "generated": 0, "skipped": 0, "errors": []}
        invoices = db.query(PurchaseInvoice).filter(
            PurchaseInvoice.company_id == company_id,
            PurchaseInvoice.id.in_(ids)
        ).order_by(PurchaseInvoice.invoice_date, PurchaseInvoice.id).all()

    generated = 0
    skipped = 0
    errors = []
    affected_periods = set()

    for inv in invoices:
        try:
            period = inv.invoice_date.strftime("%Y-%m") if inv.invoice_date else datetime.now().strftime("%Y-%m")

            # 供应商建档+采购凭证由 /api/process-all 统一处理（2026-06-10 铁律）

            # 查找或创建进项抵扣记录
            ded = db.query(InputVATDeduction).filter(
                InputVATDeduction.company_id == company_id,
                InputVATDeduction.invoice_no == inv.invoice_no
            ).first()

            if not ded:
                ded = InputVATDeduction(
                    company_id=company_id,
                    purchase_invoice_id=inv.id,
                    invoice_code=inv.invoice_code or "",
                    invoice_no=inv.invoice_no or "",
                    invoice_date=inv.invoice_date,
                    seller_name=inv.seller_name or "",
                    seller_tax_id=inv.seller_tax_no or "",
                    amount=inv.amount or 0,
                    tax_amount=inv.tax_amount or 0,
                    deductible_tax_amount=inv.tax_amount or 0,
                    total_amount=inv.total_amount or 0,
                    invoice_category=inv.invoice_category or "增值税专用发票",
                    goods_name=inv.goods_name or "",
                    deduction_period=period,
                    deduction_status="待抵扣",
                )
                db.add(ded)
                db.flush()
            else:
                if not ded.deduction_period:
                    ded.deduction_period = period
                if not ded.deduction_status:
                    ded.deduction_status = "待抵扣"
                db.flush()

            affected_periods.add(period)
            generated += 1
        except Exception as e:
            errors.append(f"发票{inv.id}({inv.invoice_no}): {str(e)}")
            db.rollback()

    # 按月汇总生成进项抵扣凭证
    voucher_count = 0
    for period in sorted(affected_periods):
        try:
            c = auto_generate_input_vat_for_period(db, company_id, period)
            voucher_count += c
        except Exception as e:
            errors.append(f"生成期间{period}凭证失败: {str(e)}")


    # 生成采购入账凭证（借：库存商品 / 贷：应付账款）
    # 只处理本次勾选的发票，避免全量重算
    if ids:
        _inv_ids = [inv.id for inv in invoices]
        purchase_count = auto_generate_purchase_journal(db, company_id, invoice_id=_inv_ids)
    else:
        purchase_count = auto_generate_purchase_journal(db, company_id)
    db.commit()
    msg = f"批量生成完成：{generated} 张发票 → 进项抵扣凭证 {voucher_count} 笔"
    if purchase_count:
        msg += f"，采购凭证 {purchase_count} 笔"
    if skipped > 0:
        msg += f"，跳过 {skipped} 张"
    if errors:
        msg += f"，{len(errors)} 项失败"
    return {"message": msg, "generated": generated, "skipped": skipped, "vouchers": voucher_count, "errors": errors}


# ==================== 记账发票 ====================

class BookkeepingInvoiceCreate(BaseModel):
    invoice_code: Optional[str] = None
    invoice_no: Optional[str] = None
    digital_invoice_no: Optional[str] = None
    seller_tax_no: Optional[str] = None
    seller_name: Optional[str] = None
    buyer_tax_no: Optional[str] = None
    buyer_name: Optional[str] = None
    invoice_date: Optional[date] = None
    tax_category_code: Optional[str] = None
    specific_business_type: Optional[str] = None
    goods_name: Optional[str] = None
    spec: Optional[str] = None
    unit: Optional[str] = None
    quantity: Optional[float] = 0
    unit_price: Optional[float] = 0
    amount: float = 0.0
    tax_rate: Optional[float] = 0.0
    tax_amount: Optional[float] = 0.0
    total_amount: Optional[float] = 0.0
    invoice_source: Optional[str] = None
    invoice_category: str = "增值税普通发票"
    status: str = "正常"
    is_positive: Optional[bool] = True
    invoice_risk_level: Optional[str] = None
    issuer: Optional[str] = None
    remark: Optional[str] = None


class BookkeepingInvoiceUpdate(BaseModel):
    invoice_code: Optional[str] = None
    digital_invoice_no: Optional[str] = None
    seller_tax_no: Optional[str] = None
    seller_name: Optional[str] = None
    buyer_tax_no: Optional[str] = None
    buyer_name: Optional[str] = None
    invoice_date: Optional[date] = None
    tax_category_code: Optional[str] = None
    specific_business_type: Optional[str] = None
    goods_name: Optional[str] = None
    spec: Optional[str] = None
    unit: Optional[str] = None
    quantity: Optional[float] = None
    unit_price: Optional[float] = None
    amount: Optional[float] = None
    tax_rate: Optional[float] = None
    tax_amount: Optional[float] = None
    total_amount: Optional[float] = None
    invoice_source: Optional[str] = None
    invoice_category: Optional[str] = None
    status: Optional[str] = None
    is_positive: Optional[bool] = None
    invoice_risk_level: Optional[str] = None
    issuer: Optional[str] = None
    remark: Optional[str] = None


@app.get("/api/bookkeeping-invoices")
def list_bookkeeping_invoices(
    company_id: int = Query(...),
    invoice_category: Optional[str] = None,
    status: Optional[str] = None,
    keyword: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    is_posted: Optional[bool] = None,
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=500),
    db: Session = Depends(get_db)
):
    q = db.query(BookkeepingInvoice).filter(BookkeepingInvoice.company_id == company_id)
    if invoice_category:
        q = q.filter(BookkeepingInvoice.invoice_category == invoice_category)
    if status:
        q = q.filter(BookkeepingInvoice.status == status)
    if date_from:
        q = q.filter(BookkeepingInvoice.invoice_date >= date_from)
    if date_to:
        q = q.filter(BookkeepingInvoice.invoice_date <= date_to)
    if is_posted is True:
        q = q.filter(BookkeepingInvoice.voucher_no.isnot(None))
    elif is_posted is False:
        q = q.filter(BookkeepingInvoice.voucher_no.is_(None))
    if keyword:
        q = q.filter(or_(
            BookkeepingInvoice.invoice_no.contains(keyword),
            BookkeepingInvoice.invoice_code.contains(keyword),
            BookkeepingInvoice.digital_invoice_no.contains(keyword),
            BookkeepingInvoice.seller_name.contains(keyword),
            BookkeepingInvoice.goods_name.contains(keyword)
        ))
    invoices = q.order_by(BookkeepingInvoice.invoice_date.desc()).offset(skip).limit(limit).all()
    t = q.count()
    return {
        "total": t,
        "items": [{
            "id": inv.id,
            "invoice_code": inv.invoice_code or "",
            "invoice_no": inv.invoice_no,
            "digital_invoice_no": inv.digital_invoice_no or "",
            "seller_tax_no": inv.seller_tax_no or "",
            "seller_name": inv.seller_name or "",
            "buyer_tax_no": inv.buyer_tax_no or "",
            "buyer_name": inv.buyer_name or "",
            "invoice_date": str(inv.invoice_date) if inv.invoice_date else "",
            "tax_category_code": inv.tax_category_code or "",
            "specific_business_type": inv.specific_business_type or "",
            "goods_name": inv.goods_name or "",
            "spec": inv.spec or "",
            "unit": inv.unit or "",
            "quantity": inv.quantity or 0,
            "unit_price": inv.unit_price or 0,
            "amount": inv.amount or 0,
            "tax_rate": inv.tax_rate or 0,
            "tax_amount": inv.tax_amount or 0,
            "total_amount": inv.total_amount or 0,
            "invoice_source": inv.invoice_source or "",
            "invoice_category": inv.invoice_category or "增值税普通发票",
            "status": inv.status,
            "is_positive": inv.is_positive if inv.is_positive is not None else True,
            "invoice_risk_level": inv.invoice_risk_level or "",
            "issuer": inv.issuer or "",
            "remark": inv.remark or "",
            "voucher_no": inv.voucher_no or "",
            "created_at": str(inv.created_at) if inv.created_at else ""
        } for inv in invoices]
    }


@app.post("/api/bookkeeping-invoices")
def create_bookkeeping_invoice(data: BookkeepingInvoiceCreate, company_id: int = Query(...), db: Session = Depends(get_db)):
    inv = BookkeepingInvoice(company_id=company_id, **data.model_dump())
    db.add(inv)
    db.commit()
    db.refresh(inv)
    return {"id": inv.id, "invoice_no": inv.invoice_no, "message": "记账发票创建成功"}


@app.get("/api/bookkeeping-invoices/stats")
def bookkeeping_invoice_stats(company_id: int = Query(...), tab: str = Query("all"), is_posted: Optional[bool] = None, db: Session = Depends(get_db)):
    base = db.query(BookkeepingInvoice).filter(BookkeepingInvoice.company_id == company_id)
    if is_posted is True:
        base = base.filter(BookkeepingInvoice.voucher_no.isnot(None))
    elif is_posted is False:
        base = base.filter(BookkeepingInvoice.voucher_no.is_(None))
    if tab == "zpt":
        base = base.filter(BookkeepingInvoice.invoice_category.contains("专用发票"))
    elif tab == "ppt":
        base = base.filter(BookkeepingInvoice.invoice_category.contains("普通发票"))
    total_count = base.count()
    total_amt = base.with_entities(func.sum(func.coalesce(BookkeepingInvoice.amount, 0))).scalar() or 0
    total_amount = base.with_entities(func.sum(func.coalesce(BookkeepingInvoice.total_amount, 0))).scalar() or 0
    total_raw_tax = base.with_entities(func.sum(func.coalesce(BookkeepingInvoice.tax_amount, 0))).scalar() or 0
    normal_count = base.filter(BookkeepingInvoice.status == "正常").count()
    void_count = base.filter(BookkeepingInvoice.status.like("%作废%")).count()
    red_count = base.filter(BookkeepingInvoice.status.like("%红冲%")).count()
    return {
        "total_count": total_count, "total_amt": round(total_amt, 2),
        "total_amount": round(total_amount, 2),
        "total_raw_tax": round(total_raw_tax, 2),
        "normal_count": normal_count, "void_count": void_count,
        "red_count": red_count
    }


@app.get("/api/bookkeeping-invoices/{invoice_id}")
def get_bookkeeping_invoice(invoice_id: int, company_id: int = Query(...), db: Session = Depends(get_db)):
    inv = db.query(BookkeepingInvoice).filter(BookkeepingInvoice.company_id == company_id, BookkeepingInvoice.id == invoice_id).first()
    if not inv:
        raise HTTPException(404, detail="发票不存在")
    return {
        "id": inv.id,
        "invoice_code": inv.invoice_code or "",
        "invoice_no": inv.invoice_no,
        "digital_invoice_no": inv.digital_invoice_no or "",
        "seller_tax_no": inv.seller_tax_no or "",
        "seller_name": inv.seller_name or "",
        "buyer_tax_no": inv.buyer_tax_no or "",
        "buyer_name": inv.buyer_name or "",
        "invoice_date": str(inv.invoice_date) if inv.invoice_date else "",
        "tax_category_code": inv.tax_category_code or "",
        "specific_business_type": inv.specific_business_type or "",
        "goods_name": inv.goods_name or "",
        "spec": inv.spec or "",
        "unit": inv.unit or "",
        "quantity": inv.quantity or 0,
        "unit_price": inv.unit_price or 0,
        "amount": inv.amount or 0,
        "tax_rate": inv.tax_rate or 0,
        "tax_amount": inv.tax_amount or 0,
        "total_amount": inv.total_amount or 0,
        "invoice_source": inv.invoice_source or "",
        "invoice_category": inv.invoice_category or "增值税普通发票",
        "status": inv.status,
        "is_positive": inv.is_positive if inv.is_positive is not None else True,
        "invoice_risk_level": inv.invoice_risk_level or "",
        "issuer": inv.issuer or "",
        "remark": inv.remark or "",
        "created_at": str(inv.created_at) if inv.created_at else "",
        "updated_at": str(inv.updated_at) if inv.updated_at else ""
    }


@app.put("/api/bookkeeping-invoices/{invoice_id}")
def update_bookkeeping_invoice(invoice_id: int, data: BookkeepingInvoiceUpdate, company_id: int = Query(...), db: Session = Depends(get_db)):
    inv = db.query(BookkeepingInvoice).filter(BookkeepingInvoice.company_id == company_id, BookkeepingInvoice.id == invoice_id).first()
    if not inv:
        raise HTTPException(404, detail="发票不存在")
    for k, v in data.model_dump(exclude_unset=True).items():
        setattr(inv, k, v)
    inv.updated_at = datetime.now()
    db.commit()
    return {"message": "更新成功"}


@app.delete("/api/bookkeeping-invoices/{invoice_id}")
def delete_bookkeeping_invoice(invoice_id: int, company_id: int = Query(...), db: Session = Depends(get_db)):
    inv = db.query(BookkeepingInvoice).filter(BookkeepingInvoice.company_id == company_id, BookkeepingInvoice.id == invoice_id).first()
    if not inv:
        raise HTTPException(404, detail="发票不存在")
    db.delete(inv)
    db.commit()
    return {"message": "删除成功"}


@app.post("/api/bookkeeping-invoices/auto-voucher")
def bookkeeping_invoices_auto_voucher(company_id: int = Query(...), db: Session = Depends(get_db)):
    """未记账发票一键生成凭证"""
    count = auto_generate_bookkeeping_journal(db, company_id)
    db.commit()
    return {"message": f"自动生成 {count} 张未记账发票凭证", "generated": count}


@app.post("/api/bookkeeping-invoices/batch-delete")
def batch_delete_bookkeeping_invoices(ids: list[int], company_id: int = Query(...), only_unposted: bool = Query(False), db: Session = Depends(get_db)):
    """批量删除记账发票。only_unposted=True时仅删除未记账的（voucher_no为空）"""
    q = db.query(BookkeepingInvoice).filter(
        BookkeepingInvoice.company_id == company_id,
        BookkeepingInvoice.id.in_(ids)
    )
    if only_unposted:
        q = q.filter(or_(BookkeepingInvoice.voucher_no == None, BookkeepingInvoice.voucher_no == ""))
    deleted = q.delete(synchronize_session=False)
    db.commit()
    return {"message": f"成功删除 {deleted} 条记录", "deleted": deleted}


@app.post("/api/bookkeeping-invoices/batch-generate-voucher")
def batch_generate_bookkeeping_voucher(ids: list[int], company_id: int = Query(...), period: Optional[str] = Query(None), db: Session = Depends(get_db)):
    """批量生成记账发票凭证。period指定记账期间（YYYY-MM），为空则按发票日期自动分组"""
    invoices = db.query(BookkeepingInvoice).filter(
        BookkeepingInvoice.company_id == company_id,
        BookkeepingInvoice.id.in_(ids)
    ).all()
    if not invoices:
        raise HTTPException(400, "未找到匹配的发票")
    
    # 确定记账期间
    if period:
        # 统一记入指定期间
        period_groups = {period: invoices}
    else:
        # 按发票日期分组
        period_groups = {}
        for inv in invoices:
            if not inv.invoice_date:
                continue
            p = str(inv.invoice_date)[:7]
            period_groups.setdefault(p, []).append(inv)
    
    if not period_groups:
        raise HTTPException(400, "所选发票均无开票日期")
    
    posted_count = 0
    for period, invs in period_groups.items():
        # 在期间内按发票三号分组，每组一个凭证号
        key_groups = {}
        for inv in invs:
            key = (inv.invoice_code or "") + "|" + (inv.invoice_no or "") + "|" + (inv.digital_invoice_no or "")
            key_groups.setdefault(key, []).append(inv)
        voucher_no = _next_voucher_no(db, company_id, period, "记")
        for key, group in key_groups.items():
            for inv in group:
                # 借方：费用科目（根据品名智能分类）
                debit_account, debit_account_name = _classify_purchase_debit(db, company_id, inv)
                amount = float(inv.amount or 0)
                tax_amount = float(inv.tax_amount or 0)
                
                # 费用分录（借方）
                db.add(JournalEntry(
                    company_id=company_id, entry_date=inv.invoice_date, period=period,
                    voucher_word="记", voucher_no=voucher_no,
                    account_code=debit_account,
                    debit_amount=amount, credit_amount=0,
                    summary=f"{inv.invoice_date} {inv.seller_name or '供应商'} {inv.goods_name or '发票'} 入账"
                ))
                # 进项税额（专票才有）
                if inv.invoice_category and "专用发票" in inv.invoice_category and tax_amount > 0:
                    db.add(JournalEntry(
                        company_id=company_id, entry_date=inv.invoice_date, period=period,
                        voucher_word="记", voucher_no=voucher_no,
                        account_code="221001002",
                        debit_amount=tax_amount, credit_amount=0,
                        summary=f"{inv.invoice_date} {inv.seller_name or '供应商'} 进项税额"
                    ))
                # 应付账款（贷方）
                db.add(JournalEntry(
                    company_id=company_id, entry_date=inv.invoice_date, period=period,
                    voucher_word="记", voucher_no=voucher_no,
                    account_code="2202",
                    debit_amount=0, credit_amount=amount + (tax_amount if (inv.invoice_category and "专用发票" in inv.invoice_category) else 0),
                    summary=f"{inv.invoice_date} {inv.seller_name or '供应商'} {inv.goods_name or '发票'}"
                ))
                # 标记已记账
                inv.voucher_no = f"记-{voucher_no}"
                posted_count += 1
            voucher_no += 1
    
    # 记账后同步锁定取得发票（通过三号匹配，标记 skip_accounting）
    for inv in invoices:
        if inv.invoice_no or inv.invoice_code or inv.digital_invoice_no:
            db.query(PurchaseInvoice).filter(
                PurchaseInvoice.company_id == company_id,
                PurchaseInvoice.invoice_code == inv.invoice_code,
                PurchaseInvoice.invoice_no == inv.invoice_no,
                PurchaseInvoice.digital_invoice_no == inv.digital_invoice_no,
            ).update({"skip_accounting": True}, synchronize_session=False)
    
    db.commit()
    return {"message": f"成功生成凭证，{posted_count} 条发票已记账", "posted": posted_count}


# ==================== 银行流水规则库 ====================

class BankRuleCreate(BaseModel):
    keyword: str
    account_code: str
    account_name: Optional[str] = None
    transaction_type: str = "全部"  # 收入 / 支出 / 全部
    direction: str = "auto"  # debit / credit / auto
    priority: int = 0

class BankRuleUpdate(BaseModel):
    keyword: Optional[str] = None
    account_code: Optional[str] = None
    account_name: Optional[str] = None
    transaction_type: Optional[str] = None
    direction: Optional[str] = None
    priority: Optional[int] = None
    is_active: Optional[int] = None

# ==================== 银行配置 ====================

class BankConfigCreate(BaseModel):
    bank_name: str
    account_number: Optional[str] = None
    account_name: Optional[str] = None
    column_mapping: Optional[str] = None  # JSON string

class BankConfigUpdate(BaseModel):
    bank_name: Optional[str] = None
    account_number: Optional[str] = None
    account_name: Optional[str] = None
    column_mapping: Optional[str] = None
    is_active: Optional[bool] = None


@app.get("/api/bank-configs")
def list_bank_configs(company_id: int = Query(...), db: Session = Depends(get_db)):
    configs = db.query(BankConfig).filter(
        BankConfig.company_id == company_id, BankConfig.is_active == True
    ).order_by(BankConfig.bank_name).all()
    return [{
        "id": c.id, "bank_name": c.bank_name,
        "account_number": c.account_number or "",
        "account_name": c.account_name or "",
        "column_mapping": c.column_mapping or "{}",
        "created_at": str(c.created_at) if c.created_at else ""
    } for c in configs]


@app.post("/api/bank-configs")
def create_bank_config(data: BankConfigCreate, company_id: int = Query(...), db: Session = Depends(get_db)):
    cfg = BankConfig(company_id=company_id, **data.model_dump())
    db.add(cfg)
    db.commit()
    db.refresh(cfg)
    return {"id": cfg.id, "message": "银行配置创建成功"}


@app.put("/api/bank-configs/{config_id}")
def update_bank_config(config_id: int, data: BankConfigUpdate, company_id: int = Query(...), db: Session = Depends(get_db)):
    cfg = db.query(BankConfig).filter(BankConfig.company_id == company_id, BankConfig.id == config_id).first()
    if not cfg:
        raise HTTPException(404, detail="银行配置不存在")
    for k, v in data.model_dump(exclude_unset=True).items():
        setattr(cfg, k, v)
    cfg.updated_at = datetime.now()
    db.commit()
    return {"message": "更新成功"}


@app.delete("/api/bank-configs/{config_id}")
def delete_bank_config(config_id: int, company_id: int = Query(...), db: Session = Depends(get_db)):
    cfg = db.query(BankConfig).filter(BankConfig.company_id == company_id, BankConfig.id == config_id).first()
    if not cfg:
        raise HTTPException(404, detail="银行配置不存在")
    cfg.is_active = False
    db.commit()
    return {"message": "已停用"}


# ==================== 银行流水规则库 ====================

@app.get("/api/bank-rules")
def list_bank_rules(
    company_id: int = Query(...),
    transaction_type: Optional[str] = None,
    db: Session = Depends(get_db)
):
    q = db.query(BankRule).filter(BankRule.company_id == company_id, BankRule.is_active == 1)
    if transaction_type and transaction_type != "全部":
        q = q.filter(or_(BankRule.transaction_type == transaction_type, BankRule.transaction_type == "全部"))
    rules = q.order_by(BankRule.priority.desc(), BankRule.id.asc()).all()
    return [{
        "id": r.id, "keyword": r.keyword, "account_code": r.account_code,
        "account_name": r.account_name, "transaction_type": r.transaction_type,
        "direction": r.direction, "priority": r.priority, "is_active": r.is_active,
    } for r in rules]


@app.post("/api/bank-rules")
def create_bank_rule(data: BankRuleCreate, company_id: int = Query(...), db: Session = Depends(get_db)):
    # 确保科目存在
    acct = db.query(Account).filter(Account.company_id == company_id, Account.code == data.account_code).first()
    rule = BankRule(
        company_id=company_id,
        keyword=data.keyword,
        account_code=data.account_code,
        account_name=acct.name if acct else data.account_name,
        transaction_type=data.transaction_type,
        direction=data.direction,
        priority=data.priority,
        is_active=1,
    )
    db.add(rule)
    db.commit()
    return {"message": "规则已添加", "id": rule.id}


@app.put("/api/bank-rules/{rule_id}")
def update_bank_rule(rule_id: int, data: BankRuleUpdate, company_id: int = Query(...), db: Session = Depends(get_db)):
    rule = db.query(BankRule).filter(BankRule.company_id == company_id, BankRule.id == rule_id).first()
    if not rule:
        raise HTTPException(404, detail="规则不存在")
    for field in ["keyword", "account_code", "account_name", "transaction_type", "direction", "priority", "is_active"]:
        val = getattr(data, field, None)
        if val is not None:
            if field == "account_code":
                acct = db.query(Account).filter(Account.company_id == company_id, Account.code == val).first()
                if acct:
                    setattr(rule, "account_name", acct.name)
            setattr(rule, field, val)
    rule.updated_at = datetime.now()
    db.commit()
    return {"message": "规则已更新"}


@app.delete("/api/bank-rules/{rule_id}")
def delete_bank_rule(rule_id: int, company_id: int = Query(...), db: Session = Depends(get_db)):
    rule = db.query(BankRule).filter(BankRule.company_id == company_id, BankRule.id == rule_id).first()
    if not rule:
        raise HTTPException(404, detail="规则不存在")
    rule.is_active = 0
    rule.updated_at = datetime.now()
    db.commit()
    return {"message": "规则已删除"}


# ==================== 银行流水 ====================

class BankTransactionCreate(BaseModel):
    bank_config_id: Optional[int] = None
    transaction_date: date
    transaction_time: Optional[str] = None
    application_date: Optional[date] = None
    voucher_no: Optional[str] = None
    debit_amount: Optional[float] = 0.0
    credit_amount: Optional[float] = 0.0
    balance: Optional[float] = 0.0
    counterparty_name: Optional[str] = None
    counterparty_account: Optional[str] = None
    counterparty_bank: Optional[str] = None
    transaction_serial_no: Optional[str] = None
    voucher_seq: Optional[str] = None
    record_status: Optional[str] = None
    summary: Optional[str] = None
    transaction_remark: Optional[str] = None
    account_type: Optional[str] = None
    # 旧字段（向后兼容）
    amount: Optional[float] = 0.0
    transaction_type: Optional[str] = "支出"
    payment_method: Optional[str] = None
    reference_no: Optional[str] = None
    raw_data: Optional[str] = None
    remark: Optional[str] = None


class BankTransactionUpdate(BaseModel):
    bank_config_id: Optional[int] = None
    transaction_date: Optional[date] = None
    transaction_time: Optional[str] = None
    application_date: Optional[date] = None
    voucher_no: Optional[str] = None
    debit_amount: Optional[float] = None
    credit_amount: Optional[float] = None
    balance: Optional[float] = None
    counterparty_name: Optional[str] = None
    counterparty_account: Optional[str] = None
    counterparty_bank: Optional[str] = None
    transaction_serial_no: Optional[str] = None
    voucher_seq: Optional[str] = None
    record_status: Optional[str] = None
    summary: Optional[str] = None
    transaction_remark: Optional[str] = None
    account_type: Optional[str] = None
    # 旧字段（向后兼容）
    amount: Optional[float] = None
    transaction_type: Optional[str] = None
    payment_method: Optional[str] = None
    reference_no: Optional[str] = None
    raw_data: Optional[str] = None
    remark: Optional[str] = None


@app.get("/api/bank-transactions")
def list_bank_transactions(
    company_id: int = Query(...),
    bank_config_id: Optional[int] = None,
    transaction_type: Optional[str] = None,
    keyword: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=500),
    db: Session = Depends(get_db)
):
    q = db.query(BankTransaction).filter(BankTransaction.company_id == company_id)
    if bank_config_id:
        q = q.filter(BankTransaction.bank_config_id == bank_config_id)
    if transaction_type:
        q = q.filter(BankTransaction.transaction_type == transaction_type)
    if date_from:
        q = q.filter(BankTransaction.transaction_date >= date_from)
    if date_to:
        q = q.filter(BankTransaction.transaction_date <= date_to)
    if keyword:
        q = q.filter(or_(
            BankTransaction.counterparty_name.contains(keyword),
            BankTransaction.summary.contains(keyword),
            BankTransaction.reference_no.contains(keyword)
        ))
    txs = q.order_by(BankTransaction.transaction_date.desc(), BankTransaction.id.desc()).offset(skip).limit(limit).all()

    # 动态查询凭证号：按 summary + 1002科目匹配（双分录中银行存款侧即代表该凭证）
    # 银行流水可能被多个模块匹配生成凭证（银行流水自动生成 / 社保缴纳 / 公积金缴纳），
    # summary 格式可能不同，需全部收集后匹配
    voucher_map = {}
    if txs:
        summaries = []
        # 收集所有可能的 summary 格式
        for tx in txs:
            cp = tx.counterparty_name or tx.summary or "银行流水"
            summaries.append(f"银行流水-#{tx.id}-{cp}")
            summaries.append(f"社保缴纳-#{tx.id}")
            summaries.append(f"公积金缴纳-#{tx.id}")
        summaries = list(set(summaries))
        bank_jes = db.query(JournalEntry).filter(
            JournalEntry.company_id == company_id,
            JournalEntry.account_code == "1002",
            JournalEntry.summary.in_(summaries)
        ).all() if summaries else []
        # 构建 summary → 凭证号 映射
        summary_to_voucher = {}
        for je in bank_jes:
            summary_to_voucher[je.summary] = f"{je.voucher_word}-{je.voucher_no}"
        for tx in txs:
            cp = tx.counterparty_name or tx.summary or "银行流水"
            target = f"银行流水-#{tx.id}-{cp}"
            voucher_no = summary_to_voucher.get(target, "")
            if not voucher_no:
                # 尝试社保缴纳/公积金缴纳 summary 格式
                for alt_fmt in [f"社保缴纳-#{tx.id}", f"公积金缴纳-#{tx.id}"]:
                    alt = summary_to_voucher.get(alt_fmt)
                    if alt:
                        voucher_no = alt
                        break
            voucher_map[tx.id] = voucher_no

    return [{
        "id": tx.id, "bank_config_id": tx.bank_config_id,
        "transaction_date": str(tx.transaction_date) if tx.transaction_date else "",
        "transaction_time": str(tx.transaction_time) if tx.transaction_time else "",
        "application_date": str(tx.application_date) if tx.application_date else "",
        "voucher_no": tx.voucher_no or "",
        "debit_amount": tx.debit_amount or 0,
        "credit_amount": tx.credit_amount or 0,
        "balance": tx.balance or 0,
        "counterparty_account": tx.counterparty_account or "",
        "counterparty_name": tx.counterparty_name or "",
        "counterparty_bank": tx.counterparty_bank or "",
        "transaction_serial_no": tx.transaction_serial_no or "",
        "voucher_seq": tx.voucher_seq or "",
        "record_status": tx.record_status or "",
        "summary": tx.summary or "",
        "transaction_remark": tx.transaction_remark or "",
        "account_type": tx.account_type or "",
        # 旧字段（向后兼容）
        "amount": tx.amount or 0,
        "transaction_type": tx.transaction_type,
        "payment_method": tx.payment_method or "",
        "reference_no": tx.reference_no or "",
        "raw_data": tx.raw_data or "{}",
        "remark": tx.remark or "",
        # 凭证号：优先读 DB 存储值（所有凭证生成路径都会回写此字段），动态匹配兜底
        "journal_voucher_no": tx.journal_voucher_no or voucher_map.get(tx.id, ""),
        "created_at": str(tx.created_at) if tx.created_at else ""
    } for tx in txs]


@app.post("/api/bank-transactions")
def create_bank_transaction(data: BankTransactionCreate, company_id: int = Query(...), db: Session = Depends(get_db)):
    tx = BankTransaction(company_id=company_id, **data.model_dump())
    db.add(tx)
    db.commit()
    db.refresh(tx)
    return {"id": tx.id, "message": "银行流水创建成功"}


@app.get("/api/bank-transactions/stats")
def bank_transaction_stats(
    company_id: int = Query(...),
    bank_config_id: Optional[int] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    db: Session = Depends(get_db)
):
    base = db.query(BankTransaction).filter(BankTransaction.company_id == company_id)
    if bank_config_id:
        base = base.filter(BankTransaction.bank_config_id == bank_config_id)
    if date_from:
        base = base.filter(BankTransaction.transaction_date >= date_from)
    if date_to:
        base = base.filter(BankTransaction.transaction_date <= date_to)

    total_count = base.count()
    income_base = base.filter(BankTransaction.transaction_type == "收入")
    expense_base = base.filter(BankTransaction.transaction_type == "支出")

    total_income = income_base.with_entities(func.sum(BankTransaction.credit_amount)).scalar() or 0
    total_expense = expense_base.with_entities(func.sum(BankTransaction.debit_amount)).scalar() or 0
    # 新字段口径：借方=支出, 贷方=收入
    total_debit = base.with_entities(func.sum(BankTransaction.debit_amount)).scalar() or 0
    total_credit = base.with_entities(func.sum(BankTransaction.credit_amount)).scalar() or 0
    income_count = income_base.count()
    expense_count = expense_base.count()

    # 最新余额（取最后一条）
    last_tx = base.order_by(BankTransaction.transaction_date.desc(), BankTransaction.id.desc()).first()
    last_balance = last_tx.balance if last_tx else 0

    return {
        "total_count": total_count,
        "total_income": round(total_income, 2),
        "total_expense": round(total_expense, 2),
        "income_count": income_count,
        "expense_count": expense_count,
        "last_balance": round(last_balance, 2)
    }


class BatchDeleteRequest(BaseModel):
    ids: List[int]


@app.post("/api/bank-transactions/batch-delete")
def batch_delete_bank_transactions(req: BatchDeleteRequest, company_id: int = Query(...), db: Session = Depends(get_db)):
    deleted = db.query(BankTransaction).filter(
        BankTransaction.company_id == company_id,
        BankTransaction.id.in_(req.ids)
    ).delete(synchronize_session=False)
    db.commit()
    return {"message": f"成功删除 {deleted} 条流水记录", "count": deleted}


@app.get("/api/bank-transactions/{tx_id}")
def get_bank_transaction(tx_id: int, company_id: int = Query(...), db: Session = Depends(get_db)):
    tx = db.query(BankTransaction).filter(BankTransaction.company_id == company_id, BankTransaction.id == tx_id).first()
    if not tx:
        raise HTTPException(404, detail="流水记录不存在")
    return {
        "id": tx.id, "bank_config_id": tx.bank_config_id,
        "transaction_date": str(tx.transaction_date) if tx.transaction_date else "",
        "amount": tx.amount or 0, "balance": tx.balance or 0,
        "counterparty_name": tx.counterparty_name or "",
        "counterparty_account": tx.counterparty_account or "",
        "counterparty_bank": tx.counterparty_bank or "",
        "summary": tx.summary or "",
        "transaction_type": tx.transaction_type,
        "payment_method": tx.payment_method or "",
        "voucher_no": tx.voucher_no or "",
        "reference_no": tx.reference_no or "",
        "raw_data": tx.raw_data or "{}",
        "remark": tx.remark or "",
        "created_at": str(tx.created_at) if tx.created_at else ""
    }


@app.put("/api/bank-transactions/{tx_id}")
def update_bank_transaction(tx_id: int, data: BankTransactionUpdate, company_id: int = Query(...), db: Session = Depends(get_db)):
    tx = db.query(BankTransaction).filter(BankTransaction.company_id == company_id, BankTransaction.id == tx_id).first()
    if not tx:
        raise HTTPException(404, detail="流水记录不存在")
    for k, v in data.model_dump(exclude_unset=True).items():
        setattr(tx, k, v)
    db.commit()
    return {"message": "更新成功"}


@app.delete("/api/bank-transactions/{tx_id}")
def delete_bank_transaction(tx_id: int, company_id: int = Query(...), db: Session = Depends(get_db)):
    tx = db.query(BankTransaction).filter(BankTransaction.company_id == company_id, BankTransaction.id == tx_id).first()
    if not tx:
        raise HTTPException(404, detail="流水记录不存在")
    db.delete(tx)
    db.commit()
    return {"message": "删除成功"}


@app.post("/api/bank-transactions/{tx_id}/to-journal")
def bank_transaction_to_journal(tx_id: int, company_id: int = Query(...), db: Session = Depends(get_db)):
    """为单条银行流水生成双分录记账凭证（使用与批量生成相同的智能分类逻辑）"""
    tx = db.query(BankTransaction).filter(
        BankTransaction.id == tx_id,
        BankTransaction.company_id == company_id
    ).first()
    if not tx:
        raise HTTPException(404, "流水记录不存在")

    cp = tx.counterparty_name or tx.summary or "银行流水"
    summary_tag = f"银行流水-#{tx_id}-{cp}"

    # 去重：已生成凭证则跳过
    if tx.journal_voucher_no:
        raise HTTPException(400, f"该流水已生成凭证：{tx.journal_voucher_no}")

    period = tx.transaction_date.strftime("%Y-%m") if tx.transaction_date else datetime.now().strftime("%Y-%m")
    max_no = db.query(JournalEntry.voucher_no).filter(
        JournalEntry.company_id == company_id,
        JournalEntry.period == period,
        JournalEntry.voucher_word == "记"
    ).order_by(JournalEntry.voucher_no.desc()).first()
    next_voucher_no = (max_no[0] + 1) if max_no and max_no[0] else 1
    date_str = tx.transaction_date.strftime("%Y-%m-%d") if tx.transaction_date else period + "-01"

    # 确保科目存在（复用 _generate_bank_journals 的依赖）
    _ensure_account(db, company_id, "1002", "银行存款", "资产", "借")
    _ensure_account(db, company_id, "1122", "应收账款", "资产", "借")
    _ensure_account(db, company_id, "1123", "预付账款", "资产", "借")
    _ensure_account(db, company_id, "2202", "应付账款", "负债", "贷")
    _ensure_account(db, company_id, "1221", "其他应收款", "资产", "借")
    _ensure_account(db, company_id, "2241", "其他应付款", "负债", "贷")
    _ensure_account(db, company_id, "4001", "实收资本", "权益", "贷")
    _ensure_account(db, company_id, "410401", "利润分配-应付股利", "权益", "贷")
    _ensure_account(db, company_id, "221101", "应付职工薪酬-工资", "负债", "贷")

    is_debit = tx.amount is not None and tx.amount < 0
    amount = abs(float(tx.amount) if tx.amount else 0)

    # 使用与批量生成相同的智能分类逻辑（跨实体匹配：股东/人员/客户/供应商）
    result = _classify_bank_tx(db, company_id, tx)
    if result is None:
        raise HTTPException(400, "无法自动分类该银行流水，请完善银行流水规则后再生成凭证")
    cp_code, cp_name, match_type, contact_name = result

    # contact_project：人员匹配时用员工规范姓名，其余用原始对方名称
    contact_proj = contact_name if contact_name else cp

    # 摘要修正
    if match_type == "customer_deposit":
        summary_tag = f"银行流水-#{tx_id}-{contact_proj}（保证金）"
    elif match_type == "prepaid_supplier":
        summary_tag = f"银行流水-#{tx_id}-{contact_proj}（预付供应商，待发票冲销）"

    # 工资匹配上月数据（老邓 2026-06-10）
    if match_type == "salary":
        salary_note = ""
        tx_date_val = tx.transaction_date
        if tx_date_val and hasattr(tx_date_val, 'year'):
            prev_month = tx_date_val.month - 1
            prev_year = tx_date_val.year
            if prev_month == 0:
                prev_month = 12
                prev_year -= 1
            prev_period = f"{prev_year}-{prev_month:02d}"
            salary_records = db.query(SalaryRecord).filter(
                SalaryRecord.company_id == company_id,
                SalaryRecord.period == prev_period
            ).all()
            if salary_records:
                total_net = sum(float(sr.net_salary or 0) for sr in salary_records)
                amt = float(amount)
                if abs(amt - total_net) < 0.02:
                    salary_note = "（已匹配上月工资表）"
                elif amt < total_net - 0.02:
                    salary_note = "（支付<计提，存有工资未发放）"
                else:
                    salary_note = "（支付>计提，可能存在工资未计提）"
            else:
                salary_note = "（无上月工资表数据）"
        summary_tag = f"{summary_tag}{salary_note}"

    if is_debit:
        # 付款：借 对方科目  贷 银行存款
        db.add(JournalEntry(
            company_id=company_id,
            entry_date=datetime.strptime(date_str, "%Y-%m-%d").date(),
            period=period, voucher_word="记", voucher_no=next_voucher_no,
            summary=summary_tag,
            account_code=cp_code, account_name=cp_name,
            debit_amount=amount, credit_amount=0,
            contact_project=contact_proj, source="银行流水", ref_id=tx_id
        ))
        db.add(JournalEntry(
            company_id=company_id,
            entry_date=datetime.strptime(date_str, "%Y-%m-%d").date(),
            period=period, voucher_word="记", voucher_no=next_voucher_no,
            summary=summary_tag,
            account_code="1002", account_name="银行存款",
            debit_amount=0, credit_amount=amount,
            contact_project=contact_proj, source="银行流水", ref_id=tx_id
        ))
    else:
        # 收款：借 银行存款  贷 对方科目
        db.add(JournalEntry(
            company_id=company_id,
            entry_date=datetime.strptime(date_str, "%Y-%m-%d").date(),
            period=period, voucher_word="记", voucher_no=next_voucher_no,
            summary=summary_tag,
            account_code="1002", account_name="银行存款",
            debit_amount=amount, credit_amount=0,
            contact_project=contact_proj, source="银行流水", ref_id=tx_id
        ))
        db.add(JournalEntry(
            company_id=company_id,
            entry_date=datetime.strptime(date_str, "%Y-%m-%d").date(),
            period=period, voucher_word="记", voucher_no=next_voucher_no,
            summary=summary_tag,
            account_code=cp_code, account_name=cp_name,
            debit_amount=0, credit_amount=amount,
            contact_project=contact_proj, source="银行流水", ref_id=tx_id
        ))

    # 老邓 2026-06-10：自动建档供应商/客户（即时创建，不依赖双源）
    entity_name = contact_name or (tx.counterparty_name or "").strip()
    if entity_name and len(entity_name) >= 2:
        _NON_ENTITY = ("手续费", "金库", "公积金", "待处理", "出售凭证", "业务收入", "国家金库", "税务", "国库")
        if not any(kw in entity_name for kw in _NON_ENTITY):
            if match_type in ("supplier", "supplier_invoice", "supplier_payment"):
                norm = _normalize_customer_name(entity_name)
                existing = db.query(Supplier).filter(
                    Supplier.company_id == company_id,
                    Supplier._fingerprint == norm
                ).first()
                if not existing:
                    max_code = db.query(Supplier.code).filter(
                        Supplier.company_id == company_id,
                        Supplier.code.like("GYS%")
                    ).order_by(Supplier.code.desc()).first()
                    next_num = 1
                    if max_code and max_code[0] and max_code[0].startswith("GYS"):
                        try:
                            next_num = int(max_code[0][3:]) + 1
                        except ValueError:
                            pass
                    code = f"GYS{next_num:03d}"
                    db.add(Supplier(
                        company_id=company_id, code=code,
                        name=entity_name, _fingerprint=norm, is_active=True,
                    ))
                    db.flush()
            elif match_type in ("customer", "customer_invoice", "customer_fallback", "customer_deposit", "customer_deposit_refund"):
                norm = _normalize_customer_name(entity_name)
                existing = db.query(Customer).filter(
                    Customer.company_id == company_id,
                    Customer._fingerprint == norm
                ).first()
                if not existing:
                    max_code = db.query(Customer.code).filter(
                        Customer.company_id == company_id,
                        Customer.code.like("KH%")
                    ).order_by(Customer.code.desc()).first()
                    next_num = 1
                    if max_code and max_code[0] and max_code[0].startswith("KH"):
                        try:
                            next_num = int(max_code[0][2:]) + 1
                        except ValueError:
                            pass
                    code = f"KH{next_num:03d}"
                    db.add(Customer(
                        company_id=company_id, code=code,
                        name=entity_name, _fingerprint=norm, is_active=True,
                    ))
                    db.flush()

    voucher_str = f"记-{next_voucher_no}"
    tx.journal_voucher_no = voucher_str
    db.commit()
    return {"message": f"已生成凭证：{voucher_str}（匹配类型：{match_type}，科目：{cp_name}）", "voucher_no": voucher_str, "period": period}


# ==================== 序时账 ====================

class JournalEntryCreate(BaseModel):
    entry_date: date
    period: str
    voucher_word: str = "记"
    voucher_no: int
    attach_count: Optional[int] = 0
    summary: Optional[str] = None
    account_code: str
    account_name: Optional[str] = None
    debit_amount: Optional[float] = 0.0
    credit_amount: Optional[float] = 0.0
    prepared_by: Optional[str] = None
    reviewed_by: Optional[str] = None
    is_reviewed: Optional[bool] = False
    remark: Optional[str] = None
    contact_project: Optional[str] = None
    spec_model: Optional[str] = None
    quantity: Optional[float] = 0.0
    unit: Optional[str] = None
    unit_price: Optional[float] = 0.0


class JournalEntryUpdate(BaseModel):
    entry_date: Optional[date] = None
    period: Optional[str] = None
    voucher_word: Optional[str] = None
    voucher_no: Optional[int] = None
    attach_count: Optional[int] = None
    summary: Optional[str] = None
    account_code: Optional[str] = None
    account_name: Optional[str] = None
    debit_amount: Optional[float] = None
    credit_amount: Optional[float] = None
    prepared_by: Optional[str] = None
    reviewed_by: Optional[str] = None
    is_reviewed: Optional[bool] = None
    remark: Optional[str] = None
    contact_project: Optional[str] = None
    spec_model: Optional[str] = None
    quantity: Optional[float] = None
    unit: Optional[str] = None
    unit_price: Optional[float] = None


@app.get("/api/journal-entries")
def list_journal_entries(
    company_id: int = Query(...),
    period: Optional[str] = None,
    period_from: Optional[str] = None,
    period_to: Optional[str] = None,
    voucher_word: Optional[str] = None,
    keyword: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    is_reviewed: Optional[bool] = None,
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=500),
    db: Session = Depends(get_db)
):
    q = db.query(JournalEntry).filter(JournalEntry.company_id == company_id)
    if period:
        q = q.filter(JournalEntry.period == period)
    if period_from:
        q = q.filter(JournalEntry.period >= period_from)
    if period_to:
        q = q.filter(JournalEntry.period <= period_to)
    if voucher_word:
        q = q.filter(JournalEntry.voucher_word == voucher_word)
    if is_reviewed is not None:
        q = q.filter(JournalEntry.is_reviewed == is_reviewed)
    if date_from:
        q = q.filter(JournalEntry.entry_date >= date_from)
    if date_to:
        q = q.filter(JournalEntry.entry_date <= date_to)
    if keyword:
        q = q.filter(or_(
            JournalEntry.summary.contains(keyword),
            JournalEntry.account_name.contains(keyword),
            JournalEntry.account_code.contains(keyword),
        ))
    total = q.count()

    # 辅助：判断是否同凭证
    def _same_voucher(a, b):
        return a.period == b.period and a.voucher_word == b.voucher_word and a.voucher_no == b.voucher_no

    def _same_voucher_filter(v):
        return and_(
            JournalEntry.period == v.period,
            JournalEntry.voucher_word == v.voucher_word,
            JournalEntry.voucher_no == v.voucher_no,
        )

    entries = q.order_by(JournalEntry.voucher_no.asc(), JournalEntry.id.asc()).offset(skip).limit(limit).all()
    effective_consumed = limit  # 本页在 DB 中实际消耗了多少条记录

    if entries:
        # --- 处理开头：如果首页分录的凭证从前一页延续过来，补全该凭证全部前置分录 ---
        if skip > 0:
            first = entries[0]
            prior_count = q.filter(
                _same_voucher_filter(first),
                JournalEntry.id < first.id,
            ).count()
            if prior_count > 0:
                prior_entries = q.filter(
                    _same_voucher_filter(first),
                    JournalEntry.id < first.id,
                ).order_by(JournalEntry.id.asc()).all()
                entries = prior_entries + entries

        # --- 处理末尾：如果末尾凭证还有分录在下一页 ---
        last = entries[-1]
        in_batch = sum(1 for e in entries if _same_voucher(e, last))
        total_same_voucher = q.filter(_same_voucher_filter(last)).count()
        remaining = total_same_voucher - in_batch

        if remaining > 0:
            if total_same_voucher > limit:
                # 大凭证（分录数 > 单页上限）：取剩余分录补全到本页
                remaining_entries = q.filter(
                    _same_voucher_filter(last),
                    JournalEntry.id > last.id,
                ).order_by(JournalEntry.id.asc()).all()
                entries = entries + remaining_entries
                effective_consumed = limit + len(remaining_entries)
            else:
                # 小凭证：排除该凭证全部分录，推到下一页
                entries = [e for e in entries if not _same_voucher(e, last)]
                effective_consumed = limit - in_batch

    next_skip = skip + effective_consumed

    hierarchy = _build_account_hierarchy(db, company_id)
    return {
        "total": total,
        "next_skip": next_skip,
        "items": [{
            "id": e.id, "entry_date": str(e.entry_date), "period": e.period,
            "voucher_word": e.voucher_word, "voucher_no": e.voucher_no,
            "attach_count": e.attach_count or 0, "summary": e.summary or "",
            "account_code": e.account_code, "account_name": e.account_name or "",
            "account_full_name": hierarchy.get(e.account_code, e.account_name or ""),
            "debit_amount": e.debit_amount or 0, "credit_amount": e.credit_amount or 0,
            "prepared_by": e.prepared_by or "", "reviewed_by": e.reviewed_by or "",
            "is_reviewed": e.is_reviewed, "remark": e.remark or "",
            "contact_project": e.contact_project or "",
            "spec_model": e.spec_model or "",
            "quantity": e.quantity or 0, "unit": e.unit or "",
            "unit_price": e.unit_price or 0,
            "source": e.source or "手动录入",
            "created_at": str(e.created_at) if e.created_at else None,
        } for e in entries]
    }


@app.get("/api/journal-entries/stats")
def journal_entry_stats(
    company_id: int = Query(...),
    period: Optional[str] = None,
    db: Session = Depends(get_db)
):
    base = db.query(JournalEntry).filter(JournalEntry.company_id == company_id)
    if period:
        base = base.filter(JournalEntry.period == period)
    total_count = base.count()
    total_debit = base.with_entities(func.sum(JournalEntry.debit_amount)).scalar() or 0
    total_credit = base.with_entities(func.sum(JournalEntry.credit_amount)).scalar() or 0
    reviewed_count = base.filter(JournalEntry.is_reviewed == True).count()
    unreviewed_count = base.filter(JournalEntry.is_reviewed == False).count()
    period_count = base.with_entities(func.count(func.distinct(JournalEntry.period))).scalar() or 0
    return {
        "total_count": total_count,
        "total_debit": round(total_debit, 2),
        "total_credit": round(total_credit, 2),
        "reviewed_count": reviewed_count,
        "unreviewed_count": unreviewed_count,
        "period_count": period_count,
    }


# ==================== 公用余额计算函数 ====================

def _prev_period(period: str) -> str:
    """计算上一个会计期间。'2025-03' → '2025-02'，'2025-01' → '2024-12'"""
    y, m = map(int, period.split("-"))
    m -= 1
    if m == 0:
        m = 12
        y -= 1
    return f"{y:04d}-{m:02d}"


def _compute_period_balances(company_id: int, period_from, period_to, db) -> dict:
    """
    公用函数：计算指定期间范围内的科目借贷方发生额。
    单数据源：所有报表均通过此函数从序时账取数，避免重复查询。
    period_from: str 如 '2025-01' 或 None（无下限）
    period_to:   str 如 '2025-12' 或 None（无上限）
    返回: {account_code: {"debit": float, "credit": float}}
    """
    q = db.query(JournalEntry).filter(JournalEntry.company_id == company_id)
    if period_from is not None:
        q = q.filter(JournalEntry.period >= period_from)
    if period_to is not None:
        q = q.filter(JournalEntry.period <= period_to)
    result = {}
    for e in q.all():
        c = result.setdefault(e.account_code, {"debit": 0.0, "credit": 0.0})
        c["debit"] += float(e.debit_amount or 0)
        c["credit"] += float(e.credit_amount or 0)
    return result


def _build_trial_balance_tree(company_id, period_raw, cum_raw, db):
    """
    公用函数：基于 period_raw / cum_raw，构建科目余额表的树形汇总结果列表。
    与科目余额表前端返回格式一致。
    """
    accounts = db.query(Account).filter(
        Account.company_id == company_id,
        Account.is_active == True
    ).order_by(Account.code).all()
    acc_map = {a.code: a for a in accounts}

    children_map = {}
    for a in accounts:
        if a.parent_code:
            children_map.setdefault(a.parent_code, []).append(a.code)

    def aggregate(code, data_map):
        total = dict(data_map.get(code, {"debit": 0.0, "credit": 0.0}))
        for child in children_map.get(code, []):
            child_data = aggregate(child, data_map)
            total["debit"] += child_data["debit"]
            total["credit"] += child_data["credit"]
        return total

    period_agg = {a.code: aggregate(a.code, period_raw) for a in accounts}
    cum_agg = {a.code: aggregate(a.code, cum_raw) for a in accounts}

    display_codes = set()
    for a in accounts:
        pt = period_agg[a.code]
        ct = cum_agg[a.code]
        if pt["debit"] != 0 or pt["credit"] != 0 or ct["debit"] != 0 or ct["credit"] != 0:
            current = a.code
            while current:
                display_codes.add(current)
                parent = acc_map[current].parent_code if current in acc_map else None
                current = parent if parent else None

    result = []
    for acc in accounts:
        if acc.code not in display_codes:
            continue
        pt = period_agg[acc.code]
        ct = cum_agg[acc.code]
        pdr = round(pt["debit"], 2)
        pcr = round(pt["credit"], 2)
        cdr = round(ct["debit"], 2)
        ccr = round(ct["credit"], 2)
        direction = acc.balance_direction
        if direction == "借":
            net = round(cdr - ccr, 2)
            end_debit = net if net >= 0 else 0
            end_credit = round(-net, 2) if net < 0 else 0
        else:
            net = round(ccr - cdr, 2)
            end_credit = round(net, 2) if net >= 0 else 0
            end_debit = round(-net, 2) if net < 0 else 0
        result.append({
            "account_code": acc.code,
            "account_name": acc.name,
            "category": acc.category,
            "balance_direction": direction,
            "level": acc.level,
            "parent_code": acc.parent_code,
            "has_children": acc.code in children_map and len(children_map[acc.code]) > 0,
            "begin_debit": 0,
            "begin_credit": 0,
            "period_debit": pdr,
            "period_credit": pcr,
            "cumulative_debit": cdr,
            "cumulative_credit": ccr,
            "end_debit": end_debit,
            "end_credit": end_credit,
        })
    return result


# ==================== 科目余额表 ====================
@app.get("/api/trial-balance")
def trial_balance(
    company_id: int = Query(...),
    period: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """科目余额表：调用统一计算函数"""
    # 本期发生额
    period_raw = _compute_period_balances(company_id, period, period, db)
    # 累计发生额（年初 → 当前期间）
    cum_raw = {}
    if period:
        year = period.split("-")[0]
        cum_raw = _compute_period_balances(company_id, f"{year}-01", period, db)
    else:
        cum_raw = dict(period_raw)

    return _build_trial_balance_tree(company_id, period_raw, cum_raw, db)


# ==================== 总账 ====================
@app.get("/api/ledger/general")
def general_ledger(
    company_id: int = Query(...),
    period_from: str = Query(...),
    period_to: str = Query(...),
    db: Session = Depends(get_db)
):
    """总账：调用统一余额计算函数，树形汇总显示全级次"""
    prev = _prev_period(period_from)
    period_raw  = _compute_period_balances(company_id, period_from, period_to, db)
    cum_raw    = _compute_period_balances(company_id, None, period_to, db)
    open_raw   = _compute_period_balances(company_id, None, prev, db)

    accounts = db.query(Account).filter(
        Account.company_id == company_id,
        Account.is_active == True
    ).order_by(Account.code).all()
    acc_map = {a.code: a for a in accounts}

    # 构建层级名称链（纯名称，不带科目编码）
    def _get_name_chain(acct):
        parts = [acct.name]
        cur = acct
        while cur.parent_code and cur.parent_code in acc_map:
            cur = acc_map[cur.parent_code]
            parts.append(cur.name)
        parts.reverse()
        return " / ".join(parts)
    name_map = {a.code: _get_name_chain(a) for a in accounts}

    # 树形汇总：父级 = 自身 + 所有子级合计
    children_map = {}
    for a in accounts:
        if a.parent_code:
            children_map.setdefault(a.parent_code, []).append(a.code)

    def aggregate(code, data_map):
        total = dict(data_map.get(code, {"debit": 0.0, "credit": 0.0}))
        for child in children_map.get(code, []):
            child_data = aggregate(child, data_map)
            total["debit"] += child_data["debit"]
            total["credit"] += child_data["credit"]
        return total

    period_agg = {a.code: aggregate(a.code, period_raw) for a in accounts}
    cum_agg = {a.code: aggregate(a.code, cum_raw) for a in accounts}

    # 全级次过滤：聚合后有数据的科目 + 其所有父级链
    display_codes = set()
    for a in accounts:
        pt = period_agg[a.code]
        ct = cum_agg[a.code]
        if pt["debit"] != 0 or pt["credit"] != 0 or ct["debit"] != 0 or ct["credit"] != 0:
            current = a.code
            while current:
                display_codes.add(current)
                parent = acc_map[current].parent_code if current in acc_map else None
                current = parent if parent else None

    result = []
    for acc in accounts:
        if acc.code not in display_codes:
            continue
        p = period_agg[acc.code]
        c = cum_agg[acc.code]
        o = open_raw.get(acc.code, {"debit": 0.0, "credit": 0.0})
        direction = acc.balance_direction or "借"
        # 期初余额：从前期累计发生额推算
        if direction == "借":
            ob = round(o["debit"] - o["credit"], 2)
        else:
            ob = round(o["credit"] - o["debit"], 2)
        # 期末余额
        if direction == "借":
            balance = round(c["debit"] - c["credit"], 2)
        else:
            balance = round(c["credit"] - c["debit"], 2)
        # 期初方向：余额>0与科目方向一致，<0相反
        if ob >= 0:
            opening_direction = direction
        else:
            opening_direction = "贷" if direction == "借" else "借"
        result.append({
            "account_code": acc.code,
            "account_name": name_map.get(acc.code, acc.name),
            "level": acc.level,
            "opening_balance": round(ob, 2),
            "opening_direction": opening_direction,
            "total_debit": round(p["debit"], 2),
            "total_credit": round(p["credit"], 2),
            "end_balance": balance,
            "end_direction": direction,
        })
    return result


# ==================== 明细账 ====================
@app.get("/api/ledger/detail")
def detail_ledger(
    company_id: int = Query(...),
    account_code: str = Query(...),
    period_from: str = Query(...),
    period_to: str = Query(...),
    db: Session = Depends(get_db)
):
    """明细账：调用统一余额计算函数获取期初余额，交易明细仍从序时账取"""
    account = db.query(Account).filter(
        Account.company_id == company_id,
        Account.code == account_code
    ).first()
    if not account:
        raise HTTPException(status_code=404, detail="科目不存在")

    # 期初余额 = 截止上期期末的累计净额
    prev = _prev_period(period_from)
    opening_raw = _compute_period_balances(company_id, None, prev, db)
    ob = opening_raw.get(account_code, {"debit": 0.0, "credit": 0.0})
    if account.balance_direction == "借":
        opening_balance = round(ob["debit"] - ob["credit"], 2)
    else:
        opening_balance = round(ob["credit"] - ob["debit"], 2)

    # 本期交易明细（仍需逐笔，无法从余额表获取）
    entries = db.query(JournalEntry).filter(
        JournalEntry.company_id == company_id,
        JournalEntry.account_code == account_code,
        JournalEntry.period >= period_from,
        JournalEntry.period <= period_to
    ).order_by(JournalEntry.entry_date, JournalEntry.voucher_no, JournalEntry.id).all()

    rows = []
    balance = float(opening_balance)
    for e in entries:
        dr = float(e.debit_amount or 0)
        cr = float(e.credit_amount or 0)
        if account.balance_direction == "借":
            balance += dr - cr
        else:
            balance += cr - dr
        rows.append({
            "voucher_date": str(e.entry_date) if e.entry_date else "",
            "voucher_no": (e.voucher_word or '记') + '-' + str(e.voucher_no).zfill(4) if e.voucher_no else "",
            "summary": e.summary or "",
            "debit_amount": dr,
            "credit_amount": cr,
            "balance": round(balance, 2),
        })

    return {
        "account_code": account.code,
        "account_name": account.name,
        "balance_direction": account.balance_direction,
        "opening_balance": round(opening_balance, 2),
        "rows": rows,
    }


# ==================== 往来明细账（人员/客户/供应商） ====================

# 往来科目映射（每类仅用一个主科目）
_CONTACT_ACCOUNTS = {
    "employee": ["1221"],   # 其他应收款（人员）
    "customer": ["1122"],   # 应收账款（客户）
    "supplier": ["2202"],   # 应付账款（供应商）
}


def _sub_ledger_by_contact(company_id: int, account_codes: list, contact_name: str,
                           period_from: str, period_to: str, db: Session):
    """共用往来明细账计算函数

    返回：{ contact_name, opening_balance, rows: [{date, voucher_no, summary, account_code, account_name, debit, credit, balance}] }
    """
    entries_all = db.query(JournalEntry).filter(
        JournalEntry.company_id == company_id,
        JournalEntry.account_code.in_(account_codes),
        JournalEntry.contact_project == contact_name,
        JournalEntry.period <= period_to
    ).order_by(JournalEntry.entry_date, JournalEntry.voucher_no, JournalEntry.id).all()

    # 期初余额：period_from 之前的累计净额
    opening_balance = 0.0
    for e in entries_all:
        if e.period < period_from:
            opening_balance += float(e.debit_amount or 0) - float(e.credit_amount or 0)

    # 本期明细
    rows = []
    balance = opening_balance
    for e in entries_all:
        if e.period < period_from:
            continue
        dr = float(e.debit_amount or 0)
        cr = float(e.credit_amount or 0)
        balance += dr - cr
        rows.append({
            "voucher_date": str(e.entry_date) if e.entry_date else "",
            "voucher_no": (e.voucher_word or '记') + '-' + str(e.voucher_no).zfill(4) if e.voucher_no else "",
            "summary": e.summary or "",
            "account_code": e.account_code,
            "account_name": e.account_name or "",
            "debit_amount": dr,
            "credit_amount": cr,
            "balance": round(balance, 2),
        })

    return {
        "contact_name": contact_name,
        "opening_balance": round(opening_balance, 2),
        "rows": rows,
    }


def _contact_list(company_id: int, account_codes: list, db: Session):
    """提取往来项目列表（从序时账 contact_project 中汇总）"""
    results = db.query(
        JournalEntry.contact_project,
        func.sum(JournalEntry.debit_amount).label("total_debit"),
        func.sum(JournalEntry.credit_amount).label("total_credit"),
    ).filter(
        JournalEntry.company_id == company_id,
        JournalEntry.account_code.in_(account_codes),
        JournalEntry.contact_project.isnot(None),
        JournalEntry.contact_project != "",
    ).group_by(JournalEntry.contact_project).all()

    contacts = []
    for r in results:
        name = r[0]
        td = round(r[1] or 0, 2)
        tc = round(r[2] or 0, 2)
        contacts.append({
            "name": name,
            "total_debit": td,
            "total_credit": tc,
            "net": round(td - tc, 2),
        })
    contacts.sort(key=lambda c: c["name"])
    return contacts


@app.get("/api/ledger/employee-contacts")
def employee_contacts(company_id: int = Query(...), db: Session = Depends(get_db)):
    return _contact_list(company_id, _CONTACT_ACCOUNTS["employee"], db)


@app.get("/api/ledger/customer-contacts")
def customer_contacts(company_id: int = Query(...), db: Session = Depends(get_db)):
    return _contact_list(company_id, _CONTACT_ACCOUNTS["customer"], db)


@app.get("/api/ledger/supplier-contacts")
def supplier_contacts(company_id: int = Query(...), db: Session = Depends(get_db)):
    return _contact_list(company_id, _CONTACT_ACCOUNTS["supplier"], db)


@app.get("/api/ledger/employee-detail")
def employee_detail(
    company_id: int = Query(...),
    contact_name: str = Query(...),
    period_from: str = Query(...),
    period_to: str = Query(...),
    db: Session = Depends(get_db)
):
    return _sub_ledger_by_contact(company_id, _CONTACT_ACCOUNTS["employee"], contact_name, period_from, period_to, db)


@app.get("/api/ledger/customer-detail")
def customer_detail(
    company_id: int = Query(...),
    contact_name: str = Query(...),
    period_from: str = Query(...),
    period_to: str = Query(...),
    db: Session = Depends(get_db)
):
    return _sub_ledger_by_contact(company_id, _CONTACT_ACCOUNTS["customer"], contact_name, period_from, period_to, db)


@app.get("/api/ledger/supplier-detail")
def supplier_detail(
    company_id: int = Query(...),
    contact_name: str = Query(...),
    period_from: str = Query(...),
    period_to: str = Query(...),
    db: Session = Depends(get_db)
):
    return _sub_ledger_by_contact(company_id, _CONTACT_ACCOUNTS["supplier"], contact_name, period_from, period_to, db)


# ==================== 利润表（企业会计准则一般企业—会企02号） ====================
def _pl_net(balances, code_prefix, is_credit_nature=True):
    """汇总指定前缀科目的净额：收入/收益类=贷-借，费用/损失类=借-贷"""
    total_dr = 0.0; total_cr = 0.0
    for code, bal in balances.items():
        if code.startswith(code_prefix):
            total_dr += bal["debit"]; total_cr += bal["credit"]
    return round(total_cr - total_dr, 2) if is_credit_nature else round(total_dr - total_cr, 2)

def _pl_row(label, current=0.0, prior=0.0, bold=False, highlight=False, indent=0):
    return {"label": label, "current": current, "prior": prior, "bold": bold, "highlight": highlight, "indent": indent}

def _build_pl(company_id, from_period, to_period, db):
    """构建单期利润表数据"""
    b = _compute_period_balances(company_id, from_period, to_period, db)
    # 一、营业收入
    rev = _pl_net(b, "6001") + _pl_net(b, "6051")
    cost = _pl_net(b, "6401", False) + _pl_net(b, "6402", False)
    tax_sur = _pl_net(b, "6403", False)
    sell_exp = _pl_net(b, "6601", False)
    admin_exp = _pl_net(b, "6602", False)
    rd_exp = _pl_net(b, "6604", False)
    fin_exp = _pl_net(b, "6603", False)
    inv_inc = _pl_net(b, "6111")
    credit_loss = _pl_net(b, "6701", False)
    asset_impair = _pl_net(b, "6702", False)
    asset_disp = _pl_net(b, "6712")  # 资产处置收益（贷余）
    other_inc = _pl_net(b, "6301")
    other_exp = _pl_net(b, "6711", False)
    income_tax = _pl_net(b, "6801", False)
    # 中间计算
    gross_p = round(rev - cost - tax_sur, 2)
    # 营业利润 = 毛利 - 期间费用 + 投资收益 + 资产处置收益 - 减值损失
    # 注：营业外收入(6301)/营业外支出(6711)不属于营业利润，在利润总额中加减
    op_p = round(gross_p - sell_exp - admin_exp - rd_exp - fin_exp + inv_inc + asset_disp - credit_loss - asset_impair, 2)
    total_p = round(op_p + other_inc - other_exp, 2)
    net_p = round(total_p - income_tax, 2)
    items = [
        _pl_row("一、营业收入", rev, bold=True),
        _pl_row("  减：营业成本", cost, indent=1),
        _pl_row("  减：税金及附加", tax_sur, indent=1),
        _pl_row("  减：销售费用", sell_exp, indent=1),
        _pl_row("  减：管理费用", admin_exp, indent=1),
        _pl_row("  减：研发费用", rd_exp, indent=1),
        _pl_row("  减：财务费用", fin_exp, indent=1),
        _pl_row("  加：其他收益", 0.0, indent=1),
        _pl_row("  加：投资收益", inv_inc, indent=1),
        _pl_row("  加：资产处置收益", asset_disp, indent=1),
        _pl_row("  减：信用减值损失", credit_loss, indent=1),
        _pl_row("  减：资产减值损失", asset_impair, indent=1),
        _pl_row("二、营业利润", op_p, bold=True, highlight=True),
        _pl_row("  加：营业外收入", other_inc, indent=1),
        _pl_row("  减：营业外支出", other_exp, indent=1),
        _pl_row("三、利润总额", total_p, bold=True, highlight=True),
        _pl_row("  减：所得税费用", income_tax, indent=1),
        _pl_row("四、净利润", net_p, bold=True, highlight=True),
        _pl_row("五、其他综合收益的税后净额", 0.0, bold=True),
        _pl_row("六、综合收益总额", net_p, bold=True, highlight=True),
        _pl_row("七、每股收益", 0.0, bold=True),
        _pl_row("  （一）基本每股收益", 0.0, indent=1),
        _pl_row("  （二）稀释每股收益", 0.0, indent=1),
    ]
    return items

def _prior_same_period(period_from: str, period_to: str):
    """计算上年同期：如 2026-01→2026-03 → 2025-01→2025-03"""
    yf, mf = map(int, period_from.split("-"))
    yt, mt = map(int, period_to.split("-"))
    return f"{yf-1}-{mf:02d}", f"{yt-1}-{mt:02d}"

@app.get("/api/reports/profit-loss")
def profit_loss_report(
    company_id: int = Query(...),
    period_from: str = Query(...),
    period_to: str = Query(...),
    db: Session = Depends(get_db)
):
    """利润表（会企02号）：本期金额 + 上期金额"""
    current_items = _build_pl(company_id, period_from, period_to, db)
    prior_from, prior_to = _prior_same_period(period_from, period_to)
    prior_items = _build_pl(company_id, prior_from, prior_to, db)
    prior_map = {it["label"]: it["current"] for it in prior_items}
    for it in current_items:
        it["prior"] = prior_map.get(it["label"], 0.0)
    return {"items": current_items, "period_from": period_from, "period_to": period_to}


# ==================== 资产负债表（企业会计准则一般企业—会企01号） ====================
def _bs_year_begin(period: str):
    """年初期间：2026-03 → 2025-12"""
    y = int(period.split("-")[0])
    return f"{y-1}-12"

def _opening_balance_dict(company_id: int, db: Session):
    """将会计科目的期初金额转为 _bs_net 可用的 balances 字典格式"""
    accounts = db.query(Account).filter(
        Account.company_id == company_id,
        Account.is_active == True
    ).all()
    result = {}
    for a in accounts:
        ob = a.opening_balance or 0
        if a.balance_direction == "借":
            if ob >= 0:
                result[a.code] = {"debit": ob, "credit": 0}
            else:
                result[a.code] = {"debit": 0, "credit": abs(ob)}
        else:
            if ob >= 0:
                result[a.code] = {"debit": 0, "credit": ob}
            else:
                result[a.code] = {"debit": abs(ob), "credit": 0}
    return result

def _bs_net(balances, code_prefix, is_debit_nature=True):
    """资产类=借-贷，负债/权益类=贷-借"""
    total_dr = 0.0; total_cr = 0.0
    for code, bal in balances.items():
        if code.startswith(code_prefix):
            total_dr += bal["debit"]; total_cr += bal["credit"]
    return round(total_dr - total_cr, 2) if is_debit_nature else round(total_cr - total_dr, 2)

def _bs_row(label, end=0.0, begin=0.0, bold=False, highlight=False, indent=0):
    return {"label": label, "end": end, "begin": begin, "bold": bold, "highlight": highlight, "indent": indent}

def _build_bs_side(balances, side):
    """构建资产负债表一侧（资产 或 负债+权益）"""
    r = _bs_row
    b = balances
    if side == "assets":
        # 流动资产
        cash = _bs_net(b, "1001") + _bs_net(b, "1002") + _bs_net(b, "1003")
        fin_asset = _bs_net(b, "1101")
        notes_recv = _bs_net(b, "1121")
        ar = _bs_net(b, "1122")
        ar_fin = _bs_net(b, "1124")
        prepay = _bs_net(b, "1123")
        other_recv = _bs_net(b, "1221")
        inventory = _bs_net(b, "1403") + _bs_net(b, "1405") + _bs_net(b, "1406") + _bs_net(b, "1408") + _bs_net(b, "1411")
        contract_asset = _bs_net(b, "1401")
        held_for_sale_a = _bs_net(b, "1501")
        noncurr_due_1y = _bs_net(b, "1502")
        other_current_a = _bs_net(b, "1503")
        total_current = round(cash + fin_asset + notes_recv + ar + ar_fin + prepay + other_recv + inventory + contract_asset + held_for_sale_a + noncurr_due_1y + other_current_a, 2)
        # 非流动资产
        debt_inv = _bs_net(b, "1504")
        other_debt_inv = _bs_net(b, "1505")
        lt_recv = _bs_net(b, "1511")
        lt_equity = _bs_net(b, "1512")
        other_equity = _bs_net(b, "1513")
        other_nc_fin = _bs_net(b, "1514")
        invest_prop = _bs_net(b, "1521")
        fixed_asset = _bs_net(b, "1601")
        accum_depr = _bs_net(b, "1602", False)
        cip = _bs_net(b, "1604")
        bio_asset = _bs_net(b, "1621")
        oil_gas = _bs_net(b, "1631")
        rou_asset = _bs_net(b, "1641")
        intangible = _bs_net(b, "1701")
        dev_exp = _bs_net(b, "1702")
        goodwill = _bs_net(b, "1711")
        lt_deferred = _bs_net(b, "1801")
        def_tax_asset = _bs_net(b, "1811")
        other_nc_a = _bs_net(b, "1901")
        total_nc = round(debt_inv + other_debt_inv + lt_recv + lt_equity + other_equity + other_nc_fin + invest_prop + (fixed_asset - accum_depr) + cip + bio_asset + oil_gas + rou_asset + intangible + dev_exp + goodwill + lt_deferred + def_tax_asset + other_nc_a, 2)
        total_assets = round(total_current + total_nc, 2)
        return [
            r("流动资产：", bold=True),
            r("  货币资金", cash, indent=1), r("  交易性金融资产", fin_asset, indent=1),
            r("  应收票据", notes_recv, indent=1), r("  应收账款", ar, indent=1),
            r("  应收款项融资", ar_fin, indent=1), r("  预付款项", prepay, indent=1),
            r("  其他应收款", other_recv, indent=1), r("  存货", inventory, indent=1),
            r("  合同资产", contract_asset, indent=1), r("  持有待售资产", held_for_sale_a, indent=1),
            r("  一年内到期的非流动资产", noncurr_due_1y, indent=1), r("  其他流动资产", other_current_a, indent=1),
            r("流动资产合计", total_current, bold=True, highlight=True),
            r("非流动资产：", bold=True),
            r("  债权投资", debt_inv, indent=1), r("  其他债权投资", other_debt_inv, indent=1),
            r("  长期应收款", lt_recv, indent=1), r("  长期股权投资", lt_equity, indent=1),
            r("  其他权益工具投资", other_equity, indent=1), r("  其他非流动金融资产", other_nc_fin, indent=1),
            r("  投资性房地产", invest_prop, indent=1),
            r("  固定资产", round(fixed_asset - accum_depr, 2) if fixed_asset else 0.0, indent=1),
            r("  在建工程", cip, indent=1), r("  生产性生物资产", bio_asset, indent=1),
            r("  使用权资产", rou_asset, indent=1), r("  无形资产", intangible, indent=1),
            r("  开发支出", dev_exp, indent=1), r("  商誉", goodwill, indent=1),
            r("  长期待摊费用", lt_deferred, indent=1), r("  递延所得税资产", def_tax_asset, indent=1),
            r("  其他非流动资产", other_nc_a, indent=1),
            r("", 0), r("", 0),
            r("非流动资产合计", total_nc, bold=True, highlight=True),
            r("资产总计", total_assets, bold=True, highlight=True),
        ]
    else:
        # 流动负债
        st_loan = _bs_net(b, "2001", False)
        fin_liab = _bs_net(b, "2101", False)
        notes_pay = _bs_net(b, "2201", False)
        ap = _bs_net(b, "2202", False)
        advance_rcv = _bs_net(b, "2203", False)
        contract_liab = _bs_net(b, "2204", False)
        payroll = _bs_net(b, "2211", False)
        taxes = _bs_net(b, "2210", False)
        other_pay = _bs_net(b, "2241", False)
        held_for_sale_l = _bs_net(b, "2242", False)
        nc_due_1y_l = _bs_net(b, "2243", False)
        other_current_l = _bs_net(b, "2244", False)
        total_current_l = round(st_loan + fin_liab + notes_pay + ap + advance_rcv + contract_liab + payroll + taxes + other_pay + held_for_sale_l + nc_due_1y_l + other_current_l, 2)
        # 非流动负债
        lt_loan = _bs_net(b, "2501", False)
        bonds_pay = _bs_net(b, "2502", False)
        lease_liab = _bs_net(b, "2601", False)
        lt_pay = _bs_net(b, "2701", False)
        estimated_liab = _bs_net(b, "2801", False)
        deferred_inc = _bs_net(b, "2901", False)
        def_tax_liab = _bs_net(b, "2902", False)
        other_nc_l = _bs_net(b, "2903", False)
        total_nc_l = round(lt_loan + bonds_pay + lease_liab + lt_pay + estimated_liab + deferred_inc + def_tax_liab + other_nc_l, 2)
        total_liab = round(total_current_l + total_nc_l, 2)
        # 所有者权益
        paid_in = _bs_net(b, "4001", False)
        other_equity_instr = _bs_net(b, "4002", False)
        capital_surplus = _bs_net(b, "4003", False)
        treasury_stock = _bs_net(b, "4004")
        oci = _bs_net(b, "4005", False)
        special_reserve = _bs_net(b, "4101", False)
        surplus = _bs_net(b, "4103", False)
        retained = round(_bs_net(b, "4104", False) + _bs_net(b, "4103", False), 2)
        total_equity = round(paid_in + other_equity_instr + capital_surplus - treasury_stock + oci + special_reserve + surplus + retained, 2)
        total_right = round(total_liab + total_equity, 2)
        return [
            r("流动负债：", bold=True),
            r("  短期借款", st_loan, indent=1), r("  交易性金融负债", fin_liab, indent=1),
            r("  应付票据", notes_pay, indent=1), r("  应付账款", ap, indent=1),
            r("  预收款项", advance_rcv, indent=1), r("  合同负债", contract_liab, indent=1),
            r("  应付职工薪酬", payroll, indent=1), r("  应交税费", taxes, indent=1),
            r("  其他应付款", other_pay, indent=1), r("  持有待售负债", held_for_sale_l, indent=1),
            r("  一年内到期的非流动负债", nc_due_1y_l, indent=1), r("  其他流动负债", other_current_l, indent=1),
            r("流动负债合计", total_current_l, bold=True, highlight=True),
            r("非流动负债：", bold=True),
            r("  长期借款", lt_loan, indent=1), r("  应付债券", bonds_pay, indent=1),
            r("  租赁负债", lease_liab, indent=1), r("  长期应付款", lt_pay, indent=1),
            r("  预计负债", estimated_liab, indent=1), r("  递延收益", deferred_inc, indent=1),
            r("  递延所得税负债", def_tax_liab, indent=1), r("  其他非流动负债", other_nc_l, indent=1),
            r("非流动负债合计", total_nc_l, bold=True, highlight=True),
            r("负债合计", total_liab, bold=True, highlight=True),
            r("所有者权益（或股东权益）：", bold=True),
            r("  实收资本（或股本）", paid_in, indent=1), r("  其他权益工具", other_equity_instr, indent=1),
            r("  资本公积", capital_surplus, indent=1), r("  减：库存股", treasury_stock, indent=1),
            r("  其他综合收益", oci, indent=1), r("  专项储备", special_reserve, indent=1),
            r("  盈余公积", surplus, indent=1), r("  未分配利润", retained, indent=1),
            r("所有者权益合计", total_equity, bold=True, highlight=True),
            r("负债和所有者权益总计", total_right, bold=True, highlight=True),
        ]

@app.get("/api/reports/balance-sheet")
def balance_sheet_report(
    company_id: int = Query(...),
    period: str = Query(...),
    db: Session = Depends(get_db)
):
    """资产负债表（会企01号）：期末余额 + 年初余额"""
    end_balances = _compute_period_balances(company_id, None, period, db)
    # 年初余额根据会计科目的期初金额确定
    begin_balances = _opening_balance_dict(company_id, db)
    assets = _build_bs_side(end_balances, "assets")
    liab_eq = _build_bs_side(end_balances, "liab_eq")
    # 年初余额单独计算
    assets_begin = _build_bs_side(begin_balances, "assets")
    liab_eq_begin = _build_bs_side(begin_balances, "liab_eq")
    begin_map_a = {r["label"]: r["end"] for r in assets_begin}
    begin_map_le = {r["label"]: r["end"] for r in liab_eq_begin}
    for r in assets:
        r["begin"] = begin_map_a.get(r["label"], 0.0)
    for r in liab_eq:
        r["begin"] = begin_map_le.get(r["label"], 0.0)
    return {"assets": assets, "liabilities_equity": liab_eq, "period": period}


# ==================== 现金流量表（企业会计准则一般企业—会企03号） ====================
def _prior_period_year(period: str):
    """上年同期：2026 → 2025"""
    y = int(period.split("-")[0])
    return str(y - 1)

def _cf_net_cash_by_accounts(company_id, period_from, period_to, cash_codes, db, inflow=True):
    """计算涉及现金科目的对方科目发生额（直接法）— 使用SQL聚合"""
    from sqlalchemy import func
    entries = db.query(JournalEntry).filter(
        JournalEntry.company_id == company_id,
        JournalEntry.period >= period_from,
        JournalEntry.period <= period_to,
        JournalEntry.voucher_no.in_(
            db.query(JournalEntry.voucher_no).filter(
                JournalEntry.company_id == company_id,
                JournalEntry.period >= period_from,
                JournalEntry.period <= period_to,
                JournalEntry.account_code.startswith(cash_codes[0])
            ).union(
                *[db.query(JournalEntry.voucher_no).filter(
                    JournalEntry.company_id == company_id,
                    JournalEntry.period >= period_from,
                    JournalEntry.period <= period_to,
                    JournalEntry.account_code.startswith(c)
                ) for c in cash_codes[1:]]
            )
        )
    ).all()
    # 按凭证号分组
    vouchers = {}
    for e in entries:
        vouchers.setdefault(e.voucher_no, []).append(e)
    total = 0.0
    for vno, lines in vouchers.items():
        for l in lines:
            if l.account_code and any(l.account_code.startswith(c) for c in cash_codes):
                if inflow:
                    total += float(l.credit_amount or 0)  # 现金流入：贷现金
                else:
                    total += float(l.debit_amount or 0)  # 现金流出：借现金
    return round(total, 2)


def _cf_op_classified(company_id, period_from, period_to, cash_codes, activity_codes, db, is_inflow=True):
    """按对方科目对经营现金流分类（SQL优化版）"""
    cash_cond = or_(*[JournalEntry.account_code.startswith(c) for c in cash_codes])
    activity_cond = or_(*[JournalEntry.account_code.startswith(a) for a in activity_codes])

    cash_vnos = db.query(JournalEntry.voucher_no).filter(
        JournalEntry.company_id == company_id,
        JournalEntry.period >= period_from,
        JournalEntry.period <= period_to,
        cash_cond
    ).distinct().subquery()

    activity_vnos = db.query(JournalEntry.voucher_no).filter(
        JournalEntry.company_id == company_id,
        JournalEntry.period >= period_from,
        JournalEntry.period <= period_to,
        activity_cond
    ).distinct().subquery()

    target_vnos = db.query(cash_vnos.c.voucher_no).join(
        activity_vnos, cash_vnos.c.voucher_no == activity_vnos.c.voucher_no
    ).subquery()

    if is_inflow:
        amt = func.coalesce(JournalEntry.credit_amount, 0)
    else:
        amt = func.coalesce(JournalEntry.debit_amount, 0)

    total = db.query(func.coalesce(func.sum(amt), 0)).filter(
        JournalEntry.company_id == company_id,
        JournalEntry.period >= period_from,
        JournalEntry.period <= period_to,
        JournalEntry.voucher_no.in_(db.query(target_vnos.c.voucher_no)),
        cash_cond
    ).scalar()

    return round(float(total or 0), 2)


def _cf_activity(company_id, period_from, period_to, cash_codes, activity_codes, db, is_inflow=True):
    """按对方科目分类计算特定活动的现金流量（SQL优化版）
    activity_codes: 对方科目前缀列表（如投资活动的固定资产科目）
    is_inflow: True=流入, False=流出
    """
    cash_cond = or_(*[JournalEntry.account_code.startswith(c) for c in cash_codes])
    activity_cond = or_(*[JournalEntry.account_code.startswith(a) for a in activity_codes])

    # 子查询：同时涉及现金科目和活动科目的凭证号
    cash_vnos = db.query(JournalEntry.voucher_no).filter(
        JournalEntry.company_id == company_id,
        JournalEntry.period >= period_from,
        JournalEntry.period <= period_to,
        cash_cond
    ).distinct().subquery()

    activity_vnos = db.query(JournalEntry.voucher_no).filter(
        JournalEntry.company_id == company_id,
        JournalEntry.period >= period_from,
        JournalEntry.period <= period_to,
        activity_cond
    ).distinct().subquery()

    target_vnos = db.query(cash_vnos.c.voucher_no).join(
        activity_vnos, cash_vnos.c.voucher_no == activity_vnos.c.voucher_no
    ).subquery()

    if is_inflow:
        amt = func.coalesce(JournalEntry.credit_amount, 0)
    else:
        amt = func.coalesce(JournalEntry.debit_amount, 0)

    total = db.query(func.coalesce(func.sum(amt), 0)).filter(
        JournalEntry.company_id == company_id,
        JournalEntry.period >= period_from,
        JournalEntry.period <= period_to,
        JournalEntry.voucher_no.in_(db.query(target_vnos.c.voucher_no)),
        cash_cond
    ).scalar()

    return round(float(total or 0), 2)


@app.get("/api/reports/cash-flow")
def cash_flow_report(
    company_id: int = Query(...),
    period_from: str = Query(...),
    period_to: str = Query(...),
    db: Session = Depends(get_db)
):
    """现金流量表（会企03号）：直接法"""
    cash_codes = ["1001", "1002", "1003"]  # 库存现金、银行存款、其他货币资金

    def cf_row(label, current=0.0, prior=0.0, bold=False, highlight=False, indent=0):
        return {"label": label, "current": current, "prior": prior, "bold": bold, "highlight": highlight, "indent": indent}

    # 期初/期末现金余额
    begin_period = period_from[:4] + "-01"
    balances_end = _compute_period_balances(company_id, None, period_to, db)
    balances_begin = _compute_period_balances(company_id, None, _prev_period(begin_period), db)
    cash_end = sum(_bs_net(balances_end, c) for c in cash_codes)
    cash_begin = sum(_bs_net(balances_begin, c) for c in cash_codes)

    # 经营活动 — 按对方科目精细分类（直接法）
    # 销售商品、提供劳务收到的现金：现金流入 + 凭证中涉及收入/应收科目
    revenue_codes = ["6001", "6051", "1122", "1123"]
    sales_cash = _cf_op_classified(company_id, period_from, period_to, cash_codes, revenue_codes, db, is_inflow=True)
    # 购买商品、接受劳务支付的现金：现金流出 + 凭证中涉及成本/存货/应付科目
    purchase_codes = ["1401", "1402", "1403", "1404", "1405", "1406", "1407", "1408", "6401", "6402", "6403", "2202"]
    purchase_cash = _cf_op_classified(company_id, period_from, period_to, cash_codes, purchase_codes, db, is_inflow=False)
    # 支付给职工以及为职工支付的现金：现金流出 + 凭证中涉及应付职工薪酬
    employee_codes = ["221101", "221102", "221103", "221104", "221105", "221106", "221107", "221108"]
    employee_cash = _cf_op_classified(company_id, period_from, period_to, cash_codes, employee_codes, db, is_inflow=False)
    # 支付的各项税费：现金流出 + 凭证中涉及应交税费
    tax_codes = ["221009", "221010", "221011", "221012", "221013", "221014", "221015"]
    tax_cash = _cf_op_classified(company_id, period_from, period_to, cash_codes, tax_codes, db, is_inflow=False)
    # 其他：总经营现金流中扣除以上各项
    total_inflow = _cf_net_cash_by_accounts(company_id, period_from, period_to, cash_codes, db, inflow=True)
    total_outflow = _cf_net_cash_by_accounts(company_id, period_from, period_to, cash_codes, db, inflow=False)

    # 投资/筹资活动
    invest_codes = ["1601", "1602", "1604", "1701", "1702", "1511", "1512"]
    invest_inflow = _cf_activity(company_id, period_from, period_to, cash_codes, invest_codes, db, is_inflow=True)
    invest_outflow = _cf_activity(company_id, period_from, period_to, cash_codes, invest_codes, db, is_inflow=False)

    finance_codes = ["4001", "4002", "2001", "2501", "2701"]
    finance_inflow = _cf_activity(company_id, period_from, period_to, cash_codes, finance_codes, db, is_inflow=True)
    finance_outflow = _cf_activity(company_id, period_from, period_to, cash_codes, finance_codes, db, is_inflow=False)

    # 经营项中扣除投资/筹资的现金部分，再从中扣出已分类的，剩余为"其他"
    op_inflow = round(total_inflow - invest_inflow - finance_inflow, 2)
    op_outflow = round(total_outflow - invest_outflow - finance_outflow, 2)
    other_op_inflow = round(op_inflow - sales_cash, 2)
    other_op_outflow = round(op_outflow - purchase_cash - employee_cash - tax_cash, 2)
    op_net = round(op_inflow - op_outflow, 2)
    invest_net = round(invest_inflow - invest_outflow, 2)
    finance_net = round(finance_inflow - finance_outflow, 2)
    total_net = round(op_net + invest_net + finance_net, 2)

    items = [
        cf_row("一、经营活动产生的现金流量：", bold=True),
        cf_row("  销售商品、提供劳务收到的现金", sales_cash, indent=1),
        cf_row("  收到的税费返还", 0.0, indent=1),
        cf_row("  收到其他与经营活动有关的现金", other_op_inflow, indent=1),
        cf_row("经营活动现金流入小计", op_inflow, bold=True, highlight=True),
        cf_row("  购买商品、接受劳务支付的现金", purchase_cash, indent=1),
        cf_row("  支付给职工以及为职工支付的现金", employee_cash, indent=1),
        cf_row("  支付的各项税费", tax_cash, indent=1),
        cf_row("  支付其他与经营活动有关的现金", other_op_outflow, indent=1),
        cf_row("经营活动现金流出小计", op_outflow, bold=True, highlight=True),
        cf_row("经营活动产生的现金流量净额", op_net, bold=True, highlight=True),
        cf_row("二、投资活动产生的现金流量：", bold=True),
        cf_row("  收回投资收到的现金", 0.0, indent=1),
        cf_row("  取得投资收益收到的现金", 0.0, indent=1),
        cf_row("  处置固定资产、无形资产收回的现金净额", invest_inflow, indent=1),
        cf_row("  处置子公司及其他营业单位收到的现金净额", 0.0, indent=1),
        cf_row("  收到其他与投资活动有关的现金", 0.0, indent=1),
        cf_row("投资活动现金流入小计", invest_inflow, bold=True, highlight=True),
        cf_row("  购建固定资产、无形资产支付的现金", invest_outflow, indent=1),
        cf_row("  投资支付的现金", 0.0, indent=1),
        cf_row("  取得子公司及其他营业单位支付的现金净额", 0.0, indent=1),
        cf_row("  支付其他与投资活动有关的现金", 0.0, indent=1),
        cf_row("投资活动现金流出小计", invest_outflow, bold=True, highlight=True),
        cf_row("投资活动产生的现金流量净额", invest_net, bold=True, highlight=True),
        cf_row("三、筹资活动产生的现金流量：", bold=True),
        cf_row("  吸收投资收到的现金", finance_inflow, indent=1),
        cf_row("  取得借款收到的现金", 0.0, indent=1),
        cf_row("  收到其他与筹资活动有关的现金", 0.0, indent=1),
        cf_row("筹资活动现金流入小计", finance_inflow, bold=True, highlight=True),
        cf_row("  偿还债务支付的现金", finance_outflow, indent=1),
        cf_row("  分配股利、利润或偿付利息支付的现金", 0.0, indent=1),
        cf_row("  支付其他与筹资活动有关的现金", 0.0, indent=1),
        cf_row("筹资活动现金流出小计", finance_outflow, bold=True, highlight=True),
        cf_row("筹资活动产生的现金流量净额", finance_net, bold=True, highlight=True),
        cf_row("四、汇率变动对现金的影响", 0.0),
        cf_row("五、现金及现金等价物净增加额", total_net, bold=True, highlight=True),
        cf_row("  加：期初现金及现金等价物余额", cash_begin, indent=1),
        cf_row("六、期末现金及现金等价物余额", cash_end, bold=True, highlight=True),
    ]
    return {"items": items, "period_from": period_from, "period_to": period_to, "cash_begin": cash_begin, "cash_end": cash_end}


# ==================== 所有者权益变动表（企业会计准则一般企业—会企04号） ====================
ZERO9 = [0.0]*9           # 9 列零值
def _eq9(*indices_vals):  # (idx, val, ...) → 9 列数组
    a = [0.0]*9
    for i in range(0, len(indices_vals), 2):
        a[indices_vals[i]] = round(indices_vals[i+1], 2)
    return a

@app.get("/api/reports/equity-changes")
def equity_changes_report(
    company_id: int = Query(...),
    period: str = Query(...),
    db: Session = Depends(get_db)
):
    """所有者权益变动表（会企04号标准格式）"""
    yb = _bs_year_begin(period)
    begin_b = _compute_period_balances(company_id, None, yb, db)
    end_b = _compute_period_balances(company_id, None, period, db)

    py = period.split("-")[0]
    pl_items = _build_pl(company_id, f"{py}-01", period, db)
    net_profit = next((it["current"] for it in pl_items if it["label"] == "四、净利润"), 0.0)

    def eq_val(balances, prefix):
        d = sum(v["debit"] for code, v in balances.items() if code.startswith(prefix))
        c = sum(v["credit"] for code, v in balances.items() if code.startswith(prefix))
        return round(c - d, 2)  # 权益类：贷-借

    prefixes = ["4001", "4002", "4003", "4004", "4005", "4101", "4103", "4104"]
    begin_each = [eq_val(begin_b, p) if p else 0.0 for p in prefixes]
    end_each = [eq_val(end_b, p) if p else 0.0 for p in prefixes]
    # 未分配利润期末：直接取自科目余额表（而非 年初+净利润 简化公式）
    # 避免因前期差错更正、利润分配等调整导致的偏差
    end_each[7] = eq_val(end_b, "4104")
    # 合计辅助
    def total9(arr): return round(sum(arr), 2)
    begin9 = begin_each + [total9(begin_each)]
    end9 = end_each + [total9(end_each)]
    chg9 = [round(end9[i] - begin9[i], 2) for i in range(9)]

    cols = ["实收资本", "其他权益工具", "资本公积", "库存股", "其他综合收益", "专项储备", "盈余公积", "未分配利润", "所有者权益合计"]

    # 净利润只影响 未分配利润(7) 和 合计(8)
    np9 = [0.0]*9; np9[7] = net_profit; np9[8] = net_profit

    items = [
        {"label": "一、上年年末余额", "vals": begin9, "bold": True, "indent": 0, "highlight": False},
        {"label": "  加：会计政策变更", "vals": ZERO9, "bold": False, "indent": 1, "highlight": False},
        {"label": "  前期差错更正", "vals": ZERO9, "bold": False, "indent": 1, "highlight": False},
        {"label": "  其他", "vals": ZERO9, "bold": False, "indent": 1, "highlight": False},
        {"label": "二、本年年初余额", "vals": begin9, "bold": True, "indent": 0, "highlight": False},
        {"label": "三、本年增减变动金额（减少以\"-\"号填列）", "vals": chg9, "bold": True, "indent": 0, "highlight": False},
        {"label": "  （一）综合收益总额", "vals": np9, "bold": False, "indent": 1, "highlight": True},
        {"label": "  （二）所有者投入和减少资本", "vals": ZERO9, "bold": False, "indent": 1, "highlight": False},
        {"label": "    1. 所有者投入的普通股", "vals": ZERO9, "bold": False, "indent": 2, "highlight": False},
        {"label": "    2. 其他权益工具持有者投入资本", "vals": ZERO9, "bold": False, "indent": 2, "highlight": False},
        {"label": "    3. 股份支付计入所有者权益的金额", "vals": ZERO9, "bold": False, "indent": 2, "highlight": False},
        {"label": "    4. 其他", "vals": ZERO9, "bold": False, "indent": 2, "highlight": False},
        {"label": "  （三）利润分配", "vals": ZERO9, "bold": False, "indent": 1, "highlight": False},
        {"label": "    1. 提取盈余公积", "vals": ZERO9, "bold": False, "indent": 2, "highlight": False},
        {"label": "    2. 对所有者（或股东）的分配", "vals": ZERO9, "bold": False, "indent": 2, "highlight": False},
        {"label": "    3. 其他", "vals": ZERO9, "bold": False, "indent": 2, "highlight": False},
        {"label": "  （四）所有者权益内部结转", "vals": ZERO9, "bold": False, "indent": 1, "highlight": False},
        {"label": "    1. 资本公积转增资本（或股本）", "vals": ZERO9, "bold": False, "indent": 2, "highlight": False},
        {"label": "    2. 盈余公积转增资本（或股本）", "vals": ZERO9, "bold": False, "indent": 2, "highlight": False},
        {"label": "    3. 盈余公积弥补亏损", "vals": ZERO9, "bold": False, "indent": 2, "highlight": False},
        {"label": "    4. 设定受益计划变动额结转留存收益", "vals": ZERO9, "bold": False, "indent": 2, "highlight": False},
        {"label": "    5. 其他综合收益结转留存收益", "vals": ZERO9, "bold": False, "indent": 2, "highlight": False},
        {"label": "    6. 其他", "vals": ZERO9, "bold": False, "indent": 2, "highlight": False},
        {"label": "  （五）专项储备", "vals": ZERO9, "bold": False, "indent": 1, "highlight": False},
        {"label": "    1. 本期提取", "vals": ZERO9, "bold": False, "indent": 2, "highlight": False},
        {"label": "    2. 本期使用", "vals": ZERO9, "bold": False, "indent": 2, "highlight": False},
        {"label": "  （六）其他", "vals": ZERO9, "bold": False, "indent": 1, "highlight": False},
        {"label": "四、本年年末余额", "vals": end9, "bold": True, "indent": 0, "highlight": True},
    ]
    return {"columns": cols, "items": items, "period": period}


@app.post("/api/journal-entries")
def create_journal_entry(data: JournalEntryCreate, company_id: int = Query(...), db: Session = Depends(get_db)):
    e = JournalEntry(company_id=company_id, **data.model_dump())
    db.add(e)
    db.commit()
    db.refresh(e)
    return {"id": e.id, "message": "序时账记录创建成功"}


@app.get("/api/journal-entries/by-voucher")
def get_voucher_detail(voucher_word: str = Query(...), voucher_no: int = Query(...), company_id: int = Query(...), db: Session = Depends(get_db)):
    """按凭证字+凭证号查询凭证详情（所有分录）"""
    entries = db.query(JournalEntry).filter(
        JournalEntry.company_id == company_id,
        JournalEntry.voucher_word == voucher_word,
        JournalEntry.voucher_no == voucher_no
    ).order_by(JournalEntry.id.asc()).all()
    if not entries:
        raise HTTPException(404, detail="凭证不存在")
    first = entries[0]
    total_debit = sum(e.debit_amount or 0 for e in entries)
    total_credit = sum(e.credit_amount or 0 for e in entries)
    hierarchy = _build_account_hierarchy(db, company_id)
    return {
        "voucher_word": voucher_word,
        "voucher_no": voucher_no,
        "voucher_full": f"{voucher_word}-{voucher_no}",
        "period": first.period,
        "entry_date": str(first.entry_date),
        "source": first.source or "手动录入",
        "total_debit": round(total_debit, 2),
        "total_credit": round(total_credit, 2),
        "is_balanced": abs(total_debit - total_credit) < 0.01,
        "entry_count": len(entries),
        "entries": [
            {
                "id": e.id,
                "summary": e.summary or "",
                "account_code": e.account_code,
                "account_name": e.account_name or "",
                "account_full_name": hierarchy.get(e.account_code, e.account_name or ""),
                "debit_amount": e.debit_amount or 0,
                "credit_amount": e.credit_amount or 0,
            }
            for e in entries
        ]
    }


@app.get("/api/journal-entries/{entry_id}")
def get_journal_entry(entry_id: int, company_id: int = Query(...), db: Session = Depends(get_db)):
    e = db.query(JournalEntry).filter(JournalEntry.company_id == company_id, JournalEntry.id == entry_id).first()
    if not e:
        raise HTTPException(404, detail="记录不存在")
    return {
        "id": e.id, "entry_date": str(e.entry_date), "period": e.period,
        "voucher_word": e.voucher_word, "voucher_no": e.voucher_no,
        "attach_count": e.attach_count or 0, "summary": e.summary or "",
        "account_code": e.account_code, "account_name": e.account_name or "",
        "debit_amount": e.debit_amount or 0, "credit_amount": e.credit_amount or 0,
        "prepared_by": e.prepared_by or "", "reviewed_by": e.reviewed_by or "",
        "is_reviewed": e.is_reviewed, "remark": e.remark or "",
        "contact_project": e.contact_project or "",
        "spec_model": e.spec_model or "",
        "quantity": e.quantity or 0, "unit": e.unit or "",
        "unit_price": e.unit_price or 0,
        "created_at": str(e.created_at) if e.created_at else None,
    }


@app.put("/api/journal-entries/{entry_id}")
def update_journal_entry(entry_id: int, data: JournalEntryUpdate, company_id: int = Query(...), db: Session = Depends(get_db)):
    e = db.query(JournalEntry).filter(JournalEntry.company_id == company_id, JournalEntry.id == entry_id).first()
    if not e:
        raise HTTPException(404, detail="记录不存在")
    submitted = data.model_dump(exclude_unset=True)
    old_voucher_no = e.voucher_no
    old_voucher_word = e.voucher_word
    for k, v in submitted.items():
        setattr(e, k, v)
    # 凭证号或凭证字变化 → 同步业务表
    if ('voucher_no' in submitted or 'voucher_word' in submitted) and (e.voucher_no != old_voucher_no or e.voucher_word != old_voucher_word):
        _sync_biz_voucher_no(db, company_id, e, f"{e.voucher_word}-{e.voucher_no}")
    db.commit()
    return {"message": "更新成功"}


@app.delete("/api/journal-entries/{entry_id}")
def delete_journal_entry(entry_id: int, company_id: int = Query(...), db: Session = Depends(get_db)):
    e = db.query(JournalEntry).filter(JournalEntry.company_id == company_id, JournalEntry.id == entry_id).first()
    if not e:
        raise HTTPException(404, detail="记录不存在")
    period, vw = e.period, e.voucher_word
    # 删除前清除关联业务记录的凭证号（银行流水 journal_voucher_no / 进项抵扣 voucher_no 等）
    _clear_source_voucher_no(db, company_id, e)
    db.flush()
    db.delete(e)
    db.flush()
    _renumber_vouchers(db, company_id, period, vw)
    db.commit()
    return {"message": "删除成功"}


def _renumber_archive(db, company_id, model_cls, prefix):
    """删除后自动整理档案编码，使其连续不断号"""
    entries = db.query(model_cls).filter(
        model_cls.company_id == company_id,
        model_cls.code.like(prefix + '%')
    ).order_by(model_cls.code).all()
    prefix_len = len(prefix)
    for i, entry in enumerate(entries, 1):
        new_code = f"{prefix}{i:03d}"
        if entry.code != new_code:
            entry.code = new_code
    db.flush()


def _sync_biz_voucher_no(db, company_id, entry, new_voucher_str):
    """同步更新单条分录关联的业务表凭证号
    注意：仅 bank_transactions / input_vat_deductions / fixed_assets / intangible_assets
    有凭证号字段；purchase/sales/salary/ss/hf 等表没有"""
    if not entry.ref_id or not entry.source:
        return
    if entry.source == "银行流水":
        db.query(BankTransaction).filter(
            BankTransaction.company_id == company_id,
            BankTransaction.id == entry.ref_id
        ).update({"journal_voucher_no": new_voucher_str}, synchronize_session=False)
    elif entry.source == "进项抵扣":
        db.query(InputVATDeduction).filter(
            InputVATDeduction.company_id == company_id,
            InputVATDeduction.id == entry.ref_id
        ).update({"voucher_no": new_voucher_str}, synchronize_session=False)
    # 取得发票 / 销项发票 / 工资 / 社保 / 公积金 — 这些表没有 voucher_no 字段，无需同步


def _renumber_vouchers(db, company_id, period, voucher_word):
    """删除后自动重排同一期间+凭证字下的凭证号，并同步业务表"""
    entries = db.query(JournalEntry).filter(
        JournalEntry.company_id == company_id,
        JournalEntry.period == period,
        JournalEntry.voucher_word == voucher_word,
    ).order_by(JournalEntry.voucher_no.asc(), JournalEntry.id.asc()).all()
    if not entries:
        return
    # 按 voucher_no 分组
    groups = {}
    for e in entries:
        groups.setdefault(e.voucher_no, []).append(e)
    # 重新分配 voucher_no: 按原有顺序从1开始，同步业务表
    new_no = 1
    for old_no in sorted(groups.keys()):
        voucher_str_new = f"{voucher_word}-{new_no}"
        for e in groups[old_no]:
            e.voucher_no = new_no
            _sync_biz_voucher_no(db, company_id, e, voucher_str_new)
        new_no += 1
    db.flush()


def _clear_source_voucher_no(db, company_id, entry):
    """删除序时账凭证时，同步清除关联业务记录的凭证号，防止残留"""
    if not entry.source:
        return
    voucher_str = f"{entry.voucher_word}-{entry.voucher_no}"

    # ── 银行流水：双保险清除 ──
    # ① ref_id 精确匹配（优先）
    if entry.source == "银行流水" and entry.ref_id:
        db.query(BankTransaction).filter(
            BankTransaction.company_id == company_id,
            BankTransaction.id == entry.ref_id
        ).update({"journal_voucher_no": None}, synchronize_session=False)
    # ② 凭证号反向匹配（兜底，不限 source，覆盖 CCF/社保/公积金等所有来源）
    db.query(BankTransaction).filter(
        BankTransaction.company_id == company_id,
        BankTransaction.journal_voucher_no == voucher_str
    ).update({"journal_voucher_no": None}, synchronize_session=False)

    # ── 进项抵扣 ──
    if entry.source == "进项抵扣" and entry.ref_id:
        db.query(InputVATDeduction).filter(
            InputVATDeduction.company_id == company_id,
            InputVATDeduction.id == entry.ref_id
        ).update({"voucher_no": None}, synchronize_session=False)
    db.query(InputVATDeduction).filter(
        InputVATDeduction.company_id == company_id,
        InputVATDeduction.voucher_no == voucher_str
    ).update({"voucher_no": None}, synchronize_session=False)

    # ── 记账发票（删除凭证后回退到未记账状态）──
    # 先查出将被清除的BI的三号key，用于后续解锁PI
    affected_bis = db.query(BookkeepingInvoice.invoice_code, BookkeepingInvoice.invoice_no,
                            BookkeepingInvoice.digital_invoice_no).filter(
        BookkeepingInvoice.company_id == company_id,
        BookkeepingInvoice.voucher_no == voucher_str
    ).all()
    bi_keys = set((c or "", n or "", d or "") for c, n, d in affected_bis)
    
    db.query(BookkeepingInvoice).filter(
        BookkeepingInvoice.company_id == company_id,
        BookkeepingInvoice.voucher_no == voucher_str
    ).update({"voucher_no": None}, synchronize_session=False)
    db.flush()

    # ── 取得发票：解锁对应的 skip_accounting ──
    if bi_keys:
        pis = db.query(PurchaseInvoice).filter(
            PurchaseInvoice.company_id == company_id,
            PurchaseInvoice.skip_accounting == True
        ).all()
        for pi in pis:
            pi_key = (pi.invoice_code or "", pi.invoice_no or "", pi.digital_invoice_no or "")
            if pi_key in bi_keys:
                pi.skip_accounting = False
        db.flush()

    # 取得发票 / 销项发票 / 工资 / 社保 / 公积金 — 这些表没有 voucher_no 字段，无需清除


@app.post("/api/journal-entries/batch-delete")
def batch_delete_journal_entries(req: BatchDeleteRequest, company_id: int = Query(...), db: Session = Depends(get_db)):
    # 先查出被删记录的 (period, voucher_word) 组合
    deleted_records = db.query(JournalEntry).filter(
        JournalEntry.company_id == company_id,
        JournalEntry.id.in_(req.ids)
    ).all()
    # 删除前清除关联业务记录的凭证号
    for e in deleted_records:
        _clear_source_voucher_no(db, company_id, e)
    db.flush()
    combos = set((e.period, e.voucher_word) for e in deleted_records)
    deleted = db.query(JournalEntry).filter(
        JournalEntry.company_id == company_id,
        JournalEntry.id.in_(req.ids)
    ).delete(synchronize_session=False)
    db.flush()
    for period, vw in combos:
        _renumber_vouchers(db, company_id, period, vw)
    db.commit()
    return {"message": f"成功删除 {deleted} 条记录", "count": deleted}


@app.post("/api/sales-invoices/batch-to-journal")
def sales_invoice_batch_to_journal(
    body: dict = Body(default=None),
    company_id: int = Query(...),
    db=Depends(get_db)
):
    """一键生成勾选发票的记账凭证"""
    ids = body.get("ids", []) if body else []
    if not ids:
        return {"message": "未选择任何发票", "generated": 0, "skipped": 0, "errors": []}

    invoices = db.query(SalesInvoice).filter(
        SalesInvoice.company_id == company_id,
        SalesInvoice.id.in_(ids)
    ).order_by(SalesInvoice.invoice_date, SalesInvoice.id).all()

    generated = 0
    skipped = 0
    errors = []

    for inv in invoices:
        try:
            existing = db.query(JournalEntry).filter(
                JournalEntry.company_id == company_id,
                JournalEntry.source == "销项发票",
                JournalEntry.ref_id == inv.id
            ).first()
            if existing:
                skipped += 1
                continue

            from database import auto_generate_single_invoice
            auto_generate_single_invoice(db, inv)
            generated += 1
        except Exception as e:
            errors.append(f"发票{inv.id}({inv.invoice_no}): {str(e)}")

    db.commit()
    msg = f"批量生成完成：生成 {generated} 笔凭证"
    if skipped > 0:
        msg += f"，跳过 {skipped} 笔（已有凭证）"
    if errors:
        msg += f"，{len(errors)} 笔失败"
        print("Batch journal errors:", errors)
    return {"message": msg, "generated": generated, "skipped": skipped, "errors": errors}


@app.post("/api/sales-invoices/auto-voucher")
def sales_invoice_auto_voucher(company_id: int = Query(...), db=Depends(get_db)):
    """导入后自动为所有未生成凭证的销项发票生成序时账"""
    # 查询已有凭证的发票ID（通过 JournalEntry.source=销项发票 + ref_id 判断）
    existing_ids = set(r[0] for r in db.query(JournalEntry.ref_id).filter(
        JournalEntry.company_id == company_id,
        JournalEntry.source == "销项发票",
        JournalEntry.ref_id.isnot(None)
    ).all())
    
    invoices = db.query(SalesInvoice).filter(
        SalesInvoice.company_id == company_id,
        ~SalesInvoice.id.in_(existing_ids) if existing_ids else True
    ).order_by(SalesInvoice.invoice_date, SalesInvoice.id).all()
    if not invoices:
        return {"message": "无待生成凭证的发票", "generated": 0}
    
    generated = 0
    errors = []
    for inv in invoices:
        try:
            from database import auto_generate_single_invoice
            auto_generate_single_invoice(db, inv)
            generated += 1
        except Exception as e:
            errors.append(f"发票{inv.id}({inv.invoice_no}): {str(e)}")
    
    db.commit()
    msg = f"自动生成 {generated} 笔凭证"
    if errors:
        msg += f"，{len(errors)} 笔失败"
    return {"message": msg, "generated": generated, "errors": errors}


@app.post("/api/input-vat-deductions/auto-voucher")
def input_vat_auto_voucher(company_id: int = Query(...), db=Depends(get_db)):
    """导入进项抵扣后自动生成序时账凭证"""
    # 查找所有未生成凭证的进项抵扣记录，按期分组
    unprocessed = db.query(InputVATDeduction).filter(
        InputVATDeduction.company_id == company_id,
        or_(InputVATDeduction.voucher_no == None, InputVATDeduction.voucher_no == "")
    ).all()
    if not unprocessed:
        return {"message": "无待生成凭证的进项抵扣", "generated": 0}
    
    periods = set()
    for d in unprocessed:
        p = d.deduction_period
        if not p and d.invoice_date:
            p = str(d.invoice_date)[:7]  # 从发票日期推导
        if p:
            periods.add(p)
    total = 0
    for period in periods:
        total += auto_generate_input_vat_for_period(db, company_id, period)
    
    db.commit()
    return {"message": f"自动生成 {total} 条进项抵扣凭证（共 {len(periods)} 个期间）", "generated": total}


@app.post("/api/sales-invoices/{invoice_id}/to-journal")
def sales_invoice_to_journal(invoice_id: int, company_id: int = Query(...), db=Depends(get_db)):
    """将单张销项发票生成记账凭证（分录）到序时账（允许重新生成，先删旧凭证）"""
    inv = db.query(SalesInvoice).filter(SalesInvoice.company_id == company_id, SalesInvoice.id == invoice_id).first()
    if not inv:
        raise HTTPException(404, "发票不存在")
    def get_full_name(code):
        """构建科目的全级次名称，如 221001001 → 应交税费/应交增值税/销项税额"""
        parts = []
        cur = code
        while cur:
            acc = db.query(Account).filter(
                Account.company_id == inv.company_id,
                Account.code == cur
            ).first()
            if not acc:
                break
            parts.insert(0, acc.name)
            cur = acc.parent_code
        return "/".join(parts) if parts else code

    def ensure_revenue_sub(goods_name):
        """确保主营业务收入下存在对应货物的子科目，返回 (code, full_name)"""
        if not goods_name:
            return ("6001", get_full_name("6001"))
        existing = db.query(Account).filter(
            Account.company_id == inv.company_id,
            Account.parent_code == "6001",
            Account.name == goods_name
        ).first()
        if existing:
            return (existing.code, get_full_name(existing.code))
        max_sub = db.query(Account.code).filter(
            Account.company_id == inv.company_id,
            Account.parent_code == "6001"
        ).order_by(Account.code.desc()).first()
        next_num = int(max_sub[0][4:7]) + 1 if (max_sub and max_sub[0] and len(max_sub[0]) >= 6) else 1
        # 科目编码规则：6001 下级为 600101/600102/...（6位，2位序号）
        new_code = f"6001{next_num:02d}"
        new_acc = Account(
            company_id=inv.company_id,
            code=new_code,
            name=goods_name,
            category="收入",
            balance_direction="贷",
            level=2,
            parent_code="6001",
        )
        db.add(new_acc)
        db.flush()
        return (new_code, get_full_name(new_code))

    period = inv.invoice_date.strftime("%Y-%m") if inv.invoice_date else datetime.now().strftime("%Y-%m")

    max_no = db.query(JournalEntry.voucher_no).filter(
        JournalEntry.company_id == inv.company_id,
        JournalEntry.period == period,
        JournalEntry.voucher_word == "记"
    ).order_by(JournalEntry.voucher_no.desc()).first()
    next_voucher_no = (max_no[0] + 1) if max_no and max_no[0] else 1

    date_str = inv.invoice_date.strftime("%Y-%m-%d") if inv.invoice_date else period + "-01"
    buyer = inv.buyer_name or "客户"
    goods = inv.goods_name or ""
    summary = f"销售{goods or '货物'}给{buyer}"

    # 先删旧凭证（允许重新生成）
    db.query(JournalEntry).filter(
        JournalEntry.company_id == inv.company_id,
        JournalEntry.source == "销项发票",
        JournalEntry.ref_id == inv.id
    ).delete(synchronize_session=False)
    db.flush()

    rev_code, rev_name = ensure_revenue_sub(goods)

    entries = [
        JournalEntry(
            company_id=inv.company_id,
            entry_date=datetime.strptime(date_str, "%Y-%m-%d").date(),
            period=period,
            voucher_word="记",
            voucher_no=next_voucher_no,
            summary=summary,
            account_code="1122",
            account_name=get_full_name("1122"),
            debit_amount=inv.total_amount,
            credit_amount=0,
            contact_project=buyer,
            spec_model=inv.spec or "",
            quantity=inv.quantity or 0,
            unit=inv.unit or "",
            unit_price=inv.unit_price or 0,
            source="销项发票", ref_id=inv.id,
        ),
        JournalEntry(
            company_id=inv.company_id,
            entry_date=datetime.strptime(date_str, "%Y-%m-%d").date(),
            period=period,
            voucher_word="记",
            voucher_no=next_voucher_no,
            summary=summary,
            account_code=rev_code,
            account_name=rev_name,
            debit_amount=0,
            credit_amount=inv.amount,
            contact_project="",
            spec_model=inv.spec or "",
            quantity=inv.quantity or 0,
            unit=inv.unit or "",
            unit_price=inv.unit_price or 0,
            source="销项发票", ref_id=inv.id,
        ),
        JournalEntry(
            company_id=inv.company_id,
            entry_date=datetime.strptime(date_str, "%Y-%m-%d").date(),
            period=period,
            voucher_word="记",
            voucher_no=next_voucher_no,
            summary=f"{summary}（增值税）",
            account_code="221001001",
            account_name=get_full_name("221001001"),
            debit_amount=0,
            credit_amount=inv.tax_amount,
            contact_project="",
            spec_model=inv.spec or "",
            quantity=inv.quantity or 0,
            unit=inv.unit or "",
            unit_price=inv.unit_price or 0,
            source="销项发票", ref_id=inv.id,
        ),
    ]
    for e in entries:
        db.add(e)
    db.commit()
    return {"message": f"已生成凭证，凭证号：记-{next_voucher_no}", "voucher_no": next_voucher_no, "period": period}


# ==================== 进项抵扣 ====================

class InputVATDeductionCreate(BaseModel):
    purchase_invoice_id: Optional[int] = None
    check_status: Optional[str] = None
    invoice_source: Optional[str] = None
    domestic_sale_cert_no: Optional[str] = None
    digital_invoice_no: Optional[str] = None
    invoice_code: Optional[str] = None
    invoice_no: Optional[str] = None
    invoice_date: Optional[date] = None
    seller_tax_id: Optional[str] = None
    seller_name: Optional[str] = None
    amount: float = 0.0
    tax_amount: float = 0.0
    deductible_tax_amount: float = 0.0
    invoice_category: Optional[str] = None
    invoice_category_label: Optional[str] = None
    invoice_status: Optional[str] = "正常"
    check_time: Optional[datetime] = None
    risk_level: Optional[str] = "正常"
    # 保留字段
    goods_name: Optional[str] = None
    total_amount: float = 0.0
    tax_rate: Optional[float] = 0.0
    deducted_tax_amount: Optional[float] = 0.0
    deduction_period: Optional[str] = None
    deduction_status: str = "待抵扣"
    certification_date: Optional[date] = None
    deduction_date: Optional[date] = None
    deduction_method: str = "凭票抵扣"
    voucher_no: Optional[str] = None
    remark: Optional[str] = None


class InputVATDeductionUpdate(BaseModel):
    check_status: Optional[str] = None
    invoice_source: Optional[str] = None
    domestic_sale_cert_no: Optional[str] = None
    digital_invoice_no: Optional[str] = None
    invoice_code: Optional[str] = None
    invoice_no: Optional[str] = None
    invoice_date: Optional[date] = None
    seller_tax_id: Optional[str] = None
    seller_name: Optional[str] = None
    amount: Optional[float] = None
    tax_amount: Optional[float] = None
    deductible_tax_amount: Optional[float] = None
    invoice_category: Optional[str] = None
    invoice_category_label: Optional[str] = None
    invoice_status: Optional[str] = None
    check_time: Optional[datetime] = None
    risk_level: Optional[str] = None
    # 保留字段
    goods_name: Optional[str] = None
    total_amount: Optional[float] = None
    tax_rate: Optional[float] = None
    deducted_tax_amount: Optional[float] = None
    deduction_period: Optional[str] = None
    deduction_status: Optional[str] = None
    certification_date: Optional[date] = None
    deduction_date: Optional[date] = None
    deduction_method: Optional[str] = None
    voucher_no: Optional[str] = None
    remark: Optional[str] = None


@app.get("/api/input-vat-deductions")
def list_input_vat_deductions(
    company_id: int = Query(...),
    invoice_status: Optional[str] = None,
    check_status: Optional[str] = None,
    risk_level: Optional[str] = None,
    deduction_period: Optional[str] = None,
    keyword: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    db: Session = Depends(get_db)
):
    q = db.query(InputVATDeduction).filter(InputVATDeduction.company_id == company_id)
    if invoice_status:
        q = q.filter(InputVATDeduction.invoice_status == invoice_status)
    if check_status:
        q = q.filter(InputVATDeduction.check_status == check_status)
    if risk_level:
        q = q.filter(InputVATDeduction.risk_level == risk_level)
    if deduction_period:
        q = q.filter(InputVATDeduction.deduction_period == deduction_period)
    if date_from:
        q = q.filter(InputVATDeduction.invoice_date >= date_from)
    if date_to:
        q = q.filter(InputVATDeduction.invoice_date <= date_to)
    if keyword:
        q = q.filter(or_(
            InputVATDeduction.invoice_no.contains(keyword),
            InputVATDeduction.digital_invoice_no.contains(keyword),
            InputVATDeduction.invoice_code.contains(keyword),
            InputVATDeduction.seller_name.contains(keyword),
            InputVATDeduction.seller_tax_id.contains(keyword),
        ))
    items = q.order_by(InputVATDeduction.invoice_date.desc(), InputVATDeduction.check_time.desc()).all()
    # 构建凭证号映射（进项抵扣 → 序时账，按期间匹配 source="进项抵扣" 的汇总凭证）
    # 期间取值与凭证生成逻辑一致：deduction_period 优先，fallback 到 invoice_date 年月
    def _effective_period(it):
        if it.deduction_period:
            return it.deduction_period
        if it.invoice_date:
            return it.invoice_date.strftime("%Y-%m")
        return None
    periods_set = list(set(_effective_period(it) for it in items if _effective_period(it)))
    period_vouchers = {}
    if periods_set:
        for je in db.query(JournalEntry).filter(
            JournalEntry.company_id == company_id,
            JournalEntry.source == "进项抵扣",
            JournalEntry.period.in_(periods_set),
            JournalEntry.account_code == "221001002"
        ).all():
            period_vouchers[je.period] = f"{je.voucher_word}-{je.voucher_no}"
    voucher_map = {}
    for it in items:
        ep = _effective_period(it)
        if ep and ep in period_vouchers:
            voucher_map[it.id] = period_vouchers[ep]
    return [{
        "id": it.id, "purchase_invoice_id": it.purchase_invoice_id,
        "check_status": it.check_status or "",
        "invoice_source": it.invoice_source or "",
        "domestic_sale_cert_no": it.domestic_sale_cert_no or "",
        "digital_invoice_no": it.digital_invoice_no or "",
        "invoice_code": it.invoice_code or "",
        "invoice_no": it.invoice_no or "",
        "invoice_date": str(it.invoice_date) if it.invoice_date else "",
        "seller_tax_id": it.seller_tax_id or "",
        "seller_name": it.seller_name or "",
        "amount": it.amount or 0,
        "tax_amount": it.tax_amount or 0,
        "deductible_tax_amount": it.deductible_tax_amount or 0,
        "invoice_category": it.invoice_category or "",
        "invoice_category_label": it.invoice_category_label or "",
        "invoice_status": it.invoice_status or "正常",
        "check_time": str(it.check_time) if it.check_time else "",
        "risk_level": it.risk_level or "正常",
        "goods_name": it.goods_name or "",
        "total_amount": it.total_amount or 0,
        "tax_rate": it.tax_rate or 0,
        "deducted_tax_amount": it.deducted_tax_amount or 0,
        "deduction_period": it.deduction_period or "",
        "deduction_status": it.deduction_status or "",
        "certification_date": str(it.certification_date) if it.certification_date else "",
        "deduction_date": str(it.deduction_date) if it.deduction_date else "",
        "deduction_method": it.deduction_method or "",
        "voucher_no": it.voucher_no or "", "remark": it.remark or "",
        "journal_voucher_no": voucher_map.get(it.id, ""),
        "import_batch_id": it.import_batch_id or "",
        "created_at": str(it.created_at) if it.created_at else ""
    } for it in items]


@app.post("/api/input-vat-deductions")
def create_input_vat_deduction(data: InputVATDeductionCreate, company_id: int = Query(...), db: Session = Depends(get_db)):
    item = InputVATDeduction(company_id=company_id, **data.model_dump())
    db.add(item)
    db.commit()
    db.refresh(item)
    # 自动生成序时账凭证（按期间汇总）
    try:
        period = item.deduction_period
        if not period:
            period = item.invoice_date.strftime("%Y-%m") if item.invoice_date else datetime.now().strftime("%Y-%m")
        auto_generate_input_vat_for_period(db, company_id, period)
    except Exception as e:
        db.rollback()
    return {"id": item.id, "message": "进项抵扣记录创建成功"}


@app.get("/api/input-vat-deductions/stats")
def input_vat_deduction_stats(company_id: int = Query(...), db: Session = Depends(get_db)):
    base = db.query(InputVATDeduction).filter(InputVATDeduction.company_id == company_id)
    total_count = base.count()
    total_tax = base.with_entities(func.sum(InputVATDeduction.tax_amount)).scalar() or 0
    total_deductible = base.with_entities(func.sum(InputVATDeduction.deductible_tax_amount)).scalar() or 0
    total_amount = base.with_entities(func.sum(InputVATDeduction.amount)).scalar() or 0
    unchecked_count = base.filter(InputVATDeduction.check_status == "未勾选").count()
    checked_count = base.filter(InputVATDeduction.check_status == "已勾选").count()
    abnormal_count = base.filter(InputVATDeduction.risk_level.in_(["疑点", "异常", "失控"])).count()
    return {
        "total_count": total_count,
        "total_amount": round(total_amount, 2),
        "total_tax": round(total_tax, 2),
        "total_deductible": round(total_deductible, 2),
        "unchecked_count": unchecked_count,
        "checked_count": checked_count,
        "abnormal_count": abnormal_count
    }


@app.get("/api/input-vat-deductions/{item_id}")
def get_input_vat_deduction(item_id: int, company_id: int = Query(...), db: Session = Depends(get_db)):
    it = db.query(InputVATDeduction).filter(InputVATDeduction.company_id == company_id, InputVATDeduction.id == item_id).first()
    if not it:
        raise HTTPException(404, detail="抵扣记录不存在")
    return {
        "id": it.id, "purchase_invoice_id": it.purchase_invoice_id,
        "check_status": it.check_status or "",
        "invoice_source": it.invoice_source or "",
        "domestic_sale_cert_no": it.domestic_sale_cert_no or "",
        "digital_invoice_no": it.digital_invoice_no or "",
        "invoice_code": it.invoice_code or "",
        "invoice_no": it.invoice_no or "",
        "invoice_date": str(it.invoice_date) if it.invoice_date else "",
        "seller_tax_id": it.seller_tax_id or "",
        "seller_name": it.seller_name or "",
        "amount": it.amount or 0,
        "tax_amount": it.tax_amount or 0,
        "deductible_tax_amount": it.deductible_tax_amount or 0,
        "invoice_category": it.invoice_category or "",
        "invoice_category_label": it.invoice_category_label or "",
        "invoice_status": it.invoice_status or "正常",
        "check_time": str(it.check_time) if it.check_time else "",
        "risk_level": it.risk_level or "正常",
        "goods_name": it.goods_name or "",
        "total_amount": it.total_amount or 0,
        "tax_rate": it.tax_rate or 0,
        "deducted_tax_amount": it.deducted_tax_amount or 0,
        "deduction_period": it.deduction_period or "",
        "deduction_status": it.deduction_status or "",
        "certification_date": str(it.certification_date) if it.certification_date else "",
        "deduction_date": str(it.deduction_date) if it.deduction_date else "",
        "deduction_method": it.deduction_method or "",
        "voucher_no": it.voucher_no or "", "remark": it.remark or "",
        "created_at": str(it.created_at) if it.created_at else "",
        "updated_at": str(it.updated_at) if it.updated_at else ""
    }


@app.put("/api/input-vat-deductions/{item_id}")
def update_input_vat_deduction(item_id: int, data: InputVATDeductionUpdate, company_id: int = Query(...), db: Session = Depends(get_db)):
    it = db.query(InputVATDeduction).filter(InputVATDeduction.company_id == company_id, InputVATDeduction.id == item_id).first()
    if not it:
        raise HTTPException(404, detail="抵扣记录不存在")
    for k, v in data.model_dump(exclude_unset=True).items():
        setattr(it, k, v)
    it.updated_at = datetime.now()
    db.commit()
    # 自动更新序时账凭证（按期间汇总）
    try:
        period = it.deduction_period
        if not period:
            period = it.invoice_date.strftime("%Y-%m") if it.invoice_date else datetime.now().strftime("%Y-%m")
        auto_generate_input_vat_for_period(db, company_id, period)
    except Exception as e:
        db.rollback()
    return {"message": "更新成功"}


@app.delete("/api/input-vat-deductions/{item_id}")
def delete_input_vat_deduction(item_id: int, company_id: int = Query(...), db: Session = Depends(get_db)):
    it = db.query(InputVATDeduction).filter(InputVATDeduction.company_id == company_id, InputVATDeduction.id == item_id).first()
    if not it:
        raise HTTPException(404, detail="抵扣记录不存在")
    db.delete(it)
    db.commit()
    return {"message": "删除成功"}


@app.post("/api/input-vat-deductions/batch-delete")
def batch_delete_input_vat_deductions(ids: list[int], company_id: int = Query(...), db: Session = Depends(get_db)):
    deleted = db.query(InputVATDeduction).filter(
        InputVATDeduction.company_id == company_id,
        InputVATDeduction.id.in_(ids)
    ).delete(synchronize_session=False)
    db.commit()
    return {"message": f"已删除 {deleted} 条记录", "deleted": deleted}


@app.post("/api/input-vat-deductions/batch-certify")
def batch_certify_input_vat_deductions(ids: list[int], company_id: int = Query(...), db: Session = Depends(get_db)):
    """批量认证：将选中记录标记为已勾选，设置勾选时间/认证日期"""
    now = datetime.now()
    today = date.today()
    
    items = db.query(InputVATDeduction).filter(
        InputVATDeduction.company_id == company_id,
        InputVATDeduction.id.in_(ids)
    ).all()
    
    if not items:
        raise HTTPException(400, detail="未找到可认证记录")
    
    certified = 0
    affected_periods = set()
    for it in items:
        changed = False
        if it.check_status != "已勾选":
            it.check_status = "已勾选"
            it.check_time = now
            changed = True
        if not it.certification_date:
            it.certification_date = today
            changed = True
        if it.deduction_status in (None, "", "待认证"):
            it.deduction_status = "待抵扣"
            changed = True
        if changed:
            it.updated_at = now
            certified += 1
            period = it.deduction_period
            if not period and it.invoice_date:
                period = it.invoice_date.strftime("%Y-%m")
            if period:
                affected_periods.add(period)
    
    db.commit()
    
    # 为受影响的期间重新生成进项抵扣凭证
    voucher_count = 0
    for period in affected_periods:
        try:
            voucher_count += auto_generate_input_vat_for_period(db, company_id, period)
        except Exception:
            pass
    db.commit()
    
    return {
        "message": f"已认证 {certified} 条记录" + (f"，生成 {voucher_count} 笔汇总凭证" if voucher_count else ""),
        "certified": certified,
        "voucher_count": voucher_count
    }


@app.post("/api/input-vat-deductions/batch-to-journal")
def input_vat_batch_to_journal(ids: Optional[List[int]] = Body(None), company_id: int = Query(...), db: Session = Depends(get_db)):
    """按指定进项抵扣记录的期间批量生成/重生成凭证；不传 ids 则处理全部"""
    q = db.query(InputVATDeduction).filter(InputVATDeduction.company_id == company_id)
    if ids:
        q = q.filter(InputVATDeduction.id.in_(ids))
    deductions = q.all()
    periods = set()
    for d in deductions:
        period = d.deduction_period or (d.invoice_date.strftime("%Y-%m") if d.invoice_date else None)
        if period:
            periods.add(period)
    total = 0
    vouchers = []
    for period in sorted(periods):
        try:
            c = auto_generate_input_vat_for_period(db, company_id, period)
            total += c
            if c:
                vouchers.append(period)
        except Exception:
            pass
    db.commit()
    return {
        "message": f"已为 {len(periods)} 个期间生成进项抵扣凭证，共 {total} 条",
        "periods": sorted(periods),
        "vouchers": vouchers,
        "total": total
    }



@app.post("/api/bank-transactions/batch-to-journal")
def bank_transactions_batch_to_journal(ids: Optional[List[int]] = Body(None), company_id: int = Query(...), db: Session = Depends(get_db)):
    """为指定银行流水批量生成记账凭证；不传 ids 则处理全部"""
    result = _generate_bank_journals(db, company_id, ids)
    db.commit()
    return {
        "message": f"已生成 {result['generated']} 条银行流水凭证，跳过 {result['skipped']} 条",
        "generated": result["generated"],
        "skipped": result["skipped"],
        "errors": result["errors"],
        "infos": result.get("infos", []),
    }


@app.post("/api/bank-transactions/auto-voucher")
def bank_transactions_auto_voucher(company_id: int = Query(...), db: Session = Depends(get_db)):
    """导入银行流水后自动全链路处理：
    0. 档案缺失补齐（序时账有往来科目但档案缺失 → 自动建档，零容忍gap）
    0.5 档案信息补全（从发票/银行流水提取税号/银行账号等，第一时间填补缺失字段）
    1. 双源供应商智能建档（发票∩银行流水 → 供应商档案）
    2. 常规银行流水凭证生成（_classify_bank_tx 11级分类）
    3. 社保缴纳匹配
    4. 国家金库税费组合缴纳匹配（含单税兜底）
    5. 公积金缴纳匹配
    """
    result = {"generated": 0, "suppliers_created": 0, "customers_fixed": 0, "suppliers_fixed": 0,
              "customers_enriched": 0, "suppliers_enriched": 0, "infos": []}

    # 第0步：档案缺失补齐（序时账有往来但档案缺失 → 自动建档）
    # 这是确定性规则：有明细账就必需有档案，零容忍gap
    try:
        gap_result = _close_archive_gap(db, company_id)
        result["customers_fixed"] = gap_result.get("customer_created", 0)
        result["suppliers_fixed"] = gap_result.get("supplier_created", 0)
    except Exception:
        pass

    # 第0.5步：档案信息补全（从发票/银行流水提取缺失字段）
    try:
        enrich_result = _enrich_archive_info(db, company_id)
        result["customers_enriched"] = enrich_result.get("customer_enriched", 0)
        result["suppliers_enriched"] = enrich_result.get("supplier_enriched", 0)
    except Exception:
        pass

    # 第1步：双源供应商智能建档（发票∩银行 → 正式供应商）
    try:
        supp_result = _do_auto_create_suppliers(db, company_id)
        result["suppliers_created"] = supp_result.get("created", 0)
        if supp_result.get("infos"):
            result["infos"].extend(supp_result["infos"])
    except Exception:
        pass

    # 第1步：常规银行流水凭证生成
    try:
        bk_result = _generate_bank_journals(db, company_id, None)
        result["generated"] += bk_result.get("generated", 0)
        if bk_result.get("infos"):
            result["infos"].extend(bk_result["infos"])
    except Exception:
        pass

    # 第2步：社保缴纳匹配
    try:
        ss_result = _match_ss_payment_journals(db, company_id)
        result["generated"] += ss_result.get("generated", 0)
    except Exception:
        pass

    # 第3步：国家金库税费组合缴纳匹配
    try:
        tax_result = _match_tax_payment_journals(db, company_id)
        result["generated"] += tax_result.get("generated", 0)
    except Exception:
        pass

    # 第4步：公积金缴纳匹配
    try:
        hf_result = _match_hf_payment_journals(db, company_id)
        result["generated"] += hf_result.get("generated", 0)
    except Exception:
        pass

    db.commit()
    return {
        "message": f"自动生成 {result['generated']} 条凭证，新建 {result['suppliers_created']} 个供应商",
        "generated": result["generated"],
        "suppliers_created": result["suppliers_created"],
        "detail": result
    }


@app.post("/api/bank-transactions/classify")
def classify_bank_transactions(ids: List[int] = Body(...), company_id: int = Query(...), db: Session = Depends(get_db)):
    """预览银行流水凭证分类结果（不生成凭证），返回每条流水的建议科目"""
    txs = db.query(BankTransaction).filter(
        BankTransaction.company_id == company_id,
        BankTransaction.id.in_(ids)
    ).all()
    results = []
    # 预建跨实体索引
    entity_index = _build_entity_index(db, company_id)
    for tx in txs:
        result = _classify_bank_tx(db, company_id, tx, entity_index)
        if result is None:
            results.append({
                "tx_id": tx.id,
                "summary": tx.summary or tx.counterparty_name or "银行流水",
                "amount": abs(float(tx.amount) if tx.amount else 0),
                "is_debit": tx.amount is not None and tx.amount < 0,
                "debit_account": "", "debit_name": "",
                "credit_account": "", "credit_name": "",
                "match_type": "unclassified",
            })
            continue
        other_code, other_name, match_type = result
        is_debit = tx.amount is not None and tx.amount < 0
        amount = abs(float(tx.amount) if tx.amount else 0)
        # 确定借贷方向
        if match_type == "internal_transfer":
            results.append({
                "tx_id": tx.id,
                "summary": tx.summary or tx.counterparty_name or "银行流水",
                "amount": amount,
                "is_debit": is_debit,
                "debit_account": "1002", "debit_name": "银行存款",
                "credit_account": "1002", "credit_name": "银行存款(内部转账)",
                "match_type": match_type,
            })
        elif is_debit:
            results.append({
                "tx_id": tx.id,
                "summary": tx.summary or tx.counterparty_name or "银行流水",
                "amount": amount,
                "is_debit": True,
                "debit_account": other_code, "debit_name": other_name,
                "credit_account": "1002", "credit_name": "银行存款",
                "match_type": match_type,
            })
        else:
            results.append({
                "tx_id": tx.id,
                "summary": tx.summary or tx.counterparty_name or "银行流水",
                "amount": amount,
                "is_debit": False,
                "debit_account": "1002", "debit_name": "银行存款",
                "credit_account": other_code, "credit_name": other_name,
                "match_type": match_type,
            })
    return {"results": results}


@app.post("/api/input-vat-deductions/{item_id}/to-journal")
def input_vat_deduction_to_journal(item_id: int, company_id: int = Query(...), db: Session = Depends(get_db)):
    """为进项抵扣记录所在期间重新生成凭证"""
    it = db.query(InputVATDeduction).filter(
        InputVATDeduction.id == item_id,
        InputVATDeduction.company_id == company_id
    ).first()
    if not it:
        raise HTTPException(404, "抵扣记录不存在")

    period = it.deduction_period or (it.invoice_date.strftime("%Y-%m") if it.invoice_date else datetime.now().strftime("%Y-%m"))
    count = auto_generate_input_vat_for_period(db, company_id, period)
    db.commit()
    return {"message": f"已为 {period} 生成进项抵扣凭证 ({count} 条)", "period": period}


# ==================== 列映射模板 ====================

class ColumnTemplateCreate(BaseModel):
    module: str
    template_name: str
    bank_config_id: Optional[int] = None
    column_mapping: Optional[str] = None
    is_default: bool = False


class ColumnTemplateUpdate(BaseModel):
    template_name: Optional[str] = None
    column_mapping: Optional[str] = None
    is_default: Optional[bool] = None


@app.get("/api/column-templates")
def list_column_templates(
    company_id: int = Query(...),
    module: Optional[str] = None,
    bank_config_id: Optional[int] = None,
    db: Session = Depends(get_db)
):
    q = db.query(ColumnTemplate).filter(ColumnTemplate.company_id == company_id)
    if module:
        q = q.filter(ColumnTemplate.module == module)
    if bank_config_id is not None:
        q = q.filter(ColumnTemplate.bank_config_id == bank_config_id)
    templates = q.order_by(ColumnTemplate.module, ColumnTemplate.template_name).all()
    return [{
        "id": t.id, "module": t.module, "template_name": t.template_name,
        "bank_config_id": t.bank_config_id,
        "column_mapping": t.column_mapping or "{}",
        "is_default": t.is_default,
        "created_at": str(t.created_at) if t.created_at else ""
    } for t in templates]


@app.post("/api/column-templates")
def create_column_template(data: ColumnTemplateCreate, company_id: int = Query(...), db: Session = Depends(get_db)):
    tpl = ColumnTemplate(company_id=company_id, **data.model_dump())
    db.add(tpl)
    db.commit()
    db.refresh(tpl)
    return {"id": tpl.id, "message": "模板创建成功"}


@app.put("/api/column-templates/{tpl_id}")
def update_column_template(tpl_id: int, data: ColumnTemplateUpdate, company_id: int = Query(...), db: Session = Depends(get_db)):
    tpl = db.query(ColumnTemplate).filter(ColumnTemplate.company_id == company_id, ColumnTemplate.id == tpl_id).first()
    if not tpl:
        raise HTTPException(404, detail="模板不存在")
    for k, v in data.model_dump(exclude_unset=True).items():
        setattr(tpl, k, v)
    tpl.updated_at = datetime.now()
    db.commit()
    return {"message": "更新成功"}


@app.delete("/api/column-templates/{tpl_id}")
def delete_column_template(tpl_id: int, company_id: int = Query(...), db: Session = Depends(get_db)):
    tpl = db.query(ColumnTemplate).filter(ColumnTemplate.company_id == company_id, ColumnTemplate.id == tpl_id).first()
    if not tpl:
        raise HTTPException(404, detail="模板不存在")
    db.delete(tpl)
    db.commit()
    return {"message": "删除成功"}


# ==================== 文件上传 - 表头分析 ====================

@app.post("/api/file/analyze-headers")
async def analyze_file_headers(
    file: UploadFile = File(...),
    module: str = Form("bank-transaction"),
    bank_config_id: Optional[int] = Form(None)
):
    """上传文件，返回表头列表供用户做列映射"""
    fname = file.filename or "unknown"
    ext = os.path.splitext(fname)[1].lower()
    if ext not in ALLOWED_EXTENSIONS:
        raise HTTPException(400, f"不支持的文件类型: {ext}，仅接受 xlsx/xls/csv/pdf/txt")
    try:
        content_bytes = await file.read()
        if len(content_bytes) > MAX_UPLOAD_SIZE:
            raise HTTPException(400, f"文件过大（{len(content_bytes)/1024/1024:.1f}MB），上限10MB")

        headers = []
        preview_rows = []

        if ext in (".xlsx", ".xls"):
            wb = openpyxl.load_workbook(io.BytesIO(content_bytes), data_only=True)
            ws = wb.active
            # 第一行作为表头
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=col)
                headers.append(str(cell.value).strip() if cell.value is not None else f"列{col}")
            # 预览前3行
            for row in range(2, min(ws.max_row + 1, 5)):
                vals = {}
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row, column=col)
                    vals[headers[col - 1]] = str(cell.value) if cell.value is not None else ""
                preview_rows.append(vals)
            total_rows = ws.max_row - 1
        elif ext == ".csv":
            text = content_bytes.decode("utf-8-sig")
            reader = csv.reader(io.StringIO(text))
            rows = list(reader)
            if rows:
                headers = [h.strip() for h in rows[0]]
                for row in rows[1:4]:
                    vals = {}
                    for i, h in enumerate(headers):
                        vals[h] = row[i] if i < len(row) else ""
                    preview_rows.append(vals)
                total_rows = len(rows) - 1
            else:
                headers, total_rows = [], 0
        else:
            return {"error": f"不支持的文件格式：{ext}。请上传 xlsx 或 csv 文件。"}

        # 获取已知的列映射模板
        field_groups = {}
        field_order = None
        if module == "sales-invoice":
            # 严格按开具发票表头26列顺序，一一平铺
            field_order = [
                "invoice_code", "invoice_no", "digital_invoice_no",
                "seller_tax_no", "seller_name",
                "buyer_tax_no", "buyer_name",
                "invoice_date", "tax_category_code", "specific_business_type",
                "goods_name", "spec", "unit", "quantity", "unit_price",
                "amount", "tax_rate", "tax_amount", "total_amount",
                "invoice_source", "invoice_category", "status", "is_positive", "invoice_risk_level",
                "issuer", "remark"
            ]
        elif module == "purchase-invoice":
            # 取得发票26列
            field_order = [
                "invoice_code", "invoice_no", "digital_invoice_no",
                "seller_tax_no", "seller_name",
                "buyer_tax_no", "buyer_name",
                "invoice_date", "tax_category_code", "specific_business_type",
                "goods_name", "spec", "unit", "quantity", "unit_price",
                "amount", "tax_rate", "tax_amount", "total_amount",
                "invoice_source", "invoice_category", "status", "is_positive", "invoice_risk_level",
                "issuer", "remark"
            ]
        elif module == "bookkeeping-invoice":
            # 记账发票25列（无认证信息列）
            field_order = [
                "invoice_code", "invoice_no", "digital_invoice_no",
                "seller_tax_no", "seller_name",
                "buyer_tax_no", "buyer_name",
                "invoice_date", "tax_category_code", "specific_business_type",
                "goods_name", "spec", "unit", "quantity", "unit_price",
                "amount", "tax_rate", "tax_amount", "total_amount",
                "invoice_source", "invoice_category", "status", "is_positive", "invoice_risk_level",
                "issuer", "remark"
            ]
        elif module == "bank-transaction":
            field_order = [
                "transaction_date", "transaction_time", "application_date",
                "voucher_no", "debit_amount", "credit_amount", "balance",
                "counterparty_account", "counterparty_name", "counterparty_bank",
                "transaction_serial_no", "voucher_seq", "record_status",
                "summary", "transaction_remark", "account_type"
            ]
        elif module == "input-vat-deduction":
            field_order = [
                "check_status", "invoice_source", "domestic_sale_cert_no",
                "digital_invoice_no", "invoice_code", "invoice_no",
                "invoice_date", "seller_tax_id", "seller_name",
                "amount", "tax_amount", "deductible_tax_amount",
                "invoice_category", "invoice_category_label", "invoice_status",
                "check_time", "risk_level"
            ]
        elif module == "employee":
            field_order = [
                "name", "id_card"
            ]
        elif module == "customer":
            field_order = [
                "name", "uscc"
            ]
        elif module == "supplier":
            field_order = [
                "name", "uscc"
            ]
        elif module == "department":
            field_order = [
                "code", "name"
            ]

        return {
            "file_name": fname,
            "headers": headers,
            "preview_rows": preview_rows,
            "total_rows": total_rows,
            "module": module,
            "field_groups": field_groups,
            "field_order": field_order
        }
    except HTTPException:
        raise
    except Exception as e:
        return {"error": f"文件分析失败：{str(e)}"}


@app.post("/api/file/import-with-mapping")
async def import_file_with_mapping(  # v2026-06-04-simplify: 进项发票改为单步导入
    file: UploadFile = File(...),
    module: str = Form("bank-transaction"),
    bank_config_id: Optional[int] = Form(None),
    column_mapping: str = Form(...),  # JSON: {标准字段: 文件列名}
    company_id: int = Form(...),
    force: str = Form("false"),
    db: Session = Depends(get_db)
):
    """根据列映射导入文件数据"""
    try:
        fname = file.filename or "unknown"
        ext = os.path.splitext(fname)[1].lower()
        if ext not in ALLOWED_EXTENSIONS:
            raise HTTPException(400, f"不支持的文件类型: {ext}，仅接受 xlsx/xls/csv/pdf/txt")
        content_bytes = await file.read()
        if len(content_bytes) > MAX_UPLOAD_SIZE:
            raise HTTPException(400, f"文件过大（{len(content_bytes)/1024/1024:.1f}MB），上限10MB")
        mapping = json.loads(column_mapping)
        force_dup = (force == "true")

        # 读取数据行
        rows_data = []
        if ext in (".xlsx", ".xls"):
            wb = openpyxl.load_workbook(io.BytesIO(content_bytes), data_only=True)
            ws = wb.active
            headers_file = []
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=col)
                headers_file.append(str(cell.value).strip() if cell.value is not None else f"列{col}")
            for row in range(2, ws.max_row + 1):
                row_dict = {}
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row, column=col)
                    if cell.value is None:
                        row_dict[headers_file[col - 1]] = ""
                    elif isinstance(cell.value, datetime):
                        row_dict[headers_file[col - 1]] = cell.value.strftime("%Y-%m-%d %H:%M:%S")
                    elif isinstance(cell.value, (int, float)):
                        # 数字直接转字符串（openpyxl data_only=True 已自动把真正的日期转 datetime，
                        # 此处 int/float 就是纯数字如金额、数量，误当日期序列号会销毁金额数据）
                        row_dict[headers_file[col - 1]] = str(cell.value)
                    else:
                        row_dict[headers_file[col - 1]] = str(cell.value).strip()
                # 跳过完全空行
                if any(v.strip() for v in row_dict.values()):
                    rows_data.append(row_dict)
        elif ext == ".csv":
            text = content_bytes.decode("utf-8-sig")
            reader = csv.reader(io.StringIO(text))
            all_rows = list(reader)
            if all_rows:
                headers_file = [h.strip() for h in all_rows[0]]
                for row in all_rows[1:]:
                    row_dict = {}
                    for i, h in enumerate(headers_file):
                        row_dict[h] = row[i] if i < len(row) else ""
                    if any(v.strip() for v in row_dict.values()):
                        rows_data.append(row_dict)

        # 根据映射转换并导入
        import_batch_id = str(uuid.uuid4())
        imported = 0
        errors = []
        infos = []  # 非错误提示（如自动创建客户档案）

        new_customers = {}  # {(tax_no, name): True} — 自动添加客户档案
        new_invoices = []  # 收集新创建的发票，导入完成后自动生成凭证
        new_deductions = []  # 收集新创建的进项抵扣，导入完成后自动生成凭证
        new_bank_tx_ids = []  # 收集新创建的银行流水ID，导入完成后自动生成凭证
        for i, row in enumerate(rows_data):
            try:
                mapped = {}
                extra = {}
                for std_field, file_col in mapping.items():
                    if file_col and file_col in row:
                        mapped[std_field] = row[file_col].strip()

                # 收集额外列（未映射的）
                for col_name, val in row.items():
                    if col_name not in mapping.values():
                        extra[col_name] = val.strip()


                if module == "bank-transaction":
                    # 解析日期
                    tx_date = None
                    date_str = mapped.get("transaction_date", "")
                    if date_str:
                        for fmt in ["%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y%m%d", "%m/%d/%Y", "%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S"]:
                            try:
                                tx_date = datetime.strptime(date_str, fmt).date()
                                break
                            except Exception: pass
                    if not tx_date:
                        errors.append(f"第{i+2}行: 无法解析日期")
                        continue

                    # 解析交易时间
                    tx_time = None
                    time_str = mapped.get("transaction_time", "")
                    if time_str:
                        for tf in ["%H:%M:%S", "%H:%M", "%H:%M:%S.%f"]:
                            try:
                                tx_time = datetime.strptime(time_str, tf).time()
                                break
                            except Exception: pass

                    # 解析申请日期
                    app_date = None
                    app_date_str = mapped.get("application_date", "")
                    if app_date_str:
                        for fmt in ["%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y%m%d"]:
                            try:
                                app_date = datetime.strptime(app_date_str, fmt).date()
                                break
                            except Exception: pass

                    # 解析借方/贷方金额
                    def parse_amt(key):
                        v = mapped.get(key, "0").replace(",", "").replace("￥", "").replace("¥", "")
                        try: return float(v) if v else 0.0
                        except Exception: return 0.0
                    debit_amount = parse_amt("debit_amount")
                    credit_amount = parse_amt("credit_amount")

                    # 余额
                    bal_str = mapped.get("balance", "0").replace(",", "").replace("￥", "").replace("¥", "")
                    balance = 0.0
                    try: balance = float(bal_str) if bal_str else 0.0
                    except Exception: pass

                    # 全行指纹去重
                    bt_fp_values = (
                        str(company_id), str(bank_config_id if bank_config_id else ""),
                        str(tx_date) if tx_date else "", str(tx_time) if tx_time else "",
                        str(app_date) if app_date else "",
                        str(mapped.get("voucher_no", "")), str(debit_amount), str(credit_amount),
                        str(balance),
                        str(mapped.get("counterparty_account", "")),
                        str(mapped.get("counterparty_name", "")),
                        str(mapped.get("counterparty_bank", "")),
                        str(mapped.get("transaction_serial_no", "")),
                        str(mapped.get("voucher_seq", "")),
                        str(mapped.get("record_status", "")),
                        str(mapped.get("summary", "")),
                        str(mapped.get("transaction_remark", "")),
                        str(mapped.get("account_type", "")),
                    )
                    bt_fp_raw = "|".join(bt_fp_values)
                    bt_fp = hashlib.sha256(bt_fp_raw.encode("utf-8")).hexdigest()
                    existing_bt = db.query(BankTransaction).filter(
                        BankTransaction.company_id == company_id,
                        BankTransaction._fingerprint == bt_fp
                    ).first()
                    if existing_bt and not force_dup:
                        errors.append(f"第{i+2}行: 数据重复，已跳过")
                        continue

                    tx = BankTransaction(
                        company_id=company_id,
                        bank_config_id=bank_config_id,
                        transaction_date=tx_date,
                        transaction_time=tx_time,
                        application_date=app_date,
                        voucher_no=mapped.get("voucher_no", ""),
                        debit_amount=debit_amount,
                        credit_amount=credit_amount,
                        balance=balance,
                        counterparty_account=mapped.get("counterparty_account", ""),
                        counterparty_name=mapped.get("counterparty_name", ""),
                        counterparty_bank=mapped.get("counterparty_bank", ""),
                        transaction_serial_no=mapped.get("transaction_serial_no", ""),
                        voucher_seq=mapped.get("voucher_seq", ""),
                        record_status=mapped.get("record_status", ""),
                        summary=mapped.get("summary", ""),
                        transaction_remark=mapped.get("transaction_remark", ""),
                        account_type=mapped.get("account_type", ""),
                        # 旧字段兼容
                        amount=credit_amount - debit_amount,
                        transaction_type="收入" if credit_amount > 0 else "支出",
                        raw_data=json.dumps(extra, ensure_ascii=False) if extra else "{}",
                        remark=mapped.get("remark", ""),
                        _fingerprint=bt_fp,
                    )
                    db.add(tx)
                    db.flush()
                    new_bank_tx_ids.append(tx.id)

                elif module in ("sales-invoice", "purchase-invoice", "bookkeeping-invoice"):
                    inv_date = None
                    date_str = mapped.get("invoice_date", "")
                    if date_str:
                        for fmt in ["%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y%m%d",
                                    "%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S",
                                    "%Y-%m-%d %H:%M", "%Y/%m/%d %H:%M"]:
                            try:
                                inv_date = datetime.strptime(date_str, fmt).date()
                                break
                            except Exception: pass

                    # 空值保留为 None，不拦截
                    inv_no = mapped.get("invoice_no", "") or None

                    
                    # 安全转浮点数——兼容千分位/百分号/空值/文本
                    # nullable=True 时源文件为空则返回 None，保留空白不填 0
                    def safe_float(val, default=0.0, nullable=False):
                        if val is None or str(val).strip() == "":
                            return None if nullable else default
                        s = str(val).strip().replace(",", "").replace("%", "").replace("￥", "").replace("¥", "").replace("元", "").replace(" ", "")
                        try:
                            return float(s)
                        except (ValueError, TypeError):
                            return None if nullable else default

                    amt = safe_float(mapped.get("amount"))
                    tax_amt = safe_float(mapped.get("tax_amount"), nullable=True)
                    total = safe_float(mapped.get("total_amount"), nullable=True)
                    # 三个字段互推：任意两个有值就能算出第三个
                    if amt is not None and tax_amt is not None and total is None:
                        total = round(amt + tax_amt, 2)
                    elif amt is not None and total is not None and tax_amt is None:
                        tax_amt = round(total - amt, 2)
                    elif tax_amt is not None and total is not None and amt is None:
                        amt = round(total - tax_amt, 2)
                    if total is None:
                        total = 0.0
                    if tax_amt is None:
                        tax_amt = 0.0
                    if amt is None:
                        amt = 0.0
                    qty = safe_float(mapped.get("quantity"), 0, nullable=True)
                    uprice = safe_float(mapped.get("unit_price"), 0, nullable=True)
                    tr = safe_float(mapped.get("tax_rate"))

                    if module == "sales-invoice":
                        # 计算全行指纹（仅用于审计，去重以票号为准）
                        fp_values = (
                            str(company_id), str(inv_no or ""), str(mapped.get("invoice_code", "")),
                            str(mapped.get("digital_invoice_no", "")),
                            str(mapped.get("seller_tax_no", "")), str(mapped.get("seller_name", "")),
                            str(mapped.get("buyer_tax_no", "")), str(mapped.get("buyer_name", "")),
                            str(inv_date) if inv_date else "",
                            str(mapped.get("tax_category_code", "")), str(mapped.get("specific_business_type", "")),
                            str(mapped.get("goods_name", "")), str(mapped.get("spec", "")),
                            str(mapped.get("unit", "")), str(qty), str(uprice),
                            str(amt), str(tr), str(tax_amt), str(total),
                            str(mapped.get("invoice_source", "")),
                            str(mapped.get("invoice_category", "增值税专用发票")),
                            str(mapped.get("status", "正常")),
                            str(mapped.get("is_positive", "是")),
                            str(mapped.get("invoice_risk_level", "")),
                            str(mapped.get("issuer", "")),
                            str(mapped.get("remark", "")),
                        )
                        fp_raw = "|".join(fp_values)
                        fp = hashlib.sha256(fp_raw.encode("utf-8")).hexdigest()

                        # ── 按票号去重（数电发票号 > 发票代码+号码） ──
                        _si_digital = str(mapped.get("digital_invoice_no", "")).strip()
                        _si_code = str(mapped.get("invoice_code", "")).strip()
                        _si_no = str(inv_no or "").strip()
                        if _si_digital:
                            _dup = db.query(SalesInvoice).filter(
                                    SalesInvoice.company_id == company_id,
                                    SalesInvoice.digital_invoice_no == _si_digital
                                ).first()
                        elif _si_no:
                            _q = db.query(SalesInvoice).filter(
                                    SalesInvoice.company_id == company_id,
                                    SalesInvoice.invoice_no == _si_no
                                )
                            if _si_code:
                                _q = _q.filter(SalesInvoice.invoice_code == _si_code)
                            _dup = _q.first()
                        else:
                            _dup = None
                        if _dup and not force_dup:
                            errors.append(f"第{i+2}行: 发票票号重复，已跳过")
                            continue
                        inv = SalesInvoice(
                            company_id=company_id, invoice_no=inv_no,
                            invoice_code=mapped.get("invoice_code", ""),
                            digital_invoice_no=mapped.get("digital_invoice_no", ""),
                            seller_tax_no=mapped.get("seller_tax_no", ""),
                            seller_name=mapped.get("seller_name", ""),
                            buyer_tax_no=mapped.get("buyer_tax_no", ""),
                            buyer_name=mapped.get("buyer_name", ""),
                            invoice_date=inv_date,
                            tax_category_code=mapped.get("tax_category_code", ""),
                            specific_business_type=mapped.get("specific_business_type", ""),
                            goods_name=mapped.get("goods_name", ""),
                            spec=mapped.get("spec", ""),
                            unit=mapped.get("unit", ""),
                            quantity=qty, unit_price=uprice,
                            amount=amt, tax_rate=tr, tax_amount=tax_amt,
                            total_amount=total,
                            invoice_source=mapped.get("invoice_source", ""),
                            invoice_category=mapped.get("invoice_category", "增值税专用发票"),
                            status=mapped.get("status", "正常"),
                            is_positive=mapped.get("is_positive", "是") in ("是", "true", "True", "1", True),
                            invoice_risk_level=mapped.get("invoice_risk_level", ""),
                            issuer=mapped.get("issuer", ""),
                            remark=mapped.get("remark", ""),
                            raw_data=json.dumps(extra) if extra else None,
                            _fingerprint=fp,
                        )
                        db.add(inv)
                        db.flush()
                        new_invoices.append(inv)
                        # 收集购买方信息，导入后自动添加客户档案
                        buyer_nm = mapped.get("buyer_name", "").strip()
                        buyer_tn = mapped.get("buyer_tax_no", "").strip()
                        if buyer_nm:
                            new_customers[(buyer_tn, buyer_nm)] = True
                    else:  # purchase-invoice
                        # ── 取得发票全指纹去重 ──
                        pi_fp_values = (
                            str(company_id), str(inv_no or ""), str(mapped.get("invoice_code", "")),
                            str(mapped.get("digital_invoice_no", "")),
                            str(mapped.get("seller_tax_no", "")), str(mapped.get("seller_name", "")),
                            str(mapped.get("buyer_tax_no", "")), str(mapped.get("buyer_name", "")),
                            str(inv_date) if inv_date else "",
                            str(mapped.get("tax_category_code", "")), str(mapped.get("specific_business_type", "")),
                            str(mapped.get("goods_name", "")), str(mapped.get("spec", "")),
                            str(mapped.get("unit", "")), str(qty), str(uprice),
                            str(amt), str(tr), str(tax_amt), str(total),
                            str(mapped.get("invoice_source", "")),
                            str(mapped.get("invoice_category", "增值税专用发票")),
                            str(mapped.get("status", "正常")),
                            str(mapped.get("is_positive", "是")),
                            str(mapped.get("invoice_risk_level", "")),
                            str(mapped.get("issuer", "")),
                            str(mapped.get("remark", "")),
                        )
                        pi_fp_raw = "|".join(pi_fp_values)
                        pi_fp = hashlib.sha256(pi_fp_raw.encode("utf-8")).hexdigest()
                        existing_pi = db.query(PurchaseInvoice).filter(
                            PurchaseInvoice.company_id == company_id,
                            PurchaseInvoice._fingerprint == pi_fp
                        ).first()
                        if existing_pi and not force_dup:
                            errors.append(f"第{i+2}行: 全指纹重复，已跳过")
                            continue
                        inv = PurchaseInvoice(
                            company_id=company_id, invoice_no=inv_no,
                            invoice_code=mapped.get("invoice_code", ""),
                            digital_invoice_no=mapped.get("digital_invoice_no", ""),
                            seller_tax_no=mapped.get("seller_tax_no", ""),
                            seller_name=mapped.get("seller_name", ""),
                            buyer_tax_no=mapped.get("buyer_tax_no", ""),
                            buyer_name=mapped.get("buyer_name", ""),
                            invoice_date=inv_date,
                            tax_category_code=mapped.get("tax_category_code", ""),
                            specific_business_type=mapped.get("specific_business_type", ""),
                            goods_name=mapped.get("goods_name", ""),
                            spec=mapped.get("spec", ""),
                            unit=mapped.get("unit", ""),
                            quantity=qty, unit_price=uprice,
                            amount=amt, tax_rate=tr, tax_amount=tax_amt,
                            total_amount=total,
                            invoice_source=mapped.get("invoice_source", ""),
                            invoice_category=mapped.get("invoice_category", "增值税专用发票"),
                            status=mapped.get("status", "正常"),
                            is_positive=mapped.get("is_positive", "是") in ("是", "true", "True", "1", True),
                            invoice_risk_level=mapped.get("invoice_risk_level", ""),
                            issuer=mapped.get("issuer", ""),
                            remark=mapped.get("remark", ""),
                            raw_data=json.dumps(extra) if extra else None,
                            _fingerprint=pi_fp,
                        )
                        db.add(inv)
                        db.flush()
                        new_invoices.append(inv)

                elif module == "bookkeeping-invoice":
                    # 记账发票导入（无认证相关字段）
                    bi_fp_values = (
                        str(company_id), str(inv_no or ""), str(mapped.get("invoice_code", "")),
                        str(mapped.get("digital_invoice_no", "")),
                        str(mapped.get("seller_tax_no", "")), str(mapped.get("seller_name", "")),
                        str(mapped.get("buyer_tax_no", "")), str(mapped.get("buyer_name", "")),
                        str(inv_date) if inv_date else "",
                        str(mapped.get("tax_category_code", "")), str(mapped.get("specific_business_type", "")),
                        str(mapped.get("goods_name", "")), str(mapped.get("spec", "")),
                        str(mapped.get("unit", "")), str(qty), str(uprice),
                        str(amt), str(tr), str(tax_amt), str(total),
                        str(mapped.get("invoice_source", "")),
                        str(mapped.get("invoice_category", "增值税普通发票")),
                        str(mapped.get("status", "正常")),
                        str(mapped.get("is_positive", "是")),
                        str(mapped.get("invoice_risk_level", "")),
                        str(mapped.get("issuer", "")),
                        str(mapped.get("remark", "")),
                    )
                    bi_fp_raw = "|".join(bi_fp_values)
                    bi_fp = hashlib.sha256(bi_fp_raw.encode("utf-8")).hexdigest()
                    existing_bi = db.query(BookkeepingInvoice).filter(
                        BookkeepingInvoice.company_id == company_id,
                        BookkeepingInvoice._fingerprint == bi_fp
                    ).first()
                    if existing_bi and not force_dup:
                        errors.append(f"第{i+2}行: 数据重复，已跳过")
                        continue
                    inv = BookkeepingInvoice(
                        company_id=company_id, invoice_no=inv_no,
                        invoice_code=mapped.get("invoice_code", ""),
                        digital_invoice_no=mapped.get("digital_invoice_no", ""),
                        seller_tax_no=mapped.get("seller_tax_no", ""),
                        seller_name=mapped.get("seller_name", ""),
                        buyer_tax_no=mapped.get("buyer_tax_no", ""),
                        buyer_name=mapped.get("buyer_name", ""),
                        invoice_date=inv_date,
                        tax_category_code=mapped.get("tax_category_code", ""),
                        specific_business_type=mapped.get("specific_business_type", ""),
                        goods_name=mapped.get("goods_name", ""),
                        spec=mapped.get("spec", ""),
                        unit=mapped.get("unit", ""),
                        quantity=qty, unit_price=uprice,
                        amount=amt, tax_rate=tr, tax_amount=tax_amt,
                        total_amount=total,
                        invoice_source=mapped.get("invoice_source", ""),
                        invoice_category=mapped.get("invoice_category", "增值税普通发票"),
                        status=mapped.get("status", "正常"),
                        is_positive=mapped.get("is_positive", "是") in ("是", "true", "True", "1", True),
                        invoice_risk_level=mapped.get("invoice_risk_level", ""),
                        issuer=mapped.get("issuer", ""),
                        remark=mapped.get("remark", ""),
                        raw_data=json.dumps(extra) if extra else None,
                        _fingerprint=bi_fp,
                    )
                    db.add(inv)
                    db.flush()
                    new_invoices.append(inv)

                elif module == "input-vat-deduction":
                    # 进项抵扣导入：解析日期
                    inv_date = None
                    date_str = mapped.get("invoice_date", "")
                    if date_str:
                        for fmt in ["%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y%m%d",
                                    "%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S"]:
                            try:
                                inv_date = datetime.strptime(date_str, fmt).date()
                                break
                            except Exception: pass

                    check_time = None
                    ct_str = mapped.get("check_time", "")
                    if ct_str:
                        for fmt in ["%Y-%m-%dT%H:%M", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M",
                                    "%Y/%m/%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S"]:
                            try:
                                check_time = datetime.strptime(ct_str, fmt)
                                break
                            except Exception: pass

                    amt = mapped.get("amount", "0").replace(",", "").replace("￥", "").replace("¥", "")
                    try: amt = float(amt) if amt else 0.0
                    except (ValueError, TypeError): amt = 0.0
                    tax_amt = mapped.get("tax_amount", "0").replace(",", "").replace("￥", "").replace("¥", "")
                    try: tax_amt = float(tax_amt) if tax_amt else 0.0
                    except (ValueError, TypeError): tax_amt = 0.0
                    deductible = mapped.get("deductible_tax_amount", "0").replace(",", "").replace("￥", "").replace("¥", "")
                    try: deductible = float(deductible) if deductible else 0.0
                    except (ValueError, TypeError): deductible = 0.0

                    # 全行指纹去重
                    ivd_fp_values = (
                        str(company_id),
                        str(mapped.get("check_status", "")),
                        str(mapped.get("invoice_source", "")),
                        str(mapped.get("domestic_sale_cert_no", "")),
                        str(mapped.get("digital_invoice_no", "")),
                        str(mapped.get("invoice_code", "")),
                        str(mapped.get("invoice_no", "")),
                        str(inv_date) if inv_date else "",
                        str(mapped.get("seller_tax_id", "")),
                        str(mapped.get("seller_name", "")),
                        str(amt), str(tax_amt), str(deductible),
                        str(mapped.get("invoice_category", "")),
                        str(mapped.get("invoice_category_label", "")),
                        str(mapped.get("invoice_status", "正常")),
                        str(check_time) if check_time else "",
                        str(mapped.get("risk_level", "正常")),
                    )
                    ivd_fp_raw = "|".join(ivd_fp_values)
                    ivd_fp = hashlib.sha256(ivd_fp_raw.encode("utf-8")).hexdigest()
                    existing_ivd = db.query(InputVATDeduction).filter(
                        InputVATDeduction.company_id == company_id,
                        InputVATDeduction._fingerprint == ivd_fp
                    ).first()
                    if existing_ivd and not force_dup:
                        errors.append(f"第{i+2}行: 数据重复，已跳过")
                        continue

                    inv = InputVATDeduction(
                        company_id=company_id,
                        check_status=mapped.get("check_status", ""),
                        invoice_source=mapped.get("invoice_source", ""),
                        domestic_sale_cert_no=mapped.get("domestic_sale_cert_no", ""),
                        digital_invoice_no=mapped.get("digital_invoice_no", ""),
                        invoice_code=mapped.get("invoice_code", ""),
                        invoice_no=mapped.get("invoice_no", ""),
                        invoice_date=inv_date,
                        seller_tax_id=mapped.get("seller_tax_id", ""),
                        seller_name=mapped.get("seller_name", ""),
                        amount=amt,
                        tax_amount=tax_amt,
                        deductible_tax_amount=deductible,
                        invoice_category=mapped.get("invoice_category", ""),
                        invoice_category_label=mapped.get("invoice_category_label", ""),
                        invoice_status=mapped.get("invoice_status", "正常"),
                        check_time=check_time,
                        risk_level=mapped.get("risk_level", "正常"),
                        remark=mapped.get("remark", ""),
                        raw_data=json.dumps(extra) if extra else None,
                        import_batch_id=import_batch_id,
                        _fingerprint=ivd_fp,
                    )
                    db.add(inv)
                    db.flush()
                    new_deductions.append(inv)

                elif module == "employee":
                    name = mapped.get("name", "").strip()
                    if not name:
                        errors.append(f"第{i+2}行: 姓名不能为空")
                        continue
                    # P1-4: 通用导入检查 id_card 去重
                    id_card = mapped.get("id_card", "").strip() or None
                    # 编码自动生成 RY001 格式：首次查DB取最大code，后续内存递增
                    if 'emp_code_counter' not in locals():
                        existing_codes = db.query(Employee.code).filter(
                            Employee.company_id == company_id,
                            Employee.code.like('RY%')
                        ).all()
                        emp_code_counter = 0
                        for c in existing_codes:
                            try:
                                num = int(c[0][2:])
                                if num > emp_code_counter:
                                    emp_code_counter = num
                            except Exception: pass
                    emp_code_counter += 1
                    code = f"RY{emp_code_counter:03d}"
                    emp = Employee(
                        company_id=company_id, code=code, name=name,
                        id_card=mapped.get("id_card", "") or None
                    )
                    db.add(emp)
                    db.flush()

                elif module == "customer":
                    name = mapped.get("name", "").strip()
                    if not name:
                        errors.append(f"第{i+2}行: 客户名称不能为空")
                        continue
                    # 编码自动生成 KH001 格式：首次查DB取最大code，后续内存递增
                    if 'cust_code_counter' not in locals():
                        existing_codes = db.query(Customer.code).filter(
                            Customer.company_id == company_id,
                            Customer.code.like('KH%')
                        ).all()
                        cust_code_counter = 0
                        for c in existing_codes:
                            try:
                                num = int(c[0][2:])
                                if num > cust_code_counter:
                                    cust_code_counter = num
                            except Exception: pass
                    cust_code_counter += 1
                    code = f"KH{cust_code_counter:03d}"
                    uscc = mapped.get("uscc", "").strip() or None
                    # 计算全行指纹
                    fp_values = (
                        str(company_id),
                        str(code),
                        str(name),
                        str(uscc or ""),
                        str(mapped.get("tax_no", "") or ""),
                        str(mapped.get("contact", "") or ""),
                        str(mapped.get("phone", "") or ""),
                        str(mapped.get("address", "") or ""),
                        str(mapped.get("credit_limit", "") or ""),
                        str(mapped.get("payment_terms", "") or ""),
                        str(mapped.get("bank_name", "") or ""),
                        str(mapped.get("bank_account", "") or ""),
                        str(True),  # is_active 默认为 True
                        str(mapped.get("remark", "") or "")
                    )
                    fp_raw = "|".join(fp_values)
                    fp = hashlib.sha256(fp_raw.encode("utf-8")).hexdigest()
                    # 去重检查
                    existing = db.query(Customer).filter(
                        Customer.company_id == company_id,
                        Customer._fingerprint == fp
                    ).first()
                    if existing and not force_dup:
                        errors.append(f"第{i+2}行: 数据重复，已跳过")
                        continue
                    cust = Customer(
                        company_id=company_id, code=code, name=name,
                        uscc=uscc,
                        tax_no=mapped.get("tax_no", "") or None,
                        contact=mapped.get("contact", "") or None,
                        phone=mapped.get("phone", "") or None,
                        address=mapped.get("address", "") or None,
                        credit_limit=float(mapped.get("credit_limit", 0) or 0),
                        payment_terms=int(mapped.get("payment_terms", 30) or 30),
                        bank_name=mapped.get("bank_name", "") or None,
                        bank_account=mapped.get("bank_account", "") or None,
                        is_active=True,
                        remark=mapped.get("remark", "") or None,
                        _fingerprint=fp
                    )
                    db.add(cust)
                    db.flush()

                elif module == "supplier":
                    name = mapped.get("name", "").strip()
                    if not name:
                        errors.append(f"第{i+2}行: 供应商名称不能为空")
                        continue
                                        # 编码自动生成 GYS001 格式：首次查DB取最大code，后续内存递增
                    if 'supp_code_counter' not in locals():
                        existing_codes = db.query(Supplier.code).filter(
                            Supplier.company_id == company_id,
                            Supplier.code.like('GYS%')
                        ).all()
                        supp_code_counter = 0
                        for c in existing_codes:
                            try:
                                num = int(c[0][3:])
                                if num > supp_code_counter:
                                    supp_code_counter = num
                            except Exception: pass
                    supp_code_counter += 1
                    code = f"GYS{supp_code_counter:03d}"
                    uscc = mapped.get("uscc", "") or None
                    supp = Supplier(
                        company_id=company_id, code=code, name=name,
                        uscc=uscc
                    )
                    db.add(supp)
                    db.flush()

                elif module == "department":
                    name = mapped.get("name", "").strip()
                    if not name:
                        errors.append(f"第{i+2}行: 部门名称不能为空")
                        continue
                                        # 编码：优先用导入的编码，为空则自动生成 BM001 格式
                    code = mapped.get("code", "").strip()
                    if not code:
                        if 'dept_code_counter' not in locals():
                            existing_codes = db.query(Department.code).filter(
                                Department.company_id == company_id,
                                Department.code.like('BM%')
                            ).all()
                            dept_code_counter = 0
                            for c in existing_codes:
                                try:
                                    num = int(c[0][2:])
                                    if num > dept_code_counter:
                                        dept_code_counter = num
                                except Exception: pass
                        dept_code_counter += 1
                        code = f"BM{dept_code_counter:03d}"
                    # 编码去重：同编码覆盖更新
                    existing = db.query(Department).filter(
                        Department.company_id == company_id, Department.code == code
                    ).first()
                    if existing:
                        existing.name = name
                    else:
                        db.add(Department(
                            company_id=company_id, code=code, name=name
                        ))
                    db.flush()

                imported += 1
            except Exception as e:
                errors.append(f"第{i+2}行: {str(e)}")

        # 自动添加客户档案（仅开具发票导入时）
        if module == "sales-invoice" and new_customers:
            customer_added = 0
            for (tax_no, name) in new_customers:
                # 先按税号匹配，再按名称匹配
                existing = None
                if tax_no:
                    existing = db.query(Customer).filter(
                        Customer.company_id == company_id,
                        Customer.tax_no == tax_no
                    ).first()
                if not existing:
                    existing = db.query(Customer).filter(
                        Customer.company_id == company_id,
                        Customer.name == name
                    ).first()
                if not existing:
                    # 自动生成编码（用局部变量递增，避免同事务内重复）
                    if 'next_cust_idx' not in locals():
                        existing_count = db.query(Customer).filter(
                            Customer.company_id == company_id
                        ).count()
                        next_cust_idx = existing_count + 1
                    code = f"KH{next_cust_idx:03d}"
                    next_cust_idx += 1
                    cust = Customer(
                        company_id=company_id,
                        code=code,
                        name=name,
                        uscc=tax_no or None,   # 购方识别号 → 统一社会信用代码
                        tax_no=tax_no or None
                    )
                    db.add(cust)
                    customer_added += 1
            if customer_added > 0:
                infos.append(f"自动新增 {customer_added} 个客户到客户档案")

        db.commit()
        return {
            "imported": imported,
            "total": len(rows_data),
            "skipped": len(rows_data) - imported,
            "errors": errors[:20],
            "infos": infos[:20],
            "message": "成功导入 " + str(imported) + "/" + str(len(rows_data)) + " 条记录"
        }
    except Exception as e:
        db.rollback()
        return {"error": f"导入失败：{str(e)}"}


# ==================== 信息真实性校验（公用的校验工具） ====================

def validate_uscc(code: str) -> tuple:
    """校验统一社会信用代码 (GB 32100-2015)"""
    if not code or not code.strip():
        return True, ""
    code = code.strip().upper()
    if len(code) != 18:
        return False, "统一社会信用代码必须为18位"
    if not re.match(r'^[0-9A-HJ-NPQRTUWXY]{2}\d{6}[0-9A-HJ-NPQRTUWXY]{10}$', code):
        return False, "统一社会信用代码格式不正确（应为：2位登记管理机关+6位组织机构代码+9位主体标识码+1位校验码）"
    char_map = '0123456789ABCDEFGHJKLMNPQRTUWXY'
    weights = [1, 3, 9, 27, 19, 26, 16, 17, 20, 29, 25, 13, 8, 24, 10, 30, 28]
    total = sum(char_map.index(code[i]) * weights[i] for i in range(17))
    check_idx = (31 - total % 31) % 31
    expected = char_map[check_idx]
    if code[17] != expected:
        return False, f"校验码不正确，应为 '{expected}'"
    return True, ""


def validate_id_card(card_no: str) -> tuple:
    """校验中国居民身份证号码 (GB 11643-1999)"""
    if not card_no or not card_no.strip():
        return True, ""
    card_no = card_no.strip().upper()
    if len(card_no) != 18:
        return False, "身份证号码必须为18位"
    if not re.match(r'^\d{17}[\dX]$', card_no):
        return False, "身份证号码前17位必须为数字，第18位为数字或X"
    try:
        birth_str = card_no[6:14]
        birth = date(int(birth_str[0:4]), int(birth_str[4:6]), int(birth_str[6:8]))
        if birth >= date.today():
            return False, "身份证号码中的出生日期不能晚于当前日期"
    except ValueError:
        return False, "身份证号码中的出生日期无效（应为YYYYMMDD格式）"
    weights = [7, 9, 10, 5, 8, 4, 2, 1, 6, 3, 7, 9, 10, 5, 8, 4, 2]
    check_chars = '10X98765432'
    total = sum(int(card_no[i]) * weights[i] for i in range(17))
    expected = check_chars[total % 11]
    if card_no[17] != expected:
        return False, f"身份证校验码不正确，应为 '{expected}'"
    return True, ""



# ==================== Chat AI 助手模块 ====================
from chat import router as chat_router
app.include_router(chat_router)

# ==================== 涉税风险规则：从浏览器 localStorage 导出到服务器 ====================
import json as _json
from pathlib import Path as _Path
from fastapi import Request

@app.post("/api/tax-risk-rules/save-local")
async def tax_risk_rules_save_local(request: Request):
    """接收浏览器 localStorage 中的涉税风险规则 JSON，保存到服务器文件"""
    dst = _Path("static/tax_risk_rules_local_export.json")
    try:
        data = await request.json()
        dst.write_text(_json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
        return {"ok": True, "count": len(data), "path": str(dst)}
    except Exception as e:
        return {"ok": False, "error": str(e)}

# ==================== 涉税风险规则审计 API ====================
@app.post("/api/tax-risk-rules/audit")
async def tax_risk_rules_audit(request: Request):
    """接收当前规则 JSON，返回 8 层质量审计报告"""
    try:
        data = await request.json()
    except Exception:
        return {"ok": False, "error": "无效的 JSON 数据"}

    from difflib import SequenceMatcher as _SeqMatcher
    from collections import Counter as _Counter
    import re as _re

    report = {"ok": True, "total": len(data), "layers": [], "summary": {}}
    issues_found = []

    # --- 第1层: ID和名称精确去重 ---
    ids = [r["id"] for r in data]
    dup_ids = [i for i in ids if ids.count(i) > 1]
    items = [r["item"] for r in data]
    dup_names = {k: v for k, v in _Counter(items).items() if v > 1}
    layer1 = {"name": "ID/名称精确去重", "pass": not dup_ids and not dup_names}
    if dup_ids:
        layer1["detail"] = f"重复ID: {list(set(dup_ids))}"
    if dup_names:
        layer1["detail"] = f"重复名称: {dup_names}"
    report["layers"].append(layer1)
    if not layer1["pass"]:
        issues_found.append("ID/名称去重")

    # --- 第2层: 名称相似度 (>=85%) ---
    sim_names = []
    for i in range(len(data)):
        for j in range(i + 1, len(data)):
            ratio = _SeqMatcher(None, data[i]["item"], data[j]["item"]).ratio()
            if ratio >= 0.85:
                sim_names.append({
                    "ratio": round(ratio, 2),
                    "a": data[i]["item"], "a_cat": data[i]["category"],
                    "b": data[j]["item"], "b_cat": data[j]["category"]
                })
    layer2 = {"name": "名称相似度检查 (≥85%)", "pass": len(sim_names) == 0}
    if sim_names:
        layer2["detail"] = sim_names
        issues_found.append("名称相似度")
    report["layers"].append(layer2)

    # --- 第3层: detail 相似度 (>=80%) ---
    by_cat = {}
    for r in data:
        by_cat.setdefault(r["category"], []).append(r)
    sim_detail = []
    # 同分类
    for cat, rules in by_cat.items():
        for i in range(len(rules)):
            for j in range(i + 1, len(rules)):
                ratio = _SeqMatcher(None, rules[i]["detail"], rules[j]["detail"]).ratio()
                if ratio >= 0.80:
                    sim_detail.append({
                        "type": "同分类", "cat": cat, "ratio": round(ratio, 2),
                        "a": rules[i]["item"], "b": rules[j]["item"]
                    })
    # 跨分类
    for i in range(len(data)):
        for j in range(i + 1, len(data)):
            if data[i]["category"] != data[j]["category"]:
                ratio = _SeqMatcher(None, data[i]["detail"], data[j]["detail"]).ratio()
                if ratio >= 0.80:
                    sim_detail.append({
                        "type": "跨分类", "ratio": round(ratio, 2),
                        "a": f"{data[i]['item']}({data[i]['category']})",
                        "b": f"{data[j]['item']}({data[j]['category']})"
                    })
    layer3 = {"name": "detail 相似度检查 (≥80%)", "pass": len(sim_detail) == 0}
    if sim_detail:
        layer3["detail"] = sim_detail
        issues_found.append("detail相似度")
    report["layers"].append(layer3)

    # --- 第4层: 语义同类跨分类扫描 ---
    keyword_groups = {
        "零申报/零税额": ["零申报", "零税额"],
        "留抵退税/留抵": ["留抵退税", "留抵", "进项留抵"],
        "红冲/作废": ["红冲", "作废"],
        "开票限额/顶额": ["顶额", "开票限额"],
        "进项转出": ["进项转出", "进项税额转出"],
        "发票跨期": ["跨期", "跨年"],
        "税负率": ["税负率"],
        "咨询费/服务费": ["咨询", "服务费"],
        "资金回流": ["资金回流"],
    }
    sem_overlaps = []
    for group_name, keywords in keyword_groups.items():
        matches = []
        seen = set()
        for kw in keywords:
            for r in data:
                combined = r["item"] + r["detail"]
                if kw in combined and r["item"] not in seen:
                    seen.add(r["item"])
                    matches.append({"item": r["item"], "category": r["category"]})
        cats = set(m["category"] for m in matches)
        if len(matches) > 1 and len(cats) > 1:
            sem_overlaps.append({"group": group_name, "categories": list(cats), "count": len(matches), "items": matches})
    layer4 = {"name": "语义同类跨分类扫描", "pass": True}
    if sem_overlaps:
        layer4["detail"] = sem_overlaps
    report["layers"].append(layer4)

    # --- 第5层: 碎片分类 (<2条) ---
    cats = _Counter(r["category"] for r in data)
    fragments = {cat: cnt for cat, cnt in cats.items() if cnt < 2}
    layer5 = {"name": "碎片分类检测 (<2条)", "pass": len(fragments) == 0}
    if fragments:
        frag_list = []
        for cat, cnt in fragments.items():
            citems = [r["item"] for r in data if r["category"] == cat]
            frag_list.append({"category": cat, "count": cnt, "items": citems})
        layer5["detail"] = frag_list
        issues_found.append("碎片分类")
    report["layers"].append(layer5)

    # --- 第6层: 归类不当 ---
    # tax_map: 税种关键词 → 允许的分类列表
    # 判断逻辑：如果规则detail/suggestion中出现某税种关键词，但分类不在允许列表中 → 标记为归类不当
    # 以下已根据实际业务关系做了合理豁免：
    #   - 城建税必然关联增值税；资金往来/隐匿虚增必然关联个税；
    #   - 税负水平关联所有税种；征管风险常涉及进项税额；
    #   - 发票深度分析影响多个税种；经营实质涉及增值税认定；
    #   - 企业所得税分类中未分配利润规则涉及规避股东个税。
    tax_map = {
        "增值税": ["增值税专项", "申报比对", "发票合规", "发票异常", "发票深度", "税负水平", "城建税", "经营实质"],
        "进项税额": ["增值税专项", "申报比对", "发票合规", "发票异常", "发票深度", "交易特征", "征管风险"],
        "销项税额": ["增值税专项", "申报比对", "发票合规", "发票异常", "发票深度"],
        "企业所得税": ["企业所得税", "纳税调整", "成本结构", "财务健康", "税负水平", "发票深度"],
        "汇算清缴": ["企业所得税", "纳税调整", "个人所得税"],
        "纳税调增": ["企业所得税", "纳税调整", "成本结构"],
        "个人所得税": ["个人所得税", "企业所得税"],
        "个税": ["个人所得税", "薪酬福利", "资金往来", "隐匿虚增", "发票深度", "企业所得税"],
        "代扣代缴": ["个人所得税"],
    }
    mismatches = []
    for r in data:
        detail = r["detail"] + r["suggestion"]
        for tax_kw, allowed_cats in tax_map.items():
            if tax_kw in detail and r["category"] not in allowed_cats:
                mismatches.append({"item": r["item"], "category": r["category"], "keyword": tax_kw})
                break
    layer6 = {"name": "归类不当检测", "pass": len(mismatches) == 0}
    if mismatches:
        layer6["detail"] = mismatches
        issues_found.append("归类不当")
    report["layers"].append(layer6)

    # --- 第7层: level 一致性 ---
    valid_levels = {"高风险", "中风险", "低风险", "良好"}
    bad_levels = []
    for r in data:
        if r["level"] not in valid_levels:
            bad_levels.append({"item": r["item"], "level": r["level"]})
    layer7 = {"name": "level 字段一致性", "pass": len(bad_levels) == 0}
    if bad_levels:
        layer7["detail"] = bad_levels
        issues_found.append("level不一致")
    report["layers"].append(layer7)

    # --- 第8层: 评分跨度 ---
    by_cat2 = {}
    for r in data:
        by_cat2.setdefault(r["category"], []).append(r["score"])
    wide_cats = []
    for cat, scores in sorted(by_cat2.items()):
        if len(scores) > 1 and max(scores) - min(scores) >= 5:
            wide_cats.append({"category": cat, "min": min(scores), "max": max(scores), "spread": max(scores) - min(scores)})
    layer8 = {"name": "同分类评分跨度检查 (≥5分)", "pass": len(wide_cats) == 0}
    if wide_cats:
        layer8["detail"] = wide_cats
    report["layers"].append(layer8)

    # --- 汇总 ---
    cats_all = _Counter(r["category"] for r in data)
    levels_all = _Counter(r["level"] for r in data)
    scores_all = [r["score"] for r in data]
    report["summary"] = {
        "total_rules": len(data),
        "total_categories": len(cats_all),
        "level_distribution": dict(levels_all),
        "score_range": f"{min(scores_all)}~{max(scores_all)}",
        "avg_score": round(sum(scores_all) / len(scores_all), 1),
        "category_distribution": dict(cats_all.most_common()),
        "issues_found": issues_found,
        "all_clear": len(issues_found) == 0
    }
    return report

# ==================== 涉税风险规则自动修复 API ====================
@app.post("/api/tax-risk-rules/fix")
async def tax_risk_rules_fix(request: Request):
    """接收当前规则 JSON，自动修复可修复的问题，返回修复后规则"""
    try:
        data = await request.json()
    except Exception:
        return {"ok": False, "error": "无效的 JSON 数据"}

    from difflib import SequenceMatcher as _SeqMatcher
    from collections import Counter as _Counter
    import copy as _copy

    rules = _copy.deepcopy(data)
    fixes = []
    skipped = []

    # ========== 修复1: 碎片分类 → 合并到语义最相关的分类 ==========
    cat_counts = _Counter(r["category"] for r in rules)
    fragments = {cat: cnt for cat, cnt in cat_counts.items() if cnt < 2}

    # 碎片合并映射表：碎片分类 → 最相关的目标分类
    fragment_merge_map = {
        "印花税": "税负水平",
        "行业专项": "经营实质",
        "城建税": "税负水平",
        "房产税": "税负水平",
        "客户穿透": "交易特征",
        "供应商穿透": "交易特征",
        "政策执行": "征管风险",
    }

    if fragments:
        for frag_cat in fragments:
            target = None
            if frag_cat in fragment_merge_map:
                target = fragment_merge_map[frag_cat]
            else:
                # 默认：按名称相似度找最匹配的非碎片分类
                best = (0, None)
                for cat in cat_counts:
                    if cat != frag_cat and cat_counts[cat] >= 2:
                        ratio = _SeqMatcher(None, frag_cat, cat).ratio()
                        if ratio > best[0]:
                            best = (ratio, cat)
                if best[1]:
                    target = best[1]

            if target:
                cnt = 0
                for r in rules:
                    if r["category"] == frag_cat:
                        r["category"] = target
                        cnt += 1
                fixes.append(f"碎片合并: {frag_cat}({cnt}条) → {target}")

    # ========== 修复2: 归类不当 → 重新分配 ==========
    tax_map = {
        "增值税": ["增值税专项", "申报比对", "发票合规", "发票异常", "发票深度", "税负水平", "城建税", "经营实质"],
        "进项税额": ["增值税专项", "申报比对", "发票合规", "发票异常", "发票深度", "交易特征", "征管风险"],
        "销项税额": ["增值税专项", "申报比对", "发票合规", "发票异常", "发票深度"],
        "企业所得税": ["企业所得税", "纳税调整", "成本结构", "财务健康", "税负水平", "发票深度"],
        "汇算清缴": ["企业所得税", "纳税调整", "个人所得税"],
        "纳税调增": ["企业所得税", "纳税调整", "成本结构"],
        "个人所得税": ["个人所得税", "企业所得税"],
        "个税": ["个人所得税", "薪酬福利", "资金往来", "隐匿虚增", "发票深度", "企业所得税"],
        "代扣代缴": ["个人所得税"],
    }

    # 关键词→首选分类映射（当多个允许时选第一个）
    keyword_preferred = {
        "增值税": "增值税专项",
        "进项税额": "增值税专项",
        "销项税额": "增值税专项",
        "企业所得税": "企业所得税",
        "汇算清缴": "纳税调整",
        "纳税调增": "纳税调整",
        "个人所得税": "个人所得税",
        "个税": "个人所得税",
        "代扣代缴": "个人所得税",
    }

    for r in rules:
        detail = r["detail"] + r["suggestion"]
        for tax_kw, allowed_cats in tax_map.items():
            if tax_kw in detail and r["category"] not in allowed_cats:
                # 找到关键词 → 选首选分类
                preferred = keyword_preferred.get(tax_kw, allowed_cats[0])
                old_cat = r["category"]
                r["category"] = preferred
                fixes.append(f"归类纠正: '{r['item'][:30]}' {old_cat} → {preferred} (关键词: {tax_kw})")
                break  # 只修第一个触发的

    # ========== 修复3: level 标准化 ==========
    level_map = {
        "高": "高风险", "中": "中风险", "低": "低风险",
        "较高": "高风险", "较低": "低风险", "中等风险": "中风险",
        "高危": "高风险",
    }
    for r in rules:
        if r["level"] in level_map:
            old = r["level"]
            r["level"] = level_map[old]
            fixes.append(f"级别标准化: '{r['item'][:30]}' {old} → {r['level']}")

    # ========== 重新生成审计报告 ==========
    # 轻量审计（仅检查是否还有问题）
    cat_counts2 = _Counter(r["category"] for r in rules)
    fragments2 = {cat: cnt for cat, cnt in cat_counts2.items() if cnt < 2}
    mismatches2 = []
    for r in rules:
        detail = r["detail"] + r["suggestion"]
        for tax_kw, allowed_cats in tax_map.items():
            if tax_kw in detail and r["category"] not in allowed_cats:
                mismatches2.append(r["item"])
                break

    remaining = []
    if fragments2:
        remaining.append(f"还有 {len(fragments2)} 个碎片分类需手动处理")
        skipped.extend([f"{cat}({cnt}条)" for cat, cnt in fragments2.items()])
    if mismatches2:
        remaining.append(f"还有 {len(mismatches2)} 项归类不当需手动处理")
        skipped.extend(mismatches2[:5])

    all_fixed = len(fragments2) == 0 and len(mismatches2) == 0

    return {
        "ok": True,
        "fixed_rules": rules,
        "fixes_applied": fixes,
        "fixes_count": len(fixes),
        "remaining_issues": remaining,
        "skipped_items": skipped,
        "all_fixed": all_fixed,
        "summary": {
            "total": len(rules),
            "categories": len(cat_counts2),
            "category_distribution": dict(cat_counts2.most_common()),
        }
    }


# =================== 涉税风险规则报告解析 API ===================
def _parse_tax_report_text(report_text: str):
    """核心解析逻辑，供文本和文件上传两个端点共用"""
    import re as _re
    import uuid as _uuid
    from difflib import SequenceMatcher as _SeqMatcher

    text = report_text

    # === 第1步：文本预处理 ===
    lines = text.split('\n')
    cleaned = []
    for line in lines:
        line = line.strip()
        if not line:
            cleaned.append('')
        elif _re.match(r'^[\d\s\-–—_・•·]+$', line):
            continue
        elif len(line) <= 3 and _re.match(r'^\d+$', line):
            continue
        else:
            cleaned.append(line)
    text = '\n'.join(cleaned)

    # === 第2步：智能分段 ===
    paragraphs = []
    current_para = []
    for line in text.split('\n'):
        line = line.strip()
        if not line:
            if current_para:
                paragraphs.append('\n'.join(current_para))
                current_para = []
            continue
        is_new_rule = False
        if _re.match(r'^[\(]?\d+[\)\.]?[\、\.)]', line):
            is_new_rule = True
        elif _re.match(r'^[一二三四五六七八九十]+[、\.]', line):
            is_new_rule = True
        elif any(kw in line for kw in ['风险点', '风险分析']) and len(line) < 40:
            is_new_rule = True
        elif '问题' in line and len(line) < 30:
            is_new_rule = True
        if is_new_rule and current_para:
            paragraphs.append('\n'.join(current_para))
            current_para = [line]
        else:
            current_para.append(line)
    if current_para:
        paragraphs.append('\n'.join(current_para))

    if len(paragraphs) < 2:
        sentences = _re.split(r'[。\.\!\?！？]+', text)
        paragraphs = []
        chunk = []
        for s in sentences:
            s = s.strip()
            if s:
                chunk.append(s)
                if len(chunk) >= 2:
                    paragraphs.append('。'.join(chunk) + '。')
                    chunk = []
        if chunk:
            paragraphs.append('。'.join(chunk))

    # === 第3步：提取规则 ===
    category_keywords = {
        "发票合规": ["发票", "进项", "销项", "税号", "全电", "数电", "红冲", "作废"],
        "发票异常": ["顶额", "作废", "红冲", "异常", "失控", "虚开"],
        "发票深度": ["油费", "运输费", "生活用品", "水电", "能耗", "进销"],
        "增值税专项": ["增值税", "留抵", "退税", "简易计税", "免税"],
        "企业所得税": ["所得税", "折旧", "摊销", "准备金", "不征税"],
        "纳税调整": ["招待费", "广告费", "业务招待", "纳税调增"],
        "个人所得税": ["个税", "工资", "薪金", "分红", "股东", "借款"],
        "成本结构": ["成本", "收入", "费用", "毛利率", "占比"],
        "经营实质": ["经营能力", "开票量", "注册地址"],
        "资金往来": ["公户", "私户", "转账", "资金回流"],
        "隐匿虚增": ["其他应收", "其他应付", "挂账", "隐瞒"],
        "财务健康": ["现金流", "偿债", "净资产", "利润率"],
        "征管风险": ["欠税", "走逃", "失联", "D级", "非正常户"],
        "申报比对": ["零申报", "比对", "未申报", "漏申报"],
        "税负水平": ["税负率", "印花税", "行业税负"],
        "交易特征": ["整数", "大额", "频繁", "同一", "回流"],
        "账务数据": ["借贷", "凭证", "序时账", "记账"],
    }

    def _auto_classify(text_content):
        best = ("其他", 0)
        for cat, kws in category_keywords.items():
            score = sum(1 for kw in kws if kw in text_content)
            if score > best[1]:
                best = (cat, score)
        return best[0] if best[1] > 0 else "其他"

    def _estimate_score(text_content):
        m = _re.search(r'评分[：:\s]*(\d+)', text_content)
        if m:
            return int(m.group(1))
        m = _re.search(r'(\d+)\s*分', text_content)
        if m and int(m.group(1)) <= 10:
            return int(m.group(1))
        high_kws = ['虚开', '偷税', '逃税', '隐瞒', '涉嫌', '不得', '禁止']
        mid_kws = ['异常', '偏高', '偏低', '超标', '不匹配', '未', '漏']
        if any(kw in text_content for kw in high_kws):
            return 8
        elif any(kw in text_content for kw in mid_kws):
            return 5
        return 5

    def _level_from_score(score):
        if score >= 7:
            return "高风险"
        elif score >= 4:
            return "中风险"
        elif score > 0:
            return "低风险"
        return "良好"

    cat_icon_map = {
        "发票合规": "🧾", "发票异常": "⚠️", "发票深度": "🔍",
        "增值税专项": "🧮", "企业所得税": "💰", "纳税调整": "⚖️",
        "个人所得税": "👤", "成本结构": "📐", "经营实质": "🏭",
        "资金往来": "💸", "隐匿虚增": "🫥", "财务健康": "💊",
        "征管风险": "🚨", "申报比对": "📊", "税负水平": "📉",
        "交易特征": "🔗", "账务数据": "📊", "其他": "📋",
    }

    rules = []
    seen_items = set()
    for para in paragraphs:
        if len(para) < 10:
            continue
        lines_para = [l.strip() for l in para.split('\n') if l.strip()]
        if not lines_para:
            continue
        first_line = _re.sub(r'^[\(\[\d]+[\)\.\、\.]?\s*', '', lines_para[0])
        first_line = _re.sub(r'^[一二三四五六七八九十]+[、\.\s]*', '', first_line)
        item = first_line[:40] if len(first_line) > 5 else first_line
        if not item or len(item) < 3:
            continue
        is_dup = False
        for seen in seen_items:
            if _SeqMatcher(None, item, seen).ratio() > 0.8:
                is_dup = True
                break
        if is_dup:
            continue
        seen_items.add(item)
        category = _auto_classify(para)
        _override_map = {
            "零申报": "申报比对", "留抵退税": "增值税专项", "出口退税": "增值税专项",
            "油费": "发票深度", "运输费": "发票深度", "水电": "发票深度",
            "走逃": "征管风险", "非正常户": "征管风险", "D级": "征管风险",
            "生活用品": "发票深度", "能耗": "发票深度", "进销": "发票深度",
            "印花税": "税负水平", "个税": "个人所得税",
            "税负率": "税负水平",
        }
        for _kw, _correct_cat in _override_map.items():
            if _kw in para and category != _correct_cat:
                category = _correct_cat
                break
        score = _estimate_score(para)
        level = _level_from_score(score)
        detail = para[:200] + ('...' if len(para) > 200 else '')
        suggestion = ""
        sug_match = _re.search(r'建议[：:\s]*(.+)', para)
        if sug_match:
            suggestion = sug_match.group(1)[:150]
        rules.append({
            "id": str(_uuid.uuid4()),
            "category": category,
            "categoryIcon": cat_icon_map.get(category, "📋"),
            "item": item,
            "detail": detail,
            "score": score,
            "level": level,
            "suggestion": suggestion,
            "urgency": "提醒" if score < 5 else ("紧急" if score >= 8 else "高"),
            "evidence": "",
            "dataSource": "报告解析",
            "remark": f"从报告解析（{len(para)}字）"
        })

    return {
        "ok": True,
        "rules": rules,
        "count": len(rules),
        "paragraphs_found": len(paragraphs),
        "text_length": len(report_text)
    }


# ══════════════════════════════════════════════════════════════
#  涉税内容相关性检测
# ══════════════════════════════════════════════════════════════
def _check_tax_relevance(text: str):
    """检测文本是否与涉税内容相关，返回相关性评分和详情"""
    import re as _re

    if not text or len(text.strip()) < 30:
        return {
            "is_tax_related": False, "score": 0,
            "keywords_found": [],
            "message": "文本过短，无法判断是否涉税内容"
        }

    # 涉税关键词体系（三层权重：强/中/弱）
    tax_keywords = {
        "strong": [
            "增值税", "企业所得税", "个人所得税", "消费税", "印花税",
            "房产税", "契税", "土地增值税", "城建税", "教育费附加",
            "进项税额", "销项税额", "进项税", "销项税", "留抵退税",
            "纳税申报", "税务稽查", "税务风险", "税收优惠", "税前扣除",
            "发票管理", "增值税专用发票", "普通发票", "数电发票",
            "应交税费", "税金及附加", "递延所得税", "文化事业建设费",
            "代扣代缴", "源泉扣缴", "税务登记", "小规模纳税人", "一般纳税人",
        ],
        "medium": [
            "税率", "税额", "税负", "纳税", "缴税", "退税", "征税",
            "免税", "扣税", "抵税", "完税", "涉税", "税务",
            "发票", "抵扣", "进项", "销项", "认证", "红冲", "作废",
            "申报", "预缴", "汇算", "清算", "留抵",
            "个税", "所得税", "流转税", "财产税",
            "纳税调整", "加计扣除", "加速折旧", "不征税收入",
            "查账征收", "核定征收", "税号",
            "进项转出", "不得抵扣", "视同销售",
            "减免税", "即征即退", "先征后退", "出口退税",
        ],
        "weak": [
            "财务报表", "利润表", "资产负债表", "现金流量表",
            "主营业务收入", "营业收入", "营业成本", "利润总额",
            "社保", "公积金", "工资薪金", "劳务报酬", "稿酬",
            "财产租赁", "财产转让", "股息红利",
            "稽查", "罚款", "滞纳金",
            "转让定价", "关联交易", "同期资料",
            "毛利率", "成本结构", "费用率", "应收账款", "应付账款",
            "其他应收款", "存货", "固定资产", "无形资产",
        ]
    }

    found_strong = [kw for kw in tax_keywords["strong"] if kw in text]
    found_medium = [kw for kw in tax_keywords["medium"] if kw in text]
    found_weak = [kw for kw in tax_keywords["weak"] if kw in text]

    total_found = len(found_strong) + len(found_medium) + len(found_weak)

    # 评分
    if total_found == 0:
        score = 0
    else:
        strong_score = len(found_strong) * 15
        medium_score = len(found_medium) * 5
        weak_score = len(found_weak) * 2
        raw_score = strong_score + medium_score + weak_score
        # 密度修正：200字出现1个关键词为基准
        text_len = len(text)
        density = total_found / max(text_len / 200, 1)
        density_factor = min(density, 2.0)
        score = min(int(raw_score * density_factor), 100)
        # 有强信号保底
        if len(found_strong) >= 1 and score < 20:
            score = max(score, 25)

    is_tax_related = score >= 20

    all_keywords = found_strong + found_medium + found_weak
    unique_keywords = list(dict.fromkeys(all_keywords))

    return {
        "is_tax_related": is_tax_related,
        "score": score,
        "strong_count": len(found_strong),
        "medium_count": len(found_medium),
        "weak_count": len(found_weak),
        "total_keywords": total_found,
        "keywords_found": unique_keywords[:40],
        "text_length": len(text),
    }


@app.post("/api/tax-risk-rules/parse-report")
async def tax_risk_rules_parse_report(request: Request):
    """接收税务报告/文章内容，智能提取风险规则"""
    try:
        body = await request.json()
        report_text = body.get("text", "")
        if not report_text:
            return {"ok": False, "error": "报告内容不能为空"}
    except Exception:
        return {"ok": False, "error": "无效的请求数据"}
    return _parse_tax_report_text(report_text)


@app.post("/api/tax-risk-rules/upload-report")
async def tax_risk_rules_upload_report(request: Request):
    """接收上传的报告文件（PDF/Word/TXT），提取文本并解析为规则"""
    import io as _io
    import os as _os

    try:
        form = await request.form()
        file = form.get("file")
        if not file:
            return {"ok": False, "error": "未找到上传文件"}
    except Exception:
        return {"ok": False, "error": "无效的文件上传请求"}

    filename = (file.filename or "").lower()
    content_bytes = await file.read()

    if not content_bytes:
        return {"ok": False, "error": "文件内容为空"}

    extracted_text = ""
    source_desc = ""

    if filename.endswith('.txt'):
        try:
            extracted_text = content_bytes.decode('utf-8')
        except UnicodeDecodeError:
            try:
                extracted_text = content_bytes.decode('gbk')
            except Exception:
                return {"ok": False, "error": "无法解码TXT文件编码"}
        source_desc = f"TXT文件 ({filename})"

    elif filename.endswith('.docx'):
        try:
            from docx import Document as _Document
            doc = _Document(_io.BytesIO(content_bytes))
            paragraphs_text = [p.text for p in doc.paragraphs]
            extracted_text = '\n\n'.join(paragraphs_text)
            source_desc = f"Word文档 ({filename})"
        except Exception as e:
            return {"ok": False, "error": f"Word文档解析失败: {str(e)}"}

    elif filename.endswith('.pdf'):
        try:
            from PyPDF2 import PdfReader as _PdfReader
            reader = _PdfReader(_io.BytesIO(content_bytes))
            pages_text = []
            for page in reader.pages:
                t = page.extract_text()
                if t:
                    pages_text.append(t.strip())
            extracted_text = '\n\n'.join(pages_text)
            source_desc = f"PDF文件 ({filename}, {len(reader.pages)}页)"
        except Exception as e:
            return {"ok": False, "error": f"PDF解析失败: {str(e)}"}

    else:
        return {"ok": False, "error": f"不支持的文件格式 (.{filename.split('.')[-1]})，仅支持 PDF/Word/TXT"}

    if not extracted_text.strip():
        return {"ok": False, "error": "未能从文件中提取到文本内容"}

    result = _parse_tax_report_text(extracted_text.strip())
    result["source_file"] = source_desc
    # 附加涉税相关性检测
    relevance = _check_tax_relevance(extracted_text.strip())
    result["relevance"] = relevance
    return result

if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True)


@app.post("/api/tax-risk-rules/check-relevance")
async def tax_risk_rules_check_relevance(request: Request):
    """检测文本内容是否为涉税相关（输入文字/上传报告前预检）"""
    try:
        body = await request.json()
        text = body.get("text", "")
        if not text or not text.strip():
            return {"ok": False, "error": "文本内容不能为空"}
    except Exception:
        return {"ok": False, "error": "无效的请求数据"}
    result = _check_tax_relevance(text)
    result["ok"] = True
    return result

# ========== 开发模式：强制无缓存 ==========
@app.middleware('http')
async def add_cache_headers(request, call_next):
    '''强制浏览器不用本地缓存，每次都重新验证资源'''
    response = await call_next(request)
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response

